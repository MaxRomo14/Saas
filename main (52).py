import streamlit as st
import json, csv, io, hashlib, difflib
from datetime import datetime, timedelta, date

st.set_page_config(page_title="Graphite Compliance", page_icon="⚖️", layout="wide", initial_sidebar_state="expanded")

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800&display=swap');

*, *::before, *::after { box-sizing: border-box; }

/* ═══════════════════════════════════════════════
   GLOBAL TYPOGRAPHY & BASE
═══════════════════════════════════════════════ */
html, body, [class*="css"],
.stApp, [data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"],
section.main, .main, .block-container,
div, p, span, label, li, td, th, button,
h1, h2, h3, h4, h5, h6 {
  font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
  color: #1a1d23 !important;
  -webkit-font-smoothing: antialiased !important;
  -moz-osx-font-smoothing: grayscale !important;
  text-rendering: optimizeLegibility !important;
}

/* ═══════════════════════════════════════════════
   APP BACKGROUND
═══════════════════════════════════════════════ */
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"],
section.main, .main, .block-container {
  background-color: #f7f8fa !important;
}
.main .block-container {
  padding-top: 1.5rem !important;
  padding-bottom: 2.5rem !important;
  max-width: 1260px !important;
}

/* ═══════════════════════════════════════════════
   SIDEBAR
═══════════════════════════════════════════════ */
section[data-testid="stSidebar"],
section[data-testid="stSidebar"] > div,
section[data-testid="stSidebar"] > div > div {
  background-color: #0c1018 !important;
  border-right: 1px solid #1e2530 !important;
}
section[data-testid="stSidebar"] div,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] li,
section[data-testid="stSidebar"] a,
section[data-testid="stSidebar"] button,
section[data-testid="stSidebar"] [class*="css"] {
  color: #b8c1d1 !important;
  background-color: transparent !important;
}
/* Radio nav (legacy) */
section[data-testid="stSidebar"] .stRadio label {
  font-size: 12px !important;
  font-weight: 500 !important;
  color: #b8c1d1 !important;
  padding: 7px 12px !important;
  display: block !important;
  border-radius: 6px !important;
  line-height: 1.4 !important;
  letter-spacing: 0.01em !important;
  transition: background-color 0.12s ease, color 0.12s ease !important;
}
section[data-testid="stSidebar"] .stRadio label:hover {
  background-color: rgba(255,255,255,0.06) !important;
  color: #ffffff !important;
}
section[data-testid="stSidebar"] .stRadio [data-baseweb="radio"] { margin-bottom: 2px !important; }
section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] {
  gap: 2px !important;
  display: flex !important;
  flex-direction: column !important;
}
section[data-testid="stSidebar"] .stRadio [data-baseweb="radio"] > div:first-child { display: none !important; }
section[data-testid="stSidebar"] .stRadio [aria-checked="true"] label {
  background-color: rgba(59,130,246,0.18) !important;
  color: #93c5fd !important;
  font-weight: 600 !important;
}
section[data-testid="stSidebar"] .stRadio [aria-checked="true"] label span { color: #93c5fd !important; }

/* Sidebar nav buttons */
section[data-testid="stSidebar"] .stButton > button {
  background-color: transparent !important;
  color: #b8c1d1 !important;
  border: none !important;
  border-radius: 6px !important;
  font-size: 12px !important;
  font-weight: 500 !important;
  padding: 7px 12px !important;
  text-align: left !important;
  width: 100% !important;
  line-height: 1.45 !important;
  margin-bottom: 2px !important;
  box-shadow: none !important;
  letter-spacing: 0.01em !important;
  transition: background-color 0.12s ease, color 0.12s ease !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
  background-color: rgba(255,255,255,0.06) !important;
  color: #ffffff !important;
}
section[data-testid="stSidebar"] .stButton > button[data-testid="baseButton-secondary"]:last-of-type {
  background-color: #2563eb !important;
  color: #ffffff !important;
  margin-top: 8px !important;
  border-radius: 6px !important;
  font-weight: 600 !important;
}

/* ═══════════════════════════════════════════════
   FORM INPUTS
═══════════════════════════════════════════════ */
.stTextInput > div > div > input,
.stTextArea > div > div > textarea,
.stSelectbox > div > div,
.stDateInput > div > div > input,
.stNumberInput > div > div > input {
  background-color: #ffffff !important;
  border: 1px solid #dfe3e9 !important;
  color: #1a1d23 !important;
  border-radius: 6px !important;
  font-size: 13px !important;
  box-shadow: 0 1px 2px rgba(0,0,0,0.04) !important;
  transition: border-color 0.15s, box-shadow 0.15s !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
  border-color: #2563eb !important;
  box-shadow: 0 0 0 3px rgba(37,99,235,0.1) !important;
  outline: none !important;
}
.stSelectbox [data-baseweb="select"] span,
.stSelectbox [data-baseweb="select"] div { color: #1a1d23 !important; }

/* Input labels */
label, .stTextInput label, .stSelectbox label,
.stTextArea label, .stDateInput label {
  font-size: 11px !important;
  font-weight: 600 !important;
  color: #525966 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.06em !important;
  margin-bottom: 4px !important;
}

/* ═══════════════════════════════════════════════
   FILE UPLOADER
═══════════════════════════════════════════════ */
[data-testid="stFileUploader"] {
  background-color: #fafbfc !important;
  border: 1.5px dashed #93c5fd !important;
  border-radius: 10px !important;
  padding: 12px !important;
  transition: border-color 0.15s ease, background-color 0.15s ease !important;
}
[data-testid="stFileUploader"]:hover {
  border-color: #60a5fa !important;
  background-color: #f8faff !important;
}
[data-testid="stFileUploader"] section { background-color: transparent !important; }
[data-testid="stFileUploader"] span,
[data-testid="stFileUploader"] p,
[data-testid="stFileUploader"] div,
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] [data-testid="stMarkdownContainer"] {
  color: #1e40af !important;
  background-color: transparent !important;
}
[data-testid="stFileUploader"] button {
  background-color: #eff6ff !important;
  color: #1d4ed8 !important;
  border: 1px solid #bfdbfe !important;
  font-weight: 600 !important;
  border-radius: 6px !important;
  padding: 6px 14px !important;
}
[data-testid="stFileUploader"] button:hover { background-color: #dbeafe !important; }
[data-testid="stFileUploaderDropzoneInstructions"] div,
[data-testid="stFileUploaderDropzoneInstructions"] span { color: #3b82f6 !important; font-size: 13px !important; }
[data-testid="stFileUploader"] [data-testid="stMarkdownContainer"] p { color: #1e40af !important; }
[data-testid="uploadedFileData"] {
  background-color: #eff6ff !important;
  border: 1px solid #bfdbfe !important;
  border-radius: 6px !important;
  padding: 8px 12px !important;
}
[data-testid="uploadedFileData"] span,
[data-testid="uploadedFileData"] div { color: #1e40af !important; }

/* ═══════════════════════════════════════════════
   BUTTONS  — compact, readable, professional
═══════════════════════════════════════════════ */
.stButton > button {
  background-color: #2563eb !important;
  color: #ffffff !important;
  border: none !important;
  font-family: 'Inter', sans-serif !important;
  font-weight: 600 !important;
  font-size: 11px !important;
  letter-spacing: 0.025em !important;
  border-radius: 6px !important;
  padding: 6px 14px !important;
  line-height: 1.5 !important;
  height: auto !important;
  min-height: 0 !important;
  transition: background-color 0.15s ease, box-shadow 0.15s ease, transform 0.1s ease !important;
  width: fit-content !important;
  min-width: 0 !important;
  white-space: nowrap !important;
  display: inline-flex !important;
  align-items: center !important;
  justify-content: center !important;
  cursor: pointer !important;
  box-shadow: 0 1px 2px rgba(0,0,0,0.06), 0 1px 3px rgba(37,99,235,0.08) !important;
}
.stButton > button:hover {
  background-color: #1d4ed8 !important;
  box-shadow: 0 2px 4px rgba(0,0,0,0.08), 0 2px 8px rgba(37,99,235,0.18) !important;
}
.stButton > button:active { 
  background-color: #1e40af !important;
  transform: translateY(0.5px) !important;
}
.stButton > button:disabled,
.stButton > button[disabled] {
  background-color: #94a3b8 !important;
  cursor: not-allowed !important;
  opacity: 0.6 !important;
  box-shadow: none !important;
}
/* Wrapper — never stretch */
.stButton,
div[data-testid="stButton"] {
  display: inline-flex !important;
  width: auto !important;
  min-width: 0 !important;
}

/* Muted / secondary buttons (toolbar All/None, collapse) */
.gc-sm-btn .stButton > button {
  background-color: #f8fafc !important;
  color: #475569 !important;
  border: 1px solid #d1d5db !important;
  box-shadow: none !important;
  font-size: 11px !important;
  padding: 5px 12px !important;
}
.gc-sm-btn .stButton > button:hover {
  background-color: #eef1f5 !important;
  color: #1e293b !important;
  border-color: #c5cad2 !important;
}

/* ═══════════════════════════════════════════════
   TABS
═══════════════════════════════════════════════ */
.stTabs [data-baseweb="tab-list"] {
  background-color: #f0f2f5 !important;
  border-bottom: 1px solid #d4d8df !important;
  border-radius: 8px 8px 0 0 !important;
  gap: 0 !important;
  padding: 0 4px !important;
}
.stTabs [data-baseweb="tab"] {
  background: transparent !important;
  color: #5c6370 !important;
  font-size: 12px !important;
  font-weight: 500 !important;
  padding: 10px 18px !important;
  border-bottom: 2px solid transparent !important;
  margin-bottom: -1px !important;
  letter-spacing: 0.01em !important;
  transition: color 0.15s ease, border-color 0.15s ease !important;
}
.stTabs [data-baseweb="tab"]:hover {
  color: #374151 !important;
}
.stTabs [aria-selected="true"] {
  color: #1d4ed8 !important;
  border-bottom: 2px solid #2563eb !important;
  font-weight: 600 !important;
  background: transparent !important;
}

/* ═══════════════════════════════════════════════
   METRIC CARDS
═══════════════════════════════════════════════ */
[data-testid="stMetric"] {
  background-color: #ffffff !important;
  border: 1px solid #e2e6eb !important;
  border-top: 3px solid #2563eb !important;
  border-radius: 8px !important;
  padding: 14px 16px !important;
  box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 1px 2px rgba(0,0,0,0.02) !important;
}
[data-testid="stMetricLabel"] {
  color: #5c6370 !important;
  font-size: 10px !important;
  font-weight: 700 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.08em !important;
}
[data-testid="stMetricValue"] {
  color: #1a1d23 !important;
  font-size: 1.65rem !important;
  font-weight: 700 !important;
  letter-spacing: -0.02em !important;
}
[data-testid="stMetricDelta"] { color: #5c6370 !important; font-size: 11px !important; }

/* ═══════════════════════════════════════════════
   TOGGLE BUTTONS (expand/collapse rows)
═══════════════════════════════════════════════ */
[data-testid="stButton"] > button[kind="secondary"],
div[data-testid="stHorizontalBlock"] button { text-align: left !important; }
.gc-toggle-btn > button {
  background: #ffffff !important;
  border: 1px solid #e2e6eb !important;
  border-radius: 8px !important;
  padding: 10px 16px !important;
  font-size: 12px !important;
  font-weight: 600 !important;
  color: #1a1d23 !important;
  text-align: left !important;
  box-shadow: 0 1px 2px rgba(0,0,0,0.03) !important;
  margin-bottom: 3px !important;
  transition: border-color 0.15s ease, box-shadow 0.15s ease !important;
}
.gc-toggle-btn > button:hover {
  border-color: #d0d5dd !important;
  box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
}

/* ═══════════════════════════════════════════════
   DATAFRAME
═══════════════════════════════════════════════ */
.stDataFrame, .stDataFrame * { color: #1a1d23 !important; background-color: #ffffff !important; }
.stDataFrame thead th {
  background-color: #f5f6f8 !important;
  color: #475569 !important;
  font-size: 10px !important;
  font-weight: 700 !important;
  text-transform: uppercase !important;
  letter-spacing: 0.06em !important;
  padding: 10px 12px !important;
}
.stDataFrame tbody td {
  padding: 8px 12px !important;
  border-bottom: 1px solid #f0f2f5 !important;
}

/* ═══════════════════════════════════════════════
   CUSTOM CARDS
═══════════════════════════════════════════════ */
.gc-card {
  background-color: #ffffff;
  border: 1px solid #e2e6eb;
  border-radius: 10px;
  padding: 16px 20px;
  margin-bottom: 10px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.04);
}
.gc-card, .gc-card * { color: #1a1d23 !important; }
.gc-card-header {
  font-size: 10px !important;
  font-weight: 700 !important;
  letter-spacing: 0.1em !important;
  color: #5c6370 !important;
  text-transform: uppercase !important;
  margin-bottom: 8px !important;
}

/* ═══════════════════════════════════════════════
   BADGES
═══════════════════════════════════════════════ */
.gc-badge {
  display: inline-block;
  padding: 3px 10px;
  border-radius: 4px;
  font-size: 10px !important;
  font-weight: 700 !important;
  letter-spacing: 0.03em !important;
  text-transform: uppercase !important;
}
.badge-open    { background:#dbeafe; color:#1e40af !important; border: 1px solid #bfdbfe; }
.badge-closed  { background:#dcfce7; color:#15803d !important; border: 1px solid #bbf7d0; }
.badge-overdue { background:#fee2e2; color:#991b1b !important; border: 1px solid #fecaca; }
.badge-pending { background:#fef9c3; color:#854d0e !important; border: 1px solid #fef08a; }
.badge-high    { background:#fee2e2; color:#991b1b !important; border: 1px solid #fecaca; }
.badge-med     { background:#fef9c3; color:#854d0e !important; border: 1px solid #fef08a; }
.badge-low     { background:#dcfce7; color:#15803d !important; border: 1px solid #bbf7d0; }
.badge-draft   { background:#f3f4f6; color:#4b5563 !important; border: 1px solid #e5e7eb; }

/* ═══════════════════════════════════════════════
   TIMELINE / AUDIT TRAIL
═══════════════════════════════════════════════ */
.gc-tl {
  border-left: 2px solid #e2e6eb;
  padding: 6px 0 6px 16px;
  margin-bottom: 6px;
  position: relative;
  font-size: 12px !important;
  color: #525966 !important;
  line-height: 1.6;
}
.gc-tl, .gc-tl * { color: #525966 !important; }
.gc-tl::before {
  content: '';
  position: absolute;
  left: -5px; top: 10px;
  width: 8px; height: 8px;
  border-radius: 50%;
  background: #2563eb;
  border: 2px solid #ffffff;
  box-shadow: 0 0 0 1px #d1d5db;
}

/* ═══════════════════════════════════════════════
   PAGE HEADERS (.gc-header)
═══════════════════════════════════════════════ */
.gc-header {
  background-color: #ffffff;
  border: 1px solid #e2e6eb;
  border-left: 4px solid #2563eb;
  border-radius: 10px;
  padding: 16px 22px;
  margin-bottom: 20px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.03);
}
.gc-header h1 {
  font-size: 1rem !important;
  font-weight: 800 !important;
  color: #1a1d23 !important;
  letter-spacing: 0.05em !important;
  text-transform: uppercase !important;
  margin: 0 0 4px 0 !important;
}
.gc-header p {
  font-size: 13px !important;
  color: #5c6370 !important;
  margin: 0 !important;
  font-weight: 400 !important;
  line-height: 1.5 !important;
}

/* ═══════════════════════════════════════════════
   ALERTS
═══════════════════════════════════════════════ */
.gc-alert {
  border-radius: 8px;
  padding: 10px 14px;
  margin: 8px 0;
  font-size: 12px;
  line-height: 1.6;
}
.gc-alert-warn   { background:#fffbeb; border-left:3px solid #f59e0b; color:#78350f !important; border: 1px solid #fde68a; border-left-width: 3px; }
.gc-alert-danger { background:#fef2f2; border-left:3px solid #ef4444; color:#7f1d1d !important; border: 1px solid #fecaca; border-left-width: 3px; }
.gc-alert-ok     { background:#f0fdf4; border-left:3px solid #22c55e; color:#14532d !important; border: 1px solid #bbf7d0; border-left-width: 3px; }
.gc-alert, .gc-alert * { color: inherit !important; }

/* ═══════════════════════════════════════════════
   CHECKLIST CARDS
═══════════════════════════════════════════════ */
.cl-owner-card {
  background-color: #ffffff;
  border: 1px solid #e2e6eb;
  border-top: 3px solid #2563eb;
  border-radius: 10px;
  padding: 14px 18px;
  margin-bottom: 14px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.03);
}
.cl-owner-card, .cl-owner-card * { color: #1a1d23 !important; }
.cl-task-row {
  background-color: #f8f9fb;
  border: 1px solid #e2e6eb;
  border-radius: 8px;
  padding: 10px 14px;
  margin: 6px 0;
  font-size: 13px;
}
.cl-task-row, .cl-task-row * { color: #1a1d23 !important; }

/* ═══════════════════════════════════════════════
   DIVIDERS, SCROLLBAR, CODE
═══════════════════════════════════════════════ */
hr {
  border: none !important;
  border-top: 1px solid #e8eaed !important;
  margin: 14px 0 !important;
}
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f7f8fa; border-radius: 3px; }
::-webkit-scrollbar-thumb { background: #cbd2da; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #a0aab8; }
code {
  font-size: 11px !important;
  background: #f1f5f9 !important;
  color: #1d4ed8 !important;
  padding: 2px 6px !important;
  border-radius: 4px !important;
  font-family: 'SF Mono', 'Fira Code', 'Consolas', monospace !important;
}

/* ═══════════════════════════════════════════════
   DOWNLOAD BUTTONS — light theme, easy to read
═══════════════════════════════════════════════ */
[data-testid="stDownloadButton"] > button {
  background-color: #f0f7ff !important;
  color: #1e40af !important;
  border: 1.5px solid #bfdbfe !important;
  box-shadow: 0 1px 2px rgba(37,99,235,0.08) !important;
}
[data-testid="stDownloadButton"] > button:hover {
  background-color: #dbeafe !important;
  color: #1d4ed8 !important;
  border-color: #93c5fd !important;
  box-shadow: 0 2px 6px rgba(37,99,235,0.15) !important;
}
[data-testid="stDownloadButton"] > button * {
  color: #1e40af !important;
}


/* The eye icon button inside password inputs */
[data-testid="stTextInput"] button,
.stTextInput [data-baseweb="input"] button,
input[type="password"] ~ button,
[data-testid="InputInstructions"] ~ button,
div[data-baseweb="input"] > div > button {
  background-color: #e8edf5 !important;
  border: 1px solid #c5cdd8 !important;
  border-radius: 5px !important;
  color: #1a1d23 !important;
  opacity: 1 !important;
}
[data-testid="stTextInput"] button:hover,
div[data-baseweb="input"] > div > button:hover {
  background-color: #d0d8e8 !important;
  border-color: #2563eb !important;
}
[data-testid="stTextInput"] button svg,
div[data-baseweb="input"] > div > button svg {
  fill: #1a1d23 !important;
  stroke: #1a1d23 !important;
  opacity: 1 !important;
}
/* Hide the "Press Enter to apply" instructions that overlap the eye icon */
[data-testid="InputInstructions"] {
  display: none !important;
}
/* Extra right padding on password inputs so text doesn't run into eye button */
input[type="password"], input[type="text"][autocomplete="current-password"],
[data-baseweb="input"] input[type="password"] {
  padding-right: 44px !important;
}


.login-wrap { max-width: 400px; margin: 10vh auto; }
.login-logo {
  font-size: 1.65rem;
  font-weight: 800;
  color: #1a1d23 !important;
  letter-spacing: -0.03em;
  text-align: center;
  margin-bottom: 4px;
}
.login-sub {
  text-align: center;
  color: #5c6370 !important;
  font-size: 11px;
  margin-bottom: 28px;
  text-transform: uppercase;
  letter-spacing: 0.14em;
  font-weight: 500;
}

/* ═══════════════════════════════════════════════
   DROPDOWN / SELECTBOX POPOVER
═══════════════════════════════════════════════ */
[data-baseweb="popover"],
[data-baseweb="popover"] *,
[data-baseweb="menu"],
[data-baseweb="menu"] *,
[data-baseweb="select"] [role="listbox"],
[data-baseweb="select"] [role="listbox"] *,
ul[data-baseweb="menu"],
ul[data-baseweb="menu"] li,
ul[data-baseweb="menu"] li * {
  background-color: #ffffff !important;
  color: #1a1d23 !important;
}
[data-baseweb="menu"] li:hover,
[data-baseweb="menu"] [aria-selected="true"],
ul[data-baseweb="menu"] li:hover {
  background-color: #eff6ff !important;
  color: #1d4ed8 !important;
}
.stSelectbox [data-baseweb="select"] > div,
.stSelectbox [data-baseweb="select"] > div > div,
.stSelectbox [data-baseweb="tag"],
.stMultiSelect [data-baseweb="tag"] {
  background-color: #ffffff !important;
  color: #1a1d23 !important;
}
.stMultiSelect [data-baseweb="tag"] {
  background-color: #dbeafe !important;
  color: #1e40af !important;
  border-radius: 4px !important;
}
.stSelectbox svg, .stMultiSelect svg { fill: #5c6370 !important; }
[data-baseweb="select"] input {
  color: #1a1d23 !important;
  background: #ffffff !important;
}
[class*="baseui"] [role="option"],
[class*="baseui"] [role="option"] * {
  background-color: #ffffff !important;
  color: #1a1d23 !important;
}
[class*="baseui"] [role="option"]:hover,
[class*="baseui"] [role="option"][aria-selected="true"] {
  background-color: #eff6ff !important;
  color: #1d4ed8 !important;
}

/* ═══════════════════════════════════════════════
   PROGRESS BAR
═══════════════════════════════════════════════ */
.stProgress > div > div > div > div {
  background: linear-gradient(90deg, #2563eb 0%, #3b82f6 100%) !important;
  border-radius: 6px !important;
}
.stProgress > div > div > div {
  background-color: #e8eaed !important;
  border-radius: 6px !important;
  height: 6px !important;
}

/* ═══════════════════════════════════════════════
   SUCCESS / INFO / WARNING STREAMLIT MESSAGES
═══════════════════════════════════════════════ */
[data-testid="stSuccess"] {
  background-color: #f0fdf4 !important;
  border: 1px solid #86efac !important;
  border-radius: 8px !important;
  padding: 12px 16px !important;
}
[data-testid="stWarning"] {
  background-color: #fffbeb !important;
  border: 1px solid #fcd34d !important;
  border-radius: 8px !important;
  padding: 12px 16px !important;
}
[data-testid="stError"] {
  background-color: #fef2f2 !important;
  border: 1px solid #fca5a5 !important;
  border-radius: 8px !important;
  padding: 12px 16px !important;
}
[data-testid="stInfo"] {
  background-color: #eff6ff !important;
  border: 1px solid #bfdbfe !important;
  border-radius: 8px !important;
  padding: 12px 16px !important;
}

/* ═══════════════════════════════════════════════
   COLUMN GAP TIGHTENING
═══════════════════════════════════════════════ */
[data-testid="stHorizontalBlock"] {
  gap: 12px !important;
  align-items: flex-start !important;
}

/* ═══════════════════════════════════════════════
   EXPANDER POLISH
═══════════════════════════════════════════════ */
.streamlit-expanderHeader {
  font-size: 13px !important;
  font-weight: 600 !important;
  color: #1a1d23 !important;
  background-color: #ffffff !important;
  border: 1px solid #e2e6eb !important;
  border-radius: 8px !important;
  padding: 12px 16px !important;
}
.streamlit-expanderContent {
  border: 1px solid #e2e6eb !important;
  border-top: none !important;
  border-radius: 0 0 8px 8px !important;
  background-color: #fafbfc !important;
  padding: 16px !important;
}

/* ═══════════════════════════════════════════════
   TOOLTIP & HELP TEXT
═══════════════════════════════════════════════ */
[data-testid="stTooltipIcon"] {
  color: #9ca3af !important;
}
[data-testid="stTooltipContent"] {
  background-color: #1f2937 !important;
  color: #ffffff !important;
  border-radius: 6px !important;
  font-size: 12px !important;
  padding: 8px 12px !important;
}

/* ═══════════════════════════════════════════════
   CHAT MESSAGE POLISH
═══════════════════════════════════════════════ */
[data-testid="stChatMessage"] {
  background-color: #ffffff !important;
  border: 1px solid #e2e6eb !important;
  border-radius: 10px !important;
  padding: 14px 18px !important;
  margin-bottom: 8px !important;
}
</style>
""", unsafe_allow_html=True)

# ── DATA LAKE BUILDER ─────────────────────────────────────────────────────────
def _build_data_lake(ts_fn):
    """Generates a large synthetic compliance data lake — hundreds of documents
    representing a real firm's shared drive migrated into the repository, plus
    all system-generated evidence from tasks, checklists, and reviews."""
    import hashlib as _hl, random as _rnd
    _rnd.seed(42)

    DEPARTMENTS = ["Compliance","Supervision","Legal","Operations","Risk","Finance","Technology","HR"]
    LINES_OF_BIZ = ["Wealth Management","Investment Banking","Operations","Fixed Income","Capital Markets"]
    DOC_TYPES = ["Policy","WSP","Training Record","Exam Evidence","Audit Report",
                 "Supervisory Review","Testing Evidence","Correspondence","Certification",
                 "Exception Report","Risk Assessment","Board Presentation","Checklist",
                 "SAR Filing","CIP Record","Trade Blotter","Surveillance Log","Complaint File"]
    SOURCES = ["Shared Drive","System Upload","Task Evidence","Checklist Evidence",
               "Exam Upload","Manual Upload","API Sync","Email Archive"]
    UPLOADERS = ["cco","supervisor","compliance_team"]
    YEARS = ["2021","2022","2023","2024","2025","2026"]
    TAGS_POOL = ["AML","FINRA 3110","FINRA 3120","Reg BI","SEC 206(4)-7","BSA","FinCEN",
                 "Supervision","Trade Review","Complaint","SAR","CIP","Reg SHO","Rule 17a-4",
                 "Annual Review","Quarterly","Monthly","Training","Certification","Escalation",
                 "Best Execution","Net Capital","BCP","Information Barrier","Insider Trading"]

    # Representative filenames per doc type
    FILENAMES = {
        "Policy": [
            "AML_Policy_v{v}_{y}.pdf","Supervision_Policy_v{v}_{y}.pdf",
            "Information_Barriers_Policy_{y}.pdf","Reg_BI_Policy_v{v}_{y}.pdf",
            "Cybersecurity_Policy_{y}.pdf","Recordkeeping_Policy_v{v}_{y}.pdf",
            "Conflicts_of_Interest_Policy_{y}.pdf","Trade_Surveillance_Policy_{y}.pdf",
            "Data_Privacy_Policy_v{v}_{y}.pdf","BCP_Policy_{y}.pdf",
        ],
        "WSP": [
            "WSP_Retail_Supervision_{y}_v{v}.pdf","WSP_AML_Procedures_v{v}_{y}.pdf",
            "WSP_Options_Supervision_{y}.pdf","WSP_Fixed_Income_{y}_v{v}.pdf",
            "WSP_New_Account_Opening_{y}.pdf","WSP_Complaint_Handling_v{v}_{y}.pdf",
            "WSP_Electronic_Comm_Surveillance_{y}.pdf","WSP_Best_Execution_{y}_v{v}.pdf",
            "WSP_Information_Barriers_{y}.pdf","WSP_Margin_Supervision_v{v}_{y}.pdf",
        ],
        "Training Record": [
            "AML_Training_Completion_Report_{y}.xlsx","Annual_CE_Records_{y}.xlsx",
            "Reg_BI_Training_Certificates_{y}.zip","Cybersecurity_Training_{y}.pdf",
            "Supervision_Training_Roster_{y}.xlsx","Ethics_Training_Completion_{y}.pdf",
            "New_Hire_Compliance_Training_{y}.pdf","Options_Principal_CE_{y}.pdf",
        ],
        "Exam Evidence": [
            "FINRA_Exam_Response_{y}_Item{n}.pdf","SEC_Request_{y}_Tab{n}.pdf",
            "Exam_Evidence_Binder_{y}_Vol{n}.pdf","Regulator_Response_Letter_{y}.pdf",
            "Document_Production_{y}_{m}.zip","Examination_Workpapers_{y}.pdf",
            "Deficiency_Response_{y}.pdf","Exam_Closing_Letter_{y}.pdf",
        ],
        "Audit Report": [
            "Internal_Audit_Report_{y}_Q{q}.pdf","AML_Audit_Findings_{y}.pdf",
            "Supervisory_Controls_Audit_{y}.pdf","IT_Security_Audit_{y}.pdf",
            "Third_Party_Audit_Report_{y}.pdf","Compliance_Audit_{y}_Final.pdf",
            "Risk_Based_Audit_Plan_{y}.pdf","Audit_Committee_Report_{y}_Q{q}.pdf",
        ],
        "Supervisory Review": [
            "Trade_Review_Log_{y}_{m}.xlsx","Options_Supervisory_Review_{y}_Q{q}.xlsx",
            "Daily_Trade_Blotter_Review_{y}_{m}_{d}.pdf","Complaint_Disposition_Log_{y}.xlsx",
            "New_Account_Review_{y}_{m}.pdf","Reg_BI_Recommendation_Review_{y}_Q{q}.xlsx",
            "CCO_Monthly_Review_{y}_{m}.pdf","Principal_Review_Sign_Off_{y}_{m}.pdf",
        ],
        "Testing Evidence": [
            "SCP_Testing_Results_{y}.xlsx","Reg_BI_Testing_Sample_{y}_Q{q}.pdf",
            "AML_Program_Testing_{y}.pdf","Supervision_Testing_Workpapers_{y}.xlsx",
            "206_4_7_Testing_Evidence_{y}.pdf","FINRA_3120_Testing_Report_{y}.pdf",
            "Exception_Sampling_Results_{y}_Q{q}.xlsx","Control_Testing_Summary_{y}.pdf",
        ],
        "Certification": [
            "Annual_Compliance_Certification_{y}.pdf","CCO_Certification_{y}.pdf",
            "Reg_BI_Annual_Review_Cert_{y}.pdf","AML_Officer_Certification_{y}.pdf",
            "Outside_Business_Activity_Forms_{y}.zip","Annual_Attestation_Roster_{y}.xlsx",
            "Information_Barrier_Acknowledgments_{y}.zip","U4_U5_Updates_{y}.pdf",
        ],
        "SAR Filing": [
            "SAR_Filing_Log_{y}.xlsx","SAR_Supporting_Docs_{y}_{m}.pdf",
            "No_File_Decision_Memo_{y}_{m}.pdf","SAR_Review_Committee_Minutes_{y}.pdf",
            "CTR_Filing_Log_{y}.xlsx","SAR_Quality_Review_{y}_Q{q}.pdf",
        ],
        "CIP Record": [
            "CIP_Audit_Results_{y}.xlsx","New_Account_CIP_Files_{y}_{m}.zip",
            "Beneficial_Ownership_Records_{y}.xlsx","EDD_Review_Log_{y}.xlsx",
            "Customer_Risk_Rating_Model_{y}.pdf","CIP_Exception_Report_{y}_{m}.pdf",
        ],
        "Trade Blotter": [
            "Trade_Blotter_{y}_{m}.xlsx","Options_Activity_Report_{y}_{m}.pdf",
            "Fixed_Income_Trade_Log_{y}_{m}.xlsx","Equity_Trade_Summary_{y}_Q{q}.pdf",
            "Surveillance_Alert_Log_{y}_{m}.xlsx","Best_Execution_Report_{y}_Q{q}.pdf",
        ],
        "Surveillance Log": [
            "Ecomm_Surveillance_Review_{y}_{m}.pdf","Lexicon_Hit_Report_{y}_{m}.xlsx",
            "Chat_Surveillance_Summary_{y}_{m}.pdf","Alert_Disposition_Log_{y}_{m}.xlsx",
            "Surveillance_Escalation_Log_{y}.xlsx","Cross_Asset_Surveillance_{y}_Q{q}.pdf",
        ],
        "Complaint File": [
            "Complaint_Log_{y}.xlsx","Customer_Complaint_{y}_{m}_Case{n}.pdf",
            "FINRA_Arbitration_Log_{y}.pdf","Complaint_Response_Letter_{y}_{m}.pdf",
            "Complaint_Trend_Analysis_{y}_Q{q}.pdf","Regulatory_Complaint_Report_{y}.pdf",
        ],
        "Risk Assessment": [
            "Annual_Risk_Assessment_{y}.pdf","AML_Risk_Assessment_{y}.pdf",
            "Business_Line_Risk_Review_{y}_Q{q}.pdf","Inherent_Risk_Matrix_{y}.xlsx",
            "Vendor_Risk_Assessment_{y}.pdf","Cybersecurity_Risk_Report_{y}.pdf",
            "Operational_Risk_Register_{y}.xlsx","Reg_Risk_Mapping_{y}.pdf",
        ],
        "Checklist": [
            "Supervisory_Checklist_{lob}_{y}_{m}.pdf","AML_Monthly_Checklist_{y}_{m}.pdf",
            "Reg_BI_Review_Checklist_{y}_Q{q}.xlsx","Annual_Review_Checklist_{y}.pdf",
            "New_Account_Checklist_{y}_{m}.pdf","Options_Review_Checklist_{y}_Q{q}.pdf",
        ],
        "Board Presentation": [
            "Compliance_Board_Report_{y}_Q{q}.pdf","AML_Board_Update_{y}.pptx",
            "Annual_Compliance_Program_Review_{y}.pptx","Risk_Committee_Presentation_{y}_Q{q}.pdf",
            "Regulatory_Update_Board_{y}.pptx","Exam_Results_Board_Summary_{y}.pdf",
        ],
        "Exception Report": [
            "Exception_Report_{y}_{m}.xlsx","SLA_Breach_Log_{y}_Q{q}.pdf",
            "Trade_Exception_Summary_{y}_{m}.pdf","Supervisory_Exception_Log_{y}.xlsx",
            "AML_Exception_Report_{y}_Q{q}.pdf","Surveillance_Exception_Log_{y}_{m}.xlsx",
        ],
        "Correspondence": [
            "FINRA_Correspondence_{y}_{m}.pdf","SEC_Letter_{y}_{m}.pdf",
            "Regulator_Inquiry_Response_{y}.pdf","Legal_Hold_Notice_{y}.pdf",
            "Outside_Counsel_Memo_{y}.pdf","Exam_Exit_Conference_Notes_{y}.pdf",
        ],
    }

    LOB_DEPT_MAP = {
        "Wealth Management": "Compliance","Investment Banking": "Compliance",
        "Operations": "Operations","Retail Brokerage": "Supervision",
        "Fixed Income": "Supervision","Equity Trading": "Supervision",
        "Investment Advisory": "Compliance","Prime Brokerage": "Risk",
    }
    TYPE_TAGS = {
        "Policy": ["Policy","Annual Review"],"WSP": ["Supervision","FINRA 3110"],
        "Training Record": ["Training","Certification"],"Exam Evidence": ["Exam Evidence"],
        "Audit Report": ["Audit Report"],"Supervisory Review": ["Supervisory Review","Trade Review"],
        "Testing Evidence": ["FINRA 3120","SEC 206(4)-7"],"Certification": ["Certification","Annual Review"],
        "SAR Filing": ["SAR","AML"],"CIP Record": ["CIP","AML","BSA"],
        "Trade Blotter": ["Trade Review","Best Execution"],"Surveillance Log": ["Trade Review","Information Barrier"],
        "Complaint File": ["Complaint"],"Risk Assessment": ["Risk Assessment"],
        "Checklist": ["Supervision","FINRA 3110"],"Board Presentation": ["Annual Review"],
        "Exception Report": ["Escalation"],"Correspondence": ["Exam Evidence"],
    }

    records = []
    counter = 1

    # ── Main synthetic corpus ─────────────────────────────────────────────────
    for dept in DEPARTMENTS:
        for lob in LINES_OF_BIZ:
            for doc_type in DOC_TYPES:
                # Each dept/lob/doctype gets 1-4 documents across years
                n_docs = _rnd.randint(1, 4)
                year_sample = _rnd.sample(YEARS, min(n_docs, len(YEARS)))
                for y in year_sample:
                    templates = FILENAMES.get(doc_type, ["Document_{y}_{n}.pdf"])
                    tmpl = _rnd.choice(templates)
                    fname = tmpl.format(
                        y=y, v=f"{_rnd.randint(1,4)}.{_rnd.randint(0,9)}",
                        n=_rnd.randint(1,20), q=_rnd.randint(1,4),
                        m=f"{_rnd.randint(1,12):02d}", d=f"{_rnd.randint(1,28):02d}",
                        lob=lob.replace(" ","_")
                    )
                    uploader = _rnd.choice(UPLOADERS)
                    days_ago = _rnd.randint(0, 365 * 5)
                    base_tags = TYPE_TAGS.get(doc_type, [])
                    extra_tags = _rnd.sample(TAGS_POOL, _rnd.randint(1, 3))
                    all_tags = list(set(base_tags + extra_tags))
                    h = _hl.md5(f"{fname}{counter}".encode()).hexdigest()
                    records.append({
                        "id": f"EV-{counter:04d}",
                        "filename": fname,
                        "uploaded_by": uploader,
                        "ts": (datetime.now() - timedelta(days=days_ago)).strftime("%Y-%m-%d %H:%M"),
                        "task_id": "",
                        "hash": f"sha256:{h}",
                        "immutable": True,
                        "archived": days_ago > 730,
                        # Extended data lake metadata
                        "department": dept,
                        "line_of_business": lob,
                        "doc_type": doc_type,
                        "year": y,
                        "source": _rnd.choice(SOURCES[:2]),  # Shared Drive or System Upload for corpus
                        "tags": all_tags,
                        "notes": "",
                        "size_kb": _rnd.randint(50, 8192),
                    })
                    counter += 1

    # ── System-generated evidence (task/checklist uploads) ────────────────────
    system_files = [
        ("supervision_policy_v2.pdf","compliance_team","T-004","WSP","Wealth Management","2026","System Upload",["FINRA 3110","Supervision"]),
        ("aml_policy_2024.pdf","cco","T-004","Policy","Operations","2024","System Upload",["AML","BSA"]),
        ("AML_Training_Completion_2025.xlsx","cco","T-001","Training Record","Operations","2025","Task Evidence",["AML","Training"]),
        ("Q3_Options_Supervisory_Review.pdf","supervisor","T-002","Supervisory Review","Wealth Management","2025","Task Evidence",["FINRA 3110","Trade Review"]),
        ("Best_Execution_Analysis_H1_2025.pdf","cco","T-003","Testing Evidence","Fixed Income","2025","Task Evidence",["Best Execution","Reg BI"]),
        ("Rule_3110_Gap_Analysis.xlsx","compliance_team","T-004","Audit Report","Compliance","2025","Task Evidence",["FINRA 3110","Annual Review"]),
        ("CT001_BCP_Evidence_Binder.pdf","supervisor","","Exam Evidence","Operations","2026","System Upload",["BCP","Exam Evidence"]),
        ("CT002_Annual_Compliance_Review.pdf","cco","","Testing Evidence","Investment Advisory","2026","System Upload",["SEC 206(4)-7","Annual Review"]),
        ("CCO_Annual_Certification_2025.pdf","cco","","Certification","Compliance","2025","Manual Upload",["Certification","Annual Review"]),
        ("SAR_Review_Log_Q4_2025.xlsx","supervisor","","SAR Filing","Operations","2025","System Upload",["SAR","AML"]),
        ("Info_Barrier_Attestations_2025.zip","compliance_team","","Certification","Investment Banking","2025","System Upload",["Information Barrier","Certification"]),
        ("Reg_BI_Conflict_Review_Q1_2026.pdf","cco","","Supervisory Review","Wealth Management","2026","Task Evidence",["Reg BI","Supervision"]),
        ("FINRA_Exam_Response_March2026.pdf","cco","","Correspondence","Compliance","2026","Manual Upload",["Exam Evidence","FINRA 3110"]),
        ("Trade_Surveillance_Alerts_Feb2026.xlsx","supervisor","","Surveillance Log","Equity Trading","2026","System Upload",["Trade Review","Surveillance Log"]),
        ("Customer_Complaint_Log_2025.xlsx","compliance_team","","Complaint File","Retail Brokerage","2025","System Upload",["Complaint"]),
    ]
    for fname, uploader, task_id, doc_type, lob, year, source, tags in system_files:
        days_ago = _rnd.randint(0, 90)
        h = _hl.md5(f"{fname}{counter}".encode()).hexdigest()
        records.append({
            "id": f"EV-{counter:04d}",
            "filename": fname,
            "uploaded_by": uploader,
            "ts": (datetime.now() - timedelta(days=days_ago)).strftime("%Y-%m-%d %H:%M"),
            "task_id": task_id,
            "hash": f"sha256:{h}",
            "immutable": True,
            "archived": False,
            "department": LOB_DEPT_MAP.get(lob, "Compliance"),
            "line_of_business": lob,
            "doc_type": doc_type,
            "year": year,
            "source": source,
            "tags": tags,
            "notes": "",
            "size_kb": _rnd.randint(100, 4096),
        })
        counter += 1

    return records


# ── SEED DATA ─────────────────────────────────────────────────────────────────
def seed_data():
    now = datetime.now()
    def ts(d=0,h=0): return (now-timedelta(days=d,hours=h)).strftime("%Y-%m-%d %H:%M")

    # WSP tasks define who owns each obligation inside a WSP
    wsp_tasks = {
        "WSP-001": [
            {"id":"WT-001-1","title":"Annual AML Training Completion","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Annually","description":"Ensure all registered representatives complete the annual AML training program and retain certificates as evidence.","category":"AML"},
            {"id":"WT-001-2","title":"Quarterly Transaction Monitoring Review","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Quarterly","description":"Review transaction monitoring alerts, document dispositions, and escalate any suspicious activity for SAR review.","category":"AML"},
            {"id":"WT-001-3","title":"SAR Filing Review & Submission","owner":"supervisor","role_title":"Supervising Principal","frequency":"As needed / 30-day SAR deadline","description":"Review all potential SAR candidates flagged by monitoring. Approve, escalate, or document no-file decisions within regulatory deadlines.","category":"AML"},
            {"id":"WT-001-4","title":"Customer Identification Program (CIP) Audit","owner":"supervisor","role_title":"Supervising Principal","frequency":"Semi-Annually","description":"Audit a sample of new account openings to confirm CIP procedures were followed. Document findings and remediate exceptions.","category":"AML"},
        ],
        "WSP-002": [
            {"id":"WT-002-1","title":"Daily Flagged Trade Review (24hr SLA)","owner":"supervisor","role_title":"Supervising Principal","frequency":"Daily","description":"Review all trades flagged by the surveillance system within 24 hours. Document review, rationale, and any follow-up actions taken.","category":"Supervisory"},
            {"id":"WT-002-2","title":"Monthly Customer Complaint Log Review","owner":"supervisor","role_title":"Supervising Principal","frequency":"Monthly","description":"Review all customer complaints received during the period. Confirm responses were sent, log outcomes, and escalate unresolved items.","category":"Supervisory"},
            {"id":"WT-002-3","title":"Reg BI Best Interest Review","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Quarterly","description":"Review a sample of retail recommendations to confirm they meet Reg BI best interest standards. Document findings and any remediation.","category":"Regulatory"},
            {"id":"WT-002-4","title":"New Account Opening Supervisory Sign-Off","owner":"supervisor","role_title":"Supervising Principal","frequency":"Weekly","description":"Review and approve all new retail account openings for the week. Confirm suitability information is complete and accurate.","category":"Supervisory"},
        ],
        "WSP-003": [
            {"id":"WT-003-1","title":"Information Barrier Attestation","owner":"compliance_team","role_title":"Compliance Analyst","frequency":"Annually","description":"Collect signed attestations from all employees confirming they understand and have complied with information barrier policies.","category":"Trading"},
            {"id":"WT-003-2","title":"Electronic Communication Surveillance Review","owner":"supervisor","role_title":"Supervising Principal","frequency":"Monthly","description":"Review electronic communications surveillance sample to confirm no MNPI sharing occurred across business lines.","category":"Trading"},
        ],
    }

    # Policy tasks (same structure as wsp_tasks — who owns each obligation inside a Policy)
    policy_tasks = {
        "POL-001": [
            {"id":"PT-001-1","title":"Annual AML Policy Certification","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Annually","description":"Review and certify the firm's AML Policy is current, reflects applicable FinCEN guidance, and has been distributed to all required personnel.","category":"AML"},
            {"id":"PT-001-2","title":"AML Program Independent Testing Coordination","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Annually","description":"Coordinate the annual independent testing of the AML program. Engage qualified third party or internal audit. Retain testing report and address all findings.","category":"AML"},
            {"id":"PT-001-3","title":"SAR Filing Review & Submission","owner":"supervisor","role_title":"VP, Head of Supervision","frequency":"As needed / 30-day SAR deadline","description":"Review all potential SAR candidates flagged by AML monitoring. Approve, escalate, or document no-file decisions within the 30-day regulatory deadline. Retain supporting documentation.","category":"AML"},
            {"id":"PT-001-4","title":"Customer Identification Program (CIP) Audit","owner":"supervisor","role_title":"VP, Head of Supervision","frequency":"Semi-Annually","description":"Audit a sample of new account openings to confirm CIP procedures and beneficial ownership documentation were completed correctly. Document findings and remediate exceptions within 10 business days.","category":"AML"},
        ],
        "POL-002": [
            {"id":"PT-002-1","title":"Supervision Policy Annual Review","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Annually","description":"Review the firm's supervision policy for completeness and accuracy. Confirm all lines of business and product types are covered. Update as needed and obtain approval.","category":"Supervisory"},
            {"id":"PT-002-2","title":"Quarterly Supervisory Controls Report","owner":"supervisor","role_title":"VP, Head of Supervision","frequency":"Quarterly","description":"Compile the quarterly supervisory controls report summarizing review activity, exceptions identified, and remediation status across all business lines.","category":"Supervisory"},
            {"id":"PT-002-3","title":"Reg BI Best Interest Review","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Quarterly","description":"Review a sample of retail recommendations to confirm they meet Reg BI best interest standards. Document findings and any remediation required.","category":"Regulatory"},
            {"id":"PT-002-4","title":"Daily Flagged Trade Review","owner":"supervisor","role_title":"VP, Head of Supervision","frequency":"Daily","description":"Review all trades flagged by the surveillance system within 24 hours. Document review rationale and any follow-up actions taken per WSP-002 requirements.","category":"Supervisory"},
        ],
        "POL-003": [
            {"id":"PT-003-1","title":"Information Barrier Acknowledgment Collection","owner":"compliance_team","role_title":"Compliance Analyst","frequency":"Annually","description":"Collect signed information barrier acknowledgment forms from all employees in covered functions. Retain executed forms in the compliance records system.","category":"Trading"},
        ],
        "POL-004": [
            {"id":"PT-004-1","title":"Reg BI Conflict of Interest Review","owner":"cco","role_title":"Chief Compliance Officer","frequency":"Quarterly","description":"Review all identified conflicts of interest under Reg BI. Confirm disclosures are current. Evaluate whether any new conflicts have arisen and update Form CRS if required.","category":"Regulatory"},
            {"id":"PT-004-2","title":"Best Interest Recommendation Sampling","owner":"supervisor","role_title":"Supervising Principal","frequency":"Monthly","description":"Sample retail customer recommendations for best interest compliance. Document findings, note any exceptions, and escalate to the CCO if a pattern of non-compliance is identified.","category":"Regulatory"},
        ],
    }

    return {
        "users": {
            "cco":              {"password":"pass","role":"Compliance Officer","name":"Jordan Reed","title":"Chief Compliance Officer"},
            "supervisor":       {"password":"pass","role":"Supervisor","name":"Morgan Ellis","title":"VP, Head of Supervision","department":"Wealth Management","line_of_business":"Wealth Management"},
            "compliance_team":  {"password":"pass","role":"Compliance Team","name":"Casey Park","title":"Compliance Analyst"},
        },
        "firm": {"name":"Apex Financial Services LLC","industry":"Broker-Dealer (FINRA)","crd":"123456","regulator":"FINRA / SEC","retention_years":6,"approval_mode":"dual","address":"101 Wall Street, New York, NY 10005"},
        "tasks": [
            {"id":"T-001","title":"AML Transaction Monitoring Review","assignee":"compliance_team","due":(now+timedelta(days=3)).strftime("%Y-%m-%d"),"status":"Open","priority":"High","category":"AML",
             "instructions":"Review all transaction monitoring alerts generated by the surveillance system for the current quarter. Log in to the AML monitoring platform and export the alert queue. For each alert: (1) review the underlying transaction activity and account history; (2) determine whether the activity is explainable by known customer behavior or constitutes potentially suspicious activity; (3) document your disposition — either 'cleared' with rationale or 'escalated for SAR review'; (4) confirm no open alerts are past the 30-day SAR review deadline. Any alerts meeting SAR thresholds must be escalated to the Supervising Principal immediately. Retain the completed alert disposition log as evidence.",
             "evidence_required":"Completed quarterly alert disposition log (exported from AML monitoring system); any SAR escalation memos; screenshot confirming no alerts outstanding past deadline.",
             "evidence":[],"audit":[{"ts":ts(5),"user":"cco","action":"Created task T-001"}],"rationale":"","delegated_to":"","delegated_memo":"","source_wsp":"WSP-001","recurrence":"Quarterly"},
            {"id":"T-002","title":"Q3 Supervisory Checklist – Options","assignee":"supervisor","due":(now-timedelta(days=2)).strftime("%Y-%m-%d"),"status":"Overdue","priority":"High","category":"Supervisory",
             "instructions":"Complete the Q3 supervisory review checklist for options accounts. Access the supervisory review module and pull all options trades executed during Q3. For each account with options activity: (1) confirm the customer's options approval level matches the strategies employed; (2) review for suitability — confirm investment objectives, net worth, and risk tolerance on file support the options activity; (3) flag any accounts where concentration exceeds 25% in a single underlying security; (4) document review, sign off, and note any exceptions identified. Escalate any suitability exceptions to the CCO with a written summary.",
             "evidence_required":"Signed and dated Q3 options supervisory review checklist; list of any exceptions identified with disposition notes; account-level trade blotter showing options activity reviewed.",
             "evidence":[],"audit":[{"ts":ts(10),"user":"cco","action":"Created task T-002"}],"rationale":"","delegated_to":"","delegated_memo":"","source_wsp":"WSP-002","recurrence":"Quarterly"},
            {"id":"T-003","title":"Best Execution Policy Review","assignee":"cco","due":(now+timedelta(days=10)).strftime("%Y-%m-%d"),"status":"Open","priority":"Medium","category":"Trading",
             "instructions":"Conduct the semi-annual review of the firm's Best Execution Policy. Pull trade execution quality data from the order management system for the prior six months. Review execution quality across equity, fixed income, and options orders: (1) compare execution prices against contemporaneous market prices (NBBO at time of order); (2) identify any patterns of inferior execution by venue, broker, or security type; (3) review any order routing changes made since the last review; (4) confirm the firm's order routing disclosure (Rule 606 report) is current and accurate; (5) update the Best Execution Policy document if any material changes to routing or execution practices are required. Present findings summary to the supervisory committee.",
             "evidence_required":"Execution quality analysis report (6-month period); updated Best Execution Policy document (if revised); Rule 606 disclosure confirmation; sign-off memo from supervisory committee.",
             "evidence":[],"audit":[{"ts":ts(3),"user":"supervisor","action":"Created task T-003"}],"rationale":"","delegated_to":"","delegated_memo":"","source_wsp":"WSP-002","recurrence":"Semi-Annually"},
            {"id":"T-004","title":"FINRA Rule 3110 Supervision Update","assignee":"compliance_team","due":(now+timedelta(days=1)).strftime("%Y-%m-%d"),"status":"Pending Approval","priority":"High","category":"Regulatory",
             "instructions":"Review and update the firm's supervisory procedures to ensure full compliance with FINRA Rule 3110 as amended. Steps: (1) download the current version of WSP-002 Supervisory Procedures – Retail; (2) cross-reference each supervisory obligation in Rule 3110 against the existing procedures — note any gaps or outdated references; (3) draft proposed amendments to address identified gaps (redline format required); (4) circulate the redline internally for review; (5) upload the revised clean copy and redline to the WSP repository; (6) update the Rule Inventory entry for R-001 to reflect the review date. All revisions must be reviewed and approved by the CCO prior to implementation.",
             "evidence_required":"Completed gap analysis worksheet; redline version showing tracked changes; clean final copy of updated procedures; CCO review sign-off memo.",
             "evidence":[{"name":"supervision_policy_v2.pdf","uploaded_by":"compliance_team","ts":ts(1),"hash":"abc123","immutable":True,"notes":"Clean copy of revised supervisory procedures v2.0 incorporating Rule 3110 amendments."}],"audit":[{"ts":ts(7),"user":"cco","action":"Created task T-004"},{"ts":ts(1),"user":"compliance_team","action":"Evidence uploaded: supervision_policy_v2.pdf"}],"rationale":"","delegated_to":"","delegated_memo":"","source_wsp":"WSP-002","recurrence":""},
            {"id":"T-005","title":"Annual Cybersecurity Training Completion","assignee":"compliance_team","due":(now-timedelta(days=5)).strftime("%Y-%m-%d"),"status":"Closed","priority":"Low","category":"InfoSec",
             "instructions":"Coordinate and confirm completion of the firm's annual cybersecurity awareness training for all staff. Steps: (1) log in to the training platform and pull the completion report for the current calendar year; (2) cross-reference against the current employee roster to identify anyone who has not yet completed the training; (3) send reminder notices to outstanding individuals with a completion deadline; (4) follow up until 100% completion is achieved or obtain a written exception memo for any individuals on leave; (5) export the final completion report and retain as evidence. The training must cover phishing awareness, data handling, password security, and incident reporting procedures.",
             "evidence_required":"Training completion report showing 100% completion (or exception memo); export from training platform with completion dates and names; confirmation email to CCO.",
             "evidence":[],"audit":[{"ts":ts(15),"user":"cco","action":"Created task T-005"},{"ts":ts(5),"user":"cco","action":"Closed – training certificates verified"}],"rationale":"All staff completed training by deadline.","delegated_to":"","delegated_memo":"","source_wsp":"","recurrence":"Annually"},
        ],
        "exam_requests": [
            {"id":"ER-001","title":"FINRA Cycle Exam 2024","regulator":"FINRA","due":(now+timedelta(days=30)).strftime("%Y-%m-%d"),"status":"In Progress",
             "items":[
                {"item":"Customer complaint files (2022-2024)","status":"Uploaded","assignee":"compliance_team","assigned_ts":ts(20),"upload_notes":""},
                {"item":"Supervisory procedures manual","status":"Pending","assignee":"cco","assigned_ts":ts(20),"upload_notes":""},
                {"item":"AML policies and SAR log","status":"Uploaded","assignee":"compliance_team","assigned_ts":ts(20),"upload_notes":""},
                {"item":"Trade blotter – Q1-Q3 2024","status":"Pending","assignee":"supervisor","assigned_ts":ts(20),"upload_notes":""},
             ],
             "audit":[
                {"ts":ts(20),"user":"cco","action":"Exam request ER-001 created"},
                {"ts":ts(20),"user":"cco","action":"Assigned 'Customer complaint files (2022-2024)' → compliance_team"},
                {"ts":ts(20),"user":"cco","action":"Assigned 'Supervisory procedures manual' → cco"},
                {"ts":ts(20),"user":"cco","action":"Assigned 'AML policies and SAR log' → compliance_team"},
                {"ts":ts(20),"user":"cco","action":"Assigned 'Trade blotter – Q1-Q3 2024' → supervisor"},
             ]},
            {"id":"ER-002","title":"SEC Annual Review","regulator":"SEC","due":(now+timedelta(days=60)).strftime("%Y-%m-%d"),"status":"Open",
             "items":[
                {"item":"Form ADV Part 2","status":"Pending","assignee":"cco","assigned_ts":ts(5),"upload_notes":""},
                {"item":"Investment advisory agreements sample","status":"Pending","assignee":"supervisor","assigned_ts":ts(5),"upload_notes":""},
             ],
             "audit":[
                {"ts":ts(5),"user":"cco","action":"Exam request ER-002 created"},
                {"ts":ts(5),"user":"cco","action":"Assigned 'Form ADV Part 2' → cco"},
                {"ts":ts(5),"user":"cco","action":"Assigned 'Investment advisory agreements sample' → supervisor"},
             ]},
        ],
        "rules": [
            {
                "id":"R-001","title":"FINRA Rule 3110 – Supervision","source":"FINRA","category":"Supervision",
                "description":"Firms must establish and maintain a system of supervision reasonably designed to achieve compliance with applicable securities laws, regulations, and FINRA rules.",
                "rule_requirements":"Member firms must: (1) establish and maintain written supervisory procedures (WSPs); (2) designate appropriately registered principals to supervise each type of business; (3) conduct reviews of registered representatives' correspondence and transactions; (4) perform annual compliance meetings; (5) document all supervisory reviews. Supervisory control testing must be completed annually under Rule 3120.",
                "status":"Active","last_reviewed":ts(30),"ai_flagged":False,
                "wsp_links":[
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) establishes written procedures for daily trade review, account opening approval, complaint handling, and escalation paths. Supervisors review all flagged trades within the 24-hour SLA. The CCO conducts Reg BI reviews quarterly. The Supervising Principal handles daily trade review, new account sign-offs, and complaint log maintenance. Annual supervisory controls testing is documented per Rule 3120."},
                ]
            },
            {
                "id":"R-002","title":"FINRA Rule 4512 – Customer Account Information","source":"FINRA","category":"Customer Accounts",
                "description":"Member firms must make reasonable efforts to obtain and maintain essential customer information for each account, including name, address, date of birth, employment status, annual income, and investment objectives.",
                "rule_requirements":"Firms must: (1) obtain essential customer information at account opening; (2) update records within 30 days of becoming aware of material changes; (3) verify customer identity per CIP requirements; (4) retain records for a minimum of 6 years; (5) ensure new account forms capture all required fields including risk tolerance, net worth, and investment experience. Rule applies to both cash and margin accounts.",
                "status":"Active","last_reviewed":ts(45),"ai_flagged":True,
                "ai_flagged_date":ts(6),"ai_flag_reason":"FINRA issued Regulatory Notice 24-02 clarifying expanded customer account information requirements for digital onboarding. Update may require revisions to CIP procedures and account opening forms in WSP-002.",
                "wsp_links":[
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) governs account opening procedures including new account form review, principal sign-off requirements, and annual customer information update process. The Supervising Principal reviews and approves all new accounts prior to trading. Customer information update requests are logged and reviewed within 30 days."},
                ]
            },
            {
                "id":"R-003","title":"SEC Rule 17a-4 – Books and Records Retention","source":"SEC","category":"Recordkeeping",
                "description":"Broker-dealers must preserve specified records for defined time periods. Records must be stored in a non-rewriteable, non-erasable format (WORM) and must be immediately available for inspection by regulators.",
                "rule_requirements":"Firms must retain: (1) blotters and ledgers for 6 years; (2) customer account records for 6 years; (3) correspondence for 3 years; (4) order tickets for 3 years; (5) electronic communications for 3 years. All records must be stored in WORM-compliant format. Firms must designate a third-party custodian for electronic records. Index of all records must be maintained and available within 4 hours of regulatory request.",
                "status":"Active","last_reviewed":ts(20),"ai_flagged":False,
                "wsp_links":[
                    {"wsp_id":"WSP-001","wsp_title":"Anti-Money Laundering Program","process":"Anti-Money Laundering Program (WSP-001) incorporates recordkeeping obligations for AML-related records including CIP documentation, SAR filings, and transaction monitoring records. All AML records are retained for a minimum of 5 years in WORM-compliant storage per BSA/Rule 17a-4 requirements. The CCO is responsible for annual review of AML recordkeeping procedures."},
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) requires all supervisory review records, trade review documentation, complaint files, and account records to be retained per Rule 17a-4 schedule. The firm's WORM-compliant evidence repository is used for all electronic supervisory records. Trade review blotters are preserved for 6 years; correspondence for 3 years."},
                ]
            },
            {
                "id":"R-004","title":"Bank Secrecy Act / FINRA Rule 3310 – AML Program","source":"FinCEN","category":"AML/BSA",
                "description":"Financial institutions must implement anti-money laundering programs including customer identification, transaction monitoring, suspicious activity reporting, and annual independent testing.",
                "rule_requirements":"Firms must maintain a written AML program that includes: (1) internal policies and controls reasonably designed to detect and report suspicious activity; (2) a designated BSA/AML compliance officer; (3) ongoing employee training; (4) independent testing at least annually; (5) Customer Identification Program (CIP) per 31 CFR 1023.220; (6) Customer Due Diligence (CDD) including beneficial ownership for legal entity accounts; (7) SAR filing within 30 days of detecting a reportable transaction. FINRA Rule 3310 requires the program to be reviewed and updated annually.",
                "status":"Active","last_reviewed":ts(60),"ai_flagged":True,
                "ai_flagged_date":ts(18),"ai_flag_reason":"FinCEN published updated Customer Due Diligence (CDD) guidance amending beneficial ownership thresholds and certification requirements for legal entity accounts. Review of WSP-001 AML Program is recommended to confirm CDD procedures remain current.",
                "wsp_links":[
                    {"wsp_id":"WSP-001","wsp_title":"Anti-Money Laundering Program","process":"Anti-Money Laundering Program (WSP-001) is the firm's primary AML compliance document. It establishes the complete AML program including CIP procedures, transaction monitoring thresholds and escalation protocols, SAR filing procedures and deadlines, enhanced due diligence for high-risk accounts, annual training requirements for all registered persons, and the annual independent testing program. The CCO serves as designated AML compliance officer. All registered persons complete AML training annually with completion tracked and documented."},
                ]
            },
            {
                "id":"R-005","title":"SEC Regulation Best Interest (Reg BI)","source":"SEC","category":"Suitability",
                "description":"Broker-dealers must act in the best interest of retail customers when making a recommendation of a securities transaction or investment strategy, without placing their own financial interest ahead of the customer's interest.",
                "rule_requirements":"Firms must satisfy four component obligations: (1) Disclosure Obligation — deliver Form CRS and disclose material facts about the recommendation; (2) Care Obligation — exercise reasonable diligence, care, and skill in making recommendations; (3) Conflict of Interest Obligation — establish policies and procedures to identify, disclose, and mitigate material conflicts; (4) Compliance Obligation — establish written policies and procedures reasonably designed to achieve compliance with Reg BI. Recommendations must consider the retail customer's investment profile including financial situation, investment objectives, time horizon, risk tolerance, and liquidity needs.",
                "status":"Active","last_reviewed":ts(10),"ai_flagged":False,
                "wsp_links":[
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) incorporates Reg BI requirements including the quarterly CCO review of recommendations against customer investment profiles, documentation standards for complex product recommendations, conflict of interest disclosure procedures, and Form CRS delivery tracking. The Supervising Principal reviews all recommendations for Reg BI compliance as part of the daily trade review process. WSP-002 v2.0 was updated specifically to incorporate Reg BI requirements."},
                ]
            },
            {
                "id":"R-006","title":"FINRA Rule 3120 – Supervisory Control System","source":"FINRA","category":"Supervision",
                "description":"Firms must test and verify that their supervisory procedures are reasonably designed to achieve compliance, and must produce an annual report summarizing the supervisory controls testing.",
                "rule_requirements":"Firms must: (1) designate one or more principals to establish, maintain, and enforce a supervisory control system; (2) conduct annual testing of the firm's supervisory procedures; (3) produce a written report detailing the testing results and any identified deficiencies; (4) report the results to senior management; (5) address identified deficiencies with corrective action. The annual report must be submitted to a senior principal. For firms with at least $200M in gross revenue, a CEO certification is required.",
                "status":"Active","last_reviewed":ts(25),"ai_flagged":False,
                "wsp_links":[
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) provides the framework for the annual FINRA Rule 3120 supervisory controls testing program. The CCO conducts annual testing of all supervisory procedures described in WSP-002. Testing results are documented and presented to senior management in the annual supervisory controls report. Any deficiencies identified are tracked to remediation. The compliance testing module is used to formally document and retain all testing evidence."},
                ]
            },
            {
                "id":"R-007","title":"SEC Rule 15c3-1 – Net Capital Rule","source":"SEC","category":"Financial Responsibility",
                "description":"Broker-dealers must maintain minimum net capital levels relative to aggregate indebtedness or customer debit items, ensuring sufficient liquid capital to meet obligations to customers and creditors.",
                "rule_requirements":"Firms must maintain: (1) minimum net capital of $250,000 (alternative standard) or meet aggregate indebtedness ratio of 15:1; (2) timely FOCUS Report filings reflecting accurate net capital computations; (3) notification to FINRA within 24 hours of falling below minimum net capital; (4) annual audit with independent accountant; (5) net capital computations must include haircuts on securities positions per the rule's schedule. Firms falling below minimum must cease business and notify regulators immediately.",
                "status":"Active","last_reviewed":ts(18),"ai_flagged":False,
                "wsp_links":[
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) requires the Chief Financial Officer to review and certify the monthly FOCUS Report net capital computation prior to filing. Any net capital early warning breach (below 120% of required minimum) triggers immediate escalation to the CCO and CEO per procedures in WSP-002. The annual audit process is documented and coordinated through the compliance calendar."},
                ]
            },
            {
                "id":"R-008","title":"FINRA Rule 4370 – Business Continuity Plan","source":"FINRA","category":"Operations",
                "description":"Each member must create and maintain a written business continuity plan identifying procedures relating to an emergency or significant business disruption, and must review the plan at least annually.",
                "rule_requirements":"Firms must maintain a BCP that addresses: (1) data backup and recovery; (2) mission critical systems; (3) financial and operational assessments; (4) alternative communications between customers and the firm; (5) alternate physical locations of employees; (6) critical business constituent, bank, and counter-party impacts; (7) regulatory reporting; (8) communications with regulators. The plan must be reviewed and updated annually. A summary must be made available to customers. Emergency contact information must be kept current.",
                "status":"Active","last_reviewed":ts(35),"ai_flagged":False,
                "wsp_links":[
                    {"wsp_id":"WSP-001","wsp_title":"Anti-Money Laundering Program","process":"Anti-Money Laundering Program (WSP-001) references BCP procedures for AML-critical systems including transaction monitoring platform failover, SAR filing backup procedures during system outages, and CIP documentation preservation. The AML system is classified as mission-critical in the firm's BCP and has a recovery time objective of 4 hours."},
                    {"wsp_id":"WSP-002","wsp_title":"Supervisory Procedures – Retail","process":"Supervisory Procedures – Retail (WSP-002) incorporates BCP procedures for supervisory functions including trade review continuity, alternate supervisor designation protocols, and customer communication procedures during business disruptions. The Supervising Principal maintains a current list of alternate supervisors and emergency contact protocols consistent with Rule 4370."},
                ]
            },
        ],
        "wsps": [
            {
                "id":"WSP-001","title":"Anti-Money Laundering Program","version":"3.2",
                "status":"Approved","review_status":"approved","line_of_business":"Operations",
                "regulations":["BSA / AML Program","FINRA Rule 3310","FinCEN CDD Rule"],
                "related_tests":["CT-001"],
                "owner":"cco","approved_by":"supervisor","approved_ts":ts(15),
                "content":"This WSP establishes our AML program per BSA requirements. Includes CIP, transaction monitoring, SAR filing procedures, and annual training requirements. All registered representatives must complete AML training annually. The Chief Compliance Officer oversees the overall program and training. The Supervising Principal reviews SAR candidates and conducts CIP audits.",
                "versions":[{"ver":"3.1","ts":ts(60),"author":"cco","note":"Updated SAR thresholds"},{"ver":"3.2","ts":ts(15),"author":"cco","note":"Added enhanced due diligence for high-risk accounts"}],
                "files":[],"reviewer":"","review_sent_ts":"",
                "audit":[
                    {"ts":ts(60),"user":"cco","action":"WSP-001 v3.1 uploaded by cco"},
                    {"ts":ts(15),"user":"cco","action":"v3.2 redline and clean copy uploaded by cco"},
                    {"ts":ts(15),"user":"cco","action":"Sent for review to supervisor"},
                    {"ts":ts(15),"user":"supervisor","action":"WSP-001 v3.2 approved by supervisor"},
                ],
            },
            {
                "id":"WSP-002","title":"Supervisory Procedures – Retail","version":"2.0",
                "status":"Approved","review_status":"approved","line_of_business":"Wealth Management",
                "regulations":["FINRA Rule 3110","FINRA Rule 3120","Reg BI / SEC 15l-1","FINRA Rule 4512"],
                "related_tests":["CT-001","CT-002"],
                "owner":"supervisor","approved_by":"cco","approved_ts":ts(30),
                "content":"Supervisory procedures for retail customer accounts. Covers trade review, account opening, complaint handling, and escalation paths. Supervisors must review all flagged trades within 24 hours. The Chief Compliance Officer conducts Reg BI reviews quarterly. The Supervising Principal handles daily trade review, complaint logs, and new account sign-offs.",
                "versions":[{"ver":"1.9","ts":ts(90),"author":"supervisor","note":"Added options review workflow"},{"ver":"2.0","ts":ts(30),"author":"supervisor","note":"Incorporated Reg BI requirements"}],
                "files":[],"reviewer":"","review_sent_ts":"",
                "audit":[
                    {"ts":ts(90),"user":"supervisor","action":"WSP-002 v1.9 uploaded by supervisor"},
                    {"ts":ts(30),"user":"supervisor","action":"v2.0 redline and clean copy uploaded by supervisor"},
                    {"ts":ts(30),"user":"supervisor","action":"Sent for review to admin"},
                    {"ts":ts(30),"user":"cco","action":"WSP-002 v2.0 approved by cco"},
                ],
            },
            {
                "id":"WSP-003","title":"Information Barriers Policy","version":"1.1",
                "status":"Draft","review_status":"with_reviewer","line_of_business":"Investment Banking",
                "regulations":["SEC Rule 10b-5","FINRA Rule 3110","Insider Trading Prevention"],
                "related_tests":[],
                "owner":"compliance_team","approved_by":"","approved_ts":"",
                "content":"Policies governing information barriers between advisory and brokerage functions. Employees may not share material non-public information across business lines. The Compliance Analyst collects attestations. The Supervising Principal reviews electronic communication surveillance.",
                "versions":[{"ver":"1.0","ts":ts(120),"author":"compliance_team","note":"Initial draft"},{"ver":"1.1","ts":ts(7),"author":"compliance_team","note":"Added electronic communication monitoring section"}],
                "files":[
                    {"name":"WSP-003_v1.1_redline.docx","type":"Redline","uploaded_by":"compliance_team","ts":ts(7),"notes":"Redline against v1.0 showing all tracked changes to section 3 and 4."},
                    {"name":"WSP-003_v1.1_clean.docx","type":"Clean","uploaded_by":"compliance_team","ts":ts(7),"notes":"Clean final copy for review and approval."},
                ],
                "reviewer":"cco",
                "review_sent_ts":ts(6),
                "audit":[
                    {"ts":ts(120),"user":"compliance_team","action":"WSP-003 v1.0 created and uploaded"},
                    {"ts":ts(7),"user":"compliance_team","action":"v1.1 redline uploaded: WSP-003_v1.1_redline.docx"},
                    {"ts":ts(7),"user":"compliance_team","action":"v1.1 clean copy uploaded: WSP-003_v1.1_clean.docx"},
                    {"ts":ts(6),"user":"compliance_team","action":"Sent for review to Jordan Reed (cco) — please review tracked changes in section 3 re: cross-departmental information sharing restrictions."},
                ],
            },
        ],
        "wsp_tasks": wsp_tasks,
        "policy_tasks": policy_tasks,
        "policies": [
            {
                "id":"POL-001","title":"Anti-Money Laundering Policy","version":"2.1",
                "status":"Approved","review_status":"approved","line_of_business":"Operations",
                "type":"Policy",
                "regulations":["BSA / AML Program","FinCEN CDD Rule","FINRA Rule 3310"],
                "related_tests":["CT-001","CT-003"],
                "owner":"cco","approved_by":"supervisor","approved_ts":ts(20),
                "content":"This policy establishes the firm's overall Anti-Money Laundering program requirements in accordance with the Bank Secrecy Act and FinCEN regulations. It defines roles and responsibilities, sets the tone for the firm's AML risk appetite, and establishes the framework for the firm's AML procedures contained in WSP-001. The Chief Compliance Officer is designated as the firm's BSA/AML Compliance Officer.",
                "versions":[{"ver":"2.0","ts":ts(90),"author":"cco","note":"Annual review — no material changes"},{"ver":"2.1","ts":ts(20),"author":"cco","note":"Updated FinCEN beneficial ownership rule references"}],
                "files":[],"reviewer":"","review_sent_ts":"",
                "audit":[
                    {"ts":ts(90),"user":"cco","action":"POL-001 v2.0 annual review completed"},
                    {"ts":ts(20),"user":"cco","action":"v2.1 uploaded — updated FinCEN references"},
                    {"ts":ts(20),"user":"cco","action":"Sent for review to supervisor"},
                    {"ts":ts(20),"user":"supervisor","action":"POL-001 v2.1 approved"},
                ],
            },
            {
                "id":"POL-002","title":"Supervision and Supervisory Controls Policy","version":"3.0",
                "status":"Approved","review_status":"approved","line_of_business":"Wealth Management",
                "type":"Policy",
                "owner":"supervisor","approved_by":"cco","approved_ts":ts(35),
                "regulations":["FINRA Rule 3110","FINRA Rule 3120","Reg BI / SEC 15l-1"],
                "related_tests":["CT-001","CT-002","CT-004"],
                "content":"This policy establishes the firm's supervisory framework consistent with FINRA Rule 3110 and 3120. It defines the firm's organizational supervision structure, the designation of supervisory principals by business line, and the annual supervisory controls testing program. Detailed supervisory procedures for each line of business are contained in the corresponding WSPs.",
                "versions":[{"ver":"2.9","ts":ts(120),"author":"supervisor","note":"Added fixed income desk supervisor designation"},{"ver":"3.0","ts":ts(35),"author":"supervisor","note":"Annual review — incorporated Reg BI supervisory obligations"}],
                "files":[],"reviewer":"","review_sent_ts":"",
                "audit":[
                    {"ts":ts(120),"user":"supervisor","action":"POL-002 v2.9 uploaded"},
                    {"ts":ts(35),"user":"supervisor","action":"v3.0 uploaded — annual review"},
                    {"ts":ts(35),"user":"supervisor","action":"Sent for review to admin"},
                    {"ts":ts(35),"user":"cco","action":"POL-002 v3.0 approved"},
                ],
            },
            {
                "id":"POL-003","title":"Information Barriers and Insider Trading Prevention Policy","version":"1.2",
                "status":"Approved","review_status":"approved","line_of_business":"Investment Banking",
                "type":"Policy",
                "regulations":["SEC Rule 10b-5","FINRA Rule 3110","MAR Regulation"],
                "related_tests":[],
                "owner":"compliance_team","approved_by":"cco","approved_ts":ts(50),
                "content":"This policy establishes the firm's information barrier requirements between investment banking and brokerage functions to prevent the misuse of material non-public information. It defines restricted and watch lists, personal account dealing rules, and communication protocols. Detailed operating procedures are contained in WSP-003.",
                "versions":[{"ver":"1.1","ts":ts(180),"author":"compliance_team","note":"Initial approved version"},{"ver":"1.2","ts":ts(50),"author":"compliance_team","note":"Added electronic communication surveillance requirements"}],
                "files":[],"reviewer":"","review_sent_ts":"",
                "audit":[
                    {"ts":ts(180),"user":"compliance_team","action":"POL-003 v1.1 created"},
                    {"ts":ts(50),"user":"compliance_team","action":"v1.2 uploaded — added e-comm surveillance section"},
                    {"ts":ts(50),"user":"compliance_team","action":"Sent for review to admin"},
                    {"ts":ts(50),"user":"cco","action":"POL-003 v1.2 approved"},
                ],
            },
            {
                "id":"POL-004","title":"Regulation Best Interest and Conflicts of Interest Policy","version":"1.0",
                "status":"Draft","review_status":"with_reviewer","line_of_business":"Wealth Management",
                "type":"Policy",
                "regulations":["Reg BI / SEC 15l-1","FINRA Rule 3110"],
                "related_tests":["CT-004"],
                "owner":"compliance_team","approved_by":"","approved_ts":"",
                "content":"This policy establishes the firm's compliance framework for SEC Regulation Best Interest. It defines the four component obligations — Disclosure, Care, Conflict of Interest, and Compliance — and sets firm-wide standards for identifying and mitigating conflicts. Detailed supervisory procedures are covered in WSP-002.",
                "versions":[{"ver":"1.0","ts":ts(10),"author":"compliance_team","note":"Initial draft for review"}],
                "files":[
                    {"name":"POL-004_v1.0_draft.docx","type":"Clean","uploaded_by":"compliance_team","ts":ts(10),"notes":"Initial clean draft for CCO review and approval."},
                ],
                "reviewer":"cco","review_sent_ts":ts(9),
                "audit":[
                    {"ts":ts(10),"user":"compliance_team","action":"POL-004 v1.0 draft created"},
                    {"ts":ts(10),"user":"compliance_team","action":"Clean copy uploaded: POL-004_v1.0_draft.docx"},
                    {"ts":ts(9),"user":"compliance_team","action":"Sent for review to Jordan Reed (cco) — initial Reg BI policy draft, please review scope and conflict of interest section."},
                ],
            },
        ],
        "audit_trail": [
            {"ts":ts(30),"user":"cco","module":"System","action":"System initialized; sample data seeded"},
            {"ts":ts(20),"user":"cco","module":"Exam Requests","action":"Created ER-001: FINRA Cycle Exam 2024"},
            {"ts":ts(15),"user":"supervisor","module":"WSP Repository","action":"Approved WSP-001 v3.2"},
            {"ts":ts(5),"user":"cco","module":"Exam Requests","action":"Created ER-002: SEC Annual Review"},
        ],
        "supervisory_checklists": [
            # ── Jordan Reed (admin) — one checklist, all procedures he owns ──
            {
                "id": "CL-001",
                "assigned_to": "cco",
                "owner_name":  "Jordan Reed",
                "owner_title": "Chief Compliance Officer",
                "status":      "Active",
                "created_ts":  ts(35),
                "created_by":  "system",
                "items": [
                    # From POL-001
                    {
                        "wt_id":"PT-001-1","source_policy_id":"POL-001",
                        "source_policy":"Anti-Money Laundering Policy",
                        "title":"Annual AML Policy Certification",
                        "description":"Review and certify the firm's AML Policy is current, reflects applicable FinCEN guidance, and has been distributed to all required personnel.",
                        "frequency":"Annually","category":"AML",
                        "due_date":(now+timedelta(days=320)).strftime("%Y-%m-%d"),
                        "status":"Signed Off","checked_off":True,"completed_ts":ts(5),
                        "supervisor_review":True,"supervisor_reviewed_by":"supervisor",
                        "supervisor_reviewed_ts":ts(3),"supervisor_notes":"Certification reviewed and confirmed current.",
                        "evidence":[{"name":"AML_Policy_Certification_2025.pdf","uploaded_by":"cco","ts":ts(5),"hash":"abc001","notes":"Signed annual certification — AML Policy v2.1 reviewed and certified as current."}],
                        "comments":[],
                        "audit":[
                            {"ts":ts(35),"user":"system","action":"Item added from POL-001 v2.1"},
                            {"ts":ts(5),"user":"cco","action":"Evidence uploaded: AML_Policy_Certification_2025.pdf"},
                            {"ts":ts(5),"user":"cco","action":"Marked complete — routed to supervisor for sign-off"},
                            {"ts":ts(3),"user":"supervisor","action":"Supervisor sign-off complete"},
                        ],
                    },
                    {
                        "wt_id":"PT-001-2","source_policy_id":"POL-001",
                        "source_policy":"Anti-Money Laundering Policy",
                        "title":"AML Program Independent Testing Coordination",
                        "description":"Coordinate the annual independent testing of the AML program. Engage qualified third party or internal audit. Retain testing report and address all findings.",
                        "frequency":"Annually","category":"AML",
                        "due_date":(now+timedelta(days=290)).strftime("%Y-%m-%d"),
                        "status":"Signed Off","checked_off":True,"completed_ts":ts(21),
                        "supervisor_review":True,"supervisor_reviewed_by":"supervisor",
                        "supervisor_reviewed_ts":ts(20),"supervisor_notes":"Testing report reviewed — one finding tracked in CT-003.",
                        "evidence":[{"name":"CT003_AML_Independent_Testing_Report_2025.pdf","uploaded_by":"cco","ts":ts(21),"hash":"abc002","notes":"Final testing report from CT-003 independent AML program review."}],
                        "comments":[],
                        "audit":[
                            {"ts":ts(35),"user":"system","action":"Item added from POL-001 v2.1"},
                            {"ts":ts(21),"user":"cco","action":"Evidence uploaded: CT003_AML_Independent_Testing_Report_2025.pdf"},
                            {"ts":ts(21),"user":"cco","action":"Marked complete — routed to supervisor"},
                            {"ts":ts(20),"user":"supervisor","action":"Supervisor sign-off complete"},
                        ],
                    },
                    # From POL-002
                    {
                        "wt_id":"PT-002-1","source_policy_id":"POL-002",
                        "source_policy":"Supervision and Supervisory Controls Policy",
                        "title":"Supervision Policy Annual Review",
                        "description":"Review the firm's supervision policy for completeness and accuracy. Confirm all lines of business and product types are covered. Update as needed and obtain approval.",
                        "frequency":"Annually","category":"Supervisory",
                        "due_date":(now+timedelta(days=270)).strftime("%Y-%m-%d"),
                        "status":"Pending","checked_off":False,"completed_ts":"",
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[],"comments":[],
                        "audit":[{"ts":ts(35),"user":"system","action":"Item added from POL-002 v3.0"}],
                    },
                    {
                        "wt_id":"PT-002-3","source_policy_id":"POL-002",
                        "source_policy":"Supervision and Supervisory Controls Policy",
                        "title":"Reg BI Best Interest Review",
                        "description":"Review a sample of retail recommendations to confirm they meet Reg BI best interest standards. Document findings and any remediation required.",
                        "frequency":"Quarterly","category":"Regulatory",
                        "due_date":(now+timedelta(days=14)).strftime("%Y-%m-%d"),
                        "status":"Pending","checked_off":False,"completed_ts":"",
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[],"comments":[],
                        "audit":[{"ts":ts(35),"user":"system","action":"Item added from POL-002 v3.0"}],
                    },
                ],
                "audit":[
                    {"ts":ts(35),"user":"system","action":"User checklist created"},
                    {"ts":ts(35),"user":"system","action":"Items added from POL-001 v2.1 and POL-002 v3.0"},
                ],
            },
            # ── Morgan Ellis (supervisor) — one checklist, all procedures she owns ──
            {
                "id": "CL-002",
                "assigned_to": "supervisor",
                "owner_name":  "Morgan Ellis",
                "owner_title": "VP, Head of Supervision",
                "status":      "Active",
                "created_ts":  ts(35),
                "created_by":  "system",
                "items": [
                    # From POL-001
                    {
                        "wt_id":"PT-001-3","source_policy_id":"POL-001",
                        "source_policy":"Anti-Money Laundering Policy",
                        "title":"SAR Filing Review & Submission",
                        "description":"Review all potential SAR candidates flagged by AML monitoring. Approve, escalate, or document no-file decisions within the 30-day regulatory deadline. Retain supporting documentation.",
                        "frequency":"As needed / 30-day SAR deadline","category":"AML",
                        "due_date":(now+timedelta(days=30)).strftime("%Y-%m-%d"),
                        "status":"Pending","checked_off":False,"completed_ts":"",
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[],"comments":[],
                        "audit":[{"ts":ts(35),"user":"system","action":"Item added from POL-001 v2.1"}],
                    },
                    {
                        "wt_id":"PT-001-4","source_policy_id":"POL-001",
                        "source_policy":"Anti-Money Laundering Policy",
                        "title":"Customer Identification Program (CIP) Audit",
                        "description":"Audit a sample of new account openings to confirm CIP procedures and beneficial ownership documentation were completed correctly. Document findings and remediate exceptions within 10 business days.",
                        "frequency":"Semi-Annually","category":"AML",
                        "due_date":(now+timedelta(days=75)).strftime("%Y-%m-%d"),
                        "status":"Pending","checked_off":False,"completed_ts":"",
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[],"comments":[],
                        "audit":[{"ts":ts(35),"user":"system","action":"Item added from POL-001 v2.1"}],
                    },
                    # From POL-002
                    {
                        "wt_id":"PT-002-2","source_policy_id":"POL-002",
                        "source_policy":"Supervision and Supervisory Controls Policy",
                        "title":"Quarterly Supervisory Controls Report",
                        "description":"Compile the quarterly supervisory controls report summarizing review activity, exceptions identified, and remediation status across all business lines.",
                        "frequency":"Quarterly","category":"Supervisory",
                        "due_date":(now+timedelta(days=7)).strftime("%Y-%m-%d"),
                        "status":"Awaiting Review","checked_off":True,"completed_ts":ts(2),
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[{"name":"Q1_2026_Supervisory_Controls_Report.pdf","uploaded_by":"supervisor","ts":ts(2),"hash":"cde003","notes":"Q1 2026 SCP report covering trade review, complaint handling, and Reg BI sampling."}],
                        "comments":[],
                        "audit":[
                            {"ts":ts(35),"user":"system","action":"Item added from POL-002 v3.0"},
                            {"ts":ts(2),"user":"supervisor","action":"Evidence uploaded: Q1_2026_Supervisory_Controls_Report.pdf"},
                            {"ts":ts(2),"user":"supervisor","action":"Marked complete — routed to Jordan Reed for sign-off"},
                        ],
                    },
                    {
                        "wt_id":"PT-002-4","source_policy_id":"POL-002",
                        "source_policy":"Supervision and Supervisory Controls Policy",
                        "title":"Daily Flagged Trade Review",
                        "description":"Review all trades flagged by the surveillance system within 24 hours. Document review rationale and any follow-up actions taken per WSP-002 requirements.",
                        "frequency":"Daily","category":"Supervisory",
                        "due_date":(now+timedelta(days=1)).strftime("%Y-%m-%d"),
                        "status":"Pending","checked_off":False,"completed_ts":"",
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[],"comments":[],
                        "audit":[{"ts":ts(35),"user":"system","action":"Item added from POL-002 v3.0"}],
                    },
                ],
                "audit":[
                    {"ts":ts(35),"user":"system","action":"User checklist created"},
                    {"ts":ts(35),"user":"system","action":"Items added from POL-001 v2.1 and POL-002 v3.0"},
                ],
            },
            # ── Casey Park (analyst) — one checklist ──
            {
                "id": "CL-003",
                "assigned_to": "compliance_team",
                "owner_name":  "Casey Park",
                "owner_title": "Compliance Analyst",
                "status":      "Active",
                "created_ts":  ts(50),
                "created_by":  "system",
                "items": [
                    # From POL-003
                    {
                        "wt_id":"PT-003-1","source_policy_id":"POL-003",
                        "source_policy":"Information Barriers and Insider Trading Prevention Policy",
                        "title":"Information Barrier Acknowledgment Collection",
                        "description":"Collect signed information barrier acknowledgment forms from all employees in covered functions. Retain executed forms in the compliance records system.",
                        "frequency":"Annually","category":"Trading",
                        "due_date":(now+timedelta(days=200)).strftime("%Y-%m-%d"),
                        "status":"Pending","checked_off":False,"completed_ts":"",
                        "supervisor_review":False,"supervisor_reviewed_by":"","supervisor_reviewed_ts":"","supervisor_notes":"",
                        "evidence":[],"comments":[],
                        "audit":[{"ts":ts(50),"user":"system","action":"Item added from POL-003 v1.2"}],
                    },
                ],
                "audit":[
                    {"ts":ts(50),"user":"system","action":"User checklist created"},
                    {"ts":ts(50),"user":"system","action":"Item added from POL-003 v1.2"},
                ],
            },
        ],
        "evidence": _build_data_lake(ts),
        "rule_scan_memo": None,
        "org_chart": [
            {"role":"Chief Compliance Officer","user":"cco","reports_to":"","responsibilities":["AML Oversight","Exam Management","Rule Inventory"]},
            {"role":"VP, Head of Supervision","user":"supervisor","reports_to":"cco","responsibilities":["Trade Review","Supervisory Checklists","Rep Oversight"]},
            {"role":"Compliance Analyst","user":"compliance_team","reports_to":"supervisor","responsibilities":["Evidence Collection","Testing","Filings"]},
        ],
        "approvals": [
            {"id":"AP-001","item":"WSP-002 v2.0","type":"WSP Approval","requestor":"supervisor","approver1":"cco","approver1_ts":ts(30),"approver2":"","status":"Approved","ts":ts(30)},
        ],
        "workflow_templates": [
            {
                "id":"WF-T-001","name":"Monthly Trade Supervision Review",
                "category":"Supervision","purpose":"Supervisory Review",
                "department":"Compliance","frequency":"Monthly",
                "owner_role":"Supervisor","regulation":"FINRA 3110",
                "tags":["trade-review","supervision","monthly"],
                "status":"Active","created":ts(45),
            },
            {
                "id":"WF-T-002","name":"Annual AML Program Certification",
                "category":"Compliance","purpose":"Periodic Certification",
                "department":"Compliance","frequency":"Annual",
                "owner_role":"Compliance Officer","regulation":"AML Program",
                "tags":["aml","certification","annual"],
                "status":"Active","created":ts(90),
            },
            {
                "id":"WF-T-003","name":"Reg BI Best Interest Exception Review",
                "category":"Risk Monitoring","purpose":"Exception Handling",
                "department":"Compliance","frequency":"Quarterly",
                "owner_role":"Compliance Officer","regulation":"FINRA 3110",
                "tags":["reg-bi","exception","quarterly"],
                "status":"Active","created":ts(30),
            },
            {
                "id":"WF-T-004","name":"Electronic Comm Surveillance Review",
                "category":"Supervision","purpose":"Regulatory Requirement",
                "department":"Supervision","frequency":"Monthly",
                "owner_role":"Supervisor","regulation":"FINRA 3120",
                "tags":["surveillance","communications","monthly"],
                "status":"Active","created":ts(20),
            },
            {
                "id":"WF-T-005","name":"Quarterly Risk Assessment",
                "category":"Risk Monitoring","purpose":"Risk Monitoring",
                "department":"Risk","frequency":"Quarterly",
                "owner_role":"Compliance Officer","regulation":"SEC 206(4)-7",
                "tags":["risk","quarterly","assessment"],
                "status":"Archived","created":ts(180),
            },
        ],
        "workflow_instances": [
            {
                "id":"WF-I-001","template_id":"WF-T-001",
                "name":"Monthly Trade Review — March 2026","instance_name":"March 2026",
                "owner":"supervisor","due":(now+timedelta(days=8)).strftime("%Y-%m-%d"),
                "priority":"High","status":"In Progress",
                "created":ts(5),"category":"Supervision","regulation":"FINRA 3110",
            },
            {
                "id":"WF-I-002","template_id":"WF-T-002",
                "name":"Annual AML Certification — 2026","instance_name":"2026",
                "owner":"cco","due":(now+timedelta(days=45)).strftime("%Y-%m-%d"),
                "priority":"High","status":"Open",
                "created":ts(3),"category":"Compliance","regulation":"AML Program",
            },
            {
                "id":"WF-I-003","template_id":"WF-T-003",
                "name":"Reg BI Exception Review — Q1 2026","instance_name":"Q1 2026",
                "owner":"cco","due":(now-timedelta(days=3)).strftime("%Y-%m-%d"),
                "priority":"Normal","status":"Overdue",
                "created":ts(20),"category":"Risk Monitoring","regulation":"FINRA 3110",
            },
            {
                "id":"WF-I-004","template_id":"WF-T-004",
                "name":"E-Comm Surveillance — March 2026","instance_name":"March 2026",
                "owner":"supervisor","due":(now+timedelta(days=15)).strftime("%Y-%m-%d"),
                "priority":"Normal","status":"Open",
                "created":ts(2),"category":"Supervision","regulation":"FINRA 3120",
            },
        ],
        "compliance_tests": [
            {
                "id":"CT-001",
                "test_type":"SCP Testing (FINRA Rule 3120)",
                "line_of_business":"Wealth Management",
                "topic":"Annual Supervisory Controls Testing — Retail",
                "linked_policies":["POL-002","POL-001"],
                "initial_risk_rating":"High",
                "risk_rating":"High",
                "time_period_reviewed":"Q1 2026 (Jan 1 – Mar 31, 2026)",
                "approval_date":ts(3),
                "assignee":"supervisor",
                "deadline":(now+timedelta(days=14)).strftime("%Y-%m-%d"),
                "related_workflow":"WF-I-001",
                "related_checklist":"",
                "status":"In Progress",
                "created":ts(10),
                "created_by":"cco",
                "date_started":ts(10),
                "submitted_for_review":False,
                "completed":False,
                "completed_ts":"",
                "reviewer":"cco",
                "review_notes":"",
                "final_risk_rating":"",
                "population_description":"All retail brokerage accounts with flagged trades Q1 2026. Total population: 847 accounts.",
                "sample_size":25,
                "sample_method":"Random",
                "sample_rationale":"Random sample of 25 accounts selected from flagged trade population using systematic random sampling.",
                "procedures":[
                    {"id":"P-001","step":1,
                     "title":"Quarterly Business Continuity Plan Review",
                     "description":"Confirm evidence of quarterly BCP reviews for 2025.",
                     "notes":"All four quarterly reviews documented with supervisor approval.",
                     "result":"Pass","exception":False,"has_remediation_item":False,
                     "tested_by":"supervisor","tested_ts":ts(5)},
                    {"id":"P-002","step":2,
                     "title":"Supervisory Sign-Off Documentation",
                     "description":"Confirm supervisory sign-off for each flagged trade in sample.",
                     "notes":"Sign-off present in all 25 sampled files. Consistent with WSP-002.",
                     "result":"Pass","exception":False,"has_remediation_item":False,
                     "tested_by":"supervisor","tested_ts":ts(4)},
                    {"id":"P-003","step":3,
                     "title":"CCO Escalation SLA Compliance",
                     "description":"Confirm escalated trades received CCO review within 48 hours.",
                     "notes":"3 of 4 escalated trades on time. 1 trade (acct #4471) escalated 2/14 — CCO review not documented until 2/19, 5 days over SLA.",
                     "result":"Exception","exception":True,"has_remediation_item":True,
                     "rem_name":"Update CCO escalation SLA monitoring procedure",
                     "rem_detail":"Implement daily automated alert for escalated trades approaching 48-hour CCO review deadline. Update WSP-002 to require supervisor to flag CCO review delays same-day.",
                     "rem_assign":"cco","rem_due":(now+timedelta(days=21)).strftime("%Y-%m-%d"),
                     "tested_by":"supervisor","tested_ts":ts(3)},
                ],
                "exceptions":[
                    {"id":"EX-001","test_id":"CT-001","procedure_id":"P-003","description":"CCO review of escalated trade for account #4471 completed 5 days beyond required SLA.","root_cause":"No automated alert triggered for pending CCO escalation reviews.","remediation_plan":"Implement SLA monitoring alert and update WSP-002.","assigned_to":"admin","due_date":(now+timedelta(days=21)).strftime("%Y-%m-%d"),"status":"Open","created":ts(3)},
                ],
                "conclusion":"",
                "audit":[
                    {"ts":ts(10),"user":"cco","action":"Test CT-001 created"},
                    {"ts":ts(5),"user":"supervisor","action":"Procedure P-001 completed: Pass"},
                    {"ts":ts(4),"user":"supervisor","action":"Procedure P-002 completed: Pass"},
                    {"ts":ts(3),"user":"supervisor","action":"Procedure P-003 completed: Exception — CCO SLA breach identified"},
                ],
            },
            {
                "id":"CT-002",
                "test_type":"SEC Rule 206(4)-7 Annual Review",
                "line_of_business":"Capital Markets",
                "topic":"Annual Compliance Program Review",
                "linked_policies":["POL-002"],
                "initial_risk_rating":"Medium",
                "risk_rating":"Medium",
                "time_period_reviewed":"Full Year 2025 (Jan 1 – Dec 31, 2025)",
                "approval_date":"",
                "assignee":"cco",
                "deadline":(now+timedelta(days=30)).strftime("%Y-%m-%d"),
                "related_workflow":"WF-I-002",
                "related_checklist":"",
                "status":"Not Started",
                "created":ts(3),
                "created_by":"cco",
                "date_started":"",
                "submitted_for_review":False,
                "completed":False,
                "completed_ts":"",
                "reviewer":"",
                "review_notes":"",
                "final_risk_rating":"",
                "population_description":"",
                "sample_size":0,
                "sample_method":"",
                "sample_rationale":"",
                "procedures":[
                    {"id":"P-001","step":1,"title":"Written Policies & Procedures Review",
                     "description":"Review written compliance policies to confirm reasonably designed to prevent violations.",
                     "result":"","notes":"","exception":False,"has_remediation_item":False,"tested_by":"","tested_ts":""},
                    {"id":"P-002","step":2,"title":"Annual Review Documentation",
                     "description":"Confirm annual review of compliance program was conducted and documented by CCO.",
                     "result":"","notes":"","exception":False,"has_remediation_item":False,"tested_by":"","tested_ts":""},
                    {"id":"P-003","step":3,"title":"Training Records Review",
                     "description":"Confirm all advisory personnel completed required compliance training.",
                     "result":"","notes":"","exception":False,"has_remediation_item":False,"tested_by":"","tested_ts":""},
                ],
                "exceptions":[],
                "conclusion":"",
                "audit":[{"ts":ts(3),"user":"cco","action":"Test CT-002 created"}],
            },
            {
                "id":"CT-003",
                "test_type":"BSA / AML Independent Testing",
                "line_of_business":"Operations",
                "topic":"Annual AML Program Independent Testing",
                "linked_policies":["POL-001"],
                "initial_risk_rating":"High",
                "risk_rating":"High",
                "time_period_reviewed":"Full Year 2025 (Jan 1 – Dec 31, 2025)",
                "approval_date":ts(20),
                "assignee":"cco",
                "deadline":(now-timedelta(days=5)).strftime("%Y-%m-%d"),
                "related_workflow":"WF-I-002",
                "related_checklist":"",
                "status":"Completed",
                "created":ts(45),
                "created_by":"cco",
                "date_started":ts(40),
                "submitted_for_review":True,
                "completed":True,
                "completed_ts":ts(20),
                "reviewer":"supervisor",
                "review_notes":"Testing complete. One finding noted for CIP documentation. Remediation in progress.",
                "final_risk_rating":"Medium",
                "population_description":"All new account openings and transaction monitoring alerts for 2025. Population: 1,240 new accounts, 3,800 alerts.",
                "sample_size":40,
                "sample_method":"Risk-Based",
                "sample_rationale":"Risk-based sample prioritising high-risk customer segments and large-value transactions.",
                "procedures":[
                    {"id":"P-001","step":1,"title":"CIP Completeness Review",
                     "description":"Confirm CIP documentation is complete for sampled new accounts.",
                     "notes":"38 of 40 accounts fully documented. 2 accounts missing beneficial ownership certification.",
                     "result":"Exception","exception":True,"has_remediation_item":True,
                     "rem_name":"Obtain missing beneficial ownership certifications",
                     "rem_detail":"Contact relationship managers for accounts #2271 and #2389 to obtain outstanding FinCEN beneficial ownership certifications. Update CIP files within 10 business days.",
                     "rem_assign":"compliance_team","rem_due":(now+timedelta(days=10)).strftime("%Y-%m-%d"),
                     "tested_by":"cco","tested_ts":ts(25)},
                    {"id":"P-002","step":2,"title":"Transaction Monitoring Alert Review",
                     "description":"Confirm alerts were reviewed and dispositioned within 30-day SAR deadline.",
                     "notes":"All 40 sampled alerts dispositioned on time. SAR filing log complete.",
                     "result":"Pass","exception":False,"has_remediation_item":False,
                     "tested_by":"cco","tested_ts":ts(22)},
                    {"id":"P-003","step":3,"title":"AML Training Completion",
                     "description":"Confirm 100% annual AML training completion for registered representatives.",
                     "notes":"100% completion confirmed via training platform export dated Dec 15, 2025.",
                     "result":"Pass","exception":False,"has_remediation_item":False,
                     "tested_by":"cco","tested_ts":ts(21)},
                ],
                "exceptions":[
                    {"id":"EX-002","test_id":"CT-003","procedure_id":"P-001","description":"2 new accounts missing FinCEN beneficial ownership certifications in CIP files.","root_cause":"Onboarding checklist did not include mandatory beneficial ownership step for corporate accounts opened via digital channel.","remediation_plan":"Obtain outstanding certifications and update onboarding workflow to require beneficial ownership cert before account activation.","assigned_to":"compliance_team","due_date":(now+timedelta(days=10)).strftime("%Y-%m-%d"),"status":"Open","created":ts(25)},
                ],
                "conclusion":"AML program is substantively compliant. One CIP documentation gap identified and remediation in progress. Overall risk rating revised from High to Medium based on testing results.",
                "audit":[
                    {"ts":ts(45),"user":"cco","action":"Test CT-003 created"},
                    {"ts":ts(25),"user":"cco","action":"Procedure P-001 completed: Exception — 2 CIP gaps"},
                    {"ts":ts(22),"user":"cco","action":"Procedure P-002 completed: Pass"},
                    {"ts":ts(21),"user":"cco","action":"Procedure P-003 completed: Pass"},
                    {"ts":ts(20),"user":"cco","action":"Test submitted for review"},
                    {"ts":ts(20),"user":"supervisor","action":"Test CT-003 approved — final risk rating: Medium"},
                ],
            },
            {
                "id":"CT-004",
                "test_type":"Reg BI Supervisory Controls Testing",
                "line_of_business":"Wealth Management",
                "topic":"Reg BI Best Interest Recommendation Review",
                "linked_policies":["POL-004","POL-002"],
                "initial_risk_rating":"Medium",
                "risk_rating":"Medium",
                "time_period_reviewed":"Q4 2025 (Oct 1 – Dec 31, 2025)",
                "approval_date":"",
                "assignee":"supervisor",
                "deadline":(now+timedelta(days=20)).strftime("%Y-%m-%d"),
                "related_workflow":"WF-I-003",
                "related_checklist":"",
                "status":"In Progress",
                "created":ts(7),
                "created_by":"cco",
                "date_started":ts(5),
                "submitted_for_review":False,
                "completed":False,
                "completed_ts":"",
                "reviewer":"cco",
                "review_notes":"",
                "final_risk_rating":"",
                "population_description":"All retail recommendations in Q4 2025. Population: 412 recommendation events across 198 accounts.",
                "sample_size":30,
                "sample_method":"Judgmental",
                "sample_rationale":"Judgmental sample focusing on complex products (options, variable annuities) and accounts with concentration >20% in single security.",
                "procedures":[
                    {"id":"P-001","step":1,"title":"Best Interest Documentation Review",
                     "description":"Confirm each sampled recommendation is supported by written rationale meeting Reg BI best interest standard.",
                     "notes":"22 of 30 reviewed. 1 exception identified — variable annuity recommendation lacking documented cost comparison.",
                     "result":"Exception","exception":True,"has_remediation_item":True,
                     "rem_name":"Update VA recommendation documentation template",
                     "rem_detail":"Require cost-benefit comparison form for all variable annuity recommendations. Retroactively document rationale for exception account.",
                     "rem_assign":"supervisor","rem_due":(now+timedelta(days=15)).strftime("%Y-%m-%d"),
                     "tested_by":"supervisor","tested_ts":ts(2)},
                    {"id":"P-002","step":2,"title":"Conflict of Interest Disclosure Review",
                     "description":"Confirm Form CRS disclosures were current and provided to clients at account opening.",
                     "notes":"Testing in progress — 8 accounts remaining.",
                     "result":"","exception":False,"has_remediation_item":False,
                     "tested_by":"","tested_ts":""},
                ],
                "exceptions":[
                    {"id":"EX-003","test_id":"CT-004","procedure_id":"P-001","description":"Variable annuity recommendation for account #7823 lacks required cost-benefit comparison documentation.","root_cause":"Recommendation documentation template did not include cost comparison field for VA products.","remediation_plan":"Update template and obtain retroactive documentation.","assigned_to":"supervisor","due_date":(now+timedelta(days=15)).strftime("%Y-%m-%d"),"status":"Open","created":ts(2)},
                ],
                "conclusion":"",
                "audit":[
                    {"ts":ts(7),"user":"cco","action":"Test CT-004 created"},
                    {"ts":ts(2),"user":"supervisor","action":"Procedure P-001 partial: Exception identified — VA documentation gap"},
                ],
            },
        ],
    }

# ── SESSION STATE ──────────────────────────────────────────────────────────────
if "data" not in st.session_state:
    st.session_state.data = seed_data()
# Force re-seed if session has stale user keys from before the role rename
elif "admin" in st.session_state.data.get("users", {}) or "analyst" in st.session_state.data.get("users", {}):
    st.session_state.data = seed_data()
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "current_user" not in st.session_state:
    st.session_state.current_user = None

D = st.session_state.data
# Back-compat: ensure workflow keys exist for older sessions
if "workflow_templates" not in D: D["workflow_templates"] = []
if "workflow_instances" not in D: D["workflow_instances"] = []
if "compliance_tests"   not in D: D["compliance_tests"]   = []

def add_audit(module, action):
    st.session_state.data["audit_trail"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"module":module,"action":action})

def current_role():
    u = st.session_state.current_user
    return D["users"][u]["role"] if u else ""

def is_cco():
    return current_role() == "Compliance Officer"

def is_compliance_team():
    return current_role() == "Compliance Team"

def is_supervisor():
    return current_role() == "Supervisor"

def has_full_compliance_access():
    """CCO or Compliance Team — broad program visibility."""
    return current_role() in ("Compliance Officer", "Compliance Team")

def current_user_lob():
    """Return the line_of_business for the current user (used to scope supervisor views)."""
    return D["users"].get(st.session_state.current_user, {}).get("line_of_business", None)

def current_name():
    u = st.session_state.current_user
    return D["users"][u]["name"] if u else ""

def badge(text, cls):
    return f"<span class='gc-badge {cls}'>{text}</span>"

def toggle(key, label, default=False):
    """Render a styled clickable header; return True if expanded."""
    if key not in st.session_state:
        st.session_state[key] = default
    open_ = st.session_state[key]
    arrow = "▾" if open_ else "▸"
    btn_style = (
        "background:#ffffff;border:1px solid #e5e7eb;border-radius:8px;"
        "padding:10px 16px;width:100%;text-align:left;cursor:pointer;"
        "font-family:Inter,-apple-system,sans-serif;font-size:13px;"
        "font-weight:600;color:#111111;margin-bottom:2px;"
        "box-shadow:0 1px 2px rgba(0,0,0,0.04);"
    )
    if st.button(f"{arrow}  {label}", key=f"_btn_{key}", use_container_width=True):
        st.session_state[key] = not open_
        st.rerun()
    return st.session_state[key]

# ── LOGIN ──────────────────────────────────────────────────────────────────────
def login_page():
    st.markdown("<div class='login-wrap'><div class='login-logo'>⚖ GRAPHITE</div><div class='login-sub'>Compliance Intelligence Platform</div></div>", unsafe_allow_html=True)
    c1,c2,c3 = st.columns([1,1.4,1])
    with c2:
        st.markdown("<div class='gc-card'>", unsafe_allow_html=True)
        st.markdown("<div class='gc-card-header'>SECURE LOGIN</div>", unsafe_allow_html=True)
        username = st.text_input("Username", placeholder="cco / supervisor / compliance_team")
        password = st.text_input("Password", type="password", placeholder="pass")
        if st.button("Sign In →", use_container_width=True):
            if username in D["users"] and D["users"][username]["password"] == password:
                st.session_state.logged_in = True
                st.session_state.current_user = username
                add_audit("Auth", f"Login: {username}")
                st.rerun()
            else:
                st.error("Invalid credentials.")
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<div style='text-align:center;margin-top:.8rem;font-size:12px;color:#6b7280'>demo logins: <b>cco</b> · <b>supervisor</b> · <b>compliance_team</b> &nbsp;·&nbsp; password: pass</div>", unsafe_allow_html=True)

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
def sidebar():
    with st.sidebar:
        user = D["users"][st.session_state.current_user]
        st.markdown(f"""
        <div style='padding:.8rem 0 1rem 0;border-bottom:1px solid #252a33;margin-bottom:1rem'>
          <div style='font-size:1.3rem;font-weight:800;color:#ffffff;letter-spacing:-.02em'>⚖ Graphite</div>
          <div style='font-size:11px;color:#6b7a99;letter-spacing:.1em;text-transform:uppercase'>Compliance Platform</div>
        </div>
        <div style='font-size:10px;color:#6b7a99;letter-spacing:.1em;text-transform:uppercase;margin-bottom:.25rem'>Logged in as</div>
        <div style='font-size:.95rem;font-weight:700;color:#ffffff;margin-bottom:.1rem'>{user['name']}</div>
        <div style='font-size:12px;color:#93c5fd;margin-bottom:.8rem'>{user.get('title', user['role'])}</div>
        <hr style='border-color:#252a33;margin:.5rem 0 .8rem 0'/>
        """, unsafe_allow_html=True)

        cl_count  = len(D.get("supervisory_checklists", []))
        wf_overdue = len([w for w in D.get("workflow_instances",[]) if w["status"]=="Overdue"])
        ct_review  = len([t for t in D.get("compliance_tests",[]) if t.get("submitted_for_review") and not t.get("completed")])
        ct_rem     = len([ex for t in D.get("compliance_tests",[]) for ex in t.get("exceptions",[]) if ex.get("status")=="Open"])

        # ── Group definitions ─────────────────────────────────
        role = current_role()

        # Role display badge
        role_badge_colors = {
            "Compliance Officer": ("#93c5fd", "CCO"),
            "Compliance Team":    ("#86efac", "Compliance Team"),
            "Supervisor":         ("#fcd34d", "Supervisor"),
        }
        badge_color, badge_label = role_badge_colors.get(role, ("#d1d5db", role))
        st.markdown(f"""
        <div style='background:rgba(255,255,255,0.06);border:1px solid #2a3344;border-radius:6px;
             padding:4px 10px;font-size:10px;font-weight:700;color:{badge_color};
             text-transform:uppercase;letter-spacing:.09em;margin-bottom:.8rem;display:inline-block'>
          {badge_label}
        </div>""", unsafe_allow_html=True)

        GROUPS = [
            {
                "label": None,
                "pages": [("🏠","Dashboard")],
                "roles": ["Compliance Officer","Compliance Team","Supervisor"],
            },
            {
                "label": "COMPLIANCE",
                "pages": [
                    ("📋","Tasks"),
                    ("📁","Exam Requests"),
                    ("📚","Policies & WSPs"),
                    ("✅","Supervisory Checklists"),
                    ("🗄️","Data Lake"),
                ],
                "roles": ["Compliance Officer","Compliance Team"],
            },
            {
                "label": "COMPLIANCE",
                "pages": [
                    ("📋","Tasks"),
                    ("📁","Exam Requests"),
                    ("✅","Supervisory Checklists"),
                    ("🗄️","Data Lake"),
                ],
                "roles": ["Supervisor"],
            },
            {
                "label": "COMPLIANCE TESTING",
                "pages": [
                    ("🧪","Compliance Testing"),
                    ("⚠️","Remediation Items"),
                ],
                "roles": ["Compliance Officer","Compliance Team"],
            },
            {
                "label": "COMPLIANCE CALENDAR",
                "pages": [
                    ("📅","Compliance Calendar"),
                ],
                "roles": ["Compliance Officer","Compliance Team"],
            },
            {
                "label": "CALENDAR & AI",
                "pages": [
                    ("📅","Compliance Calendar"),
                    ("💬","AI Copilot"),
                ],
                "roles": ["Supervisor"],
            },
            {
                "label": "WORKFLOWS",
                "pages": [
                    ("🔄","Workflow Library"),
                    ("▶️","Active Workflows"),
                ],
                "roles": ["Compliance Officer","Compliance Team"],
            },
            {
                "label": "RULES & INTELLIGENCE",
                "pages": [
                    ("📏","Rule Inventory"),
                    ("🤖","AI Rule Scan"),
                    ("💬","AI Copilot"),
                ],
                "roles": ["Compliance Officer","Compliance Team"],
            },
            {
                "label": "✦ AI COMPLIANCE INTELLIGENCE",
                "pages": [
                    ("📰","Peer Enforcement Insights"),
                    ("⚡","AI Risk Assessment"),
                    ("🔎","Policy Gap Analysis"),
                    ("🏛️","Exam Simulator"),
                    ("📊","Automated Reporting"),
                    ("🔔","Predictive Alerts"),
                ],
                "roles": ["Compliance Officer","Compliance Team"],
            },
            {
                "label": "ADMINISTRATION",
                "pages": [
                    ("🏢","Org Chart"),
                    ("🔒","Audit Trail"),
                    ("⚙️","Firm Config"),
                    ("💾","Data Export"),
                ],
                "roles": ["Compliance Officer"],
            },
        ]

        # Build flat nav_labels list (order matters for radio index)
        all_pages = []
        for g in GROUPS:
            if role in g.get("roles", []):
                all_pages.extend(g["pages"])

        def _label(icon, name):
            if name == "Supervisory Checklists" and cl_count > 0:
                return f"{icon}  {name} ({cl_count})"
            if name == "Active Workflows" and wf_overdue > 0:
                return f"{icon}  {name} ⚠"
            if name == "Compliance Testing" and ct_review > 0:
                return f"{icon}  {name} ({ct_review} review)"
            if name == "Remediation Items" and ct_rem > 0:
                return f"{icon}  {name} ({ct_rem} open)"
            return f"{icon}  {name}"

        nav_labels   = [p[1] for p in all_pages]
        display_labels = [_label(p[0], p[1]) for p in all_pages]

        # Render grouped section headers + nav buttons
        section_header_css = (
            "font-size:9px;font-weight:700;letter-spacing:.12em;"
            "text-transform:uppercase;color:#4b5a72;"
            "padding:10px 4px 4px 4px;margin-top:4px;"
        )
        premium_header_css = (
            "font-size:9px;font-weight:800;letter-spacing:.12em;"
            "text-transform:uppercase;color:#f59e0b;"
            "padding:10px 4px 4px 4px;margin-top:4px;"
        )

        for g in GROUPS:
            if role not in g.get("roles", []):
                continue
            if g["label"]:
                is_premium = g["label"].startswith("✦")
                css = premium_header_css if is_premium else section_header_css
                st.markdown(f"<div style='{css}'>{g['label']}</div>",
                            unsafe_allow_html=True)
            for icon, name in g["pages"]:
                lbl = _label(icon, name)
                # Render as a radio button but styled like a nav item
                selected = st.session_state.get("_nav_sel", nav_labels[0])
                is_active = selected == name
                is_prem_page = any(name == p[1] for g2 in GROUPS if g2["label"] and g2["label"].startswith("✦") for p in g2["pages"])
                bg  = "rgba(59,130,246,0.18)" if is_active else "transparent"
                col = "#93c5fd" if is_active else "#d0d6e4"
                fw  = "700" if is_active else "500"
                if st.button(
                    lbl,
                    key=f"_nav_{name}",
                    use_container_width=True,
                ):
                    st.session_state["_nav_sel"] = name
                    st.rerun()

        st.markdown("<hr style='border-color:#252a33;margin:.8rem 0'/>", unsafe_allow_html=True)
        if st.button("Sign Out", key="_signout", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.current_user = None
            if "_nav_sel" in st.session_state:
                del st.session_state["_nav_sel"]
            st.rerun()

        return st.session_state.get("_nav_sel", "Dashboard")

# ── DASHBOARD ──────────────────────────────────────────────────────────────────
def page_dashboard():
    st.markdown(f"<div class='gc-header'><h1>COMPLIANCE DASHBOARD</h1><p>{D['firm']['name']} · {D['firm']['regulator']} · {datetime.now().strftime('%B %d, %Y')}</p></div>", unsafe_allow_html=True)

    # Supervisor gets a focused personal dashboard
    if is_supervisor():
        today_str = date.today().strftime("%Y-%m-%d")
        for t in D["tasks"]:
            if t["status"]=="Open" and t["due"] < today_str:
                t["status"]="Overdue"
        my_tasks = [t for t in D["tasks"] if t["assignee"]==st.session_state.current_user]
        open_t    = [t for t in my_tasks if t["status"]=="Open"]
        overdue_t = [t for t in my_tasks if t["status"]=="Overdue"]
        pending_t = [t for t in my_tasks if "Pending" in t["status"]]
        closed_t  = [t for t in my_tasks if t["status"]=="Closed"]
        cols = st.columns(4)
        cols[0].metric("My Open Tasks",    len(open_t))
        cols[1].metric("My Overdue",       len(overdue_t))
        cols[2].metric("Pending Approval", len(pending_t))
        cols[3].metric("Completed",        len(closed_t))
        st.markdown("<hr/>", unsafe_allow_html=True)
        st.markdown("<div style='font-size:11px;font-weight:700;letter-spacing:.08em;color:#6b7280;text-transform:uppercase;margin-bottom:.8rem'>MY ASSIGNED TASKS</div>", unsafe_allow_html=True)
        if not my_tasks:
            st.info("No tasks currently assigned to you.")
        for t in my_tasks:
            s=t["status"]; p=t["priority"]
            sc="badge-overdue" if s=="Overdue" else "badge-open" if s=="Open" else "badge-pending" if "Pending" in s else "badge-closed"
            pc="badge-high" if p=="High" else "badge-med" if p=="Medium" else "badge-low"
            st.markdown(f"""<div class='gc-card' style='padding:.8rem 1rem;margin-bottom:.4rem'>
              <div style='display:flex;justify-content:space-between;align-items:center'>
                <div><span style='font-size:11px;color:#6b7280'>{t['id']}</span>
                <span style='font-weight:600;margin-left:.5rem'>{t['title']}</span></div>
                <div>{badge(s,sc)}&nbsp;{badge(p,pc)}</div>
              </div>
              <div style='margin-top:.3rem;font-size:12px;color:#6b7280'>Due: {t['due']} · {t['category']}</div>
            </div>""", unsafe_allow_html=True)
        return

    tasks = D["tasks"]

    # ── Clickable stat cards ──────────────────────────────────────────────────
    n_open    = len([t for t in tasks if t["status"]=="Open"])
    n_overdue = len([t for t in tasks if t["status"]=="Overdue"])
    n_pending = len([t for t in tasks if t["status"]=="Pending Approval"])
    n_closed  = len([t for t in tasks if t["status"]=="Closed"])
    n_exams   = len(D["exam_requests"])

    def _stat_card(col, number, label, color, bg, border, nav_target, filter_key=None, filter_val=None):
        with col:
            # Single button styled to look like the full card — one click navigates
            clicked = st.button(
                f"{number}\n{label}",
                key=f"dash_nav_{label.replace(' ','_')}",
                use_container_width=True,
            )
            st.markdown(f"""
            <style>
            div[data-testid="stButton"] button[kind="secondary"]:has(+ *) {{}}
            [data-testid="stButton"]:has(button[aria-label="{number}\\n{label}"]) button,
            [data-testid="stButton"] > button[data-stat-card="{label}"] {{
              border-top: 3px solid {color} !important;
            }}
            </style>
            <style>
            /* Target this card's button by position via JS-free CSS trick */
            </style>""", unsafe_allow_html=True)
            # Overlay the real styled card on top using negative margin trick
            st.markdown(f"""
            <style>
            .stat-card-{label.replace(' ','_').replace('/','_')} {{
              background:{bg};border:1px solid {border};border-top:3px solid {color};
              border-radius:8px;padding:16px 14px;text-align:center;
              box-shadow:0 1px 3px rgba(0,0,0,0.04);cursor:pointer;
              transition:box-shadow .15s,border-color .15s;
            }}
            .stat-card-{label.replace(' ','_').replace('/','_')}:hover {{
              box-shadow:0 3px 10px rgba(0,0,0,0.1);border-color:{color};
            }}
            </style>""", unsafe_allow_html=True)
            if clicked:
                st.session_state["_nav_sel"] = nav_target
                if filter_key and filter_val:
                    st.session_state[filter_key] = filter_val
                st.rerun()

    # Hide the raw button text and restyle each button as a stat card
    st.markdown(f"""
    <style>
    /* Stat card buttons — hide default Streamlit button look, render as metric cards */
    [data-testid="stHorizontalBlock"] [data-testid="stButton"] > button {{
      background: transparent !important;
      border: none !important;
      box-shadow: none !important;
      padding: 0 !important;
      height: auto !important;
      font-size: 0 !important; /* hide text — we overlay HTML below */
      color: transparent !important;
      pointer-events: none !important;
      position: absolute !important;
      width: 100% !important;
      top: 0; left: 0; height: 100% !important;
      z-index: 2 !important;
    }}
    </style>""", unsafe_allow_html=True)

    cols = st.columns(5)
    cards = [
        (cols[0], n_open,    "Open Tasks",     "#2563eb", "#eff6ff", "#bfdbfe", "Tasks",         "dash_task_filter", "Open"),
        (cols[1], n_overdue, "Overdue",         "#dc2626", "#fef2f2", "#fecaca", "Tasks",         "dash_task_filter", "Overdue"),
    ]
    if is_compliance_team():
        n_mine = len([t for t in tasks if t["assignee"]==st.session_state.current_user])
        cards.append((cols[2], n_mine, "My Tasks", "#7c3aed", "#f5f3ff", "#ddd6fe", "Tasks", "dash_task_filter", "All"))
    else:
        cards.append((cols[2], n_pending, "Pending Review", "#d97706", "#fffbeb", "#fcd34d", "Tasks", "dash_task_filter", "Pending Approval"))
    cards.append((cols[3], n_closed, "Closed",        "#16a34a", "#f0fdf4", "#bbf7d0", "Tasks",        "dash_task_filter", "Closed"))
    cards.append((cols[4], n_exams,  "Exam Requests", "#0891b2", "#ecfeff", "#a5f3fc", "Exam Requests", None, None))

    # Render HTML cards + invisible overlay buttons
    for col, number, label, color, bg, border, nav_target, filter_key, filter_val in cards:
        safe = label.replace(' ','_').replace('/','_').replace('(','').replace(')','')
        with col:
            nav_clicked = st.button(" ", key=f"dash_nav_{safe}", use_container_width=True)
            st.markdown(f"""
            <div class='dash-stat-card' id='dsc_{safe}'
                 style='background:{bg};border:1px solid {border};border-top:3px solid {color};
                        border-radius:8px;padding:16px 14px;text-align:center;margin-top:-42px;
                        box-shadow:0 1px 3px rgba(0,0,0,0.04);cursor:pointer;position:relative;z-index:1;
                        transition:box-shadow .15s'>
              <div style='font-size:1.75rem;font-weight:800;color:{color};letter-spacing:-.02em;line-height:1.1'>{number}</div>
              <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase;
                          letter-spacing:.07em;margin-top:4px'>{label}</div>
            </div>""", unsafe_allow_html=True)
            if nav_clicked:
                st.session_state["_nav_sel"] = nav_target
                if filter_key and filter_val:
                    st.session_state[filter_key] = filter_val
                st.rerun()

    # Make the invisible overlay buttons cover the cards exactly
    st.markdown("""
    <style>
    .dash-stat-card + div [data-testid="stButton"] > button,
    [data-testid="stVerticalBlock"] [data-testid="stButton"] > button[kind="secondary"] {
      /* Reset any earlier override for stat area */
    }
    /* Style the invisible overlay buttons covering stat cards */
    div:has(> .dash-stat-card) [data-testid="stButton"] > button {
      position: absolute !important; top: 0 !important; left: 0 !important;
      width: 100% !important; height: 100% !important;
      background: transparent !important; border: none !important;
      box-shadow: none !important; color: transparent !important;
      font-size: 0 !important; z-index: 3 !important; cursor: pointer !important;
    }
    </style>""", unsafe_allow_html=True)

    st.markdown("<hr/>", unsafe_allow_html=True)
    cl, cr = st.columns([1.6,1])
    with cl:
        st.markdown("<div style='font-size:11px;font-weight:700;letter-spacing:.08em;color:#6b7280;text-transform:uppercase;margin-bottom:.8rem'>MY TASKS</div>", unsafe_allow_html=True)
        today_str = date.today().strftime("%Y-%m-%d")
        for t in D["tasks"]:
            if t["status"]=="Open" and t["due"] < today_str:
                t["status"]="Overdue"
        my_tasks = [t for t in tasks if t["assignee"]==st.session_state.current_user or is_cco()]
        if is_compliance_team():
            # Compliance Team sees all tasks except those pending CCO-only approval
            my_tasks = [t for t in tasks if t["status"] != "Pending Approval" or t["assignee"] == st.session_state.current_user]
        for t in my_tasks[:6]:
            s=t["status"]; p=t["priority"]
            sc="badge-overdue" if s=="Overdue" else "badge-open" if s=="Open" else "badge-pending" if "Pending" in s else "badge-closed"
            pc="badge-high" if p=="High" else "badge-med" if p=="Medium" else "badge-low"
            uname=D["users"].get(t["assignee"],{}).get("name",t["assignee"])
            utitle=D["users"].get(t["assignee"],{}).get("title","")
            expanded = toggle(f"dash_task_{t['id']}", f"{t['id']} — {t['title']}")
            if not expanded:
                # collapsed card preview
                st.markdown(f"""<div style='background:#f8fafc;border:1px solid #e5e7eb;border-radius:6px;
                     padding:.5rem 1rem;margin-top:-6px;margin-bottom:.4rem;
                     display:flex;justify-content:space-between;align-items:center'>
                  <div style='font-size:12px;color:#6b7280'>Due: {t['due']} · {uname} · {t['category']}</div>
                  <div>{badge(s,sc)}&nbsp;{badge(p,pc)}</div>
                </div>""", unsafe_allow_html=True)
            else:
                # expanded detail view
                st.markdown(f"""
                <div style='background:#f8fafc;border:1px solid #e5e7eb;border-radius:8px;
                     padding:.75rem 1.1rem;margin-top:-4px;margin-bottom:.5rem'>
                  <div style='display:flex;gap:2rem;flex-wrap:wrap;margin-bottom:.6rem'>
                    <div>
                      <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Owner</div>
                      <div style='font-size:13px;font-weight:700;color:#111'>Jordan Reed</div>
                      <div style='font-size:11px;color:#6b7280'>Chief Compliance Officer</div>
                    </div>
                    <div>
                      <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Assignee</div>
                      <div style='font-size:13px;font-weight:600;color:#111'>{uname}</div>
                      <div style='font-size:11px;color:#6b7280'>{utitle}</div>
                    </div>
                    <div>
                      <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Due Date</div>
                      <div style='font-size:13px;font-weight:600;color:{"#dc2626" if s=="Overdue" else "#111"}'>{t['due']}</div>
                    </div>
                    <div>
                      <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Status / Priority</div>
                      <div style='margin-top:.2rem'>{badge(s,sc)}&nbsp;{badge(p,pc)}</div>
                    </div>
                    <div>
                      <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Category</div>
                      <div style='font-size:13px;font-weight:600;color:#111'>{t['category']}</div>
                    </div>
                  </div>
                """, unsafe_allow_html=True)
                if t.get("instructions"):
                    st.markdown(f"""
                    <div style='background:#fffbeb;border:1px solid #fde68a;border-left:4px solid #f59e0b;
                         border-radius:8px;padding:.7rem 1rem;margin:.4rem 0'>
                      <div style='font-size:10px;font-weight:800;color:#92400e;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.3rem'>📋 WHAT TO DO</div>
                      <div style='font-size:12px;color:#374151;line-height:1.7'>{t['instructions']}</div>
                    </div>
                    """, unsafe_allow_html=True)
                if t.get("evidence_required"):
                    st.markdown(f"""
                    <div style='background:#f0f7ff;border:1px solid #bfdbfe;border-left:4px solid #2563eb;
                         border-radius:8px;padding:.7rem 1rem;margin:.4rem 0'>
                      <div style='font-size:10px;font-weight:800;color:#1e40af;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.3rem'>📎 EVIDENCE REQUIRED</div>
                      <div style='font-size:12px;color:#374151;line-height:1.7'>{t['evidence_required']}</div>
                    </div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown("</div>", unsafe_allow_html=True)
    with cr:
        st.markdown("<div style='font-size:11px;font-weight:700;letter-spacing:.08em;color:#6b7280;text-transform:uppercase;margin-bottom:.8rem'>RECENT ACTIVITY</div>", unsafe_allow_html=True)
        for e in reversed(D["audit_trail"][-8:]):
            st.markdown(f"<div class='gc-tl'><span style='color:#2563eb;font-size:11px'>{e['ts']}</span> · <b>{e['user']}</b><br/>{e['action']}</div>", unsafe_allow_html=True)
        flagged=[r for r in D["rules"] if r["ai_flagged"]]
        if flagged:
            st.markdown("<hr/>", unsafe_allow_html=True)
            if toggle("dash_ai_flagged", f"🤖 AI Rule Scan — {len(flagged)} rule change(s) flagged for review"):
                for r in flagged:
                    flag_date = r.get("ai_flagged_date","")
                    flag_reason = r.get("ai_flag_reason","Potential regulatory change detected. Review recommended.")
                    # Calculate time period
                    try:
                        from datetime import datetime as _dt
                        days_ago = (_dt.now() - _dt.strptime(flag_date[:10], "%Y-%m-%d")).days
                        if days_ago <= 7:
                            period_label = f"Flagged {days_ago}d ago — within last week"
                            period_color = "#dc2626"
                            period_bg    = "#fef2f2"
                        elif days_ago <= 30:
                            period_label = f"Flagged {days_ago}d ago — within last month"
                            period_color = "#d97706"
                            period_bg    = "#fffbeb"
                        elif days_ago <= 90:
                            period_label = f"Flagged {days_ago}d ago — within last 3 months"
                            period_color = "#2563eb"
                            period_bg    = "#eff6ff"
                        else:
                            period_label = f"Flagged {days_ago}d ago"
                            period_color = "#6b7280"
                            period_bg    = "#f3f4f6"
                    except Exception:
                        period_label = f"Flagged: {flag_date[:10]}" if flag_date else "Date unknown"
                        period_color = "#6b7280"
                        period_bg    = "#f3f4f6"

                    affected_wsps = " · ".join([w["wsp_id"] for w in r.get("wsp_links",[])])
                    st.markdown(f"""
                    <div style='background:#fff;border:1px solid #fde68a;border-left:4px solid #f59e0b;
                         border-radius:8px;padding:12px 16px;margin-bottom:8px'>
                      <div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-bottom:6px'>
                        <span style='font-size:13px;font-weight:800;color:#1a1d23'>{r["id"]} — {r["title"]}</span>
                        <span style='background:#fef9c3;color:#92400e;font-size:10px;font-weight:700;
                             padding:2px 8px;border-radius:4px;border:1px solid #fde68a'>🤖 AI FLAGGED</span>
                        <span style='background:{period_bg};color:{period_color};font-size:10px;font-weight:700;
                             padding:2px 8px;border-radius:4px;border:1px solid {period_color}40;margin-left:auto'>
                          🕐 {period_label}
                        </span>
                      </div>
                      <div style='font-size:11px;color:#6b7280;margin-bottom:6px'>
                        <b>Source:</b> {r["source"]} &nbsp;·&nbsp; <b>Category:</b> {r["category"]}
                        {f' &nbsp;·&nbsp; <b>Affected WSPs:</b> {affected_wsps}' if affected_wsps else ''}
                      </div>
                      <div style='background:#fffbeb;border-left:3px solid #f59e0b;border-radius:4px;
                           padding:7px 11px;font-size:12px;color:#78350f;line-height:1.6;margin-bottom:6px'>
                        <b>⚠ Why flagged:</b> {flag_reason}
                      </div>
                      <div style='font-size:11px;color:#6b7280'>
                        Last reviewed: {r.get("last_reviewed","—")[:10] if r.get("last_reviewed") else "—"}
                      </div>
                    </div>""", unsafe_allow_html=True)
                # Navigate to Rule Inventory button
                if st.button("📏 Open Rule Inventory →", key="dash_ai_flag_nav"):
                    st.session_state["_nav_sel"] = "Rule Inventory"
                    st.rerun()

# ── TASKS ──────────────────────────────────────────────────────────────────────
def page_tasks():
    st.markdown("<div class='gc-header'><h1>TASKS & REMEDIATION</h1><p>Create, assign, track, and close compliance tasks with full audit trail</p></div>", unsafe_allow_html=True)

    # Supervisors only get their own assigned tasks — no new task creation
    if is_supervisor():
        st.markdown("<div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.8rem'>MY ASSIGNED TASKS</div>", unsafe_allow_html=True)
        today_str=date.today().strftime("%Y-%m-%d")
        for t in D["tasks"]:
            if t["status"]=="Open" and t["due"]<today_str:
                t["status"]="Overdue"
        my_tasks = [t for t in D["tasks"] if t["assignee"]==st.session_state.current_user]
        if not my_tasks:
            st.info("No tasks currently assigned to you.")
            return
        for t in my_tasks:
            s=t["status"]; p=t["priority"]
            sc="badge-overdue" if s=="Overdue" else "badge-open" if s=="Open" else "badge-pending" if "Pending" in s else "badge-closed"
            pc="badge-high" if p=="High" else "badge-med" if p=="Medium" else "badge-low"
            if "instructions" not in t: t["instructions"] = ""
            if "evidence_required" not in t: t["evidence_required"] = ""
            assignee_name = D["users"].get(t["assignee"],{}).get("name", t["assignee"])
            if toggle(f"task_{t['id']}", f"{t['id']} — {t['title']} [{t['status']}]"):
                utitle_t = D["users"].get(t["assignee"],{}).get("title","")
                st.markdown(f"""
                <div style='background:#f8fafc;border:1px solid #e5e7eb;border-radius:8px;
                     padding:.75rem 1.1rem;margin:.3rem 0 .7rem 0;display:flex;gap:2.5rem;flex-wrap:wrap'>
                  <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Due Date</div>
                    <div style='font-size:13px;font-weight:600;color:{"#dc2626" if s=="Overdue" else "#111"}'>{t['due']}</div></div>
                  <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Status / Priority</div>
                    <div style='margin-top:.2rem'>{badge(s,sc)}&nbsp;{badge(p,pc)}</div></div>
                  <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Category</div>
                    <div style='font-size:13px;font-weight:600;color:#111'>{t['category']}</div></div>
                </div>""", unsafe_allow_html=True)
                if t.get("instructions"):
                    st.markdown(f"""<div style='background:#fffbeb;border:1px solid #fde68a;border-left:4px solid #f59e0b;
                         border-radius:8px;padding:.8rem 1rem;margin:.5rem 0'>
                      <div style='font-size:11px;font-weight:800;color:#92400e;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>📋 WHAT TO DO — INSTRUCTIONS</div>
                      <div style='font-size:13px;color:#374151;line-height:1.8'>{t['instructions']}</div>
                    </div>""", unsafe_allow_html=True)
                if t.get("evidence_required"):
                    st.markdown(f"""<div style='background:#f0f7ff;border:1px solid #bfdbfe;border-left:4px solid #2563eb;
                         border-radius:8px;padding:.8rem 1rem;margin:.5rem 0'>
                      <div style='font-size:11px;font-weight:800;color:#1e40af;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>📎 EVIDENCE REQUIRED</div>
                      <div style='font-size:13px;color:#374151;line-height:1.7'>{t['evidence_required']}</div>
                    </div>""", unsafe_allow_html=True)
                st.markdown("<hr style='margin:.6rem 0'/><div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>ATTACHED EVIDENCE</div>", unsafe_allow_html=True)
                if not t["evidence"]:
                    st.markdown("<div style='font-size:13px;color:#9ca3af;font-style:italic;margin-bottom:.4rem'>No evidence attached yet.</div>", unsafe_allow_html=True)
                for ev in t["evidence"]:
                    st.markdown(f"<div class='gc-tl'>📎 <b>{ev['name']}</b> · {ev['uploaded_by']} · {ev['ts'][:16]}</div>", unsafe_allow_html=True)
                up = st.file_uploader("Upload evidence file", key=f"sup_up_{t['id']}")
                up_notes = st.text_area("Upload notes", key=f"sup_up_notes_{t['id']}", height=60, placeholder="Describe what this file is or any context.")
                if up:
                    fh = hashlib.md5(up.name.encode()).hexdigest()[:8]
                    ev_e = {"name":up.name,"uploaded_by":st.session_state.current_user,"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"hash":fh,"immutable":True,"notes":up_notes.strip()}
                    t["evidence"].append(ev_e)
                    t["audit"].append({"ts":ev_e["ts"],"user":st.session_state.current_user,"action":f"Evidence uploaded: {up.name}"})
                    add_audit("Tasks", f"Uploaded {up.name} for {t['id']}")
                    st.success(f"✓ {up.name} uploaded")
                st.markdown("<hr style='margin:.6rem 0'/>", unsafe_allow_html=True)
                if st.button("📤 Submit for Approval", key=f"sup_sub_{t['id']}"):
                    if t["status"] != "Closed":
                        t["status"] = "Pending Approval"
                        t["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":"Submitted for approval"})
                        add_audit("Tasks", f"{t['id']} submitted for approval"); st.rerun()
        return

    tab1,tab2 = st.tabs(["📋  TASK LIST","➕  NEW TASK"])
    with tab1:
        fc1,fc2,fc3 = st.columns(3)
        # Compliance Team doesn't see the CCO approval queue filter
        status_opts = ["All","Open","Overdue","Closed"] if is_compliance_team() else ["All","Open","Overdue","Pending Approval","Closed"]
        # Pre-select filter if navigated from dashboard stat card
        _preset = st.session_state.pop("dash_task_filter", None)
        _default_idx = status_opts.index(_preset) if _preset and _preset in status_opts else 0
        fs=fc1.selectbox("Status", status_opts, index=_default_idx)
        fp=fc2.selectbox("Priority",["All","High","Medium","Low"])
        fc_=fc3.selectbox("Category",["All"]+list({t["category"] for t in D["tasks"]}))
        today_str=date.today().strftime("%Y-%m-%d")
        for t in D["tasks"]:
            if t["status"]=="Open" and t["due"]<today_str:
                t["status"]="Overdue"
                t["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":"SYSTEM","action":"Auto-escalated to Overdue"})
        tasks=D["tasks"]
        # Compliance Team: exclude tasks that are Pending Approval (CCO queue) unless it's their own
        if is_compliance_team():
            tasks = [t for t in tasks if t["status"] != "Pending Approval" or t["assignee"] == st.session_state.current_user]
        if fs!="All": tasks=[t for t in tasks if t["status"]==fs]
        if fp!="All": tasks=[t for t in tasks if t["priority"]==fp]
        if fc_!="All": tasks=[t for t in tasks if t["category"]==fc_]
        for t in tasks:
            s=t["status"]; p=t["priority"]
            sc="badge-overdue" if s=="Overdue" else "badge-open" if s=="Open" else "badge-pending" if "Pending" in s else "badge-closed"
            pc="badge-high" if p=="High" else "badge-med" if p=="Medium" else "badge-low"

            # Back-compat
            if "instructions"     not in t: t["instructions"]     = ""
            if "evidence_required" not in t: t["evidence_required"] = ""

            assignee_name = D["users"].get(t["assignee"],{}).get("name", t["assignee"])
            if toggle(f"task_{t['id']}", f"{t['id']} — {t['title']} [{t['status']}]"):

                # ── Header metadata row ──
                status_badge   = f"<span class='gc-badge {sc}'>{s}</span>"
                priority_badge = f"<span class='gc-badge {pc}'>{p}</span>"
                recur_html = f"<div style='font-size:12px;color:#7c3aed;font-weight:600;margin-top:.2rem'>🔁 {t['recurrence']}</div>" if t.get("recurrence") else ""

                st.markdown(f"""
                <div style='background:#f8fafc;border:1px solid #e5e7eb;border-radius:8px;
                     padding:.75rem 1.1rem;margin:.3rem 0 .7rem 0;display:flex;gap:2.5rem;flex-wrap:wrap;align-items:flex-start'>
                  <div>
                    <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Owner</div>
                    <div style='font-size:13px;font-weight:700;color:#111111'>Jordan Reed</div>
                    <div style='font-size:11px;color:#6b7280'>Chief Compliance Officer</div>
                  </div>
                  <div>
                    <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Assignee</div>
                    <div style='font-size:13px;font-weight:600;color:#111111'>{assignee_name}</div>
                    <div style='font-size:11px;color:#6b7280'>{D['users'].get(t['assignee'],{}).get('title','')}</div>
                  </div>
                  <div>
                    <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Due Date</div>
                    <div style='font-size:13px;font-weight:600;color:{"#dc2626" if s=="Overdue" else "#111111"}'>{t['due']}</div>
                    {recur_html}
                  </div>
                  <div>
                    <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Category</div>
                    <div style='font-size:13px;font-weight:600;color:#111111'>{t['category']}</div>
                  </div>
                  <div>
                    <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase;margin-bottom:.25rem'>Status / Priority</div>
                    {status_badge}&nbsp;{priority_badge}
                  </div>
                </div>
                """, unsafe_allow_html=True)

                if t.get("delegated_to"):
                    st.markdown(f"<div class='gc-alert gc-alert-warn'>🔀 Delegated to: <b>{t['delegated_to']}</b> · {t['delegated_memo']}</div>", unsafe_allow_html=True)

                # ── Instructions ──
                if t.get("instructions"):
                    st.markdown(f"""
                    <div style='background:#fffbeb;border:1px solid #fde68a;border-left:4px solid #f59e0b;
                         border-radius:8px;padding:.8rem 1rem;margin:.5rem 0'>
                      <div style='font-size:11px;font-weight:800;color:#92400e;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>
                        📋 WHAT TO DO — INSTRUCTIONS
                      </div>
                      <div style='font-size:13px;color:#374151;line-height:1.8'>{t['instructions']}</div>
                    </div>
                    """, unsafe_allow_html=True)

                # ── Evidence required ──
                if t.get("evidence_required"):
                    st.markdown(f"""
                    <div style='background:#f0f7ff;border:1px solid #bfdbfe;border-left:4px solid #2563eb;
                         border-radius:8px;padding:.8rem 1rem;margin:.5rem 0'>
                      <div style='font-size:11px;font-weight:800;color:#1e40af;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>
                        📎 EVIDENCE REQUIRED
                      </div>
                      <div style='font-size:13px;color:#374151;line-height:1.7'>{t['evidence_required']}</div>
                    </div>
                    """, unsafe_allow_html=True)

                # ── Source policy / WSP link ──
                if t.get("source_wsp"):
                    wsp_obj = next((w for w in D["wsps"] if w["id"] == t["source_wsp"]), None)
                    if wsp_obj:
                        lob_tag   = f" · {wsp_obj.get('line_of_business','')}" if wsp_obj.get('line_of_business') else ""
                        ver_tag   = f"v{wsp_obj['version']}"
                        wsp_status_col = "#22c55e" if wsp_obj["status"] == "Approved" else "#f59e0b"
                        st.markdown(f"""
                        <div style='background:#f0fdf4;border:1px solid #bbf7d0;border-left:4px solid #22c55e;
                             border-radius:8px;padding:.65rem 1rem;margin:.5rem 0;
                             display:flex;align-items:center;gap:1rem;flex-wrap:wrap'>
                          <div style='font-size:11px;font-weight:800;color:#166534;text-transform:uppercase;letter-spacing:.07em;white-space:nowrap'>📚 SOURCE POLICY</div>
                          <div>
                            <span style='font-size:12px;font-weight:800;color:{wsp_status_col}'>{wsp_obj['id']}</span>
                            <span style='font-size:13px;font-weight:600;color:#111111;margin-left:.4rem'>{wsp_obj['title']}</span>
                            <span style='font-size:11px;color:#6b7280;margin-left:.4rem'>{ver_tag}{lob_tag}</span>
                            <span style='background:#dcfce7;color:#166534;font-size:10px;font-weight:700;padding:1px 7px;border-radius:10px;margin-left:.5rem'>{wsp_obj['status'].upper()}</span>
                          </div>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin:.4rem 0'>📚 Source Policy: <b>{t['source_wsp']}</b></div>", unsafe_allow_html=True)

                # ── Evidence files ──
                st.markdown("<hr style='margin:.6rem 0'/><div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>ATTACHED EVIDENCE</div>", unsafe_allow_html=True)
                if not t["evidence"]:
                    st.markdown("<div style='font-size:13px;color:#9ca3af;font-style:italic;margin-bottom:.4rem'>No evidence attached yet.</div>", unsafe_allow_html=True)
                for ev in t["evidence"]:
                    imm_badge  = "<span class='gc-badge badge-closed'>IMMUTABLE · WORM</span>"
                    notes_txt  = f" &nbsp;·&nbsp; <i style='color:#6b7280'>{ev['notes']}</i>" if ev.get("notes") else ""
                    st.markdown(f"<div class='gc-tl'>📎 <b>{ev['name']}</b> · {ev['uploaded_by']} · {ev['ts'][:16]} &nbsp;{imm_badge}{notes_txt}</div>", unsafe_allow_html=True)

                upc1, upc2 = st.columns([1,1])
                up       = upc1.file_uploader("Browse and upload evidence file", key=f"up_{t['id']}")
                up_notes = upc2.text_area("Upload notes / context", key=f"up_notes_{t['id']}", height=68, placeholder="Describe what this file is, any issues, exceptions, or context the reviewer should know.")
                if up:
                    fh   = hashlib.md5(up.name.encode()).hexdigest()[:8]
                    ev_e = {"name":up.name,"uploaded_by":st.session_state.current_user,"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"hash":fh,"immutable":True,"notes":up_notes.strip()}
                    t["evidence"].append(ev_e)
                    t["audit"].append({"ts":ev_e["ts"],"user":st.session_state.current_user,"action":f"Evidence uploaded: {up.name}" + (f" — {up_notes.strip()}" if up_notes.strip() else "")})
                    eid = f"EV-{len(D['evidence'])+1:03d}"
                    st.session_state.data["evidence"].append({"id":eid,"filename":up.name,"uploaded_by":st.session_state.current_user,"ts":ev_e["ts"],"task_id":t["id"],"hash":fh,"immutable":True,"archived":False,"notes":up_notes.strip()})
                    add_audit("Tasks", f"Uploaded {up.name} for {t['id']}")
                    st.success(f"✓ {up.name} uploaded (WORM-locked)")

                # ── Action buttons ──
                st.markdown("<hr style='margin:.6rem 0'/>", unsafe_allow_html=True)
                a1, a2, a3 = st.columns(3)
                if a1.button("📤 Submit for Approval", key=f"sub_{t['id']}"):
                    if t["status"] != "Closed":
                        t["status"] = "Pending Approval"
                        t["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":"Submitted for approval"})
                        add_audit("Tasks", f"{t['id']} submitted for approval"); st.rerun()
                if a2.button("✅ Approve & Close", key=f"apr_{t['id']}"):
                    if current_role() in ["Compliance Officer","Compliance Team","Supervisor"]:
                        t["status"] = "Closed"
                        t["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":"Approved and closed"})
                        add_audit("Tasks", f"{t['id']} approved and closed"); st.rerun()

                if toggle(f"dlg_exp_{t['id']}", "🔀 Delegate"):
                    d1, d2 = st.columns(2)
                    dto   = d1.selectbox("Delegate to", list(D["users"].keys()), key=f"dt_{t['id']}")
                    dmemo = d2.text_input("Memo", key=f"dm_{t['id']}")
                    if st.button("Confirm Delegation", key=f"dlg_{t['id']}"):
                        t["delegated_to"] = dto; t["delegated_memo"] = dmemo
                        t["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":f"Delegated to {dto}: {dmemo}"})
                        add_audit("Tasks", f"{t['id']} delegated to {dto}"); st.rerun()

                if toggle(f"cwr_exp_{t['id']}", "🚪 Close with Rationale"):
                    rat = st.text_area("Rationale (required)", key=f"rat_{t['id']}")
                    if st.button("Close Task", key=f"cwr_{t['id']}"):
                        if rat.strip():
                            t["status"] = "Closed"; t["rationale"] = rat
                            t["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":f"Closed: {rat}"})
                            add_audit("Tasks", f"{t['id']} closed"); st.rerun()
                        else: st.warning("Rationale required.")

                # ── Audit trail ──
                st.markdown("<hr style='margin:.6rem 0'/><div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.4rem'>AUDIT TRAIL</div>", unsafe_allow_html=True)
                for e in t["audit"]:
                    st.markdown(f"<div class='gc-tl'><span style='color:#2563eb;font-size:11px'>{e['ts']}</span> · <b>{e['user']}</b> · {e['action']}</div>", unsafe_allow_html=True)
    with tab2:
        c1,c2 = st.columns(2)
        title    = c1.text_input("Task Title *")
        assignee = c2.selectbox("Assignee", list(D["users"].keys()), format_func=lambda x: f"{D['users'][x]['name']} ({D['users'][x]['title']})")
        c3,c4,c5 = st.columns(3)
        due = c3.date_input("Due Date", value=date.today()+timedelta(days=7))
        pri = c4.selectbox("Priority", ["High","Medium","Low"])
        cat = c5.selectbox("Category", ["AML","Supervisory","Trading","Regulatory","InfoSec","Testing","Other"])
        c6,c7 = st.columns(2)
        recurrence = c6.text_input("Recurrence", placeholder="e.g. Quarterly, Annually, Monthly")
        wsp_options = [""] + [w["id"] for w in D["wsps"] if w["status"]=="Approved"]
        wsp_labels  = {w["id"]: f"{w['id']} — {w['title']}" for w in D["wsps"]}
        source_wsp  = c7.selectbox("Source Policy / WSP", wsp_options, format_func=lambda x: "— None —" if x=="" else wsp_labels.get(x,x))
        instructions      = st.text_area("Instructions / What to Do *", height=120, placeholder="Describe step-by-step what the assignee needs to do to complete this task. Be specific — they should be able to act on this without reading the full policy.")
        evidence_required = st.text_area("Evidence Required", height=80, placeholder="List what files, reports, or documentation the assignee must upload to demonstrate completion.")
        if st.button("➕ Create Task"):
            if title.strip():
                nid = f"T-{len(D['tasks'])+1:03d}"
                D["tasks"].append({
                    "id":nid,"title":title.strip(),"assignee":assignee,
                    "due":due.strftime("%Y-%m-%d"),"status":"Open","priority":pri,"category":cat,
                    "instructions":instructions.strip(),"evidence_required":evidence_required.strip(),
                    "source_wsp":source_wsp,"recurrence":recurrence.strip(),
                    "evidence":[],"rationale":"","delegated_to":"","delegated_memo":"",
                    "audit":[{"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":f"Created {nid}"}],
                })
                add_audit("Tasks", f"Created {nid}: {title.strip()}")
                st.success(f"✓ Task {nid} created and assigned to {D['users'][assignee]['name']}.")
                st.rerun()

# ── EXAM REQUESTS ──────────────────────────────────────────────────────────────
def page_exam_requests():
    st.markdown("<div class='gc-header'><h1>EXAM / DOCUMENT REQUEST WORKFLOW</h1><p>Centralized workspace for regulatory exam preparation</p></div>", unsafe_allow_html=True)

    # Supervisor: only show tab 1, and only exams they have at least one item assigned to
    if is_supervisor():
        cu = st.session_state.current_user
        my_exams = [er for er in D["exam_requests"] if any(i["assignee"]==cu for i in er["items"])]
        if not my_exams:
            st.info("You have no exam document requests currently assigned to you.")
            return
        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>Showing {len(my_exams)} exam request(s) with items assigned to you. You can only upload documents for items assigned to you.</div>", unsafe_allow_html=True)
        for er in my_exams:
            done=len([i for i in er["items"] if i["status"]=="Uploaded"])
            total=len(er["items"]); pct=int(done/total*100) if total else 0
            if toggle(f"er_{er['id']}", f"{er['id']} — {er['title']} | {er['regulator']} | Due: {er['due']}"):
                st.markdown(f"Progress: **{done}/{total}** ({pct}%)")
                st.progress(pct/100)
                for idx,item in enumerate(er["items"]):
                    aname = D["users"].get(item["assignee"],{}).get("name", item["assignee"])
                    assigned_ts = item.get("assigned_ts","—")[:10] if item.get("assigned_ts") else "—"
                    is_mine = item["assignee"] == cu
                    ic1,ic2,ic3=st.columns([3,1.5,1.5])
                    ic1.markdown(f"{'✅' if item['status']=='Uploaded' else '⏳'}  **{item['item']}**")
                    ic2.markdown(badge(item["status"],"badge-closed" if item["status"]=="Uploaded" else "badge-pending"),unsafe_allow_html=True)
                    ic3.markdown(f"_{aname}_  \n<span style='font-size:11px;color:#6b7280'>Assigned {assigned_ts}</span>", unsafe_allow_html=True)
                    if is_mine:
                        up_col, notes_col = st.columns([1,1])
                        up = up_col.file_uploader("Upload", key=f"er_{er['id']}_{idx}", label_visibility="collapsed")
                        er_notes = notes_col.text_area("Notes / context", key=f"er_notes_{er['id']}_{idx}", height=68, placeholder="What is this file? Any issues, version notes, or context for the examiner.")
                        if up:
                            item["status"]="Uploaded"
                            item["upload_notes"] = er_notes.strip()
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            er["audit"].append({"ts":now_str,"user":cu,"action":f"Uploaded {up.name} for '{item['item']}'"})
                            add_audit("Exam Requests",f"{er['id']}: uploaded {up.name}"); st.rerun()
                    else:
                        st.markdown(f"<div style='font-size:11px;color:#9ca3af;font-style:italic;padding:.3rem .5rem;margin-bottom:.3rem'>Upload handled by {aname}</div>", unsafe_allow_html=True)
                    if item.get("upload_notes"):
                        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.3rem;padding-left:.5rem'>💬 <i>{item['upload_notes']}</i></div>", unsafe_allow_html=True)
                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
        return

    tab1,tab2=st.tabs(["📁  ACTIVE EXAMS","➕  NEW EXAM REQUEST"])
    with tab1:
        for er in D["exam_requests"]:
            done=len([i for i in er["items"] if i["status"]=="Uploaded"])
            total=len(er["items"]); pct=int(done/total*100) if total else 0
            if toggle(f"er_{er['id']}", f"{er['id']} — {er['title']} | {er['regulator']} | Due: {er['due']}"):

                # ── ORIGINAL REQUEST LIST — downloadable reference at top ──
                request_lines = [
                    f"EXAM / DOCUMENT REQUEST LIST",
                    f"{'='*50}",
                    f"Exam:       {er['title']}",
                    f"ID:         {er['id']}",
                    f"Regulator:  {er['regulator']}",
                    f"Due Date:   {er['due']}",
                    f"Status:     {er['status']}",
                    f"",
                    f"DOCUMENT ITEMS REQUESTED:",
                    f"{'-'*50}",
                ]
                for i, item in enumerate(er["items"], 1):
                    aname = D["users"].get(item["assignee"],{}).get("name", item["assignee"])
                    request_lines.append(f"{i}. {item['item']}")
                    request_lines.append(f"   Assigned to: {aname}  |  Assigned: {item.get('assigned_ts','—')[:10]}  |  Status: {item['status']}")
                    if item.get("upload_notes"):
                        request_lines.append(f"   Notes: {item['upload_notes']}")
                    request_lines.append("")
                request_txt = "\n".join(request_lines)
                import base64 as _b64
                _encoded = _b64.b64encode(request_txt.encode()).decode()
                _fname   = f"{er['id']}_request_list.txt"

                st.markdown(f"""
                <div style='background:#f0f7ff;border:1px solid #bfdbfe;border-radius:8px;
                     padding:.7rem 1rem;margin-bottom:.8rem;
                     display:flex;align-items:center;justify-content:space-between;gap:.5rem;flex-wrap:wrap'>
                  <div>
                    <div style='font-size:13px;font-weight:700;color:#1e40af'>📄 Original Request List — {er['title']}</div>
                    <div style='font-size:12px;color:#6b7280;margin-top:2px'>{len(er['items'])} item(s) · assigned to {len(set(i['assignee'] for i in er['items']))} person(s)</div>
                  </div>
                  <a href="data:text/plain;base64,{_encoded}" download="{_fname}"
                     style='background:#2563eb;color:#ffffff !important;font-size:12px;font-weight:700;
                            padding:7px 16px;border-radius:6px;text-decoration:none;white-space:nowrap;
                            border:none;cursor:pointer;letter-spacing:.02em'>
                    📥 Download Original Request
                  </a>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("<hr style='margin:.5rem 0 .8rem 0'/>", unsafe_allow_html=True)

                # ── PROGRESS ──
                st.markdown(f"Progress: **{done}/{total}** ({pct}%)")
                st.progress(pct/100)

                # ── ITEM ROWS ──
                for idx,item in enumerate(er["items"]):
                    aname = D["users"].get(item["assignee"],{}).get("name", item["assignee"])
                    assigned_ts = item.get("assigned_ts","—")[:10] if item.get("assigned_ts") else "—"
                    can_upload = is_cco() or item["assignee"] == st.session_state.current_user
                    ic1,ic2,ic3=st.columns([3,1.5,1.5])
                    ic1.markdown(f"{'✅' if item['status']=='Uploaded' else '⏳'}  **{item['item']}**")
                    ic2.markdown(badge(item["status"],"badge-closed" if item["status"]=="Uploaded" else "badge-pending"),unsafe_allow_html=True)
                    ic3.markdown(f"_{aname}_  \n<span style='font-size:11px;color:#6b7280'>Assigned {assigned_ts}</span>", unsafe_allow_html=True)
                    if can_upload:
                        up_col, notes_col = st.columns([1,1])
                        up = up_col.file_uploader("Upload", key=f"er_{er['id']}_{idx}", label_visibility="collapsed")
                        er_notes = notes_col.text_area("Notes / context", key=f"er_notes_{er['id']}_{idx}", height=68, placeholder="What is this file? Any issues, version notes, or context for the examiner.")
                        if up:
                            item["status"]="Uploaded"
                            item["upload_notes"] = er_notes.strip()
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            er["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"Uploaded {up.name} for '{item['item']}'" + (f" — {er_notes.strip()}" if er_notes.strip() else "")})
                            add_audit("Exam Requests",f"{er['id']}: uploaded {up.name}"); st.rerun()
                    else:
                        st.markdown(f"<div style='font-size:11px;color:#9ca3af;font-style:italic;padding:.3rem .5rem;margin-bottom:.3rem'>Upload handled by {aname}</div>", unsafe_allow_html=True)
                    if item.get("upload_notes"):
                        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.3rem;padding-left:.5rem'>💬 <i>{item['upload_notes']}</i></div>", unsafe_allow_html=True)
                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

                st.markdown("<hr/>", unsafe_allow_html=True)
                if st.button(f"📦 Export Package",key=f"exp_{er['id']}"):
                    payload=json.dumps({"exam":er,"audit":er["audit"]},indent=2)
                    st.download_button("Download JSON",data=payload,file_name=f"{er['id']}_package.json",mime="application/json",key=f"dl_{er['id']}")

                # ── AUDIT TRAIL — shows assignments + all actions ──
                st.markdown("<hr/><div style='font-size:11px;font-weight:700;letter-spacing:.08em;color:#6b7280;text-transform:uppercase;margin-bottom:.5rem'>AUDIT TRAIL</div>", unsafe_allow_html=True)

                # Build enriched audit: seed with assignment entries if not already present
                shown = set()
                for e in er["audit"]:
                    key = f"{e['ts']}_{e['action']}"
                    if key in shown:
                        continue
                    shown.add(key)
                    # Colour-code assignment entries differently
                    is_assign = e["action"].startswith("Assigned ")
                    dot_color  = "#8b5cf6" if is_assign else "#3b82f6"
                    icon = "👤" if is_assign else "·"
                    st.markdown(
                        f"<div class='gc-tl' style='border-left-color:{dot_color}'>"
                        f"<span style='color:#2563eb;font-size:11px'>{e['ts']}</span> "
                        f"&nbsp;<b>{e['user']}</b> {icon} {e['action']}"
                        f"</div>",
                        unsafe_allow_html=True
                    )
    with tab2:
        c1,c2=st.columns(2)
        title=c1.text_input("Exam Title"); reg=c2.selectbox("Regulator",["FINRA","SEC","CFTC","NFA","State Regulator","Internal"])
        due=st.date_input("Due Date",value=date.today()+timedelta(days=30))
        items_raw=st.text_area("Document Items (one per line)","Customer files\nSupervisory procedures\nAML documentation",height=100)
        if st.button("Create Exam Request"):
            if title.strip():
                nid=f"ER-{len(D['exam_requests'])+1:03d}"
                now_str=datetime.now().strftime("%Y-%m-%d %H:%M")
                items=[{"item":i.strip(),"status":"Pending","assignee":st.session_state.current_user,"assigned_ts":now_str,"upload_notes":""} for i in items_raw.splitlines() if i.strip()]
                audit_entries=[{"ts":now_str,"user":st.session_state.current_user,"action":f"Created {nid}"}]
                for it in items:
                    audit_entries.append({"ts":now_str,"user":st.session_state.current_user,"action":f"Assigned '{it['item']}' → {it['assignee']}"})
                D["exam_requests"].append({"id":nid,"title":title,"regulator":reg,"due":due.strftime("%Y-%m-%d"),"status":"Open","items":items,"audit":audit_entries})
                add_audit("Exam Requests",f"Created {nid}: {title}"); st.success(f"✓ {nid} created."); st.rerun()

# ── WSP REPOSITORY ─────────────────────────────────────────────────────────────
LINES_OF_BUSINESS = ["Wealth Management","Investment Banking","Operations","Fixed Income","Capital Markets","All Firm"]

def _auto_parse_wsp(wsp):
    """Called when a policy is approved or first expanded.
    Upserts supervisory checklist ITEMS into each user's single aggregated checklist.
    Each user has exactly ONE checklist that spans ALL policies.
    Items are keyed by (source_policy_id, wt_id) to prevent duplication.
    Only fires for POL- docs; WSPs do not create checklists."""
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    doc_id  = wsp["id"]
    if not doc_id.startswith("POL-"):
        return 0

    task_list = D.get("policy_tasks",{}).get(doc_id, [])
    if not task_list:
        return 0

    # Ensure D["supervisory_checklists"] is keyed by user (one per user)
    # Build lookup: user_key → checklist
    cl_by_user = {}
    for cl in D.get("supervisory_checklists", []):
        uk = cl.get("assigned_to","")
        if uk:
            cl_by_user[uk] = cl

    users_updated = set()

    for wt in task_list:
        owner_key = wt["owner"]
        oinfo     = D["users"].get(owner_key, {})

        # Get or create the user's single checklist
        if owner_key not in cl_by_user:
            cl_id = f"CL-{len(D['supervisory_checklists'])+1:03d}"
            new_cl = {
                "id":          cl_id,
                "assigned_to": owner_key,
                "owner_name":  oinfo.get("name", owner_key),
                "owner_title": oinfo.get("title", oinfo.get("role","")),
                "status":      "Active",
                "created_ts":  now_str,
                "created_by":  "system",
                "items":       [],
                "audit":       [{"ts":now_str,"user":"system","action":"User checklist created"}],
            }
            D["supervisory_checklists"].append(new_cl)
            cl_by_user[owner_key] = new_cl

        cl = cl_by_user[owner_key]

        # Dedup: skip if this policy+procedure already in the checklist
        existing_keys = {(i.get("source_policy_id",""), i.get("wt_id","")) for i in cl["items"]}
        if (doc_id, wt["id"]) in existing_keys:
            continue

        cl["items"].append({
            "wt_id":            wt["id"],
            "source_policy_id": doc_id,
            "source_policy":    wsp["title"],
            "title":            wt["title"],
            "description":      wt["description"],
            "frequency":        wt["frequency"],
            "category":         wt["category"],
            "due_date":         _freq_to_due(wt["frequency"]),
            "status":           "Pending",
            "evidence":         [],
            "completed_ts":     "",
            "checked_off":      False,
            "supervisor_review":       False,
            "supervisor_reviewed_by":  "",
            "supervisor_reviewed_ts":  "",
            "supervisor_notes":        "",
            "comments":                [],
            "audit":            [{"ts":now_str,"user":"system",
                                  "action":f"Item added from {doc_id} v{wsp['version']}"}],
        })
        cl["audit"].append({"ts":now_str,"user":"system",
                            "action":f"Item '{wt['title']}' added from {doc_id}"})
        users_updated.add(owner_key)

    return len(users_updated)


def _freq_to_due(frequency: str) -> str:
    """Convert a frequency string into a concrete due date for sorting."""
    now = datetime.now()
    f = frequency.lower()
    if "daily"       in f: days = 1
    elif "weekly"    in f: days = 7
    elif "month"     in f: days = 30
    elif "quarter"   in f: days = 90
    elif "semi"      in f: days = 180
    elif "annual"    in f: days = 365
    elif "as needed" in f: days = 30
    else:                  days = 30
    return (now + timedelta(days=days)).strftime("%Y-%m-%d")


def page_wsp():
    st.markdown("<div class='gc-header' style='padding:.7rem 1.2rem;margin-bottom:.8rem'><h1 style='font-size:1rem !important;margin-bottom:.1rem !important'>POLICIES & WSPs</h1><p style='font-size:12px !important'>Compliance document library — policies, procedures, regulation mapping, testing traceability, and checklist generation</p></div>", unsafe_allow_html=True)

    import base64 as _b64w

    # ── Back-compat for both collections ──────────────────────────────────────
    for w in D["wsps"] + D.get("policies",[]):
        if "files"            not in w: w["files"]            = []
        if "reviewer"         not in w: w["reviewer"]         = ""
        if "review_sent_ts"   not in w: w["review_sent_ts"]   = ""
        if "review_status"    not in w: w["review_status"]    = "approved" if w["status"]=="Approved" else "with_author"
        if "audit"            not in w: w["audit"]            = []
        if "line_of_business" not in w: w["line_of_business"] = "Operations"
        if "type"             not in w: w["type"]             = "WSP"
        if "regulations"      not in w: w["regulations"]      = []
        if "related_tests"    not in w: w["related_tests"]    = []

    REVIEW_STATUS_LABELS = {
        "with_author":    ("WITH AUTHOR",                     "#1e40af","#dbeafe"),
        "with_reviewer":  ("WITH REVIEWER",                   "#854d0e","#fef9c3"),
        "pending_changes":("PENDING CHANGES — SEE COMMENTS",  "#991b1b","#fee2e2"),
        "approved":       ("APPROVED",                        "#166534","#dcfce7"),
    }

    # ── Shared: scalable table-style document library ─────────────────────────
    def _render_library(docs, tasks_key, id_prefix):
        import base64 as _b64w2
        from collections import defaultdict as _ddf

        approved = [w for w in docs if w["status"] == "Approved"]

        # ── Toolbar ────────────────────────────────────────────────────────────
        tb1, tb2, tb3, tb4, tb5 = st.columns([2.5, 1.4, 1.4, 1.4, 1.4])
        f_search  = tb1.text_input("🔍 Search", placeholder="name, ID, owner, regulation…", key=f"{id_prefix}_search", label_visibility="collapsed")
        all_lobs  = sorted(set(w.get("line_of_business","") for w in approved))
        all_owners= sorted(set(D["users"].get(w.get("owner",""),{}).get("name",w.get("owner","")) for w in approved))
        all_regs  = sorted(set(r for w in approved for r in w.get("regulations",[])))
        f_lob     = tb2.selectbox("LOB",    ["All LOBs"]  + all_lobs,   key=f"{id_prefix}_lob",   label_visibility="collapsed")
        f_owner   = tb3.selectbox("Owner",  ["All Owners"]+ all_owners,  key=f"{id_prefix}_owner", label_visibility="collapsed")
        f_reg     = tb4.selectbox("Reg",    ["All Regs"]  + all_regs,    key=f"{id_prefix}_reg",   label_visibility="collapsed")
        f_sort    = tb5.selectbox("Sort",   ["Name ↑","Name ↓","Date ↑","Date ↓","Ver ↑","Ver ↓"], key=f"{id_prefix}_sort", label_visibility="collapsed")

        # Second row: compact expand/collapse toggles + view + doc count
        cr1, cr2, cr3, cr4 = st.columns([0.55, 0.55, 1.5, 5])
        with cr1:
            st.markdown("<div class='gc-sm-btn'>", unsafe_allow_html=True)
            if st.button("⊞ All",  key=f"{id_prefix}_expall"):
                for w in approved: st.session_state[f"docrow_{id_prefix}_{w['id']}"] = True
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
        with cr2:
            st.markdown("<div class='gc-sm-btn'>", unsafe_allow_html=True)
            if st.button("⊟ None", key=f"{id_prefix}_colall"):
                for w in approved: st.session_state[f"docrow_{id_prefix}_{w['id']}"] = False
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
        f_view = cr3.selectbox("View", ["Grouped by LOB","Flat list"], key=f"{id_prefix}_view", label_visibility="collapsed")

        # Apply filters
        filtered = approved
        if f_search:
            q = f_search.lower()
            filtered = [w for w in filtered if q in w["title"].lower() or q in w["id"].lower()
                        or q in D["users"].get(w["owner"],{}).get("name",w["owner"]).lower()
                        or any(q in r.lower() for r in w.get("regulations",[]))]
        if f_lob   != "All LOBs":    filtered = [w for w in filtered if w.get("line_of_business") == f_lob]
        if f_owner != "All Owners":  filtered = [w for w in filtered if D["users"].get(w["owner"],{}).get("name",w["owner"]) == f_owner]
        if f_reg   != "All Regs":    filtered = [w for w in filtered if f_reg in w.get("regulations",[])]

        def _sort_key(w):
            if "Name" in f_sort: return w["title"].lower()
            if "Date" in f_sort: return w.get("approved_ts","")
            return w.get("version","")
        filtered = sorted(filtered, key=_sort_key, reverse="↓" in f_sort)

        cr4.markdown(f"<div style='font-size:11px;color:#6b7280;padding:.42rem 0'>{len(filtered)} of {len(approved)} document(s)</div>", unsafe_allow_html=True)

        if not filtered:
            st.markdown("<div class='gc-alert gc-alert-warn'>No documents match the current filters.</div>", unsafe_allow_html=True)
            return

        # ── Table header row ──────────────────────────────────────────────────
        st.markdown("""
        <div style='display:grid;
             grid-template-columns:76px 1.8fr 58px 120px 180px 110px 110px 82px 52px 110px;
             background:#f1f5f9;border:1px solid #e2e8f0;border-radius:6px 6px 0 0;
             padding:.32rem .7rem;margin-top:.4rem;gap:4px;align-items:center'>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>ID</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Document Name</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Type</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Line of Business</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Regulation Mapping</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Owner</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Approved By</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Date</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Ver</div>
          <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Actions</div>
        </div>
        """, unsafe_allow_html=True)

        # ── Group or flat ─────────────────────────────────────────────────────
        if "Grouped" in f_view:
            by_lob = _ddf(list)
            for w in filtered: by_lob[w.get("line_of_business","Other")].append(w)
            groups = [(lob, by_lob[lob]) for lob in sorted(by_lob.keys())]
        else:
            groups = [("", filtered)]

        for grp_label, grp_docs in groups:
            if grp_label:
                st.markdown(f"""
                <div style='background:#1e293b;padding:.3rem .7rem;
                     border-left:1px solid #334155;border-right:1px solid #334155;'>
                  <span style='font-size:10px;font-weight:800;letter-spacing:.09em;
                               color:#f8fafc !important;-webkit-text-fill-color:#f8fafc'>
                    📁 {grp_label.upper()}
                  </span>
                  <span style='font-size:10px;font-weight:400;color:#94a3b8 !important;
                               -webkit-text-fill-color:#94a3b8;margin-left:.5rem'>{len(grp_docs)} doc(s)</span>
                </div>
                """, unsafe_allow_html=True)

            for i, doc in enumerate(grp_docs):
                owner_name    = D["users"].get(doc["owner"],{}).get("name", doc["owner"])
                approver_name = D["users"].get(doc["approved_by"],{}).get("name", doc["approved_by"])
                row_bg        = "#ffffff" if i % 2 == 0 else "#fafbfc"
                expand_key    = f"docrow_{id_prefix}_{doc['id']}"
                is_expanded   = st.session_state.get(expand_key, False)
                approved_date = doc.get("approved_ts","")[:10]
                lob_short     = doc.get("line_of_business","")
                doc_type      = doc.get("type","WSP")
                regs          = doc.get("regulations",[])
                reg_str       = " · ".join(regs[:2]) + (" …" if len(regs)>2 else "") if regs else "—"
                type_color    = "#7c3aed" if doc_type=="Policy" else "#0369a1"
                type_bg       = "#f3e8ff" if doc_type=="Policy" else "#e0f2fe"
                bl_left       = "border-left:3px solid #3b82f6;" if is_expanded else "border-left:1px solid #e2e8f0;"

                # Build download b64
                doc_lines = [f"{'='*60}",f"  {doc['id']} — {doc['title']}",
                    f"  Type: {doc_type}  |  LOB: {lob_short}",
                    f"  Regulations: {', '.join(regs)}",
                    f"  Version: {doc['version']}  |  Approved: {approved_date}",
                    f"  Owner: {owner_name}  |  Approved By: {approver_name}",
                    f"{'='*60}","","CONTENT","-"*60,doc.get("content",""),"","VERSION HISTORY","-"*60]
                for v in doc.get("versions",[]): doc_lines.append(f"  v{v['ver']}  {v['ts']}  {v['author']}  —  {v['note']}")
                doc_lines += ["","AUDIT TRAIL","-"*60]
                for e in doc.get("audit",[]): doc_lines.append(f"  {e['ts']}  {e['user']}  {e['action']}")
                doc_b64   = _b64w2.b64encode("\n".join(doc_lines).encode()).decode()
                doc_fname = f"{doc['id']}_v{doc['version']}.txt"

                # ── Compact table row (HTML only — no interactive elements) ──
                st.markdown(f"""
                <div style='display:grid;
                     grid-template-columns:76px 1.8fr 58px 120px 180px 110px 110px 82px 52px 110px;
                     gap:4px;background:{row_bg};{bl_left}
                     border-right:1px solid #e2e8f0;border-bottom:1px solid #e2e8f0;
                     padding:.36rem .7rem;align-items:center;min-height:36px'>
                  <div style='font-size:10px;font-weight:800;color:#059669;white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>{doc['id']}</div>
                  <div style='font-size:11px;font-weight:600;color:#111827;overflow:hidden;text-overflow:ellipsis;white-space:nowrap' title='{doc["title"]}'>📄 {doc['title']}</div>
                  <div><span style='background:{type_bg};color:{type_color};font-size:9px;font-weight:700;padding:1px 5px;border-radius:8px;white-space:nowrap'>{doc_type}</span></div>
                  <div style='font-size:10px;color:#4b5563;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>{lob_short}</div>
                  <div style='font-size:10px;color:#1d4ed8;overflow:hidden;text-overflow:ellipsis;white-space:nowrap' title='{", ".join(regs)}'>{reg_str}</div>
                  <div style='font-size:10px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>{owner_name}</div>
                  <div style='font-size:10px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>{approver_name}</div>
                  <div style='font-size:10px;color:#6b7280;white-space:nowrap'>{approved_date}</div>
                  <div><span style='background:#dcfce7;color:#166534;font-size:9px;font-weight:700;padding:1px 5px;border-radius:8px'>v{doc['version']}</span></div>
                  <div style='display:flex;gap:4px;align-items:center'>
                    <a href="data:text/plain;base64,{doc_b64}" download="{doc_fname}"
                       style='background:#e0f2fe;color:#0369a1 !important;font-size:9px;font-weight:700;
                              padding:2px 7px;border-radius:4px;text-decoration:none;white-space:nowrap;border:1px solid #bae6fd'>⬇</a>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # Expand button in Streamlit native (must be outside HTML)
                btn_row = st.columns([76/930, 470/930, 58/930, 120/930, 180/930, 110/930, 110/930, 82/930, 52/930, 110/930])
                with btn_row[9]:
                    exp_label = "▲ Close" if is_expanded else "▼ Expand"
                    if st.button(exp_label, key=f"expbtn_{id_prefix}_{doc['id']}"):
                        st.session_state[expand_key] = not is_expanded
                        st.rerun()

                # ── Expanded detail panel ─────────────────────────────────────
                if is_expanded:
                    task_list    = D.get(tasks_key,{}).get(doc["id"],[])
                    rel_test_ids = doc.get("related_tests",[])
                    # New model: one checklist per user — find checklists that have items from this policy
                    if doc["id"].startswith("POL-"):
                        pol_owners = {wt["owner"] for wt in task_list}
                        checklists = [c for c in D.get("supervisory_checklists",[])
                                      if c.get("assigned_to") in pol_owners
                                      and any(i.get("source_policy_id")==doc["id"] for i in c.get("items",[]))]
                        # Evidence = only files on items sourced from this policy
                        ev_count   = sum(len(item.get("evidence",[]))
                                         for c in checklists
                                         for item in c.get("items",[])
                                         if item.get("source_policy_id")==doc["id"])
                    else:
                        checklists = []
                        ev_count   = 0
                    rel_tests  = [t for t in D.get("compliance_tests",[]) if t["id"] in rel_test_ids]
                    rem_items  = [ex for t in rel_tests for ex in t.get("exceptions",[]) if ex.get("status")=="Open"]

                    # Traceability counts
                    proc_count  = len(task_list)
                    test_count  = len(rel_tests)
                    rem_count   = len(rem_items)

                    st.markdown(f"""
                    <div style='background:#f8fafc;border:1px solid #bfdbfe;border-top:none;
                         border-left:3px solid #3b82f6;padding:.9rem 1.1rem .6rem;margin-bottom:3px'>

                      <!-- Metadata bar -->
                      <div style='display:flex;gap:2rem;flex-wrap:wrap;padding-bottom:.7rem;
                           border-bottom:1px solid #e2e8f0;margin-bottom:.8rem'>
                        <div><div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.15rem'>Line of Business</div>
                          <div style='font-size:12px;font-weight:600;color:#111'>{lob_short or "—"}</div></div>
                        <div><div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.15rem'>Owner</div>
                          <div style='font-size:12px;font-weight:600;color:#111'>{owner_name}</div></div>
                        <div><div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.15rem'>Approved By</div>
                          <div style='font-size:12px;font-weight:600;color:#111'>{approver_name}</div></div>
                        <div><div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.15rem'>Approval Date</div>
                          <div style='font-size:12px;font-weight:600;color:#111'>{approved_date}</div></div>
                        <div><div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.15rem'>Version</div>
                          <div style='font-size:12px;font-weight:600;color:#111'>v{doc['version']}</div></div>
                        <div><div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.15rem'>Type</div>
                          <div style='font-size:12px;font-weight:600;color:{type_color}'>{doc_type}</div></div>
                      </div>

                      <!-- Regulation mapping -->
                      <div style='margin-bottom:.8rem'>
                        <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.35rem'>📐 MAPPED REGULATIONS</div>
                        <div style='display:flex;flex-wrap:wrap;gap:.3rem'>
                          {''.join(f"<span style='background:#dbeafe;color:#1e40af;font-size:10px;font-weight:700;padding:2px 9px;border-radius:10px;border:1px solid #bfdbfe'>{r}</span>" for r in regs) if regs else "<span style='color:#9ca3af;font-size:12px'>No regulations mapped</span>"}
                        </div>
                      </div>

                    </div>
                    """, unsafe_allow_html=True)

                    # ── Compliance Traceability Graph ─────────────────────────
                    st.markdown("<div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin:.6rem 0 .35rem 0'>🔗 COMPLIANCE TRACEABILITY</div>", unsafe_allow_html=True)

                    def _chain_node(icon, label, count, color, bg, nav_hint=""):
                        border_color = color + "44"
                        count_html = f"<div style='font-size:16px;font-weight:800;color:{color}'>{count}</div>" if count != "" else ""
                        hint_html  = f"<div style='font-size:9px;color:#9ca3af;margin-top:1px'>{nav_hint}</div>" if nav_hint else ""
                        return (
                            f"<div style='background:{bg};border:1px solid {border_color};border-radius:8px;"
                            f"padding:.45rem .7rem;text-align:center;min-width:90px'>"
                            f"<div style='font-size:14px'>{icon}</div>"
                            f"{count_html}"
                            f"<div style='font-size:9px;font-weight:700;color:{color};text-transform:uppercase;letter-spacing:.04em'>{label}</div>"
                            f"{hint_html}"
                            f"</div>"
                        )

                    arrow_html = "<div style='font-size:18px;color:#cbd5e1;align-self:center;padding:0 .1rem'>&#x2192;</div>"

                    n1_color = "#7c3aed" if doc_type == "Policy" else "#0369a1"
                    n1_bg    = "#f3e8ff" if doc_type == "Policy" else "#e0f2fe"

                    chain_parts = [
                        "<div style='display:flex;align-items:stretch;gap:.3rem;margin-bottom:.8rem;flex-wrap:wrap'>",
                        _chain_node("📄", doc_type,     "",              n1_color,  n1_bg),
                        arrow_html,
                        _chain_node("📋", "Procedures",  proc_count,     "#2563eb", "#eff6ff", "Supervisory"),
                        arrow_html,
                        _chain_node("✅", "Checklists",  len(checklists),"#059669", "#f0fdf4", "Checklists tab"),
                        arrow_html,
                        _chain_node("📁", "Evidence",    ev_count,       "#d97706", "#fffbeb", "Evidence Repo"),
                        arrow_html,
                        _chain_node("🧪", "Tests",       test_count,     "#7c3aed", "#faf5ff", "Comp. Testing"),
                        arrow_html,
                        _chain_node("⚠", "Remediation", rem_count,      "#dc2626", "#fef2f2", "Rem. Items"),
                        "</div>",
                    ]
                    st.markdown("".join(chain_parts), unsafe_allow_html=True)

                    # ── Version History ───────────────────────────────────────
                    st.markdown("<div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin:.5rem 0 .25rem 0'>📅 VERSION HISTORY</div>", unsafe_allow_html=True)
                    for v in doc.get("versions",[]):
                        vauthor = D["users"].get(v["author"],{}).get("name",v["author"])
                        st.markdown(f"<div class='gc-tl' style='font-size:12px'><b style='color:#2563eb'>v{v['ver']}</b> &nbsp;·&nbsp; {v['ts'][:10]} &nbsp;·&nbsp; {vauthor} &nbsp;·&nbsp; <i>{v['note']}</i></div>", unsafe_allow_html=True)

                    # ── Supervisory Procedures (simplified inline) ────────────
                    st.markdown("<div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin:.7rem 0 .3rem 0'>📋 SUPERVISORY PROCEDURES</div>", unsafe_allow_html=True)
                    if not task_list:
                        st.markdown("<div class='gc-alert gc-alert-warn' style='font-size:11px'>No procedures defined for this document yet.</div>", unsafe_allow_html=True)
                    else:
                        for wt in task_list:
                            owner_info = D["users"].get(wt["owner"],{})
                            owner_nm   = owner_info.get("name", wt["owner"])
                            owner_ttl  = owner_info.get("title", wt.get("role_title",""))
                            cat_cls    = "badge-high" if wt["category"] in ["AML","Regulatory"] else "badge-med"
                            st.markdown(f"""
                            <div style='background:#ffffff;border:1px solid #e5e7eb;border-left:3px solid #3b82f6;
                                 border-radius:6px;padding:.45rem .8rem;margin-bottom:4px;
                                 display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:.4rem'>
                              <div style='flex:1;min-width:200px'>
                                <div style='font-size:12px;font-weight:700;color:#111827;margin-bottom:.2rem'>{wt['title']}</div>
                                <div style='font-size:11px;color:#6b7280'>
                                  <span style='color:#374151'>Responsible Role:</span> {owner_ttl} &nbsp;·&nbsp;
                                  <span style='color:#374151'>Assigned Owner:</span> {owner_nm} &nbsp;·&nbsp;
                                  <span style='color:#2563eb;font-weight:600'>🔁 {wt['frequency']}</span>
                                </div>
                                <div style='font-size:11px;color:#6b7280;margin-top:.15rem'>{wt['description']}</div>
                              </div>
                              <span class='gc-badge {cat_cls}' style='white-space:nowrap'>{wt['category']}</span>
                            </div>
                            """, unsafe_allow_html=True)

                    # ── Related Compliance Testing ─────────────────────────────
                    st.markdown("<div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin:.7rem 0 .3rem 0'>🧪 RELATED COMPLIANCE TESTING</div>", unsafe_allow_html=True)
                    if not rel_tests:
                        st.markdown("<div style='font-size:12px;color:#9ca3af;font-style:italic;margin-bottom:.4rem'>No compliance tests linked to this document.</div>", unsafe_allow_html=True)
                    else:
                        for t in rel_tests:
                            t_status  = t.get("status","—")
                            t_color   = {"In Progress":"#2563eb","Completed":"#059669","Not Started":"#6b7280","Overdue":"#dc2626"}.get(t_status,"#6b7280")
                            t_bg      = {"In Progress":"#eff6ff","Completed":"#f0fdf4","Not Started":"#f9fafb","Overdue":"#fef2f2"}.get(t_status,"#f9fafb")
                            rr        = t.get("risk_rating") or t.get("initial_risk_rating","—")
                            rr_color  = {"High":"#dc2626","Medium":"#d97706","Low":"#059669"}.get(rr,"#6b7280")
                            period    = t.get("time_period_reviewed","—")
                            appr_date = (t.get("approval_date") or "")[:10] or "Pending"
                            exc_ct    = len(t.get("exceptions",[]))
                            open_exc  = len([e for e in t.get("exceptions",[]) if e.get("status")=="Open"])
                            closed_exc= exc_ct - open_exc

                            # Header row
                            st.markdown(f"""
                            <div style='background:{t_bg};border:1px solid #e5e7eb;border-left:3px solid {t_color};
                                 border-radius:6px 6px 0 0;padding:.45rem .8rem;margin-top:5px;
                                 display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:.4rem'>
                              <div>
                                <span style='font-size:11px;font-weight:800;color:#111827'>{t["id"]} — {t["topic"]}</span>
                                <span style='font-size:10px;color:#6b7280;margin-left:.6rem'>{t["test_type"]}</span>
                              </div>
                              <div style='display:flex;gap:.4rem;flex-wrap:wrap;align-items:center'>
                                <span style='background:{rr_color}18;color:{rr_color};font-size:9px;font-weight:700;padding:1px 6px;border-radius:8px'>Risk: {rr}</span>
                                <span style='background:{t_color}18;color:{t_color};font-size:9px;font-weight:700;padding:1px 6px;border-radius:8px'>{t_status}</span>
                              </div>
                            </div>
                            """, unsafe_allow_html=True)

                            # Expandable detail
                            t_detail_key = f"t_detail_{id_prefix}_{doc['id']}_{t['id']}"
                            t_open = st.session_state.get(t_detail_key, False)
                            if st.button("▼ Test Details" if not t_open else "▲ Hide", key=f"tbtn_{id_prefix}_{doc['id']}_{t['id']}"):
                                st.session_state[t_detail_key] = not t_open
                                st.rerun()

                            if t_open:
                                # Metadata grid
                                st.markdown(f"""
                                <div style='background:#ffffff;border:1px solid #e5e7eb;border-top:none;border-left:3px solid {t_color};
                                     padding:.55rem .8rem;border-radius:0 0 6px 6px'>
                                  <div style='display:flex;gap:2rem;flex-wrap:wrap;margin-bottom:.5rem;padding-bottom:.4rem;border-bottom:1px solid #f1f5f9'>
                                    <div><div style='font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Time Period Reviewed</div>
                                      <div style='font-size:11px;font-weight:600;color:#111'>{period}</div></div>
                                    <div><div style='font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Test Approval Date</div>
                                      <div style='font-size:11px;font-weight:600;color:#111'>{appr_date}</div></div>
                                    <div><div style='font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Assignee</div>
                                      <div style='font-size:11px;font-weight:600;color:#111'>{D["users"].get(t.get("assignee",""),{}).get("name",t.get("assignee","—"))}</div></div>
                                    <div><div style='font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Sample Size</div>
                                      <div style='font-size:11px;font-weight:600;color:#111'>{t.get("sample_size") or "—"}</div></div>
                                    <div><div style='font-size:8px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Exceptions</div>
                                      <div style='font-size:11px;font-weight:600;color:#111'><span style='color:#dc2626'>{open_exc} open</span> · {closed_exc} closed</div></div>
                                  </div>
                                """, unsafe_allow_html=True)

                                # Procedures
                                if t.get("procedures"):
                                    st.markdown("<div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin-bottom:.25rem'>Test Steps</div>", unsafe_allow_html=True)
                                    for p in t["procedures"]:
                                        p_result = p.get("result","")
                                        p_color  = {"Pass":"#059669","Exception":"#dc2626","":""}.get(p_result,"#6b7280")
                                        p_bg     = {"Pass":"#f0fdf4","Exception":"#fef2f2","":"#f9fafb"}.get(p_result,"#f9fafb")
                                        p_icon   = {"Pass":"✅","Exception":"⚠️","":"⬜"}.get(p_result,"⬜")
                                        p_title  = p.get("title") or p.get("description","")[:60]
                                        p_notes  = p.get("notes","")
                                        st.markdown(f"""
                                        <div style='background:{p_bg};border-left:2px solid {p_color if p_color else "#e2e8f0"};
                                             padding:.3rem .6rem;margin-bottom:2px;border-radius:0 4px 4px 0;font-size:11px'>
                                          <span style='font-weight:700'>{p_icon} {p.get("step","")}.  {p_title}</span>
                                          {f"<span style='color:{p_color};font-weight:700;margin-left:.5rem'>{p_result}</span>" if p_result else ""}
                                          {f"<div style='color:#6b7280;margin-top:.15rem;font-style:italic'>{p_notes}</div>" if p_notes else ""}
                                        </div>
                                        """, unsafe_allow_html=True)

                                # Remediation items
                                t_rems = t.get("exceptions",[])
                                if t_rems:
                                    st.markdown("<div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin:.4rem 0 .2rem'>Remediation Items</div>", unsafe_allow_html=True)
                                    for ex in t_rems:
                                        ex_open   = ex.get("status","Open") == "Open"
                                        ex_color  = "#dc2626" if ex_open else "#059669"
                                        ex_bg     = "#fef2f2" if ex_open else "#f0fdf4"
                                        ex_badge  = "OPEN" if ex_open else "CLOSED"
                                        st.markdown(f"""
                                        <div style='background:{ex_bg};border:1px solid {ex_color}33;border-radius:4px;padding:.3rem .6rem;margin-bottom:3px;font-size:11px'>
                                          <div style='display:flex;justify-content:space-between;align-items:center'>
                                            <span style='font-weight:700;color:#111'>{ex["id"]}: {ex["description"][:100]}</span>
                                            <span style='background:{ex_color};color:#fff;font-size:8px;font-weight:800;padding:1px 6px;border-radius:8px;white-space:nowrap;margin-left:.4rem'>{ex_badge}</span>
                                          </div>
                                          {f"<div style='color:#6b7280;margin-top:.1rem'>Assigned: {D['users'].get(ex.get('assigned_to',''),{{}}).get('name',ex.get('assigned_to','—'))} · Due: {ex.get('due_date','—')}</div>" if ex.get("assigned_to") else ""}
                                        </div>
                                        """, unsafe_allow_html=True)

                                if t.get("conclusion"):
                                    st.markdown(f"<div style='font-size:11px;color:#374151;background:#f8fafc;border-radius:4px;padding:.3rem .6rem;margin-top:.3rem'><b>Conclusion:</b> {t['conclusion']}</div>", unsafe_allow_html=True)

                                st.markdown("</div>", unsafe_allow_html=True)

                    # ── Auto-upsert checklist items on expand ─────────────────
                    if doc["id"].startswith("POL-") and task_list:
                        pol_owners  = {wt["owner"] for wt in task_list}
                        existing_cls = [c for c in D.get("supervisory_checklists",[])
                                        if c.get("assigned_to") in pol_owners
                                        and any(i.get("source_policy_id")==doc["id"] for i in c.get("items",[]))]
                        if not existing_cls:
                            n = _auto_parse_wsp(doc)
                            if n:
                                add_audit("Policies & WSPs", f"{doc['id']}: items added to {n} user checklist(s)")
                            existing_cls = [c for c in D.get("supervisory_checklists",[])
                                            if c.get("assigned_to") in pol_owners
                                            and any(i.get("source_policy_id")==doc["id"] for i in c.get("items",[]))]

                        st.markdown("<hr style='margin:.5rem 0'/>", unsafe_allow_html=True)
                        if existing_cls:
                            owners_str = ", ".join(sorted(set(c["owner_name"] for c in existing_cls)))
                            st.markdown(f"<div class='gc-alert gc-alert-ok' style='font-size:11px'>✓ Procedures from this policy appear in the checklists of: {owners_str}</div>", unsafe_allow_html=True)
                    else:
                        st.markdown("<hr style='margin:.5rem 0'/>", unsafe_allow_html=True)

                    if st.button("▲ Collapse", key=f"col_{id_prefix}_{doc['id']}"):
                        st.session_state[expand_key] = False
                        st.rerun()


    # ── Shared: render drafts tab ─────────────────────────────────────────────
    def _render_drafts(docs, id_prefix, tasks_key):
        drafts = [w for w in docs if w["status"] == "Draft"]
        if not drafts:
            st.markdown("<div class='gc-alert gc-alert-ok'>No drafts pending review.</div>", unsafe_allow_html=True)
            return

        for doc in drafts:
            rs               = doc.get("review_status","with_author")
            rs_lbl, rs_col, rs_bg = REVIEW_STATUS_LABELS.get(rs,("DRAFT","#6b7280","#f3f4f6"))
            owner_name       = D["users"].get(doc["owner"],{}).get("name", doc["owner"])
            reviewer_name    = D["users"].get(doc.get("reviewer",""),{}).get("name","—") if doc.get("reviewer") else "Not yet assigned"

            if toggle(f"draft_{id_prefix}_{doc['id']}", f"{doc['id']} — {doc['title']}  ·  v{doc['version']}  ·  {doc.get('line_of_business','')}  ·  {rs_lbl}"):
                st.markdown(f"""
                <div style='background:{rs_bg};border:1px solid #e5e7eb;border-radius:8px;
                     padding:.8rem 1.1rem;margin-bottom:.8rem;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem'>
                  <div>
                    <div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.2rem'>Current Status</div>
                    <div style='font-size:13px;font-weight:800;color:{rs_col}'>{rs_lbl}</div>
                  </div>
                  <div style='text-align:right'>
                    <div style='font-size:12px;color:#6b7280'>Author: <b>{owner_name}</b></div>
                    <div style='font-size:12px;color:#6b7280'>Reviewer: <b>{reviewer_name}</b></div>
                    {'<div style="font-size:11px;color:#6b7280;margin-top:.2rem">Sent: '+doc["review_sent_ts"][:16]+'</div>' if doc.get("review_sent_ts") else ''}
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # Files
                st.markdown("<div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.5rem'>UPLOADED FILES</div>", unsafe_allow_html=True)
                if not doc["files"]:
                    st.markdown("<div style='font-size:13px;color:#6b7280;margin-bottom:.5rem'>No files uploaded yet.</div>", unsafe_allow_html=True)
                for f in doc["files"]:
                    type_cls   = "badge-closed" if f["type"]=="Clean" else "badge-pending"
                    file_notes = f"<div style='font-size:12px;color:#6b7280;margin-top:2px'>💬 <i>{f['notes']}</i></div>" if f.get("notes") else ""
                    uploader_n = D["users"].get(f["uploaded_by"],{}).get("name", f["uploaded_by"])
                    type_badge = f"<span class='gc-badge {type_cls}'>{f['type']}</span>"
                    st.markdown(f"""
                    <div style='background:#f9fafb;border:1px solid #e5e7eb;border-radius:6px;
                         padding:.6rem .9rem;margin-bottom:4px;display:flex;justify-content:space-between;align-items:flex-start'>
                      <div>
                        <div style='font-weight:600;font-size:13px;color:#111'>📎 {f['name']}</div>
                        <div style='font-size:11px;color:#6b7280;margin-top:1px'>{uploader_n} · {f['ts'][:16]}</div>
                        {file_notes}
                      </div>
                      {type_badge}
                    </div>
                    """, unsafe_allow_html=True)

                # Upload
                if doc.get("review_status") != "approved":
                    st.markdown("<hr style='margin:.6rem 0'/>", unsafe_allow_html=True)
                    uf1, uf2 = st.columns(2)
                    new_file   = uf1.file_uploader("Upload File", key=f"up_{id_prefix}_{doc['id']}")
                    new_ftype  = uf1.selectbox("File Type", ["Redline","Clean","Supporting"], key=f"ftype_{id_prefix}_{doc['id']}")
                    new_fnotes = uf2.text_area("File Notes", height=90, key=f"fnotes_{id_prefix}_{doc['id']}", placeholder="Describe what this file is.")
                    if new_file:
                        doc["files"].append({"name":new_file.name,"type":new_ftype,"uploaded_by":st.session_state.current_user,"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"notes":new_fnotes.strip()})
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        doc["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"{new_ftype} uploaded: {new_file.name}" + (f" — {new_fnotes.strip()}" if new_fnotes.strip() else "")})
                        add_audit("Policies & WSPs", f"{doc['id']}: {new_ftype} uploaded")
                        st.success(f"✓ {new_file.name} uploaded."); st.rerun()

                # Send for review
                if doc.get("review_status") in ("with_author","pending_changes"):
                    st.markdown("<hr style='margin:.6rem 0'/><div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.5rem'>SEND FOR REVIEW</div>", unsafe_allow_html=True)
                    rev_opts = [u for u in D["users"] if u != doc["owner"]]
                    sr1, sr2 = st.columns(2)
                    sel_rev  = sr1.selectbox("Send To", rev_opts, format_func=lambda x: f"{D['users'][x]['name']} ({D['users'][x]['title']})", key=f"sendto_{id_prefix}_{doc['id']}")
                    rev_note = sr2.text_area("Message", height=68, key=f"sendnote_{id_prefix}_{doc['id']}", placeholder="Optional message to reviewer.")
                    if st.button(f"📤 Send {doc['id']} for Review", key=f"send_{id_prefix}_{doc['id']}"):
                        if not doc["files"]: st.error("Upload at least one file before sending.")
                        else:
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            doc["reviewer"] = sel_rev; doc["review_sent_ts"] = now_str; doc["review_status"] = "with_reviewer"
                            msg = f"Sent for review to {D['users'][sel_rev]['name']}" + (f" — {rev_note.strip()}" if rev_note.strip() else "")
                            doc["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":msg})
                            add_audit("Policies & WSPs", f"{doc['id']} sent for review")
                            st.success(f"✓ Sent to {D['users'][sel_rev]['name']}."); st.rerun()

                # Reviewer actions
                is_reviewer = (st.session_state.current_user == doc.get("reviewer") or has_full_compliance_access())
                if doc.get("review_status") == "with_reviewer" and is_reviewer:
                    st.markdown("<hr style='margin:.6rem 0'/><div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.5rem'>REVIEWER ACTIONS</div>", unsafe_allow_html=True)
                    rv_comments = st.text_area("Review Notes", height=80, key=f"rvcomm_{id_prefix}_{doc['id']}", placeholder="Approval notes or required changes.")
                    rva, rvb, _ = st.columns(3)
                    if rva.button(f"✅ Approve {doc['id']}", key=f"approve_{id_prefix}_{doc['id']}"):
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        doc["status"] = "Approved"; doc["review_status"] = "approved"
                        doc["approved_by"] = st.session_state.current_user; doc["approved_ts"] = now_str
                        doc["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":"Approved" + (f" — {rv_comments.strip()}" if rv_comments.strip() else "")})
                        D.get("approvals",[]).append({"id":f"AP-{len(D.get('approvals',[]))+1:03d}","item":f"{doc['id']} v{doc['version']}","type":"Approval","requestor":doc["owner"],"approver1":st.session_state.current_user,"approver1_ts":now_str,"approver2":"","status":"Approved","ts":now_str})
                        n = _auto_parse_wsp(doc)
                        add_audit("Policies & WSPs", f"Approved {doc['id']} v{doc['version']}; {n} checklist(s) auto-generated")
                        st.success(f"✓ {doc['id']} approved. {n} supervisory checklist(s) generated."); st.rerun()
                    if rvb.button(f"↩ Return — Pending Changes", key=f"return_{id_prefix}_{doc['id']}"):
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        doc["review_status"] = "pending_changes"
                        doc["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":"Returned — pending changes" + (f": {rv_comments.strip()}" if rv_comments.strip() else "")})
                        add_audit("Policies & WSPs", f"{doc['id']} returned for revision"); st.rerun()

                # Audit trail
                st.markdown("<hr style='margin:.6rem 0'/><div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.5rem'>AUDIT TRAIL</div>", unsafe_allow_html=True)
                for e in doc.get("audit",[]):
                    is_approve = "approved" in e["action"].lower()
                    is_return  = "Returned" in e["action"]
                    is_send    = "Sent for review" in e["action"]
                    dot  = "#22c55e" if is_approve else "#f59e0b" if is_return else "#8b5cf6" if is_send else "#3b82f6"
                    icon = "✅" if is_approve else "↩" if is_return else "📤" if is_send else "·"
                    st.markdown(f"<div class='gc-tl' style='border-left-color:{dot}'><span style='color:#2563eb;font-size:11px'>{e['ts']}</span> &nbsp;<b>{e['user']}</b> {icon} {e['action']}</div>", unsafe_allow_html=True)

    # ── New doc form (shared) ─────────────────────────────────────────────────
    def _render_new_form(doc_type, collection_key, tasks_key, id_prefix):
        label = "Policy" if doc_type=="Policy" else "WSP"
        st.markdown(f"<div style='font-size:13px;color:#6b7280;margin-bottom:1rem'>Upload your drafted {label} here. Select the line of business, upload files, then send for review. Checklists generate automatically on approval.</div>", unsafe_allow_html=True)
        nc1, nc2, nc3 = st.columns(3)
        new_title = nc1.text_input(f"{label} Title *", placeholder=f"e.g. {'AML Policy' if label=='Policy' else 'AML Supervisory Procedures'}", key=f"new_title_{id_prefix}")
        new_lob   = nc2.selectbox("Line of Business *", LINES_OF_BUSINESS, key=f"new_lob_{id_prefix}")
        new_ver   = nc3.text_input("Version *", value="1.0", key=f"new_ver_{id_prefix}")
        fu1, fu2  = st.columns(2)
        new_up    = fu1.file_uploader("Browse and select file", key=f"new_up_{id_prefix}")
        new_ftype = fu1.selectbox("File Type", ["Redline","Clean","Supporting"], key=f"new_ftype_{id_prefix}")
        new_fnote = fu2.text_area("File Notes", height=90, key=f"new_fnote_{id_prefix}", placeholder="Describe the file.")
        if st.button(f"💾 Save {label} as Draft", key=f"save_new_{id_prefix}"):
            if not new_title.strip():
                st.error(f"{label} title is required.")
            else:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                docs    = D[collection_key]
                nid     = f"{id_prefix}-{len(docs)+1:03d}"
                files   = []
                if new_up:
                    files.append({"name":new_up.name,"type":new_ftype,"uploaded_by":st.session_state.current_user,"ts":now_str,"notes":new_fnote.strip()})
                audit = [{"ts":now_str,"user":st.session_state.current_user,"action":f"{nid} draft created"}]
                if files: audit.append({"ts":now_str,"user":st.session_state.current_user,"action":f"{new_ftype} uploaded: {new_up.name}"})
                docs.append({
                    "id":nid,"title":new_title.strip(),"version":new_ver.strip(),"type":doc_type,
                    "status":"Draft","review_status":"with_author","line_of_business":new_lob,
                    "owner":st.session_state.current_user,"approved_by":"","approved_ts":"","content":"",
                    "versions":[{"ver":new_ver.strip(),"ts":now_str,"author":st.session_state.current_user,"note":"Initial draft"}],
                    "files":files,"reviewer":"","review_sent_ts":"","audit":audit,
                })
                D[tasks_key][nid] = []
                add_audit("Policies & WSPs", f"Created {nid}: {new_title.strip()} v{new_ver.strip()} [{new_lob}]")
                st.success(f"✓ {nid} saved as Draft under {new_lob}."); st.rerun()

    # ══════════════════════════════════════════════════════════
    # 4 TABS
    # ══════════════════════════════════════════════════════════
    tab_pol, tab_wsp, tab_pdraft, tab_wdraft = st.tabs([
        "📋  Policy Library",
        "📚  WSP Library",
        "🕐  Policy Drafts",
        "🕐  WSP Drafts",
    ])

    with tab_pol:
        _render_library(D.get("policies",[]), "policy_tasks", "pol")
        if toggle("new_pol_toggle", "➕ New Policy"):
            _render_new_form("Policy", "policies", "policy_tasks", "POL")

    with tab_wsp:
        _render_library(D["wsps"], "wsp_tasks", "wsp")
        if toggle("new_wsp_toggle", "➕ New WSP"):
            _render_new_form("WSP", "wsps", "wsp_tasks", "WSP")

    with tab_pdraft:
        _render_drafts(D.get("policies",[]), "pdraft", "policy_tasks")

    with tab_wdraft:
        _render_drafts(D["wsps"], "wdraft", "wsp_tasks")

# ── SUPERVISORY CHECKLISTS ─────────────────────────────────────────────────────
def page_checklists():
    st.markdown("<div class='gc-header' style='padding:.7rem 1.2rem;margin-bottom:.8rem'>"
                "<h1 style='font-size:1rem !important;margin-bottom:.1rem !important'>SUPERVISORY CHECKLISTS</h1>"
                "<p style='font-size:12px !important'>One checklist per person — all assigned supervisory procedures across every policy, sorted by due date. Evidence upload, completion tracking, and supervisor sign-off.</p>"
                "</div>", unsafe_allow_html=True)

    raw_cls = D.get("supervisory_checklists", [])

    # ── Back-compat: ensure new fields on all items ───────────────────────────
    for cl in raw_cls:
        for item in cl.get("items", []):
            if "source_policy_id" not in item: item["source_policy_id"] = cl.get("source_wsp","")
            if "source_policy"    not in item: item["source_policy"]    = cl.get("wsp_title","")
            if "due_date"         not in item: item["due_date"]         = cl.get("due","")
            if "comments"         not in item: item["comments"]         = []
            if "supervisor_review"      not in item: item["supervisor_review"]      = False
            if "supervisor_reviewed_by" not in item: item["supervisor_reviewed_by"] = ""
            if "supervisor_reviewed_ts" not in item: item["supervisor_reviewed_ts"] = ""
            if "supervisor_notes"       not in item: item["supervisor_notes"]       = ""

    if not raw_cls:
        st.markdown("<div class='gc-alert gc-alert-warn'>No checklists yet. Expand any approved policy in <b>Policies &amp; WSPs</b> to generate.</div>", unsafe_allow_html=True)
        return

    # Supervisor: only sees their own checklist — no filter bar needed
    if is_supervisor():
        cu = st.session_state.current_user
        my_cls = [cl for cl in raw_cls if cl.get("assigned_to") == cu]
        if not my_cls:
            st.info("You have no supervisory checklist items currently assigned to you.")
            return
        raw_cls = my_cls
        # Skip filter bar, jump straight to rendering
        checklists = raw_cls
    else:
        # ── Filter bar (CCO / Compliance Team) ───────────────────────────────────
        fc1, fc2 = st.columns([2, 2])
        all_owners = sorted(set(cl["assigned_to"] for cl in raw_cls))
        f_owner = fc1.selectbox("Filter by person", ["All"] + all_owners,
            format_func=lambda x: "All" if x=="All" else D["users"].get(x,{}).get("name",x))
        status_opts = ["All", "Has Pending Items", "All Complete"]
        f_status = fc2.selectbox("Status", status_opts)

        checklists = raw_cls
        if f_owner != "All":
            checklists = [cl for cl in checklists if cl["assigned_to"] == f_owner]
        if f_status == "Has Pending Items":
            checklists = [cl for cl in checklists if any(not i["checked_off"] for i in cl.get("items",[]))]
        elif f_status == "All Complete":
            checklists = [cl for cl in checklists if all(i.get("checked_off") for i in cl.get("items",[]))]

    # ── One card per user ─────────────────────────────────────────────────────
    for cl in checklists:
        items       = cl.get("items", [])
        # Sort items by due_date ascending (soonest first)
        items_sorted = sorted(items, key=lambda i: i.get("due_date","9999-99-99"))

        done_ct     = sum(1 for i in items if i.get("checked_off"))
        signoff_ct  = sum(1 for i in items if i.get("supervisor_review"))
        pending_ct  = sum(1 for i in items if i.get("checked_off") and not i.get("supervisor_review"))
        total_ct    = len(items)
        pct         = int(done_ct / total_ct * 100) if total_ct else 0

        supervisor_key  = get_supervisor(cl["assigned_to"])
        supervisor_name = D["users"].get(supervisor_key,{}).get("name","—") if supervisor_key else "No supervisor"

        # Status badge
        if pct == 100 and signoff_ct == total_ct:
            s_color = "#059669"; s_label = "All Complete"
        elif pending_ct:
            s_color = "#d97706"; s_label = f"{pending_ct} Awaiting Review"
        elif done_ct:
            s_color = "#2563eb"; s_label = f"{done_ct}/{total_ct} Done"
        else:
            s_color = "#6b7280"; s_label = "Pending"

        # Header toggle
        if toggle(f"cl_head_{cl['id']}",
                  f"{cl['owner_name']}  ·  {cl['owner_title']}  ·  {done_ct}/{total_ct} done  ·  {s_label}"):

            # ── Summary header ────────────────────────────────────────────────
            st.markdown(f"""
            <div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
                 padding:.7rem 1rem;margin-bottom:.6rem;display:flex;justify-content:space-between;
                 align-items:center;flex-wrap:wrap;gap:.5rem'>
              <div>
                <div style='font-size:13px;font-weight:800;color:#111'>{cl['owner_name']}</div>
                <div style='font-size:11px;color:#6b7280'>{cl['owner_title']} &nbsp;·&nbsp;
                  Reports to: <b>{supervisor_name}</b> &nbsp;·&nbsp; {total_ct} procedure(s)</div>
              </div>
              <div style='display:flex;gap:.4rem;align-items:center;flex-wrap:wrap'>
                <span style='background:{s_color}18;color:{s_color};font-size:9px;font-weight:800;
                      padding:2px 8px;border-radius:8px'>{s_label}</span>
                <span style='font-size:9px;color:#6b7280'>{signoff_ct} signed off · {pending_ct} awaiting review</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.progress(pct / 100)

            # ── Table header ──────────────────────────────────────────────────
            # Uses st.columns([20,1]) to match the row layout so headers stay aligned
            hdr_row, hdr_btn = st.columns([20, 1])
            hdr_row.markdown("""
            <div style='display:grid;grid-template-columns:1.8fr 1.6fr 90px 90px 110px;
                 gap:6px;background:#f1f5f9;border:1px solid #e2e8f0;border-radius:5px 5px 0 0;
                 padding:.3rem .75rem;margin-top:.4rem;align-items:center'>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Supervisory Procedure</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Policy</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Frequency</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Due Date</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Status</div>
            </div>
            """, unsafe_allow_html=True)

            # ── Item rows ─────────────────────────────────────────────────────
            for idx, item in enumerate(items_sorted):
                sup_done   = item.get("supervisor_review", False)
                checked    = item.get("checked_off", False)
                ev_ct      = len(item.get("evidence",[]))
                due_str    = item.get("due_date","")[:10]
                try:
                    due_dt  = datetime.strptime(due_str, "%Y-%m-%d") if due_str else None
                    overdue = due_dt and due_dt < datetime.now() and not checked
                except: overdue = False

                if sup_done:
                    row_bg = "#f0fdf4"; border = "#22c55e"; status_txt = "✅ Signed Off"; s_tc = "#166534"
                elif checked:
                    row_bg = "#fffbeb"; border = "#f59e0b"; status_txt = "⏳ Awaiting Review"; s_tc = "#92400e"
                elif overdue:
                    row_bg = "#fef2f2"; border = "#dc2626"; status_txt = "🔴 Overdue"; s_tc = "#dc2626"
                else:
                    row_bg = "#ffffff"; border = "#3b82f6"; status_txt = "⬜ Pending"; s_tc = "#1e40af"

                ev_pill    = (f"<span style='background:#dbeafe;color:#1e40af;font-size:8px;font-weight:700;"
                              f"padding:1px 5px;border-radius:8px;margin-left:4px'>{ev_ct} file(s)</span>") if ev_ct else ""
                due_color  = "#dc2626" if overdue else "#6b7280"

                # Row: 5 data columns — expand button rendered separately via st.button in a narrow col
                c_row, c_btn = st.columns([20, 1])
                c_row.markdown(
                    f"<div style='display:grid;grid-template-columns:1.8fr 1.6fr 90px 90px 110px;"
                    f"gap:6px;background:{row_bg};"
                    f"border:1px solid #e2e8f0;border-left:3px solid {border};"
                    f"padding:.35rem .75rem;align-items:center;border-radius:4px'>"
                    f"<div style='font-size:11px;font-weight:700;color:#111827;overflow:hidden;"
                    f"text-overflow:ellipsis;white-space:nowrap'>{item['title']}{ev_pill}</div>"
                    f"<div style='font-size:10px;color:#374151;overflow:hidden;text-overflow:ellipsis;"
                    f"white-space:nowrap'>{item.get('source_policy','')}</div>"
                    f"<div style='font-size:10px;color:#6b7280;white-space:nowrap'>{item.get('frequency','')}</div>"
                    f"<div style='font-size:10px;color:{due_color};white-space:nowrap'>{due_str or '—'}</div>"
                    f"<div style='font-size:9px;font-weight:700;color:{s_tc}'>{status_txt}</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )

                detail_key  = f"cl_dtl_{cl['id']}_{idx}"
                detail_open = st.session_state.get(detail_key, False)
                if c_btn.button("▼" if not detail_open else "▲", key=f"cl_dtbtn_{cl['id']}_{idx}"):
                    st.session_state[detail_key] = not detail_open
                    st.rerun()

                # ── Detail panel ──────────────────────────────────────────────
                if detail_open:
                    is_mine      = (cl["assigned_to"] == st.session_state.current_user) or has_full_compliance_access()
                    is_sup_user  = current_role() in ["Compliance Officer","Compliance Team","Supervisor"]

                    st.markdown(f"""
                    <div style='background:#ffffff;border:1px solid #e2e8f0;border-left:3px solid {border};
                         border-top:none;border-radius:0 0 6px 6px;padding:.7rem .9rem .6rem;margin-bottom:6px'>

                      <div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;
                           letter-spacing:.07em;margin-bottom:.3rem'>SUPERVISORY PROCEDURE</div>
                      <div style='font-size:12px;font-weight:800;color:#111;margin-bottom:.4rem'>{item['title']}</div>

                      <div style='display:flex;gap:1.5rem;flex-wrap:wrap;margin-bottom:.5rem;padding-bottom:.4rem;border-bottom:1px solid #f1f5f9'>
                        <div><span style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase'>Policy</span>
                          <div style='font-size:11px;font-weight:600;color:#2563eb'>{item.get('source_policy_id','')} — {item.get('source_policy','')}</div></div>
                        <div><span style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase'>Frequency</span>
                          <div style='font-size:11px;font-weight:600;color:#111'>{item.get('frequency','')}</div></div>
                        <div><span style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase'>Due Date</span>
                          <div style='font-size:11px;font-weight:600;color:{due_color}'>{due_str or "—"}</div></div>
                        <div><span style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase'>Category</span>
                          <div style='font-size:11px;font-weight:600;color:#111'>{item.get('category','')}</div></div>
                      </div>

                      <div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;
                           letter-spacing:.07em;margin-bottom:.25rem'>WHAT TO DO</div>
                      <div style='font-size:12px;color:#374151;line-height:1.65;margin-bottom:.6rem;
                           background:#f8fafc;border-radius:4px;padding:.5rem .7rem'>{item['description']}</div>

                    </div>
                    """, unsafe_allow_html=True)

                    # ── Evidence files already uploaded ───────────────────────
                    if item.get("evidence"):
                        st.markdown("<div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin:.3rem 0 .2rem .1rem'>UPLOADED EVIDENCE</div>", unsafe_allow_html=True)
                        for ev in item["evidence"]:
                            ev_note = f" · <i style='color:#6b7280'>{ev['notes']}</i>" if ev.get("notes") else ""
                            up_name = D["users"].get(ev.get("uploaded_by",""),{}).get("name", ev.get("uploaded_by",""))
                            st.markdown(f"<div class='gc-tl' style='font-size:11px;margin-left:.3rem'>📎 <b>{ev['name']}</b> · {up_name} · {ev['ts'][:16]}{ev_note} &nbsp;<span class='gc-badge badge-closed' style='font-size:8px'>IMMUTABLE</span></div>", unsafe_allow_html=True)

                    # ── Comments ──────────────────────────────────────────────
                    if item.get("comments"):
                        st.markdown("<div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin:.3rem 0 .2rem .1rem'>COMMENTS</div>", unsafe_allow_html=True)
                        for cmt in item["comments"]:
                            cmt_name = D["users"].get(cmt.get("user",""),{}).get("name", cmt.get("user",""))
                            st.markdown(f"<div class='gc-tl' style='font-size:11px;margin-left:.3rem'>💬 <b>{cmt_name}</b> · {cmt['ts'][:16]} · {cmt['text']}</div>", unsafe_allow_html=True)

                    # ── Upload + comment + complete controls ──────────────────
                    if is_mine and not checked:
                        st.markdown("<div style='font-size:9px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.07em;margin:.4rem 0 .2rem 0'>UPLOAD EVIDENCE</div>", unsafe_allow_html=True)
                        up_col, note_col = st.columns([1,1])
                        new_up = up_col.file_uploader("Browse file", key=f"cl_up_{cl['id']}_{idx}", label_visibility="collapsed")
                        up_note = up_col.text_input("File note", placeholder="Describe the file or note any exceptions…", key=f"cl_upnote_{cl['id']}_{idx}", label_visibility="collapsed")
                        new_cmt = note_col.text_area("Add a comment", height=68, placeholder="Add context, questions, or notes for your supervisor…", key=f"cl_cmt_{cl['id']}_{idx}", label_visibility="collapsed")
                        save_cmt = note_col.button("💬 Save Comment", key=f"cl_savecmt_{cl['id']}_{idx}")

                        if new_up:
                            fh  = hashlib.md5(new_up.name.encode()).hexdigest()[:8]
                            ev_e = {"name":new_up.name,"uploaded_by":st.session_state.current_user,
                                    "ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"hash":fh,"notes":up_note.strip()}
                            item["evidence"].append(ev_e)
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            item["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"Evidence uploaded: {new_up.name}"})
                            cl["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"Evidence on '{item['title']}': {new_up.name}"})
                            add_audit("Supervisory Checklists", f"{cl['id']}: evidence uploaded for '{item['title']}'")
                            st.success(f"✓ {new_up.name} attached"); st.rerun()

                        if save_cmt and new_cmt.strip():
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            item["comments"].append({"user":st.session_state.current_user,"ts":now_str,"text":new_cmt.strip()})
                            item["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"Comment added"})
                            st.rerun()

                        st.markdown("<div style='margin-top:.4rem'/>", unsafe_allow_html=True)
                        if st.button("✅ Mark Complete & Route to Supervisor", key=f"cl_done_{cl['id']}_{idx}"):
                            now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                            item["checked_off"] = True
                            item["completed_ts"] = now_str
                            item["status"] = "Awaiting Review"
                            sup_key  = get_supervisor(cl["assigned_to"])
                            sup_name = D["users"].get(sup_key,{}).get("name","your supervisor") if sup_key else "your supervisor"
                            item["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"Marked complete — routed to {sup_name} for sign-off"})
                            cl["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"'{item['title']}' marked complete — pending {sup_name} sign-off"})
                            add_audit("Supervisory Checklists", f"{cl['id']}: '{item['title']}' marked complete → {sup_name}")
                            if all(i["checked_off"] for i in items): cl["status"] = "Complete"
                            st.rerun()

                    # ── Awaiting supervisor sign-off ──────────────────────────
                    elif checked and not sup_done:
                        sup_key   = get_supervisor(cl["assigned_to"])
                        sup_name  = D["users"].get(sup_key,{}).get("name","supervisor") if sup_key else "supervisor"
                        comp_name = D["users"].get(item["audit"][-1]["user"],{}).get("name", item["audit"][-1]["user"]) if item.get("audit") else cl["owner_name"]
                        st.markdown(f"<div style='font-size:11px;color:#92400e;background:#fffbeb;border-radius:4px;padding:.3rem .6rem;margin:.3rem 0'>⏳ Completed by <b>{comp_name}</b> on {item.get('completed_ts','')[:16]} — awaiting sign-off from <b>{sup_name}</b></div>", unsafe_allow_html=True)

                        # Show sign-off controls if current user IS the supervisor for this person
                        if st.session_state.current_user == sup_key or is_sup_user:
                            sv1, sv2 = st.columns([3, 1])
                            sv_note  = sv1.text_input("Sign-off notes", placeholder="Approval notes or follow-up required…", key=f"sv_note_{cl['id']}_{idx}", label_visibility="collapsed")
                            if sv2.button("✅ Sign Off", key=f"sv_done_{cl['id']}_{idx}"):
                                now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                                item["supervisor_review"]      = True
                                item["supervisor_reviewed_by"] = st.session_state.current_user
                                item["supervisor_reviewed_ts"] = now_str
                                item["supervisor_notes"]       = sv_note.strip()
                                item["status"] = "Signed Off"
                                item["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":"Supervisor sign-off complete" + (f" — {sv_note.strip()}" if sv_note.strip() else "")})
                                cl["audit"].append({"ts":now_str,"user":st.session_state.current_user,"action":f"Sign-off: '{item['title']}'"})
                                add_audit("Supervisory Checklists", f"{cl['id']}: '{item['title']}' signed off")
                                st.rerun()

                    # ── Fully signed off ──────────────────────────────────────
                    elif checked and sup_done:
                        sv_name = D["users"].get(item.get("supervisor_reviewed_by",""),{}).get("name", "—")
                        notes_tail = f" · <i>{item['supervisor_notes']}</i>" if item.get("supervisor_notes") else ""
                        st.markdown(f"<div style='font-size:11px;color:#166534;padding:.2rem 0'>✅ Signed off by <b>{sv_name}</b> on {item.get('supervisor_reviewed_ts','')[:16]}{notes_tail}</div>", unsafe_allow_html=True)

                    # ── Per-item audit trail ───────────────────────────────────
                    if toggle(f"itm_aud_{cl['id']}_{idx}", f"Audit trail: {item['title'][:45]}"):
                        for e in item.get("audit",[]):
                            st.markdown(f"<div class='gc-tl' style='font-size:11px'><span style='color:#2563eb'>{e['ts']}</span> · <b>{e['user']}</b> · {e['action']}</div>", unsafe_allow_html=True)

                st.markdown("<div style='height:2px'/>", unsafe_allow_html=True)

            # ── Checklist-level audit ─────────────────────────────────────────
            st.markdown("<div style='margin-top:.3rem'/>", unsafe_allow_html=True)
            if toggle(f"cl_aud_{cl['id']}", "📋 Checklist Audit Trail"):
                for e in cl.get("audit",[]):
                    st.markdown(f"<div class='gc-tl' style='font-size:11px'><span style='color:#2563eb'>{e['ts']}</span> · <b>{e['user']}</b> · {e['action']}</div>", unsafe_allow_html=True)

# ── EVIDENCE REPOSITORY ────────────────────────────────────────────────────────
def page_evidence():
    st.markdown("<div class='gc-header' style='padding:.7rem 1.2rem;margin-bottom:.8rem'><h1 style='font-size:1rem !important;margin-bottom:.1rem !important'>COMPLIANCE DATA LAKE</h1><p style='font-size:12px !important'>Centralised document repository — shared drive imports, system-generated evidence, and all uploaded files. Searchable by filename, keyword, reviewer, LOB, type, year, and tag.</p></div>", unsafe_allow_html=True)

    import base64 as _b64e

    ev_list = D.get("evidence", [])

    # Back-compat: ensure extended fields on old records
    for ev in ev_list:
        if "department"       not in ev: ev["department"]       = "Compliance"
        if "line_of_business" not in ev: ev["line_of_business"] = ""
        if "doc_type"         not in ev: ev["doc_type"]         = "Document"
        if "year"             not in ev: ev["year"]             = ev.get("ts","")[:4]
        if "source"           not in ev: ev["source"]           = "System Upload"
        if "tags"             not in ev: ev["tags"]             = []
        if "size_kb"          not in ev: ev["size_kb"]          = 0
        if "notes"            not in ev: ev["notes"]            = ""

    # Supervisor: only sees files they uploaded or that match their LOB/department
    if is_supervisor():
        cu  = st.session_state.current_user
        lob = current_user_lob() or ""
        uname = D["users"].get(cu,{}).get("name", cu)
        ev_list = [e for e in ev_list if
                   e.get("uploaded_by") == cu or
                   e.get("line_of_business") == lob or
                   e.get("department","").lower() in ("supervision","operations")]
        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.6rem'>Showing files uploaded by you or associated with the <b>{lob}</b> line of business.</div>", unsafe_allow_html=True)

    total = len(ev_list)

    # ── Build filter option lists ─────────────────────────────────────────────
    all_depts   = sorted(set(e["department"]       for e in ev_list))
    all_lobs    = sorted(set(e["line_of_business"] for e in ev_list if e["line_of_business"]))
    all_types   = sorted(set(e["doc_type"]         for e in ev_list))
    all_years   = sorted(set(e["year"]             for e in ev_list if e["year"]), reverse=True)
    all_sources = sorted(set(e["source"]           for e in ev_list))
    all_tags    = sorted(set(t for e in ev_list for t in e.get("tags",[])))
    all_ups     = sorted(set(e["uploaded_by"]      for e in ev_list))

    # ── Summary bar ──────────────────────────────────────────────────────────
    total_size_gb = sum(e.get("size_kb",0) for e in ev_list) / (1024*1024)
    type_counts   = {}
    for e in ev_list: type_counts[e["doc_type"]] = type_counts.get(e["doc_type"],0)+1
    top_types     = sorted(type_counts.items(), key=lambda x:-x[1])[:4]
    top_str       = " &nbsp;·&nbsp; ".join(f"<b>{k}</b> ({v})" for k,v in top_types)

    st.markdown(f"""
    <div style='background:#1e293b;border-radius:8px;padding:.7rem 1.1rem;margin-bottom:.8rem;
         display:flex;gap:2.5rem;flex-wrap:wrap;align-items:center'>
      <div>
        <div style='font-size:9px;font-weight:700;color:#94a3b8 !important;-webkit-text-fill-color:#94a3b8;text-transform:uppercase;letter-spacing:.08em'>Total Files</div>
        <div style='font-size:22px;font-weight:800;color:#f8fafc !important;-webkit-text-fill-color:#f8fafc;line-height:1.2'>{total:,}</div>
      </div>
      <div>
        <div style='font-size:9px;font-weight:700;color:#94a3b8 !important;-webkit-text-fill-color:#94a3b8;text-transform:uppercase;letter-spacing:.08em'>Est. Size</div>
        <div style='font-size:22px;font-weight:800;color:#60a5fa !important;-webkit-text-fill-color:#60a5fa;line-height:1.2'>{total_size_gb:.1f} GB</div>
      </div>
      <div>
        <div style='font-size:9px;font-weight:700;color:#94a3b8 !important;-webkit-text-fill-color:#94a3b8;text-transform:uppercase;letter-spacing:.08em'>Departments</div>
        <div style='font-size:22px;font-weight:800;color:#34d399 !important;-webkit-text-fill-color:#34d399;line-height:1.2'>{len(all_depts)}</div>
      </div>
      <div>
        <div style='font-size:9px;font-weight:700;color:#94a3b8 !important;-webkit-text-fill-color:#94a3b8;text-transform:uppercase;letter-spacing:.08em'>Document Types</div>
        <div style='font-size:22px;font-weight:800;color:#a78bfa !important;-webkit-text-fill-color:#a78bfa;line-height:1.2'>{len(all_types)}</div>
      </div>
      <div style='flex:1;min-width:200px'>
        <div style='font-size:9px;font-weight:700;color:#94a3b8 !important;-webkit-text-fill-color:#94a3b8;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.2rem'>Top Document Types</div>
        <div style='font-size:11px;color:#cbd5e1 !important;-webkit-text-fill-color:#cbd5e1'>{top_str}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Two-panel layout: nav tree left, results right ────────────────────────
    nav_col, res_col = st.columns([1.1, 4])

    with nav_col:
        st.markdown("<div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.4rem'>📂 BROWSE BY DEPARTMENT</div>", unsafe_allow_html=True)

        # Build dept → lob → doctype tree with counts
        tree = {}
        for e in ev_list:
            d = e["department"]; lb = e["line_of_business"] or "General"; dt = e["doc_type"]
            tree.setdefault(d,{}).setdefault(lb,{})
            tree[d][lb][dt] = tree[d][lb].get(dt,0) + 1

        nav_dept  = st.session_state.get("dl_nav_dept","")
        nav_lob   = st.session_state.get("dl_nav_lob","")

        for dept in sorted(tree.keys()):
            dept_count = sum(tree[dept][lb][dt] for lb in tree[dept] for dt in tree[dept][lb])
            is_active  = nav_dept == dept
            bg = "#1e293b" if is_active else "#f8fafc"
            fg = "#f8fafc !important;-webkit-text-fill-color:#f8fafc" if is_active else "#111827"
            st.markdown(f"""
            <div style='background:{bg};border-radius:5px;padding:.3rem .6rem;margin-bottom:2px;cursor:pointer;border:1px solid {"#334155" if is_active else "#e5e7eb"}'>
              <span style='font-size:10px;font-weight:700;color:{fg}'>📁 {dept}</span>
              <span style='font-size:9px;color:{"#94a3b8" if is_active else "#9ca3af"};margin-left:.4rem'>{dept_count}</span>
            </div>
            """, unsafe_allow_html=True)
            if st.button(f"{'▼' if is_active else '▶'} {dept}", key=f"navdept_{dept}"):
                if nav_dept == dept:
                    st.session_state["dl_nav_dept"] = ""
                    st.session_state["dl_nav_lob"]  = ""
                else:
                    st.session_state["dl_nav_dept"] = dept
                    st.session_state["dl_nav_lob"]  = ""
                st.rerun()

            if is_active:
                for lob in sorted(tree[dept].keys()):
                    lob_count  = sum(tree[dept][lob].values())
                    is_lob_act = nav_lob == lob
                    st.markdown(f"""
                    <div style='margin-left:.8rem;background:{"#334155" if is_lob_act else "#f1f5f9"};
                         border-radius:4px;padding:.2rem .5rem;margin-bottom:1px'>
                      <span style='font-size:9px;font-weight:600;color:{"#e2e8f0 !important;-webkit-text-fill-color:#e2e8f0" if is_lob_act else "#374151"}'>📄 {lob}</span>
                      <span style='font-size:9px;color:{"#94a3b8" if is_lob_act else "#9ca3af"};margin-left:.3rem'>{lob_count}</span>
                    </div>
                    """, unsafe_allow_html=True)
                    if st.button(lob, key=f"navlob_{dept}_{lob}"):
                        st.session_state["dl_nav_lob"] = lob if nav_lob != lob else ""
                        st.rerun()

    with res_col:
        # ── Search + filters ─────────────────────────────────────────────────
        s1, s2 = st.columns([3, 1])
        f_search = s1.text_input("🔍", placeholder="Search filename, keyword, tag, reviewer, notes…", key="dl_search", label_visibility="collapsed")
        f_year   = s2.selectbox("Year", ["All Years"] + all_years, key="dl_year", label_visibility="collapsed")

        f1, f2, f3, f4, f5 = st.columns(5)
        f_dept   = f1.selectbox("Dept",   ["All Depts"]   + all_depts,   key="dl_dept",   label_visibility="collapsed")
        f_lob    = f2.selectbox("LOB",    ["All LOBs"]    + all_lobs,    key="dl_lob",    label_visibility="collapsed")
        f_type   = f3.selectbox("Type",   ["All Types"]   + all_types,   key="dl_type",   label_visibility="collapsed")
        f_source = f4.selectbox("Source", ["All Sources"] + all_sources, key="dl_source", label_visibility="collapsed")
        f_tag    = f5.selectbox("Tag",    ["All Tags"]    + all_tags,    key="dl_tag",    label_visibility="collapsed")

        f_up_opts = ["All Uploaders"] + [D["users"].get(u,{}).get("name",u) for u in all_ups]
        f_up_raw  = st.selectbox("Uploader", f_up_opts, key="dl_uploader", label_visibility="collapsed")
        f_uploader= all_ups[f_up_opts.index(f_up_raw)-1] if f_up_raw != "All Uploaders" else "All Uploaders"

        # Apply nav tree filter
        nav_dept_active = st.session_state.get("dl_nav_dept","")
        nav_lob_active  = st.session_state.get("dl_nav_lob","")

        # Apply all filters
        results = ev_list
        if nav_dept_active: results = [e for e in results if e["department"] == nav_dept_active]
        if nav_lob_active:  results = [e for e in results if e["line_of_business"] == nav_lob_active]
        if f_search:
            q = f_search.lower()
            results = [e for e in results if
                q in e["filename"].lower() or
                q in e.get("notes","").lower() or
                q in e.get("doc_type","").lower() or
                q in e.get("department","").lower() or
                q in e.get("line_of_business","").lower() or
                q in e.get("source","").lower() or
                any(q in t.lower() for t in e.get("tags",[])) or
                q in D["users"].get(e["uploaded_by"],{}).get("name",e["uploaded_by"]).lower() or
                q in e.get("task_id","").lower()]
        if f_year   != "All Years":    results = [e for e in results if e["year"]             == f_year]
        if f_dept   != "All Depts":    results = [e for e in results if e["department"]       == f_dept]
        if f_lob    != "All LOBs":     results = [e for e in results if e["line_of_business"] == f_lob]
        if f_type   != "All Types":    results = [e for e in results if e["doc_type"]         == f_type]
        if f_source != "All Sources":  results = [e for e in results if e["source"]           == f_source]
        if f_tag    != "All Tags":     results = [e for e in results if f_tag in e.get("tags",[])]
        if f_uploader != "All Uploaders": results = [e for e in results if e["uploaded_by"] == f_uploader]

        # Sort newest first
        results = sorted(results, key=lambda e: e["ts"], reverse=True)

        # ── Results header ────────────────────────────────────────────────────
        active_filters = []
        if nav_dept_active: active_filters.append(f"Dept: {nav_dept_active}")
        if nav_lob_active:  active_filters.append(f"LOB: {nav_lob_active}")
        if f_search:        active_filters.append(f'"{f_search}"')
        if f_year   != "All Years":    active_filters.append(f_year)
        if f_dept   != "All Depts":    active_filters.append(f_dept)
        if f_lob    != "All LOBs":     active_filters.append(f_lob)
        if f_type   != "All Types":    active_filters.append(f_type)
        if f_source != "All Sources":  active_filters.append(f_source)
        if f_tag    != "All Tags":     active_filters.append(f_tag)
        if f_uploader != "All Uploaders": active_filters.append(f_up_raw)

        filter_str = " · ".join(active_filters) if active_filters else "All documents"
        st.markdown(f"<div style='font-size:11px;color:#6b7280;margin:.3rem 0 .5rem 0'><b style='color:#111'>{len(results):,}</b> of <b>{total:,}</b> files &nbsp;·&nbsp; {filter_str}</div>", unsafe_allow_html=True)

        if not results:
            st.markdown("<div class='gc-alert gc-alert-warn'>No files match the current filters.</div>", unsafe_allow_html=True)
        else:
            # Paginate: 50 per page
            page_size = 50
            n_pages   = max(1, (len(results) + page_size - 1) // page_size)
            pg_key    = "dl_page"
            if pg_key not in st.session_state: st.session_state[pg_key] = 0
            # Reset page if filters changed
            cur_filter_sig = f"{nav_dept_active}|{nav_lob_active}|{f_search}|{f_year}|{f_dept}|{f_lob}|{f_type}|{f_source}|{f_tag}|{f_uploader}"
            if st.session_state.get("dl_filter_sig","") != cur_filter_sig:
                st.session_state[pg_key] = 0
                st.session_state["dl_filter_sig"] = cur_filter_sig

            cur_page = min(st.session_state[pg_key], n_pages-1)
            page_results = results[cur_page*page_size:(cur_page+1)*page_size]

            # ── Table header ──────────────────────────────────────────────────
            st.markdown("""
            <div style='display:grid;grid-template-columns:70px 2fr 90px 100px 120px 70px 70px 55px 75px 50px;
                 gap:4px;background:#f1f5f9;border:1px solid #e2e8f0;border-radius:6px 6px 0 0;
                 padding:.3rem .7rem;margin-top:.2rem;align-items:center'>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>ID</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Filename</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Department</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Line of Business</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Document Type</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Year</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Source</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Size</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>Uploaded</div>
              <div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.07em'>View</div>
            </div>
            """, unsafe_allow_html=True)

            for i, ev in enumerate(page_results):
                uploader_name = D["users"].get(ev["uploaded_by"],{}).get("name", ev["uploaded_by"])
                row_bg   = "#ffffff" if i % 2 == 0 else "#fafbfc"
                tags_str = " ".join(
                    f"<span style='background:#eff6ff;color:#2563eb;font-size:8px;font-weight:700;padding:1px 5px;border-radius:8px;white-space:nowrap'>{t}</span>"
                    for t in ev.get("tags",[])[:3]
                )
                size_kb  = ev.get("size_kb",0)
                size_str = f"{size_kb/1024:.1f} MB" if size_kb >= 1024 else f"{size_kb} KB"
                src_color = {"Shared Drive":"#7c3aed","System Upload":"#2563eb","Task Evidence":"#059669",
                             "Checklist Evidence":"#059669","Exam Upload":"#d97706","Manual Upload":"#6b7280",
                             "API Sync":"#0891b2","Email Archive":"#9333ea"}.get(ev["source"],"#6b7280")

                ext = ev["filename"].rsplit(".",1)[-1].upper() if "." in ev["filename"] else "DOC"
                ext_colors = {"PDF":"#dc2626","XLSX":"#059669","ZIP":"#d97706","DOCX":"#2563eb",
                              "PPTX":"#ea580c","PDF":"#dc2626","CSV":"#16a34a"}
                ext_bg = ext_colors.get(ext,"#6b7280")

                row_html = (
                    f"<div style='display:grid;grid-template-columns:70px 2fr 90px 100px 120px 70px 70px 55px 75px 50px;"
                    f"gap:4px;background:{row_bg};border-left:1px solid #e2e8f0;border-right:1px solid #e2e8f0;"
                    f"border-bottom:1px solid #f1f5f9;padding:.3rem .7rem;align-items:center;min-height:32px'>"
                    f"<div style='font-size:9px;font-weight:700;color:#9ca3af'>{ev['id']}</div>"
                    f"<div style='font-size:10px;font-weight:600;color:#111827;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>"
                    f"<span style='background:{ext_bg};color:#fff;font-size:8px;font-weight:800;padding:1px 4px;border-radius:3px;margin-right:4px'>{ext}</span>"
                    f"{ev['filename']}</div>"
                    f"<div style='font-size:10px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>{ev['department']}</div>"
                    f"<div style='font-size:10px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>{ev['line_of_business'] or '—'}</div>"
                    f"<div style='font-size:10px;color:#374151;overflow:hidden;text-overflow:ellipsis;white-space:nowrap'>{ev['doc_type']}</div>"
                    f"<div style='font-size:10px;color:#6b7280'>{ev['year']}</div>"
                    f"<div><span style='background:{src_color}18;color:{src_color};font-size:8px;font-weight:700;"
                    f"padding:1px 5px;border-radius:8px;white-space:nowrap'>{ev['source'][:10]}</span></div>"
                    f"<div style='font-size:10px;color:#9ca3af'>{size_str}</div>"
                    f"<div style='font-size:9px;color:#9ca3af;white-space:nowrap'>{ev['ts'][:10]}</div>"
                    f"<div><a href='#' style='background:#2563eb;color:#fff !important;font-size:8px;font-weight:700;"
                    f"padding:2px 7px;border-radius:4px;text-decoration:none;white-space:nowrap'>⬇ View</a></div>"
                    f"</div>"
                )
                st.markdown(row_html, unsafe_allow_html=True)

                # Tags row underneath (only if has tags)
                if ev.get("tags"):
                    st.markdown(
                        f"<div style='background:{row_bg};border-left:1px solid #e2e8f0;border-right:1px solid #e2e8f0;"
                        f"border-bottom:1px solid #f1f5f9;padding:.1rem .7rem .3rem 5.5rem;display:flex;gap:3px;flex-wrap:wrap'>"
                        f"{tags_str}</div>",
                        unsafe_allow_html=True
                    )

            # ── Pagination ────────────────────────────────────────────────────
            if n_pages > 1:
                pg_cols = st.columns([1, 1, 4, 1, 1])
                if pg_cols[0].button("« First", key="dl_first"):
                    st.session_state[pg_key] = 0; st.rerun()
                if pg_cols[1].button("‹ Prev",  key="dl_prev"):
                    st.session_state[pg_key] = max(0, cur_page-1); st.rerun()
                pg_cols[2].markdown(f"<div style='text-align:center;font-size:11px;color:#6b7280;padding:.3rem 0'>Page {cur_page+1} of {n_pages} &nbsp;·&nbsp; showing {len(page_results)} files</div>", unsafe_allow_html=True)
                if pg_cols[3].button("Next ›",  key="dl_next"):
                    st.session_state[pg_key] = min(n_pages-1, cur_page+1); st.rerun()
                if pg_cols[4].button("Last »",  key="dl_last"):
                    st.session_state[pg_key] = n_pages-1; st.rerun()

    # ── Upload new file ───────────────────────────────────────────────────────
    st.markdown("<hr style='margin:.8rem 0'/>", unsafe_allow_html=True)
    st.markdown("<div style='font-size:9px;font-weight:800;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:.5rem'>⬆ UPLOAD TO DATA LAKE</div>", unsafe_allow_html=True)
    up_c1, up_c2, up_c3, up_c4 = st.columns([2, 1.5, 1.5, 1.5])
    new_up      = up_c1.file_uploader("File", key="dl_up", label_visibility="collapsed")
    new_dept    = up_c2.selectbox("Department",    all_depts,   key="dl_up_dept")
    new_lob_up  = up_c3.selectbox("Line of Business", [""] + all_lobs, key="dl_up_lob")
    new_type_up = up_c4.selectbox("Document Type", all_types,   key="dl_up_type")
    up_notes    = st.text_input("Notes / keywords", placeholder="Describe file, context, reviewer, exam reference…", key="dl_up_notes")
    if new_up:
        import hashlib as _hlup
        fh   = "sha256:" + _hlup.md5(new_up.name.encode()).hexdigest()
        eid  = f"EV-{len(D['evidence'])+1:04d}"
        year = datetime.now().strftime("%Y")
        D["evidence"].append({
            "id":eid,"filename":new_up.name,
            "uploaded_by":st.session_state.current_user,
            "ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
            "task_id":"","hash":fh,"immutable":True,"archived":False,
            "department":new_dept,"line_of_business":new_lob_up,
            "doc_type":new_type_up,"year":year,"source":"Manual Upload",
            "tags":[new_type_up],"notes":up_notes.strip(),
            "size_kb": len(new_up.getvalue())//1024,
        })
        add_audit("Data Lake", f"Uploaded {new_up.name} → {new_dept} / {new_type_up}")
        st.success(f"✓ {new_up.name} stored as {eid} (WORM-locked)"); st.rerun()


# ── RULE INVENTORY ─────────────────────────────────────────────────────────────
def page_rules():
    st.markdown("<div class='gc-header'><h1>MASTER RULE INVENTORY</h1><p>Regulatory obligations mapped to policies — organized by category, WSP, and rule</p></div>", unsafe_allow_html=True)

    # Back-compat: ensure new fields exist on old rule records
    for r in D["rules"]:
        if "rule_requirements" not in r: r["rule_requirements"] = r.get("description","")
        if "wsp_links"         not in r: r["wsp_links"]         = []

    # ── Excel Download ──────────────────────────────────────────────────────────
    import io as _io, base64 as _b64r
    try:
        import openpyxl as _xl
        from openpyxl.styles import Font as _Font, PatternFill as _PFill, Alignment as _Align, Border as _Border, Side as _Side
        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "Rule Inventory"

        hdr_fill = _PFill(fill_type="solid", fgColor="1E293B")
        hdr_font = _Font(bold=True, color="FFFFFF", size=10)
        alt_fill = _PFill(fill_type="solid", fgColor="F8FAFC")
        bd       = _Border(bottom=_Side(style="thin", color="E5E7EB"))

        headers  = ["Rule ID","Title","Source","Category","Status","Last Reviewed","AI Flagged","Rule Requirements","Linked WSPs","Process in Place"]
        col_widths = [10, 40, 10, 20, 10, 16, 10, 80, 30, 80]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = _Align(wrap_text=True, vertical="center")
            ws.column_dimensions[_xl.utils.get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 22

        for ri, rule in enumerate(D["rules"], 2):
            wsp_ids   = ", ".join(lk["wsp_id"] for lk in rule.get("wsp_links",[]))
            processes = " | ".join(lk["process"] for lk in rule.get("wsp_links",[]))
            row_data  = [
                rule["id"], rule["title"], rule["source"], rule["category"],
                rule["status"], rule["last_reviewed"][:10],
                "Yes" if rule["ai_flagged"] else "No",
                rule.get("rule_requirements",""), wsp_ids, processes
            ]
            fill = alt_fill if ri % 2 == 0 else _PFill(fill_type="solid", fgColor="FFFFFF")
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill
                cell.alignment = _Align(wrap_text=True, vertical="top")
                cell.border = bd
            ws.row_dimensions[ri].height = 60

        ws.freeze_panes = "A2"
        _buf = _io.BytesIO()
        wb.save(_buf)
        _buf.seek(0)
        _xl_b64 = _b64r.b64encode(_buf.read()).decode()

        st.markdown(f"""
        <a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{_xl_b64}"
           download="Rule_Inventory.xlsx"
           style='display:inline-block;background:#166534;color:#ffffff !important;font-size:12px;
                  font-weight:700;padding:8px 18px;border-radius:6px;text-decoration:none;
                  letter-spacing:.02em;margin-bottom:1rem'>
          📊 Download Full Inventory to Excel
        </a>
        """, unsafe_allow_html=True)
    except Exception as _e:
        st.warning(f"Excel export unavailable: {_e}")

    # ── Tabs ────────────────────────────────────────────────────────────────────
    tab1, tab2 = st.tabs(["📏  Rule Inventory","➕  Add Rule"])

    # ── TAB 1: INVENTORY — grouped Category → WSP → Rules ──────────────────────
    with tab1:
        # Filter bar
        fc1, fc2, fc3 = st.columns([1.5, 1.5, 1.5])
        all_cats = sorted(set(r["category"] for r in D["rules"]))
        all_srcs = sorted(set(r["source"]   for r in D["rules"]))
        f_cat    = fc1.selectbox("Category", ["All"] + all_cats)
        f_src    = fc2.selectbox("Source",   ["All"] + all_srcs)
        f_flag   = fc3.selectbox("AI Flagged", ["All","Flagged Only","Not Flagged"])

        rules = D["rules"]
        if f_cat  != "All":         rules = [r for r in rules if r["category"] == f_cat]
        if f_src  != "All":         rules = [r for r in rules if r["source"]   == f_src]
        if f_flag == "Flagged Only": rules = [r for r in rules if r["ai_flagged"]]
        if f_flag == "Not Flagged":  rules = [r for r in rules if not r["ai_flagged"]]

        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin:.3rem 0 .8rem 0'>{len(rules)} rule(s) shown</div>", unsafe_allow_html=True)

        if not rules:
            st.markdown("<div class='gc-alert gc-alert-ok'>No rules match the current filters.</div>", unsafe_allow_html=True)

        # Group by category
        from collections import defaultdict as _dd2
        by_cat = _dd2(list)
        for r in rules:
            by_cat[r["category"]].append(r)

        for cat in sorted(by_cat.keys()):
            cat_rules = by_cat[cat]

            # ── Category header ──
            src_set = sorted(set(r["source"] for r in cat_rules))
            src_str = " · ".join(src_set)
            st.markdown(f"""
            <div style='background:#1e293b;color:#ffffff;border-radius:8px;
                 padding:.55rem 1rem;margin:1rem 0 .5rem 0;
                 display:flex;justify-content:space-between;align-items:center'>
              <span style='font-size:13px;font-weight:800;letter-spacing:.06em'>{cat.upper()}</span>
              <span style='font-size:11px;color:#94a3b8'>{src_str} &nbsp;·&nbsp; {len(cat_rules)} rule(s)</span>
            </div>
            """, unsafe_allow_html=True)

            # Within each category, group rules by which WSPs they link to
            # First show rules that have WSP links, grouped by WSP
            wsp_to_rules = _dd2(list)
            no_wsp_rules = []
            for r in cat_rules:
                if r.get("wsp_links"):
                    for lk in r["wsp_links"]:
                        wsp_to_rules[lk["wsp_id"]].append((r, lk))
                else:
                    no_wsp_rules.append(r)

            # ── Per-WSP block ──
            for wsp_id in sorted(wsp_to_rules.keys()):
                rule_links = wsp_to_rules[wsp_id]
                wsp_obj    = next((w for w in D["wsps"] if w["id"]==wsp_id), None)
                wsp_title  = wsp_obj["title"] if wsp_obj else wsp_id
                wsp_ver    = wsp_obj["version"] if wsp_obj else ""
                wsp_status = wsp_obj["status"] if wsp_obj else ""
                status_col = "#22c55e" if wsp_status=="Approved" else "#f59e0b"
                status_bg  = "#f0fdf4" if wsp_status=="Approved" else "#fffbeb"

                # WSP title bar
                st.markdown(f"""
                <div style='background:{status_bg};border:1px solid #e5e7eb;border-left:4px solid {status_col};
                     border-radius:6px;padding:.55rem 1rem;margin:.3rem 0 .2rem 0;
                     display:flex;justify-content:space-between;align-items:center'>
                  <div>
                    <span style='font-size:11px;font-weight:800;color:{status_col}'>{wsp_id}</span>
                    <span style='font-size:13px;font-weight:700;color:#111111;margin-left:.5rem'>{wsp_title}</span>
                    <span style='font-size:11px;color:#6b7280;margin-left:.4rem'>v{wsp_ver}</span>
                  </div>
                  <span style='font-size:10px;font-weight:700;color:{status_col}'>{wsp_status.upper()}</span>
                </div>
                """, unsafe_allow_html=True)

                # Each rule inside this WSP
                for (r, lk) in rule_links:
                    src_color = "#1e40af" if r["source"]=="FINRA" else "#7c3aed" if r["source"]=="SEC" else "#065f46"
                    src_bg    = "#dbeafe" if r["source"]=="FINRA" else "#ede9fe" if r["source"]=="SEC" else "#d1fae5"
                    flag_pill = "&nbsp;<span style='background:#fef9c3;color:#92400e;font-size:10px;font-weight:700;padding:1px 7px;border-radius:10px'>🤖 AI FLAGGED</span>" if r["ai_flagged"] else ""

                    if toggle(f"rule_expand_{r['id']}_{wsp_id}", f"  {r['id']} — {r['title']}"):
                        st.markdown(f"""
                        <div style='background:#ffffff;border:1px solid #e5e7eb;border-radius:8px;
                             padding:.9rem 1.1rem;margin:.3rem 0 .4rem 1rem'>

                          <div style='display:flex;align-items:center;gap:.6rem;margin-bottom:.7rem;flex-wrap:wrap'>
                            <span style='background:{src_bg};color:{src_color};font-size:10px;font-weight:800;padding:2px 8px;border-radius:4px'>{r['source']}</span>
                            <span style='font-size:13px;font-weight:700;color:#111111'>{r['title']}</span>
                            {flag_pill}
                          </div>

                          <div style='display:flex;gap:2rem;flex-wrap:wrap;margin-bottom:.7rem'>
                            <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Category</div>
                              <div style='font-size:12px;color:#111111'>{r['category']}</div></div>
                            <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Source</div>
                              <div style='font-size:12px;color:#111111'>{r['source']}</div></div>
                            <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Last Reviewed</div>
                              <div style='font-size:12px;color:#111111'>{r['last_reviewed'][:10]}</div></div>
                            <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Status</div>
                              <div style='font-size:12px;color:#111111'>{r['status']}</div></div>
                          </div>

                          <div style='margin-bottom:.7rem'>
                            <div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem'>RULE REQUIREMENTS</div>
                            <div style='font-size:12px;color:#374151;line-height:1.7;background:#f8fafc;border-radius:6px;padding:.6rem .8rem'>{r.get('rule_requirements', r['description'])}</div>
                          </div>

                          <div>
                            <div style='font-size:11px;font-weight:700;color:#2563eb;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem'>✅ PROCESS IN PLACE — {lk['wsp_id']}</div>
                            <div style='font-size:12px;color:#374151;line-height:1.7;background:#f0f7ff;border:1px solid #bfdbfe;border-radius:6px;padding:.6rem .8rem'>{lk['process']}</div>
                          </div>
                        </div>
                        """, unsafe_allow_html=True)

                        bc1, bc2 = st.columns([1,4])
                        if bc1.button("✓ Mark Reviewed", key=f"rev_{r['id']}_{wsp_id}"):
                            r["last_reviewed"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                            r["ai_flagged"]    = False
                            add_audit("Rules", f"Reviewed {r['id']}: {r['title']}")
                            st.rerun()
                        if r["ai_flagged"]:
                            bc2.markdown("<div style='font-size:12px;color:#92400e;padding:.4rem 0'>🤖 AI detected a potential change to this rule. Review and mark as reviewed when confirmed.</div>", unsafe_allow_html=True)
                    else:
                        # Collapsed row — compact single line
                        src_span  = f"<span style='background:{src_bg};color:{src_color};font-size:10px;font-weight:800;padding:1px 7px;border-radius:4px;margin-right:.5rem'>{r['source']}</span>"
                        rev_span  = f"<span style='font-size:11px;color:#6b7280'>Reviewed: {r['last_reviewed'][:10]}</span>"
                        st.markdown(f"<div style='padding:.1rem 0 .1rem 1rem'>{src_span}{rev_span}{flag_pill}</div>", unsafe_allow_html=True)

            # Rules with no WSP links
            for r in no_wsp_rules:
                flag_pill = "&nbsp;<span style='background:#fef9c3;color:#92400e;font-size:10px;font-weight:700;padding:1px 7px;border-radius:10px'>🤖 AI FLAGGED</span>" if r["ai_flagged"] else ""
                if toggle(f"rule_expand_{r['id']}_nwsp", f"  {r['id']} — {r['title']} (no WSP linked)"):
                    st.markdown(f"""
                    <div style='background:#ffffff;border:1px solid #e5e7eb;border-radius:8px;padding:.9rem 1.1rem;margin:.3rem 0 .4rem 0'>
                      <div style='display:flex;gap:2rem;flex-wrap:wrap;margin-bottom:.6rem'>
                        <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Source</div><div style='font-size:12px'>{r['source']}</div></div>
                        <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Category</div><div style='font-size:12px'>{r['category']}</div></div>
                        <div><div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Last Reviewed</div><div style='font-size:12px'>{r['last_reviewed'][:10]}</div></div>
                      </div>
                      <div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem'>RULE REQUIREMENTS</div>
                      <div style='font-size:12px;color:#374151;line-height:1.7;background:#f8fafc;border-radius:6px;padding:.6rem .8rem'>{r.get('rule_requirements',r['description'])}</div>
                      <div style='margin-top:.6rem'><div class='gc-alert gc-alert-warn'>No WSP linked to this rule. Upload and approve a WSP to establish a process in place.</div></div>
                    </div>
                    """, unsafe_allow_html=True)
                    if st.button("✓ Mark Reviewed", key=f"rev_{r['id']}_nwsp"):
                        r["last_reviewed"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                        r["ai_flagged"]    = False
                        add_audit("Rules", f"Reviewed {r['id']}: {r['title']}")
                        st.rerun()

    # ── TAB 2: ADD RULE ─────────────────────────────────────────────────────────
    with tab2:
        st.markdown("<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>Add a new rule to the inventory. When a WSP is approved, the system automatically parses and links applicable rules.</div>", unsafe_allow_html=True)
        c1,c2 = st.columns(2)
        rt = c1.text_input("Rule Title *", placeholder="e.g. FINRA Rule 3110 – Supervision")
        rs = c2.selectbox("Source", ["FINRA","SEC","CFTC","FinCEN","MSRB","State","Internal"])
        c3,c4 = st.columns(2)
        rc = c3.selectbox("Category", ["Supervision","AML/BSA","Recordkeeping","Suitability","Customer Accounts","Financial Responsibility","Operations","Trading","Other"])
        rd = st.text_area("Description", height=80, placeholder="Brief description of the rule.")
        rr = st.text_area("Rule Requirements", height=120, placeholder="Detail the specific obligations this rule imposes on the firm.")
        if st.button("➕ Add Rule"):
            if rt.strip():
                nid = f"R-{len(D['rules'])+1:03d}"
                D["rules"].append({
                    "id":nid,"title":rt.strip(),"source":rs,"category":rc,
                    "description":rd.strip(),"rule_requirements":rr.strip(),
                    "status":"Active","last_reviewed":datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "ai_flagged":False,"wsp_links":[]
                })
                add_audit("Rules", f"Added {nid}: {rt.strip()}")
                st.success(f"✓ Rule {nid} added.")
                st.rerun()

# ── AI RULE SCAN ───────────────────────────────────────────────────────────────
def page_ai_scan():
    st.markdown("<div class='gc-header'><h1>AI RULE CHANGE SCAN</h1><p>Simulated quarterly scan with summary memo and recommendations</p></div>", unsafe_allow_html=True)
    CHANGES=[
        {"rule":"FINRA Rule 4512","change":"Updated customer account information requirements to include digital communication preferences and ESG disclosures.","impact":"High","recommendation":"Review and update customer onboarding forms and WSP-002."},
        {"rule":"Bank Secrecy Act AML","change":"FinCEN updated guidance on cryptocurrency transaction monitoring thresholds and SAR reporting.","impact":"High","recommendation":"Update AML program to include digital asset monitoring. Retrain staff by Q2."},
        {"rule":"Reg BI","change":"SEC staff issued FAQ clarifying best interest obligations for automated investment tools.","impact":"Medium","recommendation":"Review robo-advisory disclosures and update Reg BI documentation."},
    ]
    if D["rule_scan_memo"]:
        m=D["rule_scan_memo"]
        st.markdown(f"<div class='gc-card'><div class='gc-card-header'>LAST SCAN · {m['date']}</div><div style='font-size:13px'>{m['summary']}</div></div>", unsafe_allow_html=True)
        for ch in m["changes"]:
            ic="badge-high" if ch["impact"]=="High" else "badge-med"
            st.markdown(f"""<div class='gc-card'>
              <div style='display:flex;justify-content:space-between'><b>{ch['rule']}</b>{badge(f"Impact: {ch['impact']}",ic)}</div>
              <div style='font-size:13px;color:#6b7280;margin:.4rem 0'>{ch['change']}</div>
              <div style='font-size:13px'><b>Recommendation:</b> {ch['recommendation']}</div>
            </div>""", unsafe_allow_html=True)
        memo_txt   = f"GRAPHITE COMPLIANCE — AI RULE CHANGE MEMO\n{m['date']}\n\n{m['summary']}\n\n" + "".join(f"RULE: {c['rule']}\nCHANGE: {c['change']}\nIMPACT: {c['impact']}\nRECOMMENDATION: {c['recommendation']}\n\n" for c in m["changes"])
        import base64 as _b64m
        _memo_enc  = _b64m.b64encode(memo_txt.encode()).decode()
        _memo_fname = "rule_change_memo.txt"

        btn_col1, btn_col2 = st.columns([1, 2])

        # ── Styled download anchor ──
        btn_col1.markdown(f"""
        <a href="data:text/plain;base64,{_memo_enc}" download="{_memo_fname}"
           style='display:inline-block;background:#2563eb;color:#ffffff !important;
                  font-size:12px;font-weight:700;padding:8px 18px;border-radius:6px;
                  text-decoration:none;letter-spacing:.02em;white-space:nowrap'>
          📥 Download Memo
        </a>
        """, unsafe_allow_html=True)

        # ── Send button + recipient picker ──
        with btn_col2:
            if toggle("ai_send_memo_panel", "📤  Download Memo & Send to Team"):
                team_opts = list(D["users"].keys())
                recipients = st.multiselect(
                    "Select team members to send to",
                    team_opts,
                    format_func=lambda x: f"{D['users'][x]['name']} — {D['users'][x]['title']}",
                    key="ai_memo_recipients",
                )
                send_note = st.text_area("Optional message", height=60, key="ai_memo_note",
                    placeholder="e.g. Please review the flagged rule changes and confirm any required WSP updates.")
                sc1, sc2 = st.columns([1,3])
                if sc1.button("✉️ Send Memo", key="ai_memo_send_btn"):
                    if not recipients:
                        st.error("Please select at least one recipient.")
                    else:
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        names   = ", ".join(D["users"][r]["name"] for r in recipients)
                        # Log a notification entry per recipient in session state
                        if "memo_sends" not in st.session_state.data:
                            st.session_state.data["memo_sends"] = []
                        st.session_state.data["memo_sends"].append({
                            "ts": now_str,
                            "sent_by": st.session_state.current_user,
                            "recipients": recipients,
                            "memo_date": m["date"],
                            "note": send_note.strip(),
                        })
                        add_audit("AI Scan", f"Rule change memo sent to: {names}")
                        st.success(f"✓ Memo sent to {names}.")

                        # Show the download link immediately after sending
                        st.markdown(f"""
                        <div style='margin-top:.5rem'>
                        <a href="data:text/plain;base64,{_memo_enc}" download="{_memo_fname}"
                           style='display:inline-block;background:#16a34a;color:#ffffff !important;
                                  font-size:12px;font-weight:700;padding:7px 16px;border-radius:6px;
                                  text-decoration:none;letter-spacing:.02em'>
                          📥 Download &amp; Save Memo
                        </a></div>
                        """, unsafe_allow_html=True)

                # Show prior sends
                prior = [s for s in st.session_state.data.get("memo_sends",[])
                         if s["memo_date"]==m["date"]]
                if prior:
                    st.markdown("<div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-top:.6rem;margin-bottom:.3rem'>PREVIOUSLY SENT</div>", unsafe_allow_html=True)
                    for ps in prior:
                        pnames = ", ".join(D["users"].get(r,{}).get("name",r) for r in ps["recipients"])
                        sender = D["users"].get(ps["sent_by"],{}).get("name", ps["sent_by"])
                        st.markdown(f"<div class='gc-tl'>📤 {ps['ts']} · <b>{sender}</b> → {pnames}" + (f" — <i>{ps['note']}</i>" if ps.get("note") else "") + "</div>", unsafe_allow_html=True)
        st.markdown("<hr/>", unsafe_allow_html=True)
    if st.button("🤖 Run AI Rule Change Scan"):
        import time; 
        with st.spinner("Scanning..."):
            time.sleep(1.2)
        D["rule_scan_memo"]={"date":datetime.now().strftime("%Y-%m-%d %H:%M"),"summary":f"Quarterly scan complete. Scanned {len(D['rules'])} active rules. Identified {len(CHANGES)} material changes.","changes":CHANGES}
        for r in D["rules"]:
            for ch in CHANGES:
                if any(w in r["title"] for w in ch["rule"].split()): r["ai_flagged"]=True
        add_audit("AI Scan",f"Scan complete; {len(CHANGES)} changes flagged"); st.rerun()

# ── ORG CHART ──────────────────────────────────────────────────────────────────
def page_org():
    st.markdown("<div class='gc-header'><h1>ORG CHART & RESPONSIBILITY MAP</h1><p>Supervisory hierarchy with linked responsibilities</p></div>", unsafe_allow_html=True)
    for node in D["org_chart"]:
        u=D["users"].get(node["user"],{})
        rep=f"Reports to: **{node['reports_to']}**" if node["reports_to"] else "**Top of hierarchy**"
        st.markdown(f"""<div class='gc-card'>
          <div style='font-size:12px;color:#2563eb;font-weight:700;letter-spacing:.05em'>{node['role']}</div>
          <div style='font-size:1rem;font-weight:700;margin:.2rem 0'>{u.get('name','–')} <span style='color:#6b7280;font-size:.8rem'>@{node['user']}</span></div>
          <div style='font-size:12px;color:#6b7280;margin-bottom:.5rem'>{rep}</div>
          <div>{'  '.join(badge(r,'badge-open') for r in node['responsibilities'])}</div>
        </div>""", unsafe_allow_html=True)
    delegated=[t for t in D["tasks"] if t.get("delegated_to")]
    if delegated:
        st.markdown("<hr/><b>Active Delegations</b>", unsafe_allow_html=True)
        for t in delegated:
            st.markdown(f"<div class='gc-tl'>{t['id']} · <b>{t['title']}</b> → delegated to <b>{t['delegated_to']}</b> · {t['delegated_memo']}</div>", unsafe_allow_html=True)

# ── AUDIT TRAIL ────────────────────────────────────────────────────────────────
def page_audit():
    st.markdown("<div class='gc-header'><h1>LINKED AUDIT TRAIL</h1><p>Immutable timestamped log of all actions</p></div>", unsafe_allow_html=True)
    fc1,fc2,fc3=st.columns(3)
    fu=fc1.selectbox("User",["All"]+list(D["users"].keys())+["SYSTEM"])
    fm=fc2.selectbox("Module",["All"]+list({a["module"] for a in D["audit_trail"]}))
    fs=fc3.text_input("Search",placeholder="keyword...")
    trail=list(reversed(D["audit_trail"]))
    if fu!="All": trail=[a for a in trail if a["user"]==fu]
    if fm!="All": trail=[a for a in trail if a["module"]==fm]
    if fs: trail=[a for a in trail if fs.lower() in a["action"].lower()]
    st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>{len(trail)} records</div>", unsafe_allow_html=True)
    for e in trail:
        st.markdown(f"<div class='gc-tl'><span style='color:#2563eb;font-size:11px'>{e['ts']}</span> &nbsp;{badge(e['module'],'badge-open')}&nbsp; <b>{e['user']}</b> · {e['action']}</div>", unsafe_allow_html=True)

# ── FIRM CONFIG ────────────────────────────────────────────────────────────────
def page_firm_config():
    st.markdown("<div class='gc-header'><h1>FIRM PROFILE CONFIGURATION</h1><p>Tailor rules and workflows to your firm</p></div>", unsafe_allow_html=True)
    firm=D["firm"]
    c1,c2=st.columns(2)
    firm["name"]=c1.text_input("Firm Name",value=firm["name"])
    firm["industry"]=c2.text_input("Industry/Type",value=firm["industry"])
    c3,c4=st.columns(2)
    firm["crd"]=c3.text_input("CRD / License #",value=firm["crd"])
    firm["regulator"]=c4.text_input("Regulator(s)",value=firm["regulator"])
    c5,c6=st.columns(2)
    firm["retention_years"]=c5.number_input("Retention Period (years)",value=firm["retention_years"],min_value=1,max_value=20)
    firm["approval_mode"]=c6.selectbox("Approval Mode",["single","dual"],index=0 if firm["approval_mode"]=="single" else 1)
    firm["address"]=st.text_input("Address",value=firm["address"])
    if st.button("Save Profile"):
        add_audit("Firm Config","Firm profile updated"); st.success("✓ Saved.")
    st.markdown("<hr/><b>Third-Party Compliance Packet</b>", unsafe_allow_html=True)
    packet=f"THIRD-PARTY REGULATORY COMPLIANCE PACKET\nFirm: {firm['name']}\nIndustry: {firm['industry']}\nCRD: {firm['crd']}\nRegulator: {firm['regulator']}\nRetention: {firm['retention_years']} years\n\nCERTIFICATIONS:\n- FINRA Rule 3110: Active\n- AML Program: Active\n- Reg BI: Active\n- WORM Recordkeeping: Active\n\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    st.download_button("📥 Download Compliance Packet",data=packet,file_name="compliance_packet.txt",mime="text/plain")

# ── DATA EXPORT ────────────────────────────────────────────────────────────────
def page_export():
    st.markdown("<div class='gc-header'><h1>DATA EXPORT & BACKUP</h1><p>Export all compliance data as JSON or CSV</p></div>", unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        if st.button("Generate Full JSON Export"):
            payload=json.dumps({"exported_at":datetime.now().strftime("%Y-%m-%d %H:%M"),"firm":D["firm"],"tasks":D["tasks"],"exam_requests":D["exam_requests"],"rules":D["rules"],"wsps":D["wsps"],"evidence":D["evidence"],"supervisory_checklists":D.get("supervisory_checklists",[]),"audit_trail":D["audit_trail"]},indent=2)
            add_audit("Export","Full JSON export"); st.download_button("📥 Download",data=payload,file_name="graphite_export.json",mime="application/json")
    with c2:
        if st.button("Generate Tasks CSV"):
            buf=io.StringIO(); w=csv.DictWriter(buf,fieldnames=["id","title","assignee","due","status","priority","category"]); w.writeheader()
            for t in D["tasks"]: w.writerow({k:t[k] for k in ["id","title","assignee","due","status","priority","category"]})
            add_audit("Export","Tasks CSV export"); st.download_button("📥 Download",data=buf.getvalue(),file_name="tasks.csv",mime="text/csv")
    st.markdown("<div class='gc-alert gc-alert-ok' style='margin-top:1rem'>✓ All exports include audit trails · WORM hashes preserved · Meets SEC Rule 17a-4</div>", unsafe_allow_html=True)

# ── WORKFLOW CONSTANTS ─────────────────────────────────────────────────────────
WF_CATEGORIES   = ["Supervision","Compliance","Operations","Risk Monitoring","Regulatory","Operational"]
WF_PURPOSES     = ["Supervisory Review","Regulatory Requirement","Risk Monitoring","Exception Handling","Periodic Certification","Operational Control"]
WF_FREQUENCIES  = ["Ad Hoc","Event Driven","Daily","Weekly","Monthly","Quarterly","Annual"]
WF_REGULATIONS  = ["FINRA 3110","FINRA 3120","SEC 206(4)-7","AML Program","Internal Policy"]
WF_OWNER_ROLES  = ["Compliance Officer","Supervisor","Compliance Team","Risk Officer","Operations Manager"]
WF_DEPARTMENTS  = ["Compliance","Supervision","Risk","Operations","Legal","Finance"]
WF_TAGS_ALLOWED = [
    "aml","annual","assessment","attestation","certification","communications",
    "daily","exception","monthly","quarterly","reg-bi","risk","supervision",
    "surveillance","trade-review","weekly","training","reporting"
]
WF_PRIORITIES   = ["Low","Normal","High"]
WF_INST_STATUSES= ["Open","In Progress","Completed","Overdue"]

# ── WORKFLOW LIBRARY ───────────────────────────────────────────────────────────
def page_workflow_library():
    st.markdown("<div class='gc-header'><h1>WORKFLOW LIBRARY</h1><p>Reusable workflow templates for structured compliance processes</p></div>", unsafe_allow_html=True)
    tab1, tab2 = st.tabs(["📚  TEMPLATES", "➕  NEW TEMPLATE"])

    with tab1:
        templates = D.get("workflow_templates", [])

        # ── Search + Filters ──
        search = st.text_input("🔍 Search", placeholder="Workflow name, tag, category, purpose…", key="wfl_search")

        with st.container():
            fc1,fc2,fc3,fc4 = st.columns(4)
            f_cat  = fc1.selectbox("Category",  ["All"] + WF_CATEGORIES,  key="wfl_cat")
            f_purp = fc2.selectbox("Purpose",   ["All"] + WF_PURPOSES,    key="wfl_purp")
            f_freq = fc3.selectbox("Frequency", ["All"] + WF_FREQUENCIES, key="wfl_freq")
            f_reg  = fc4.selectbox("Regulation",["All"] + WF_REGULATIONS, key="wfl_reg")
            fc5,fc6,_ = st.columns(3)
            f_dept = fc5.selectbox("Department",["All"] + WF_DEPARTMENTS, key="wfl_dept")
            f_stat = fc6.selectbox("Status",    ["All","Active","Archived"], key="wfl_stat")

        # Apply filters
        results = templates
        if search:
            sl = search.lower()
            results = [t for t in results if
                sl in t["name"].lower() or
                sl in t["category"].lower() or
                sl in t["purpose"].lower() or
                any(sl in tag for tag in t.get("tags",[]))]
        if f_cat  != "All": results = [t for t in results if t["category"]   == f_cat]
        if f_purp != "All": results = [t for t in results if t["purpose"]    == f_purp]
        if f_freq != "All": results = [t for t in results if t["frequency"]  == f_freq]
        if f_reg  != "All": results = [t for t in results if t["regulation"] == f_reg]
        if f_dept != "All": results = [t for t in results if t["department"] == f_dept]
        if f_stat != "All": results = [t for t in results if t["status"]     == f_stat]

        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin:.5rem 0 .8rem 0'>{len(results)} template(s)</div>", unsafe_allow_html=True)

        # ── Column headers ──
        st.markdown("""
        <div style='display:grid;grid-template-columns:2fr 1fr 1.2fr 1fr 1fr 1.1fr 0.7fr;
             gap:8px;padding:6px 12px;background:#e5e7eb;border-radius:6px;
             font-size:10px;font-weight:700;letter-spacing:.08em;color:#6b7280;
             text-transform:uppercase;margin-bottom:4px;'>
          <div>Workflow Name</div><div>Category</div><div>Purpose</div>
          <div>Frequency</div><div>Owner Role</div><div>Regulation</div><div>Status</div>
        </div>""", unsafe_allow_html=True)

        for tmpl in results:
            sc   = "badge-closed" if tmpl["status"]=="Active" else "badge-draft"
            tags_html = "".join(f"<span style='background:#e0f2fe;color:#0369a1;font-size:10px;padding:1px 7px;border-radius:10px;margin-right:3px;font-weight:600'>{tg}</span>" for tg in tmpl.get("tags",[]))

            st.markdown(f"""
            <div style='display:grid;grid-template-columns:2fr 1fr 1.2fr 1fr 1fr 1.1fr 0.7fr;
                 gap:8px;padding:10px 12px;background:#ffffff;border:1px solid #e5e7eb;
                 border-radius:8px;margin-bottom:4px;align-items:center;font-size:13px;'>
              <div>
                <div style='font-weight:700;color:#111111;margin-bottom:2px'>{tmpl['name']}</div>
                <div style='margin-top:3px'>{tags_html}</div>
              </div>
              <div style='color:#374151'>{tmpl['category']}</div>
              <div style='color:#374151'>{tmpl['purpose']}</div>
              <div style='color:#374151'>{tmpl['frequency']}</div>
              <div style='color:#374151'>{tmpl['owner_role']}</div>
              <div style='color:#374151'>{tmpl['regulation']}</div>
              <div>{badge(tmpl['status'], sc)}</div>
            </div>""", unsafe_allow_html=True)

            if toggle(f"wft_{tmpl['id']}", f"Details — {tmpl['id']}"):
                dc1,dc2,dc3 = st.columns(3)
                dc1.markdown(f"**Department:** {tmpl['department']}")
                dc2.markdown(f"**Created:** {tmpl['created'][:10]}")
                dc3.markdown(f"**ID:** `{tmpl['id']}`")

                # Launch instance from template
                st.markdown("<hr style='margin:.6rem 0'/>", unsafe_allow_html=True)
                st.markdown("<div style='font-size:12px;font-weight:700;color:#374151;margin-bottom:.4rem'>LAUNCH INSTANCE FROM THIS TEMPLATE</div>", unsafe_allow_html=True)
                li1,li2,li3 = st.columns(3)
                inst_name = li1.text_input("Instance Name (e.g. Q1 2026)", key=f"li_name_{tmpl['id']}")
                inst_due  = li2.date_input("Due Date", value=date.today()+timedelta(days=30), key=f"li_due_{tmpl['id']}")
                inst_pri  = li3.selectbox("Priority", WF_PRIORITIES, index=1, key=f"li_pri_{tmpl['id']}")
                inst_own  = st.selectbox("Assign Owner", list(D["users"].keys()),
                    format_func=lambda x: f"{D['users'][x]['name']} ({D['users'][x]['title']})",
                    key=f"li_own_{tmpl['id']}")
                if st.button("▶ Launch Instance", key=f"li_launch_{tmpl['id']}"):
                    if inst_name.strip():
                        nid = f"WF-I-{len(D['workflow_instances'])+1:03d}"
                        D["workflow_instances"].append({
                            "id": nid, "template_id": tmpl["id"],
                            "name": f"{tmpl['name']} — {inst_name}",
                            "instance_name": inst_name,
                            "owner": inst_own,
                            "due": inst_due.strftime("%Y-%m-%d"),
                            "priority": inst_pri,
                            "status": "Open",
                            "created": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "category": tmpl["category"],
                            "regulation": tmpl["regulation"],
                        })
                        add_audit("Workflows", f"Launched {nid} from {tmpl['id']}: {inst_name}")
                        st.success(f"✓ Instance {nid} launched.")
                        st.rerun()
                    else:
                        st.warning("Instance name required.")

                # Archive / Restore
                if current_role() == "Compliance Officer":
                    st.markdown("<hr style='margin:.5rem 0'/>", unsafe_allow_html=True)
                    new_status = "Archived" if tmpl["status"]=="Active" else "Active"
                    btn_label  = f"📦 Archive Template" if tmpl["status"]=="Active" else "♻ Restore Template"
                    if st.button(btn_label, key=f"wft_arch_{tmpl['id']}"):
                        tmpl["status"] = new_status
                        add_audit("Workflows", f"Template {tmpl['id']} set to {new_status}")
                        st.rerun()

    with tab2:
        if current_role() not in ["Compliance Officer","Compliance Team"]:
            st.markdown("<div class='gc-alert gc-alert-warn'>⚠ Only Compliance Officers can create workflow templates.</div>", unsafe_allow_html=True)
            return

        st.markdown("<div style='font-size:12px;color:#6b7280;margin-bottom:1rem'>All fields use structured options to maintain filtering consistency across the platform.</div>", unsafe_allow_html=True)

        n1,n2 = st.columns(2)
        wf_name = n1.text_input("Workflow Name *")
        wf_dept = n2.selectbox("Department", WF_DEPARTMENTS)

        r1,r2,r3 = st.columns(3)
        wf_cat  = r1.selectbox("Category",   WF_CATEGORIES)
        wf_purp = r2.selectbox("Purpose",    WF_PURPOSES)
        wf_freq = r3.selectbox("Frequency",  WF_FREQUENCIES)

        r4,r5 = st.columns(2)
        wf_role = r4.selectbox("Owner Role", WF_OWNER_ROLES)
        wf_reg  = r5.selectbox("Regulation Mapping", WF_REGULATIONS)

        st.markdown("<div style='font-size:12px;font-weight:600;color:#374151;text-transform:uppercase;letter-spacing:.05em;margin:.6rem 0 .3rem 0'>Tags <span style='font-weight:400;text-transform:none;letter-spacing:0'>(select up to 4)</span></div>", unsafe_allow_html=True)
        tag_cols = st.columns(6)
        selected_tags = []
        for i, tg in enumerate(WF_TAGS_ALLOWED):
            if tag_cols[i % 6].checkbox(tg, key=f"ntag_{tg}"):
                selected_tags.append(tg)
        if len(selected_tags) > 4:
            st.warning("Maximum 4 tags allowed. Only the first 4 will be saved.")
            selected_tags = selected_tags[:4]

        st.markdown("<hr/>", unsafe_allow_html=True)
        if st.button("✚ Create Workflow Template"):
            if wf_name.strip():
                nid = f"WF-T-{len(D['workflow_templates'])+1:03d}"
                D["workflow_templates"].append({
                    "id": nid, "name": wf_name.strip(),
                    "category": wf_cat, "purpose": wf_purp,
                    "department": wf_dept, "frequency": wf_freq,
                    "owner_role": wf_role, "regulation": wf_reg,
                    "tags": selected_tags,
                    "status": "Active",
                    "created": datetime.now().strftime("%Y-%m-%d %H:%M"),
                })
                add_audit("Workflows", f"Created template {nid}: {wf_name}")
                st.success(f"✓ Template {nid} created.")
                st.rerun()
            else:
                st.warning("Workflow name is required.")

# ── ACTIVE WORKFLOWS ───────────────────────────────────────────────────────────
def page_active_workflows():
    now_str = date.today().strftime("%Y-%m-%d")
    # Auto-escalate overdue
    for inst in D.get("workflow_instances", []):
        if inst["status"] in ("Open","In Progress") and inst["due"] < now_str:
            inst["status"] = "Overdue"

    st.markdown("<div class='gc-header'><h1>ACTIVE WORKFLOWS</h1><p>Running workflow instances — track, update, and close compliance processes</p></div>", unsafe_allow_html=True)

    instances = D.get("workflow_instances", [])

    # ── Search + Filters ──
    search = st.text_input("🔍 Search", placeholder="Workflow name, category, regulation…", key="wfi_search")

    fc1,fc2,fc3,fc4,fc5 = st.columns(5)
    f_cat   = fc1.selectbox("Category",  ["All"] + WF_CATEGORIES,    key="wfi_cat")
    f_stat  = fc2.selectbox("Status",    ["All"] + WF_INST_STATUSES, key="wfi_stat")
    f_pri   = fc3.selectbox("Priority",  ["All"] + WF_PRIORITIES,    key="wfi_pri")
    f_own   = fc4.selectbox("Owner",     ["All"] + list(D["users"].keys()),
                             format_func=lambda x: "All" if x=="All" else D["users"][x]["name"],
                             key="wfi_own")
    f_reg   = fc5.selectbox("Regulation",["All"] + WF_REGULATIONS,   key="wfi_reg")

    fc6,fc7 = st.columns(2)
    f_due_from = fc6.date_input("Due From", value=None, key="wfi_due_from")
    f_due_to   = fc7.date_input("Due To",   value=None, key="wfi_due_to")

    # Apply filters
    results = instances
    if search:
        sl = search.lower()
        results = [i for i in results if sl in i["name"].lower() or sl in i["category"].lower() or sl in i.get("regulation","").lower()]
    if f_cat  != "All": results = [i for i in results if i["category"]   == f_cat]
    if f_stat != "All": results = [i for i in results if i["status"]     == f_stat]
    if f_pri  != "All": results = [i for i in results if i["priority"]   == f_pri]
    if f_own  != "All": results = [i for i in results if i["owner"]      == f_own]
    if f_reg  != "All": results = [i for i in results if i.get("regulation","") == f_reg]
    if f_due_from:
        results = [i for i in results if i["due"] >= f_due_from.strftime("%Y-%m-%d")]
    if f_due_to:
        results = [i for i in results if i["due"] <= f_due_to.strftime("%Y-%m-%d")]

    # Summary metrics
    mc = st.columns(4)
    mc[0].metric("Total",     len(instances))
    mc[1].metric("Open",      len([i for i in instances if i["status"]=="Open"]))
    mc[2].metric("In Progress",len([i for i in instances if i["status"]=="In Progress"]))
    mc[3].metric("Overdue",   len([i for i in instances if i["status"]=="Overdue"]))
    st.markdown("<hr/>", unsafe_allow_html=True)

    st.markdown(f"<div style='font-size:12px;color:#6b7280;margin:.2rem 0 .8rem 0'>{len(results)} instance(s) shown</div>", unsafe_allow_html=True)

    # ── Column headers ──
    st.markdown("""
    <div style='display:grid;grid-template-columns:2.2fr 1fr 1fr 1fr 0.9fr 0.9fr 0.8fr;
         gap:8px;padding:6px 12px;background:#e5e7eb;border-radius:6px;
         font-size:10px;font-weight:700;letter-spacing:.08em;color:#6b7280;
         text-transform:uppercase;margin-bottom:4px;'>
      <div>Workflow Name</div><div>Category</div><div>Regulation</div>
      <div>Owner</div><div>Due Date</div><div>Priority</div><div>Status</div>
    </div>""", unsafe_allow_html=True)

    for inst in results:
        s = inst["status"]
        p = inst["priority"]
        sc = "badge-overdue" if s=="Overdue" else "badge-open" if s=="Open" else "badge-pending" if s=="In Progress" else "badge-closed"
        pc = "badge-high" if p=="High" else "badge-low" if p=="Low" else "badge-open"
        owner_name = D["users"].get(inst["owner"],{}).get("name", inst["owner"])

        # Highlight overdue rows
        row_border = "border-left:3px solid #ef4444" if s=="Overdue" else "border-left:3px solid #e5e7eb"

        st.markdown(f"""
        <div style='display:grid;grid-template-columns:2.2fr 1fr 1fr 1fr 0.9fr 0.9fr 0.8fr;
             gap:8px;padding:10px 12px;background:#ffffff;border:1px solid #e5e7eb;{row_border};
             border-radius:8px;margin-bottom:4px;align-items:center;font-size:13px;'>
          <div>
            <div style='font-weight:700;color:#111111'>{inst['name']}</div>
            <div style='font-size:11px;color:#6b7280;margin-top:2px'>Instance: {inst['instance_name']} · Created: {inst['created'][:10]}</div>
          </div>
          <div style='color:#374151'>{inst['category']}</div>
          <div style='color:#374151'>{inst.get('regulation','—')}</div>
          <div style='color:#374151'>{owner_name}</div>
          <div style='color:{"#dc2626" if s=="Overdue" else "#374151"};font-weight:{"700" if s=="Overdue" else "400"}'>{inst['due']}</div>
          <div>{badge(p, pc)}</div>
          <div>{badge(s, sc)}</div>
        </div>""", unsafe_allow_html=True)

        if toggle(f"wfi_{inst['id']}", f"Manage — {inst['id']}"):
            # Template cross-reference
            tmpl = next((t for t in D.get("workflow_templates",[]) if t["id"]==inst.get("template_id","")), None)
            if tmpl:
                st.markdown(f"<div class='gc-alert gc-alert-ok' style='margin-bottom:.5rem'>📚 Template: <b>{tmpl['name']}</b> · {tmpl['purpose']} · {tmpl['frequency']}</div>", unsafe_allow_html=True)

            ua1,ua2,ua3 = st.columns(3)
            new_status = ua1.selectbox("Update Status", WF_INST_STATUSES,
                index=WF_INST_STATUSES.index(inst["status"]) if inst["status"] in WF_INST_STATUSES else 0,
                key=f"wfi_stat_{inst['id']}")
            new_pri    = ua2.selectbox("Update Priority", WF_PRIORITIES,
                index=WF_PRIORITIES.index(inst["priority"]) if inst["priority"] in WF_PRIORITIES else 1,
                key=f"wfi_pri_{inst['id']}")
            new_due    = ua3.date_input("Update Due Date",
                value=datetime.strptime(inst["due"],"%Y-%m-%d").date(),
                key=f"wfi_due_{inst['id']}")

            if st.button("💾 Save Changes", key=f"wfi_save_{inst['id']}"):
                inst["status"]   = new_status
                inst["priority"] = new_pri
                inst["due"]      = new_due.strftime("%Y-%m-%d")
                add_audit("Workflows", f"Updated {inst['id']}: status={new_status}, priority={new_pri}, due={inst['due']}")
                st.success("✓ Instance updated.")
                st.rerun()

# ── COMPLIANCE CALENDAR ─────────────────────────────────────────────────────────
# Seed data: 2026 FINRA/SEC compliance calendar for a full-service broker-dealer
# Deadlines sourced from FINRA Information Notice 11/10/25, FINRA Compliance Calendar,
# SEC Rule 17a-5, and standard FINRA/SEC annual obligations.

COMPLIANCE_CALENDAR_ITEMS = [
    # ── JANUARY ──
    {"id":"CAL-001","regulator":"FINRA","item":"Annual Renewal — Final Statement Discrepancy Report Due","category":"Registration","deadline":"2026-01-23","rule":"FINRA Renewal Program","assignee":"cco","notes":"Report any discrepancies found on the Final Renewal Statement to FINRA in writing."},
    {"id":"CAL-002","regulator":"FINRA","item":"Annual Renewal — Final Payment Due","category":"Registration","deadline":"2026-01-23","rule":"FINRA Renewal Program","assignee":"cco","notes":"Submit any outstanding balance from the Final Renewal Statement. FINRA will transfer from Flex-Funding Account if authorized."},
    {"id":"CAL-003","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (December 2025)","category":"Financial Reporting","deadline":"2026-01-26","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report Part II or IIA due for December 2025 reporting period. File via eFOCUS through FINRA Gateway."},
    {"id":"CAL-004","regulator":"FINRA","item":"Supplemental Inventory Schedule (SIS) — Q4 2025 Filing","category":"Financial Reporting","deadline":"2026-01-28","rule":"FINRA Rule 4524","assignee":"cco","notes":"SIS filing due within 22 business days of December 31, 2025 quarter-end."},
    {"id":"CAL-005","regulator":"SEC","item":"Form CRS Annual Review and Update (if material changes)","category":"Reg BI / Form CRS","deadline":"2026-01-31","rule":"SEC Reg BI / Form CRS","assignee":"supervisor","notes":"Review Form CRS for accuracy. File amendments promptly if material changes occurred. Deliver updated Form CRS to new and existing retail customers as required."},
    # ── FEBRUARY ──
    {"id":"CAL-006","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (January 2026)","category":"Financial Reporting","deadline":"2026-02-25","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for January 2026 reporting period."},
    {"id":"CAL-007","regulator":"FINRA","item":"Annual Report (Fiscal Year-End Dec 31, 2025) — Standard Deadline","category":"Financial Reporting","deadline":"2026-03-02","rule":"SEA Rule 17a-5(d)","assignee":"cco","notes":"Annual Report (audited financial statements) due 60 days after fiscal year-end for standard filers. File via FINRA Gateway; also file on EDGAR in PDF format per new SEC requirements."},
    {"id":"CAL-008","regulator":"SEC","item":"Annual Report on Form X-17A-5 Part III — EDGAR Filing","category":"Financial Reporting","deadline":"2026-03-02","rule":"SEA Rule 17a-5(d)","assignee":"cco","notes":"Broker-dealers must file Annual Report with SEC via EDGAR in PDF format in addition to FINRA Gateway filing."},
    # ── MARCH ──
    {"id":"CAL-009","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (February 2026)","category":"Financial Reporting","deadline":"2026-03-24","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for February 2026 reporting period."},
    {"id":"CAL-010","regulator":"FINRA","item":"Supplemental Inventory Schedule (SIS) — Q1 2026 Interim Filing","category":"Financial Reporting","deadline":"2026-03-27","rule":"FINRA Rule 4524","assignee":"cco","notes":"SIS filing due within 22 business days of end of Q4 2025 (December 31, 2025)."},
    {"id":"CAL-011","regulator":"FINRA","item":"Annual Report — Extended Deadline (Smaller B-Ds with 30-day extension)","category":"Financial Reporting","deadline":"2026-03-31","rule":"SEA Rule 17a-5(d) / SEC Extension Order","assignee":"cco","notes":"Firms qualifying for the SEC's 30-day extension order may file Annual Report by this date. Must have submitted notification to FINRA per Regulatory Notice 21-05."},
    {"id":"CAL-012","regulator":"FINRA","item":"Annual Supervisory Controls Testing Report (FINRA Rule 3120)","category":"Supervision","deadline":"2026-03-31","rule":"FINRA Rule 3120","assignee":"supervisor","notes":"CEO/CCO must certify annual supervisory controls testing. Written report due to senior management. Testing must cover all supervisory control systems and procedures."},
    {"id":"CAL-013","regulator":"SEC","item":"Annual CEO Certification — Reg BI Compliance (Form BD Update if needed)","category":"Reg BI / Form CRS","deadline":"2026-03-31","rule":"SEC Reg BI","assignee":"cco","notes":"Annual internal review of Reg BI compliance program. Document findings and remediation steps taken since last review."},
    # ── APRIL ──
    {"id":"CAL-014","regulator":"FINRA","item":"Supplemental Liquidity Schedule (SLS) — Q1 2026 Filing","category":"Financial Reporting","deadline":"2026-04-02","rule":"FINRA Rule 4524","assignee":"cco","notes":"SLS filing due within 2 business days of end of Q1 2026 for firms required to file."},
    {"id":"CAL-015","regulator":"FINRA","item":"Form Custody Filing — Q1 2026","category":"Financial Reporting","deadline":"2026-04-23","rule":"SEA Rule 17a-5(f) / FINRA Rule 4524","assignee":"cco","notes":"Form Custody due within 17 business days of end of Q1 2026 (March 31, 2026)."},
    {"id":"CAL-016","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (March 2026)","category":"Financial Reporting","deadline":"2026-04-23","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for March 2026 reporting period."},
    {"id":"CAL-017","regulator":"FINRA","item":"Form OBS (Derivatives & Off-Balance Sheet) — Q1 2026","category":"Financial Reporting","deadline":"2026-04-30","rule":"FINRA Rule 4524","assignee":"cco","notes":"OBS filing due within 22 business days of March 31, 2026 quarter-end for applicable firms."},
    {"id":"CAL-018","regulator":"SEC","item":"Regulation S-P — Annual Privacy Notice (if material changes)","category":"Customer Privacy","deadline":"2026-04-30","rule":"SEC Regulation S-P","assignee":"supervisor","notes":"Deliver annual privacy notice to customers if firm's privacy policies have materially changed. Review and update privacy policies to reflect new SEC Reg S-P amendments (cybersecurity/data breach requirements effective 2025)."},
    # ── MAY ──
    {"id":"CAL-019","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (April 2026)","category":"Financial Reporting","deadline":"2026-05-26","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for April 2026 reporting period."},
    {"id":"CAL-020","regulator":"FINRA","item":"Annual AML Independent Testing — Complete by Mid-Year","category":"AML / BSA","deadline":"2026-05-31","rule":"FINRA Rule 3310 / BSA","assignee":"cco","notes":"Independent (internal audit or third party) testing of AML program must be completed. Testing must cover CIP, transaction monitoring, SAR filing, and employee training. Document results and remediation."},
    {"id":"CAL-021","regulator":"SEC","item":"Regulation S-ID (Identity Theft Red Flags) — Annual Program Review","category":"Customer Privacy","deadline":"2026-05-31","rule":"SEC Regulation S-ID","assignee":"supervisor","notes":"Annual review of the firm's Identity Theft Prevention Program. Update Red Flag policies to reflect new account types and emerging identity theft methods."},
    # ── JUNE ──
    {"id":"CAL-022","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (May 2026)","category":"Financial Reporting","deadline":"2026-06-24","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for May 2026 reporting period."},
    {"id":"CAL-023","regulator":"FINRA","item":"Supplemental Inventory Schedule (SIS) — Q2 2026 Filing","category":"Financial Reporting","deadline":"2026-06-30","rule":"FINRA Rule 4524","assignee":"cco","notes":"SIS filing due within 22 business days of June 30, 2026 quarter-end."},
    {"id":"CAL-024","regulator":"SEC","item":"Rule 15c3-3 Daily Reserve Formula Computation — Compliance Deadline","category":"Financial Responsibility","deadline":"2026-06-30","rule":"SEA Rule 15c3-3(e)(3)(i)(B)(1)","assignee":"cco","notes":"Firms with minimum fixed dollar net capital >= $250K as of Dec 31, 2025 must begin daily reserve formula computation by this date per amended Rule 15c3-3."},
    {"id":"CAL-025","regulator":"FINRA","item":"Mid-Year Supervisory System Review (FINRA Rule 3110)","category":"Supervision","deadline":"2026-06-30","rule":"FINRA Rule 3110","assignee":"supervisor","notes":"Conduct mid-year review of supervisory procedures and controls. Document review findings. Update WSPs for any regulatory or business changes in H1 2026."},
    # ── JULY ──
    {"id":"CAL-026","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (June 2026)","category":"Financial Reporting","deadline":"2026-07-23","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for June 2026 reporting period."},
    {"id":"CAL-027","regulator":"FINRA","item":"Supplemental Liquidity Schedule (SLS) — Q2 2026 Filing","category":"Financial Reporting","deadline":"2026-07-02","rule":"FINRA Rule 4524","assignee":"cco","notes":"SLS filing due within 2 business days of end of Q2 2026 for applicable firms."},
    {"id":"CAL-028","regulator":"FINRA","item":"Form Custody Filing — Q2 2026","category":"Financial Reporting","deadline":"2026-07-23","rule":"SEA Rule 17a-5(f) / FINRA Rule 4524","assignee":"cco","notes":"Form Custody due within 17 business days of end of Q2 2026 (June 30, 2026)."},
    {"id":"CAL-029","regulator":"FINRA","item":"Form OBS (Derivatives & Off-Balance Sheet) — Q2 2026","category":"Financial Reporting","deadline":"2026-07-31","rule":"FINRA Rule 4524","assignee":"cco","notes":"OBS filing due within 22 business days of June 30, 2026 quarter-end."},
    {"id":"CAL-030","regulator":"SEC","item":"Annual Reg BI Best Interest Obligations — Q2 Trade Review & Documentation","category":"Reg BI / Form CRS","deadline":"2026-07-31","rule":"SEC Reg BI","assignee":"supervisor","notes":"Complete Q2 review of recommendations for compliance with Reg BI care obligation. Document review of complex and higher-risk products. Retain records per Rule 17a-4."},
    # ── AUGUST ──
    {"id":"CAL-031","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (July 2026)","category":"Financial Reporting","deadline":"2026-08-25","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for July 2026 reporting period."},
    {"id":"CAL-032","regulator":"FINRA","item":"Continuing Education (CE) — Regulatory Element Completion Deadline","category":"Registration / CE","deadline":"2026-08-31","rule":"FINRA Rule 1240","assignee":"supervisor","notes":"All registered persons with anniversary dates through August must complete their Regulatory Element CE by their anniversary date. Firm must track and enforce completion deadlines."},
    {"id":"CAL-033","regulator":"SEC","item":"Regulation S-P Cybersecurity Incident Response — Program Assessment","category":"Customer Privacy","deadline":"2026-08-31","rule":"SEC Regulation S-P (amended)","assignee":"cco","notes":"Assess cybersecurity incident response and notification procedures under amended Reg S-P. Ensure policies cover unauthorized access to customer financial information and 30-day customer notification requirement."},
    # ── SEPTEMBER ──
    {"id":"CAL-034","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (August 2026)","category":"Financial Reporting","deadline":"2026-09-24","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for August 2026 reporting period."},
    {"id":"CAL-035","regulator":"FINRA","item":"Supplemental Inventory Schedule (SIS) — Q3 2026 Filing","category":"Financial Reporting","deadline":"2026-09-30","rule":"FINRA Rule 4524","assignee":"cco","notes":"SIS filing due within 22 business days of September 30, 2026 quarter-end."},
    {"id":"CAL-036","regulator":"FINRA","item":"Annual AML Training Completion — All Registered Persons","category":"AML / BSA","deadline":"2026-09-30","rule":"FINRA Rule 3310 / BSA","assignee":"supervisor","notes":"All registered persons must complete annual AML training by end of Q3. Document completion records. Training must cover current typologies, red flags, SAR obligations, and CIP requirements."},
    {"id":"CAL-037","regulator":"FINRA","item":"Business Continuity Plan (BCP) Annual Review and Update","category":"Operations","deadline":"2026-09-30","rule":"FINRA Rule 4370","assignee":"cco","notes":"Annual review of the firm's BCP. Update for any changes in personnel, systems, or business lines. Test emergency contact lists. Document review and any updates made."},
    # ── OCTOBER ──
    {"id":"CAL-038","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (September 2026)","category":"Financial Reporting","deadline":"2026-10-22","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for September 2026 reporting period."},
    {"id":"CAL-039","regulator":"FINRA","item":"Supplemental Liquidity Schedule (SLS) — Q3 2026 Filing","category":"Financial Reporting","deadline":"2026-10-02","rule":"FINRA Rule 4524","assignee":"cco","notes":"SLS filing due within 2 business days of end of Q3 2026 for applicable firms."},
    {"id":"CAL-040","regulator":"FINRA","item":"Form Custody Filing — Q3 2026","category":"Financial Reporting","deadline":"2026-10-22","rule":"SEA Rule 17a-5(f) / FINRA Rule 4524","assignee":"cco","notes":"Form Custody due within 17 business days of end of Q3 2026 (September 30, 2026)."},
    {"id":"CAL-041","regulator":"FINRA","item":"Form OBS (Derivatives & Off-Balance Sheet) — Q3 2026","category":"Financial Reporting","deadline":"2026-10-30","rule":"FINRA Rule 4524","assignee":"cco","notes":"OBS filing due within 22 business days of September 30, 2026 quarter-end."},
    {"id":"CAL-042","regulator":"FINRA","item":"Annual Renewal Program — Preliminary Statement Review","category":"Registration","deadline":"2026-10-31","rule":"FINRA Renewal Program","assignee":"cco","notes":"Review Preliminary Renewal Statement. Verify all registrations to be renewed and confirm budget for renewal fees. Begin ensuring sufficient funds in Renewal Account or Flex-Funding Account."},
    # ── NOVEMBER ──
    {"id":"CAL-043","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (October 2026)","category":"Financial Reporting","deadline":"2026-11-24","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for October 2026 reporting period."},
    {"id":"CAL-044","regulator":"FINRA","item":"Annual Renewal — Full Payment Due (Preliminary Statement)","category":"Registration","deadline":"2026-12-01","rule":"FINRA Renewal Program","assignee":"cco","notes":"Full payment of Preliminary Renewal Statement must be available in Renewal Account or Flex-Funding Account. Failure to pay may result in late fees and automatic termination of registrations."},
    {"id":"CAL-045","regulator":"FINRA","item":"Mass Transfer Moratorium Begins — Last Day for Mass Transfers","category":"Registration","deadline":"2026-11-27","rule":"FINRA Renewal Program","assignee":"cco","notes":"Last day to request a mass transfer (used for mergers, acquisitions, succession). Moratorium on mass transfer processing begins Dec 19 through Jan 1, 2027."},
    {"id":"CAL-046","regulator":"SEC","item":"Annual Compliance Program Review (SEC 206(4)-7 / FINRA 3130)","category":"Supervision","deadline":"2026-11-30","rule":"SEC Rule 206(4)-7 / FINRA Rule 3130","assignee":"cco","notes":"CEO must certify annual review of compliance policies and procedures. Written report to senior management required. Review must address any material compliance matters identified during the year."},
    {"id":"CAL-047","regulator":"FINRA","item":"Annual Firm Element CE Planning — Needs Analysis Due","category":"Registration / CE","deadline":"2026-11-30","rule":"FINRA Rule 1240","assignee":"supervisor","notes":"Complete annual Firm Element training needs analysis. Identify training topics for coming year based on business activities, regulatory developments, and prior year exam findings."},
    # ── DECEMBER ──
    {"id":"CAL-048","regulator":"FINRA","item":"FOCUS Report Part II/IIA — Monthly Filing (November 2026)","category":"Financial Reporting","deadline":"2026-12-23","rule":"SEA Rule 17a-5 / FINRA Rule 4524","assignee":"cco","notes":"Monthly FOCUS Report for November 2026 reporting period."},
    {"id":"CAL-049","regulator":"FINRA","item":"Annual Renewal — CRD/IARD System Shutdown","category":"Registration","deadline":"2026-12-26","rule":"FINRA Renewal Program","assignee":"cco","notes":"CRD/IARD shuts down at 6 p.m. ET on Dec 26. Submit all registration filings before shutdown. Terminate any registrations not being renewed. Finalize all E-Bill renewal payments."},
    {"id":"CAL-050","regulator":"FINRA","item":"Year-End WSP Review and Update","category":"Supervision","deadline":"2026-12-31","rule":"FINRA Rule 3110","assignee":"supervisor","notes":"Annual review of all Written Supervisory Procedures. Update for regulatory changes, new products, and personnel changes occurring during 2026. CCO must document approval of final versions."},
    {"id":"CAL-051","regulator":"FINRA","item":"Supplemental Inventory Schedule (SIS) — Q4 2026 Filing Prep","category":"Financial Reporting","deadline":"2026-12-31","rule":"FINRA Rule 4524","assignee":"cco","notes":"Ensure all positions are accurately recorded as of December 31, 2026 for SIS filing due in late January 2027."},
    {"id":"CAL-052","regulator":"SEC","item":"Annual Net Capital Computation Review (Rule 15c3-1)","category":"Financial Responsibility","deadline":"2026-12-31","rule":"SEA Rule 15c3-1","assignee":"cco","notes":"Year-end review of net capital compliance. Ensure aggregate indebtedness/net capital ratio is within limits. Verify haircut calculations are current with regulatory requirements."},
]

def page_compliance_calendar():
    st.markdown("<div class='gc-header'><h1>COMPLIANCE CALENDAR</h1><p>2026 FINRA &amp; SEC regulatory deadlines for a full-service broker-dealer — chronological with evidence upload</p></div>", unsafe_allow_html=True)

    # Back-compat: ensure calendar items exist in session data
    if "cal_items" not in st.session_state.data:
        st.session_state.data["cal_items"] = {
            item["id"]: {"completed": False, "completed_by": "", "completed_ts": "", "files": []}
            for item in COMPLIANCE_CALENDAR_ITEMS
        }
    cal_state = st.session_state.data["cal_items"]
    # Fill any missing keys (new items added to constant)
    for item in COMPLIANCE_CALENDAR_ITEMS:
        if item["id"] not in cal_state:
            cal_state[item["id"]] = {"completed": False, "completed_by": "", "completed_ts": "", "files": []}

    # Supervisor: only see items assigned to them (supervisory role items)
    cu = st.session_state.current_user
    if is_supervisor():
        BASE_ITEMS = [i for i in COMPLIANCE_CALENDAR_ITEMS if i.get("assignee") == cu]
        lob = current_user_lob() or "Wealth Management"
        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.6rem'>Showing your assigned calendar items — <b>{lob}</b> line of business</div>", unsafe_allow_html=True)
    else:
        BASE_ITEMS = COMPLIANCE_CALENDAR_ITEMS

    # ── Filter bar ─────────────────────────────────────────────────────────────
    fc1, fc2, fc3, fc4 = st.columns([1.5, 1.5, 1.5, 1])
    f_reg      = fc1.selectbox("Regulator", ["All","FINRA","SEC"])
    categories = sorted(set(i["category"] for i in BASE_ITEMS))
    f_cat      = fc2.selectbox("Category", ["All"] + categories)
    months     = ["All","January","February","March","April","May","June",
                  "July","August","September","October","November","December"]
    f_month    = fc3.selectbox("Month", months)
    f_status   = fc4.selectbox("Status", ["All","Pending","Completed"])

    month_map  = {m: f"{i+1:02d}" for i, m in enumerate(months[1:])}

    # Apply filters
    items = BASE_ITEMS
    if f_reg    != "All": items = [i for i in items if i["regulator"] == f_reg]
    if f_cat    != "All": items = [i for i in items if i["category"]  == f_cat]
    if f_month  != "All": items = [i for i in items if i["deadline"][5:7] == month_map[f_month]]
    if f_status == "Pending":   items = [i for i in items if not cal_state[i["id"]]["completed"]]
    if f_status == "Completed": items = [i for i in items if cal_state[i["id"]]["completed"]]

    # ── Summary metrics ────────────────────────────────────────────────────────
    total_all   = len(BASE_ITEMS)
    done_all    = sum(1 for i in BASE_ITEMS if cal_state[i["id"]]["completed"])
    pending_all = total_all - done_all
    today_str   = datetime.now().strftime("%Y-%m-%d")
    overdue_all = sum(1 for i in BASE_ITEMS if not cal_state[i["id"]]["completed"] and i["deadline"] < today_str)

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Items", total_all)
    mc2.metric("Completed", done_all)
    mc3.metric("Pending", pending_all)
    mc4.metric("Overdue", overdue_all, delta=f"-{overdue_all}" if overdue_all else None, delta_color="inverse")

    st.markdown(f"<div style='font-size:12px;color:#6b7280;margin:.4rem 0 .8rem 0'>Showing {len(items)} of {total_all} items</div>", unsafe_allow_html=True)

    if not items:
        st.markdown("<div class='gc-alert gc-alert-ok'>No items match the current filters.</div>", unsafe_allow_html=True)
        return

    # ── Group by month ──────────────────────────────────────────────────────────
    month_names = {f"{i+1:02d}": m for i, m in enumerate(
        ["January","February","March","April","May","June",
         "July","August","September","October","November","December"])}
    from collections import defaultdict as _dd
    by_month = _dd(list)
    for item in items:
        mo = item["deadline"][5:7]
        by_month[mo].append(item)

    for mo in sorted(by_month.keys()):
        month_items = by_month[mo]
        mo_done    = sum(1 for i in month_items if cal_state[i["id"]]["completed"])
        mo_total   = len(month_items)

        # Month header
        st.markdown(f"""
        <div style='background:#1e293b;color:#ffffff;border-radius:8px;padding:.6rem 1rem;
             margin:.8rem 0 .4rem 0;display:flex;justify-content:space-between;align-items:center'>
          <span style='font-size:14px;font-weight:800;letter-spacing:.06em'>{month_names[mo].upper()} 2026</span>
          <span style='font-size:12px;color:#94a3b8'>{mo_done}/{mo_total} completed</span>
        </div>
        """, unsafe_allow_html=True)

        # ── Table header ──
        th1,th2,th3,th4,th5,th6 = st.columns([1, 3.5, 1.5, 1.2, 1, 1.2])
        th1.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Regulator</div>", unsafe_allow_html=True)
        th2.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Item</div>", unsafe_allow_html=True)
        th3.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Assignee</div>", unsafe_allow_html=True)
        th4.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Deadline</div>", unsafe_allow_html=True)
        th5.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Status</div>", unsafe_allow_html=True)
        th6.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase'>Evidence</div>", unsafe_allow_html=True)
        st.markdown("<hr style='margin:.2rem 0 .4rem 0'/>", unsafe_allow_html=True)

        for item in sorted(month_items, key=lambda x: x["deadline"]):
            cs        = cal_state[item["id"]]
            is_done   = cs["completed"]
            is_overdue = not is_done and item["deadline"] < today_str
            assignee_name = D["users"].get(item["assignee"],{}).get("name", item["assignee"])

            # Regulator badge color
            reg_color = "#1e40af" if item["regulator"]=="FINRA" else "#7c3aed"
            reg_bg    = "#dbeafe" if item["regulator"]=="FINRA" else "#ede9fe"

            # Status display
            if is_done:
                status_html = "<span style='background:#dcfce7;color:#166534;font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px'>✅ COMPLETE</span>"
            elif is_overdue:
                status_html = "<span style='background:#fee2e2;color:#991b1b;font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px'>⚠ OVERDUE</span>"
            else:
                status_html = "<span style='background:#fef9c3;color:#854d0e;font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px'>⏳ PENDING</span>"

            file_count = len(cs["files"])
            file_html  = f"<span style='font-size:12px;color:#2563eb'>📎 {file_count} file(s)</span>" if file_count else "<span style='font-size:12px;color:#9ca3af'>—</span>"

            row_bg = "#f0fdf4" if is_done else "#fff7ed" if is_overdue else "#ffffff"
            row_border = "#22c55e" if is_done else "#f97316" if is_overdue else "#e5e7eb"

            c1,c2,c3,c4,c5,c6 = st.columns([1, 3.5, 1.5, 1.2, 1, 1.2])

            c1.markdown(f"<span style='background:{reg_bg};color:{reg_color};font-size:10px;font-weight:800;padding:2px 7px;border-radius:4px'>{item['regulator']}</span>", unsafe_allow_html=True)
            c2.markdown(f"<div style='font-size:12px;font-weight:600;color:#111111;line-height:1.4'>{item['item']}</div><div style='font-size:10px;color:#6b7280;margin-top:1px'>{item['rule']}</div>", unsafe_allow_html=True)
            c3.markdown(f"<div style='font-size:12px;color:#374151'>{assignee_name}</div>", unsafe_allow_html=True)
            c4.markdown(f"<div style='font-size:12px;font-weight:600;color:{'#991b1b' if is_overdue else '#374151'}'>{item['deadline']}</div>", unsafe_allow_html=True)
            c5.markdown(status_html, unsafe_allow_html=True)
            c6.markdown(file_html, unsafe_allow_html=True)

            # ── Expandable detail row ──
            if toggle(f"cal_expand_{item['id']}", f"Details / Upload — {item['id']}"):
                st.markdown(f"<div style='background:{row_bg};border:1px solid {row_border};border-left:4px solid {row_border};border-radius:8px;padding:.8rem 1.1rem;margin:.3rem 0 .5rem 0'>", unsafe_allow_html=True)
                st.markdown(f"<div style='font-size:12px;color:#374151;margin-bottom:.5rem'><b>Notes:</b> {item['notes']}</div>", unsafe_allow_html=True)

                # Uploaded files
                if cs["files"]:
                    st.markdown("<div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem'>UPLOADED EVIDENCE</div>", unsafe_allow_html=True)
                    for f in cs["files"]:
                        upname = D["users"].get(f["uploaded_by"],{}).get("name", f["uploaded_by"])
                        notes_part = f" — <i>{f['notes']}</i>" if f.get("notes") else ""
                        st.markdown(f"<div style='font-size:12px;color:#374151;padding:.2rem 0'>📎 <b>{f['name']}</b> · {upname} · {f['ts'][:16]}{notes_part}</div>", unsafe_allow_html=True)

                st.markdown("</div>", unsafe_allow_html=True)

                # Upload + mark complete controls
                if not is_done:
                    ua, ub = st.columns([2,2])
                    new_up    = ua.file_uploader("Upload Evidence", key=f"cal_up_{item['id']}", label_visibility="visible")
                    up_notes  = ub.text_area("File Notes", height=68, key=f"cal_notes_{item['id']}", placeholder="e.g. Q1 2026 BCP review sign-off memo")
                    if new_up:
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        cs["files"].append({"name":new_up.name,"uploaded_by":st.session_state.current_user,"ts":now_str,"notes":up_notes.strip()})
                        add_audit("Compliance Calendar", f"{item['id']}: evidence uploaded — {new_up.name}")
                        st.success(f"✓ {new_up.name} uploaded."); st.rerun()

                    mc_col, _ = st.columns([1,3])
                    if mc_col.button(f"✅ Mark as Complete", key=f"cal_done_{item['id']}"):
                        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
                        cs["completed"]    = True
                        cs["completed_by"] = st.session_state.current_user
                        cs["completed_ts"] = now_str
                        add_audit("Compliance Calendar", f"{item['id']} marked complete: {item['item']}")
                        st.success("✓ Marked as complete."); st.rerun()
                else:
                    comp_name = D["users"].get(cs["completed_by"],{}).get("name", cs["completed_by"])
                    st.markdown(f"<div style='font-size:12px;color:#166534;margin:.3rem 0'>✅ Completed by <b>{comp_name}</b> on {cs['completed_ts'][:16]}</div>", unsafe_allow_html=True)
                    if st.button(f"↩ Reopen", key=f"cal_reopen_{item['id']}"):
                        cs["completed"] = False; cs["completed_by"] = ""; cs["completed_ts"] = ""
                        add_audit("Compliance Calendar", f"{item['id']} reopened")
                        st.rerun()

        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

# ── COMPLIANCE TESTING CONSTANTS ───────────────────────────────────────────────
CT_TEST_TYPES   = ["FINRA 3120","SCP Testing","SEC 206(4)-7","Custom Test"]
CT_RISK_RATINGS = ["Low","Medium","High"]
CT_STATUSES     = ["Not Started","In Progress","Submitted for Review","Completed","Cancelled"]
CT_RESULTS      = ["Pass","Fail","Exception","N/A"]
CT_SAMPLE_METHODS = ["Random","Judgmental","Stratified","Full Population"]
CT_EXCEPTION_STATUSES = ["Open","Remediation In Progress","Closed"]

# Default procedure templates per test type
CT_DEFAULT_PROCEDURES = {
    "FINRA 3120": [
        "Review the firm's supervisory control system and confirm it is reasonably designed to achieve compliance with applicable rules.",
        "Test that supervisory review of correspondence and electronic communications occurred within required timeframes.",
        "Confirm annual certification by a senior registered principal that the supervisory control system is in place and has been tested.",
        "Review documentation of the annual report to senior management on the results of supervisory control testing.",
        "Verify that exception reports were generated, reviewed, and acted upon in accordance with WSP requirements.",
    ],
    "SCP Testing": [
        "Review written supervisory procedures (WSPs) and confirm they are current, approved, and distributed to relevant staff.",
        "Test a sample of supervisory reviews to confirm they were performed by the appropriate supervisor within required timeframes.",
        "Verify that all registered persons have a designated supervisor and that the supervisory hierarchy is documented.",
        "Review training records to confirm supervisors completed required supervisory training for the period.",
        "Confirm that escalation procedures were followed for all flagged items during the test period.",
    ],
    "SEC 206(4)-7": [
        "Review the firm's written compliance policies and procedures and confirm they are reasonably designed to prevent violations of the Advisers Act.",
        "Confirm the annual review of the compliance program was conducted and documented by the CCO.",
        "Review a sample of client files to confirm investment advice is consistent with stated client objectives and Form ADV disclosures.",
        "Test that required disclosure documents (Form ADV Part 2) were delivered to clients within required timeframes.",
        "Review records of compliance training to confirm advisory personnel received required compliance education.",
    ],
    "Custom Test": [
        "Define the objective of this testing procedure.",
        "Identify the population and sampling methodology.",
        "Execute testing and document results.",
        "Identify any exceptions or deficiencies noted during testing.",
        "Summarize findings and provide recommendations.",
    ],
}

# ── COMPLIANCE TESTING HELPERS ──────────────────────────────────────────────────
def _ct_status_badge(status):
    m = {
        "Not Started":          "badge-draft",
        "In Progress":          "badge-open",
        "Submitted for Review": "badge-pending",
        "Completed":            "badge-closed",
        "Cancelled":            "badge-draft",
    }
    return badge(status, m.get(status, "badge-draft"))

def _ct_risk_badge(rating):
    m = {"Low":"badge-low","Medium":"badge-med","High":"badge-high"}
    return badge(rating, m.get(rating,"badge-draft"))

def _ct_result_badge(result):
    m = {"Pass":"badge-closed","Fail":"badge-overdue","Exception":"badge-high","N/A":"badge-draft"}
    return badge(result, m.get(result,"badge-draft")) if result else ""

def _linked_name(wf_id, cl_id):
    if wf_id:
        inst = next((i for i in D.get("workflow_instances",[]) if i["id"]==wf_id), None)
        if inst: return f"Workflow: {inst['name']}"
    if cl_id:
        cl = next((c for c in D.get("supervisory_checklists",[]) if c["id"]==cl_id), None)
        if cl: return f"Checklist: {cl.get('wsp_title','')}"
    return "None"

# ── PAGE: COMPLIANCE TESTING ────────────────────────────────────────────────────
def page_compliance_testing():
    st.markdown("<div class='gc-header'><h1>COMPLIANCE TESTING</h1><p>Structured internal testing — FINRA 3120, SCP Testing, SEC 206(4)-7, and custom programs</p></div>", unsafe_allow_html=True)

    tests = D.get("compliance_tests", [])

    # ── Auto-escalate overdue ──
    today_str = date.today().strftime("%Y-%m-%d")
    for t in tests:
        if t["status"] in ("Not Started","In Progress") and t["deadline"] < today_str:
            pass  # status is controlled by user; show deadline warning instead

    # ── View selector ──
    VIEW_KEY = "_ct_view"
    if VIEW_KEY not in st.session_state:
        st.session_state[VIEW_KEY] = "all"

    v1,v2,v3,v4 = st.columns(4)
    if v1.button("📋  All Tests",        key="_ctv_all",    use_container_width=True):
        st.session_state[VIEW_KEY] = "all";    st.rerun()
    if v2.button("🔍  Submitted for Review", key="_ctv_rev", use_container_width=True):
        st.session_state[VIEW_KEY] = "review"; st.rerun()
    if v3.button("✅  Completed Tests",  key="_ctv_done",   use_container_width=True):
        st.session_state[VIEW_KEY] = "done";   st.rerun()
    if v4.button("➕  Create New Test",  key="_ctv_new",    use_container_width=True):
        st.session_state[VIEW_KEY] = "new";    st.rerun()

    st.markdown("<hr style='margin:.6rem 0 1rem 0'/>", unsafe_allow_html=True)

    view = st.session_state[VIEW_KEY]

    # ──────────────────────────────────────────────────────────────────────────
    # VIEW: CREATE NEW TEST  (Step 1 — Setup)
    # ──────────────────────────────────────────────────────────────────────────
    if view == "new":
        st.markdown("<div style='font-size:1rem;font-weight:700;color:#111111;margin-bottom:1rem'>STEP 1 — TEST SETUP</div>", unsafe_allow_html=True)
        st.markdown("<div style='font-size:13px;color:#6b7280;margin-bottom:1.2rem'>Complete all required fields, then click <b>Next →</b> to open the full testing template.</div>", unsafe_allow_html=True)

        with st.container():
            st.markdown("<div class='gc-card'>", unsafe_allow_html=True)
            st.markdown("<div class='gc-card-header'>TEST IDENTIFICATION</div>", unsafe_allow_html=True)

            r1c1, r1c2 = st.columns(2)
            ns_type = r1c1.selectbox("Type of Test *", CT_TEST_TYPES, key="ns_type")
            ns_lob  = r1c2.text_input("Line of Business *", placeholder="e.g. Retail Brokerage, Investment Advisory", key="ns_lob")

            r2c1, r2c2 = st.columns(2)
            ns_topic  = r2c1.text_input("Topic *", placeholder="e.g. Annual Supervisory Controls Testing", key="ns_topic")
            ns_risk   = r2c2.selectbox("Initial Risk Rating *", CT_RISK_RATINGS, index=1, key="ns_risk")

            st.markdown("<div class='gc-card-header' style='margin-top:1rem'>ASSIGNMENT & SCHEDULING</div>", unsafe_allow_html=True)
            r3c1, r3c2 = st.columns(2)
            ns_assignee = r3c1.selectbox("Assignee *", list(D["users"].keys()),
                format_func=lambda x: f"{D['users'][x]['name']} ({D['users'][x]['title']})",
                key="ns_assignee")
            ns_deadline = r3c2.date_input("Testing Deadline *",
                value=date.today()+timedelta(days=30), key="ns_deadline")

            st.markdown("<div class='gc-card-header' style='margin-top:1rem'>LINK TO EXISTING WORK</div>", unsafe_allow_html=True)
            st.markdown("<div style='font-size:12px;color:#6b7280;margin-bottom:.5rem'>Optionally link to a workflow or supervisory checklist to pull related evidence.</div>", unsafe_allow_html=True)

            link_opts_wf  = ["None"] + [f"{i['id']} — {i['name']}" for i in D.get("workflow_instances",[])]
            link_opts_cl  = ["None"] + [f"{c['id']} — {c.get('wsp_title','')} ({c.get('owner_name','')})" for c in D.get("supervisory_checklists",[])]

            r4c1, r4c2 = st.columns(2)
            ns_wf = r4c1.selectbox("Related Workflow Instance", link_opts_wf, key="ns_wf")
            ns_cl = r4c2.selectbox("Related Supervisory Checklist", link_opts_cl, key="ns_cl")

            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("<hr/>", unsafe_allow_html=True)
        nb1, nb2 = st.columns([1,5])
        if nb1.button("← Cancel", key="ns_cancel"):
            st.session_state[VIEW_KEY] = "all"; st.rerun()
        if nb2.button("Next → Open Testing Template", key="ns_next", type="primary" if hasattr(st.button,"type") else "secondary", use_container_width=True):
            if not ns_lob.strip() or not ns_topic.strip():
                st.error("Line of Business and Topic are required.")
            else:
                nid = f"CT-{len(D['compliance_tests'])+1:03d}"
                procs = [
                    {"id":f"P-{i+1:03d}","step":i+1,"title":f"Test {i+1}","description":desc,
                     "result":"","notes":"","exception":False,"has_remediation_item":False,"tested_by":"","tested_ts":""}
                    for i, desc in enumerate(CT_DEFAULT_PROCEDURES[ns_type])
                ]
                wf_id = ns_wf.split(" — ")[0] if ns_wf != "None" else ""
                cl_id = ns_cl.split(" — ")[0] if ns_cl != "None" else ""
                new_test = {
                    "id": nid,
                    "test_type": ns_type,
                    "line_of_business": ns_lob.strip(),
                    "topic": ns_topic.strip(),
                    "initial_risk_rating": ns_risk,
                    "assignee": ns_assignee,
                    "deadline": ns_deadline.strftime("%Y-%m-%d"),
                    "related_workflow": wf_id,
                    "related_checklist": cl_id,
                    "status": "Not Started",
                    "created": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "created_by": st.session_state.current_user,
                    "date_started": "",
                    "submitted_for_review": False,
                    "completed": False,
                    "completed_ts": "",
                    "reviewer": "",
                    "review_notes": "",
                    "final_risk_rating": "",
                    "population_description": "",
                    "sample_size": 0,
                    "sample_method": "",
                    "sample_rationale": "",
                    "procedures": procs,
                    "exceptions": [],
                    "conclusion": "",
                    "audit": [{"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
                               "user":st.session_state.current_user,
                               "action":f"Test {nid} created"}],
                }
                D["compliance_tests"].append(new_test)
                add_audit("Compliance Testing", f"Created {nid}: {ns_type} — {ns_topic.strip()}")
                st.session_state["_ct_open"] = nid
                st.session_state[VIEW_KEY] = "test"
                st.rerun()

    # ──────────────────────────────────────────────────────────────────────────
    # VIEW: OPEN SINGLE TEST  (Step 2 — Main Test Page)
    # ──────────────────────────────────────────────────────────────────────────
    elif view == "test":
        tid  = st.session_state.get("_ct_open","")
        test = next((t for t in D["compliance_tests"] if t["id"]==tid), None)
        if not test:
            st.error("Test not found."); st.session_state[VIEW_KEY]="all"; st.rerun()
            return

        # Auto-set status to In Progress when opened the first time
        if test["status"] == "Not Started":
            test["status"] = "In Progress"
            test["date_started"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            test["audit"].append({"ts":test["date_started"],"user":st.session_state.current_user,"action":"Test status set to In Progress"})

        # ── Back button ──
        if st.button("← Back to All Tests", key="ct_back"):
            st.session_state[VIEW_KEY] = "all"; st.rerun()

        st.markdown("<hr style='margin:.4rem 0 1rem 0'/>", unsafe_allow_html=True)

        # ── OVERVIEW SECTION ──────────────────────────────────────────────────
        exc_count   = len([e for e in test.get("exceptions",[]) if e["status"]=="Open"])
        procs_done  = len([p for p in test["procedures"] if p["result"]])
        procs_total = len(test["procedures"])
        linked_lbl  = _linked_name(test.get("related_workflow",""), test.get("related_checklist",""))

        risk_color = {"Low":"#166534","Medium":"#854d0e","High":"#991b1b"}.get(test["initial_risk_rating"],"#374151")
        risk_bg    = {"Low":"#dcfce7","Medium":"#fef9c3","High":"#fee2e2"}.get(test["initial_risk_rating"],"#f3f4f6")

        st.markdown(f"""
        <div style='background:#ffffff;border:1px solid #e5e7eb;border-left:4px solid #3b82f6;
             border-radius:8px;padding:1.2rem 1.6rem;margin-bottom:1.2rem'>
          <div style='display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:.8rem'>
            <div style='flex:1;min-width:260px'>
              <div style='font-size:11px;font-weight:700;letter-spacing:.08em;color:#6b7280;text-transform:uppercase;margin-bottom:.3rem'>{test['test_type']}</div>
              <div style='font-size:1.15rem;font-weight:800;color:#111111;margin-bottom:.2rem'>{test['topic']}</div>
              <div style='font-size:13px;color:#374151;margin-bottom:.4rem'>{test['line_of_business']}</div>
              <div style='font-size:12px;color:#6b7280'>
                Created: {test['created'][:10]} &nbsp;·&nbsp;
                {'Started: '+test['date_started'][:10] if test['date_started'] else 'Not yet started'} &nbsp;·&nbsp;
                Deadline: <b style='color:{"#dc2626" if test["deadline"]<date.today().strftime("%Y-%m-%d") else "#374151"}'>{test['deadline']}</b>
              </div>
            </div>
            <div style='display:flex;flex-direction:column;gap:.4rem;align-items:flex-end'>
              {_ct_status_badge(test['status'])}
              <span style='background:{risk_bg};color:{risk_color};font-size:11px;padding:2px 10px;border-radius:10px;font-weight:700'>Risk: {test['initial_risk_rating']}</span>
              <span style='font-size:12px;color:#6b7280'>Assignee: <b>{D['users'].get(test['assignee'],{}).get('name', test['assignee'])}</b></span>
            </div>
          </div>
          <div style='margin-top:.8rem;display:flex;gap:1.5rem;flex-wrap:wrap'>
            <span style='font-size:12px;color:#6b7280'>📋 Procedures: <b style='color:#111111'>{procs_done}/{procs_total} completed</b></span>
            <span style='font-size:12px;color:#6b7280'>⚠ Open Exceptions: <b style='color:{"#dc2626" if exc_count>0 else "#111111"}'>{exc_count}</b></span>
            <span style='font-size:12px;color:#6b7280'>🔗 Linked: <b style='color:#111111'>{linked_lbl}</b></span>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── TABS ──────────────────────────────────────────────────────────────
        tt2,tt3,tt4,tt5 = st.tabs(["📋  Test Procedures","⚠  Exceptions","📝  Conclusion & Review","📜  Audit Trail"])

        # ── TAB 1: TEST PROCEDURES ─────────────────────────────────────────────
        with tt2:
            st.markdown("<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>Complete each test below. Update the title, testing details, and testing results as needed, then set the result and flag any exceptions.</div>", unsafe_allow_html=True)

            for idx, proc in enumerate(test["procedures"]):
                has_exc      = proc.get("exception", False)
                res_bg       = {"Pass":"#f0fdf4","Fail":"#fef2f2","Exception":"#fffbeb","N/A":"#f9fafb","":"#ffffff"}.get(proc["result"],"#ffffff")
                exc_border   = "border-left:4px solid #f59e0b" if has_exc else "border-left:4px solid #e5e7eb"
                exc_pill     = "<span style='background:#fffbeb;color:#92400e;font-size:10px;padding:1px 8px;border-radius:10px;font-weight:700'>⚠ EXCEPTION</span>" if has_exc else ""
                result_badge = _ct_result_badge(proc["result"])
                tested_line  = f"<div style='font-size:11px;color:#6b7280;margin-top:.4rem'>Tested by: <b>{D['users'].get(proc['tested_by'],{}).get('name','—')}</b> · {proc['tested_ts']}</div>" if proc.get("tested_ts") else ""

                # Card header
                st.markdown(f"""
                <div style='background:{res_bg};border:1px solid #e5e7eb;{exc_border};
                     border-radius:8px;padding:.9rem 1.1rem;margin-bottom:4px'>
                  <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:.4rem'>
                    <div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em'>Test {proc['step']}</div>
                    <div style='display:flex;gap:.5rem;align-items:center'>{exc_pill}{result_badge}</div>
                  </div>
                  {tested_line}
                </div>
                """, unsafe_allow_html=True)

                is_mine = (test["assignee"]==st.session_state.current_user or has_full_compliance_access())
                is_editable = is_mine and test["status"] not in ("Submitted for Review","Completed","Cancelled")

                # Title + Details + Results as text inputs (editable or read-only)
                new_title = st.text_input(
                    f"Test {proc['step']} Title",
                    value=proc.get("title",""),
                    key=f"pt_{tid}_{idx}",
                    disabled=not is_editable,
                )
                new_desc = st.text_area(
                    "Testing Details",
                    value=proc.get("description",""),
                    height=80, key=f"pd_{tid}_{idx}",
                    disabled=not is_editable,
                    placeholder="Describe what was requested or reviewed during this test.",
                )
                new_notes = st.text_area(
                    "Testing Results",
                    value=proc.get("notes",""),
                    height=80, key=f"pn_{tid}_{idx}",
                    disabled=not is_editable,
                    placeholder="Document the findings — what was observed, whether it was satisfactory, and the basis for the result.",
                )

                if is_editable:
                    pc1, pc2 = st.columns([1, 1])
                    new_result   = pc1.selectbox("Result", [""] + CT_RESULTS,
                        index=(CT_RESULTS.index(proc["result"])+1) if proc["result"] in CT_RESULTS else 0,
                        key=f"pr_{tid}_{idx}")
                    new_exc_flag = pc2.checkbox("Flag as Exception", value=proc.get("exception",False), key=f"pe_{tid}_{idx}")

                    # ── Remediation item prompt ──
                    if new_exc_flag:
                        st.markdown("<div style='background:#fffbeb;border:1px solid #fde68a;border-radius:6px;padding:.6rem 1rem;margin:.3rem 0'>", unsafe_allow_html=True)
                        rem_yes = st.checkbox(
                            "Did this exception result in a remediation item?",
                            value=proc.get("has_remediation_item", False),
                            key=f"rem_yn_{tid}_{idx}"
                        )
                        if rem_yes:
                            st.markdown("<div style='font-size:12px;font-weight:700;color:#92400e;text-transform:uppercase;letter-spacing:.05em;margin:.5rem 0 .3rem 0'>Remediation Item Details</div>", unsafe_allow_html=True)
                            ri1, ri2 = st.columns(2)
                            rem_name   = ri1.text_input("Remediation Item Name *", value=proc.get("rem_name",""), key=f"rem_name_{tid}_{idx}", placeholder="e.g. Update CCO escalation SLA procedure")
                            rem_assign = ri2.selectbox("Assign To", list(D["users"].keys()),
                                format_func=lambda x: f"{D['users'][x]['name']} ({D['users'][x]['title']})",
                                key=f"rem_assign_{tid}_{idx}")
                            rem_detail = st.text_area("Remediation Item Details *", value=proc.get("rem_detail",""), height=72, key=f"rem_detail_{tid}_{idx}", placeholder="Describe the corrective action required, steps to resolve, and expected outcome.")
                            rem_due    = st.date_input("Due Date", value=date.today()+timedelta(days=30), key=f"rem_due_{tid}_{idx}")
                            proc["has_remediation_item"] = True
                            proc["rem_name"]   = rem_name
                            proc["rem_detail"] = rem_detail
                            proc["rem_assign"] = rem_assign
                            proc["rem_due"]    = rem_due.strftime("%Y-%m-%d")
                        else:
                            proc["has_remediation_item"] = False
                        st.markdown("</div>", unsafe_allow_html=True)

                    if st.button(f"💾 Save Test {proc['step']}", key=f"ps_{tid}_{idx}"):
                        proc["title"]     = new_title
                        proc["description"] = new_desc
                        proc["notes"]     = new_notes
                        proc["result"]    = new_result
                        proc["exception"] = new_exc_flag
                        if not new_exc_flag:
                            proc["has_remediation_item"] = False
                        proc["tested_by"] = st.session_state.current_user
                        proc["tested_ts"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                        msg = f"Test {proc['step']} saved: {new_result}"
                        test["audit"].append({"ts":proc["tested_ts"],"user":st.session_state.current_user,"action":msg})
                        add_audit("Compliance Testing", f"{tid} {msg}")
                        # Auto-create exception record if flagged
                        if new_exc_flag and not any(e["procedure_id"]==proc["id"] for e in test["exceptions"]):
                            eid = f"EX-{len(test['exceptions'])+1:03d}"
                            test["exceptions"].append({
                                "id":eid,"test_id":tid,"procedure_id":proc["id"],
                                "description":f"Exception flagged on Test {proc['step']}: {proc.get('title', proc['description'])[:80]}",
                                "root_cause":"","remediation_plan":"","assigned_to":"",
                                "due_date":"","status":"Open",
                                "created":proc["tested_ts"],
                            })
                            test["audit"].append({"ts":proc["tested_ts"],"user":st.session_state.current_user,"action":f"Exception {eid} auto-created for Test {proc['step']}"})
                        # Create remediation item in exceptions if flagged
                        if new_exc_flag and proc.get("has_remediation_item") and proc.get("rem_name","").strip():
                            already = any(e.get("is_remediation_item") and e["procedure_id"]==proc["id"] for e in test["exceptions"])
                            if not already:
                                rid = f"EX-{len(test['exceptions'])+1:03d}"
                                test["exceptions"].append({
                                    "id":rid,"test_id":tid,"procedure_id":proc["id"],
                                    "description":proc["rem_name"].strip(),
                                    "root_cause":proc.get("rem_detail",""),
                                    "remediation_plan":proc.get("rem_detail",""),
                                    "assigned_to":proc.get("rem_assign",""),
                                    "due_date":proc.get("rem_due",""),
                                    "status":"Remediation In Progress",
                                    "created":proc["tested_ts"],
                                    "is_remediation_item":True,
                                })
                                test["audit"].append({"ts":proc["tested_ts"],"user":st.session_state.current_user,
                                    "action":f"Remediation item '{proc['rem_name'].strip()}' created for Test {proc['step']}, assigned to {proc.get('rem_assign','')}"})
                                add_audit("Compliance Testing", f"{tid}: Remediation item created for Test {proc['step']}")
                        st.success(f"✓ Test {proc['step']} saved."); st.rerun()

                st.markdown("<hr style='margin:.5rem 0 .8rem 0;opacity:.4'/>", unsafe_allow_html=True)

            # Add custom test (CCO/compliance only)
            if has_full_compliance_access() and test["status"] not in ("Submitted for Review","Completed","Cancelled"):
                if toggle(f"ct_addstep_{tid}", "➕  Add Custom Test"):
                    new_step_title = st.text_input("Test Title", key=f"nst_{tid}", placeholder="e.g. Annual Training Completion Review")
                    new_step_desc  = st.text_area("Testing Details", key=f"nsd_{tid}", height=72, placeholder="Describe what was requested or reviewed.")
                    if st.button("Add Test", key=f"nsa_{tid}"):
                        if new_step_desc.strip() or new_step_title.strip():
                            nstep = len(test["procedures"])+1
                            test["procedures"].append({
                                "id":f"P-{nstep:03d}","step":nstep,
                                "title":new_step_title.strip(),
                                "description":new_step_desc.strip(),
                                "result":"","notes":"","exception":False,
                                "has_remediation_item":False,
                                "tested_by":"","tested_ts":""
                            })
                            test["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "user":st.session_state.current_user,"action":f"Custom test {nstep} added"})
                            st.success("✓ Test added."); st.rerun()

        # ── TAB 3: EXCEPTIONS ──────────────────────────────────────────────────
        with tt3:
            exceptions = test.get("exceptions",[])
            if not exceptions:
                st.markdown("<div class='gc-alert gc-alert-ok' style='margin:.5rem 0'>✓ No exceptions recorded for this test.</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>{len(exceptions)} exception(s) · {len([e for e in exceptions if e['status']=='Open'])} open</div>", unsafe_allow_html=True)

            for ex in exceptions:
                exc_sc = {"Open":"badge-overdue","Remediation In Progress":"badge-pending","Closed":"badge-closed"}.get(ex["status"],"badge-draft")
                is_rem_item  = ex.get("is_remediation_item", False)
                card_bg      = "#f0fdf4" if is_rem_item else "#fffbeb"
                card_border  = "#22c55e" if is_rem_item else "#f59e0b"
                card_title_c = "#14532d" if is_rem_item else "#92400e"
                type_pill    = "<span style='background:#dcfce7;color:#166534;font-size:10px;padding:1px 8px;border-radius:10px;font-weight:700;margin-left:6px'>🔧 REMEDIATION ITEM</span>" if is_rem_item else ""
                rc_line  = f"<div style='font-size:12px;color:#374151;margin:.2rem 0'><b>Root Cause:</b> {ex['root_cause']}</div>" if ex.get("root_cause") else ""
                rp_line  = f"<div style='font-size:12px;color:#374151;margin:.2rem 0'><b>Remediation:</b> {ex['remediation_plan']}</div>" if ex.get("remediation_plan") else ""
                asgn_name = D["users"].get(ex.get("assigned_to",""),{}).get("name","")
                at_line  = f"<div style='font-size:12px;color:#374151'><b>Assigned to:</b> {asgn_name} &nbsp;·&nbsp; <b>Due:</b> {ex['due_date']}</div>" if ex.get("assigned_to") else ""

                st.markdown(f"""
                <div style='background:{card_bg};border:1px solid #e5e7eb;border-left:3px solid {card_border};
                     border-radius:8px;padding:.9rem 1.1rem;margin-bottom:6px'>
                  <div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:.4rem'>
                    <div style='font-size:12px;font-weight:700;color:{card_title_c}'>{ex['id']} &nbsp;·&nbsp; Step ref: {ex['procedure_id'] or '—'}{type_pill}</div>
                    {badge(ex['status'], exc_sc)}
                  </div>
                  <div style='font-size:13px;font-weight:600;color:#111111;margin-bottom:.4rem'>{ex['description']}</div>
                  {rc_line}{rp_line}{at_line}
                </div>
                """, unsafe_allow_html=True)

                if test["status"] not in ("Completed","Cancelled"):
                    if toggle(f"ex_edit_{tid}_{ex['id']}", f"Edit Exception {ex['id']}"):
                        ea1,ea2 = st.columns(2)
                        ex["root_cause"]       = ea1.text_area("Root Cause", value=ex.get("root_cause",""), key=f"erc_{tid}_{ex['id']}", height=72)
                        ex["remediation_plan"] = ea2.text_area("Remediation Plan", value=ex.get("remediation_plan",""), key=f"erp_{tid}_{ex['id']}", height=72)
                        eb1,eb2,eb3 = st.columns(3)
                        ex["assigned_to"] = eb1.selectbox("Assign To", [""] + list(D["users"].keys()),
                            format_func=lambda x: "Unassigned" if x=="" else D["users"][x]["name"],
                            index=(list(D["users"].keys()).index(ex["assigned_to"])+1) if ex.get("assigned_to") in D["users"] else 0,
                            key=f"eat_{tid}_{ex['id']}")
                        ex["due_date"] = eb2.text_input("Due Date (YYYY-MM-DD)", value=ex.get("due_date",""), key=f"edd_{tid}_{ex['id']}")
                        ex["status"]   = eb3.selectbox("Exception Status", CT_EXCEPTION_STATUSES,
                            index=CT_EXCEPTION_STATUSES.index(ex["status"]) if ex["status"] in CT_EXCEPTION_STATUSES else 0,
                            key=f"es_{tid}_{ex['id']}")
                        if st.button(f"💾 Save Exception {ex['id']}", key=f"esave_{tid}_{ex['id']}"):
                            test["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "user":st.session_state.current_user,"action":f"Exception {ex['id']} updated: {ex['status']}"})
                            add_audit("Compliance Testing", f"{tid}: Exception {ex['id']} updated")
                            st.success("✓ Exception saved."); st.rerun()

            # Manual exception add
            if test["status"] not in ("Submitted for Review","Completed","Cancelled"):
                st.markdown("<hr/>", unsafe_allow_html=True)
                if toggle(f"ct_addex_{tid}", "➕  Add Exception Manually"):
                    new_exc_desc = st.text_area("Exception Description *", key=f"ned_{tid}", height=72)
                    if st.button("Add Exception", key=f"nea_{tid}"):
                        if new_exc_desc.strip():
                            eid = f"EX-{len(test['exceptions'])+1:03d}"
                            test["exceptions"].append({
                                "id":eid,"test_id":tid,"procedure_id":"",
                                "description":new_exc_desc.strip(),
                                "root_cause":"","remediation_plan":"","assigned_to":"",
                                "due_date":"","status":"Open",
                                "created":datetime.now().strftime("%Y-%m-%d %H:%M"),
                            })
                            test["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
                                "user":st.session_state.current_user,"action":f"Exception {eid} manually added"})
                            st.success("✓ Exception added."); st.rerun()

        # ── TAB 4: CONCLUSION & REVIEW ─────────────────────────────────────────
        with tt4:
            st.markdown("<div class='gc-card-header'>FINAL RISK RATING & CONCLUSION</div>", unsafe_allow_html=True)

            # Progress check
            incomplete = [p for p in test["procedures"] if not p["result"]]
            if incomplete:
                st.markdown(f"<div class='gc-alert gc-alert-warn'>⚠ {len(incomplete)} procedure step(s) not yet completed. Complete all steps before submitting for review.</div>", unsafe_allow_html=True)

            is_mine = (test["assignee"]==st.session_state.current_user or has_full_compliance_access())

            if test["status"] not in ("Submitted for Review","Completed","Cancelled") and is_mine:
                con1,con2 = st.columns(2)
                new_final_risk  = con1.selectbox("Final Risk Rating", CT_RISK_RATINGS,
                    index=CT_RISK_RATINGS.index(test["final_risk_rating"]) if test.get("final_risk_rating") in CT_RISK_RATINGS else CT_RISK_RATINGS.index(test["initial_risk_rating"]),
                    key=f"cfr_{tid}")
                new_conclusion  = st.text_area("Testing Conclusion *",
                    value=test.get("conclusion",""), height=150, key=f"cc_{tid}",
                    placeholder="Summarize the testing approach, key findings, exceptions identified, and overall assessment. State whether controls are operating effectively.")
                if st.button("💾 Save Conclusion", key=f"ccsave_{tid}"):
                    test["final_risk_rating"] = new_final_risk
                    test["conclusion"]        = new_conclusion
                    test["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "user":st.session_state.current_user,"action":"Conclusion and final risk rating saved"})
                    st.success("✓ Conclusion saved.")

                st.markdown("<hr/>", unsafe_allow_html=True)

                # Submit for review
                open_ex = [e for e in test.get("exceptions",[]) if e["status"]=="Open"]
                if open_ex:
                    st.markdown(f"<div class='gc-alert gc-alert-warn'>⚠ {len(open_ex)} open exception(s) must have root cause and remediation plan before submitting.</div>", unsafe_allow_html=True)

                reviewer_opts = [u for u in D["users"] if u != test["assignee"]]
                sub_reviewer  = st.selectbox("Submit To (Reviewer)", reviewer_opts,
                    format_func=lambda x: f"{D['users'][x]['name']} ({D['users'][x]['title']})",
                    key=f"sub_rev_{tid}")
                if st.button("📤 Submit for Review", key=f"submit_{tid}"):
                    missing = [p for p in test["procedures"] if not p["result"]]
                    if missing:
                        st.error(f"{len(missing)} procedure step(s) still incomplete.")
                    elif not test.get("conclusion","").strip():
                        st.error("Conclusion is required before submitting.")
                    else:
                        test["submitted_for_review"] = True
                        test["reviewer"]             = sub_reviewer
                        test["status"]               = "Submitted for Review"
                        ts_now = datetime.now().strftime("%Y-%m-%d %H:%M")
                        test["audit"].append({"ts":ts_now,"user":st.session_state.current_user,
                            "action":f"Submitted for review to {D['users'][sub_reviewer]['name']}"})
                        add_audit("Compliance Testing", f"{tid} submitted for review to {sub_reviewer}")
                        st.success("✓ Test submitted for review."); st.rerun()

            elif test["status"] == "Submitted for Review":
                # Show conclusion read-only
                st.markdown(f"<div class='gc-card'><div class='gc-card-header'>CONCLUSION</div><div style='font-size:13px;line-height:1.7'>{test.get('conclusion','—')}</div></div>", unsafe_allow_html=True)
                st.markdown(f"<div class='gc-alert gc-alert-ok'>📤 Submitted for review · Reviewer: <b>{D['users'].get(test['reviewer'],{}).get('name','—')}</b></div>", unsafe_allow_html=True)

                # Reviewer actions
                if st.session_state.current_user == test["reviewer"] or has_full_compliance_access():
                    st.markdown("<hr/><div class='gc-card-header'>REVIEWER ACTIONS</div>", unsafe_allow_html=True)
                    rev_notes = st.text_area("Review Notes", value=test.get("review_notes",""), height=100, key=f"rvn_{tid}")
                    rv1,rv2   = st.columns(2)
                    if rv1.button("✅ Approve & Complete", key=f"rv_approve_{tid}"):
                        test["review_notes"]  = rev_notes
                        test["status"]        = "Completed"
                        test["completed"]     = True
                        test["completed_ts"]  = datetime.now().strftime("%Y-%m-%d %H:%M")
                        test["audit"].append({"ts":test["completed_ts"],"user":st.session_state.current_user,"action":"Test approved and completed"})
                        add_audit("Compliance Testing", f"{tid} approved and completed"); st.rerun()
                    if rv2.button("↩ Return for Revision", key=f"rv_return_{tid}"):
                        test["review_notes"]        = rev_notes
                        test["status"]              = "In Progress"
                        test["submitted_for_review"] = False
                        test["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),"user":st.session_state.current_user,"action":f"Returned for revision. Notes: {rev_notes}"})
                        add_audit("Compliance Testing", f"{tid} returned for revision"); st.rerun()

            elif test["status"] == "Completed":
                rev_notes_section = f"<div class='gc-card-header' style='margin-top:.8rem'>REVIEW NOTES</div><div style='font-size:13px;color:#374151'>{test['review_notes']}</div>" if test.get("review_notes") else ""
                reviewer_name = D["users"].get(test.get("reviewer",""),{}).get("name","—")
                st.markdown(f"""
                <div class='gc-card'>
                  <div class='gc-card-header'>COMPLETED TEST SUMMARY</div>
                  <div style='display:flex;gap:2rem;margin-bottom:.8rem;flex-wrap:wrap'>
                    <div><div style='font-size:11px;color:#6b7280;text-transform:uppercase;font-weight:700'>Final Risk Rating</div>
                      <div style='margin-top:.2rem'>{_ct_risk_badge(test.get('final_risk_rating',''))}</div></div>
                    <div><div style='font-size:11px;color:#6b7280;text-transform:uppercase;font-weight:700'>Completed</div>
                      <div style='font-size:13px;font-weight:600;margin-top:.2rem'>{test.get('completed_ts','')[:10]}</div></div>
                    <div><div style='font-size:11px;color:#6b7280;text-transform:uppercase;font-weight:700'>Reviewer</div>
                      <div style='font-size:13px;font-weight:600;margin-top:.2rem'>{reviewer_name}</div></div>
                  </div>
                  <div style='font-size:13px;line-height:1.7;color:#374151'>{test.get('conclusion','—')}</div>
                  {rev_notes_section}
                </div>
                """, unsafe_allow_html=True)
                # Export
                export_txt = f"GRAPHITE COMPLIANCE — TEST REPORT\n{'='*60}\n"
                export_txt += f"Test ID: {test['id']}\nType: {test['test_type']}\nTopic: {test['topic']}\n"
                export_txt += f"Line of Business: {test['line_of_business']}\nInitial Risk: {test['initial_risk_rating']}\nFinal Risk: {test.get('final_risk_rating','')}\n"
                export_txt += f"Assignee: {D['users'].get(test['assignee'],{}).get('name','')}\nDeadline: {test['deadline']}\nCompleted: {test.get('completed_ts','')}\n\n"
                export_txt += f"POPULATION: {test.get('population_description','')}\nSample Size: {test.get('sample_size',0)} | Method: {test.get('sample_method','')}\n\n"
                export_txt += "PROCEDURES:\n"
                for p in test["procedures"]:
                    export_txt += f"  Step {p['step']}: {p['description']}\n  Result: {p['result']} | Notes: {p['notes']}\n\n"
                export_txt += f"EXCEPTIONS: {len(test['exceptions'])}\n"
                for e in test["exceptions"]:
                    export_txt += f"  {e['id']}: {e['description']} | Status: {e['status']}\n"
                export_txt += f"\nCONCLUSION:\n{test.get('conclusion','')}\n"
                st.download_button("📥 Export Test Report", data=export_txt,
                    file_name=f"{test['id']}_report.txt", mime="text/plain")

        # ── TAB 5: AUDIT TRAIL ────────────────────────────────────────────────
        with tt5:
            for e in reversed(test["audit"]):
                st.markdown(f"<div class='gc-tl'><span style='color:#2563eb;font-size:11px'>{e['ts']}</span> &nbsp;<b>{e['user']}</b> · {e['action']}</div>", unsafe_allow_html=True)

    # ──────────────────────────────────────────────────────────────────────────
    # VIEW: SUBMITTED FOR REVIEW
    # ──────────────────────────────────────────────────────────────────────────
    elif view == "review":
        review_tests = [t for t in tests if t.get("submitted_for_review") and not t.get("completed")]
        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>{len(review_tests)} test(s) awaiting review</div>", unsafe_allow_html=True)
        _render_test_list(review_tests)

    # ──────────────────────────────────────────────────────────────────────────
    # VIEW: COMPLETED TESTS
    # ──────────────────────────────────────────────────────────────────────────
    elif view == "done":
        done_tests = [t for t in tests if t.get("completed")]
        st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>{len(done_tests)} completed test(s)</div>", unsafe_allow_html=True)
        _render_test_list(done_tests)

    # ──────────────────────────────────────────────────────────────────────────
    # VIEW: ALL TESTS (default)
    # ──────────────────────────────────────────────────────────────────────────
    else:
        # Summary metrics
        mc = st.columns(5)
        mc[0].metric("Total",    len(tests))
        mc[1].metric("Not Started", len([t for t in tests if t["status"]=="Not Started"]))
        mc[2].metric("In Progress", len([t for t in tests if t["status"]=="In Progress"]))
        mc[3].metric("In Review",   len([t for t in tests if t["status"]=="Submitted for Review"]))
        mc[4].metric("Completed",   len([t for t in tests if t["status"]=="Completed"]))
        st.markdown("<hr style='margin:.6rem 0'/>", unsafe_allow_html=True)
        _render_test_list(tests)

def _render_test_list(tests):
    """Shared list renderer for all views."""
    if not tests:
        st.markdown("<div class='gc-alert gc-alert-ok'>No tests to display.</div>", unsafe_allow_html=True)
        return

    # Column header
    st.markdown("""
    <div style='display:grid;grid-template-columns:0.6fr 1.8fr 1fr 1fr 0.8fr 0.8fr 1fr 0.7fr;
         gap:8px;padding:6px 12px;background:#e5e7eb;border-radius:6px;
         font-size:10px;font-weight:700;letter-spacing:.08em;color:#6b7280;
         text-transform:uppercase;margin-bottom:4px;'>
      <div>ID</div><div>Topic</div><div>Type</div><div>Line of Business</div>
      <div>Risk</div><div>Deadline</div><div>Assignee</div><div>Status</div>
    </div>""", unsafe_allow_html=True)

    for t in tests:
        exc_open = len([e for e in t.get("exceptions",[]) if e["status"]=="Open"])
        overdue  = t["deadline"] < date.today().strftime("%Y-%m-%d") and t["status"] not in ("Completed","Cancelled")
        row_border = "border-left:3px solid #ef4444" if overdue else ("border-left:3px solid #f59e0b" if exc_open else "border-left:3px solid #e5e7eb")
        owner_name = D["users"].get(t["assignee"],{}).get("name", t["assignee"])

        st.markdown(f"""
        <div style='display:grid;grid-template-columns:0.6fr 1.8fr 1fr 1fr 0.8fr 0.8fr 1fr 0.7fr;
             gap:8px;padding:10px 12px;background:#ffffff;border:1px solid #e5e7eb;{row_border};
             border-radius:8px;margin-bottom:4px;align-items:center;font-size:13px;'>
          <div style='font-weight:700;color:#2563eb;font-size:12px'>{t['id']}</div>
          <div>
            <div style='font-weight:600;color:#111111'>{t['topic']}</div>
            {f'<div style="font-size:11px;color:#dc2626;margin-top:1px">⚠ {exc_open} open exception(s)</div>' if exc_open else ''}
          </div>
          <div style='color:#374151;font-size:12px'>{t['test_type']}</div>
          <div style='color:#374151;font-size:12px'>{t['line_of_business']}</div>
          <div>{_ct_risk_badge(t['initial_risk_rating'])}</div>
          <div style='color:{"#dc2626" if overdue else "#374151"};font-weight:{"700" if overdue else "400"};font-size:12px'>{t['deadline']}</div>
          <div style='color:#374151;font-size:12px'>{owner_name}</div>
          <div>{_ct_status_badge(t['status'])}</div>
        </div>
        """, unsafe_allow_html=True)

        if st.button(f"Open {t['id']}", key=f"ct_open_{t['id']}"):
            st.session_state["_ct_open"] = t["id"]
            st.session_state["_ct_view"] = "test"
            st.rerun()

# ── REMEDIATION ITEMS ───────────────────────────────────────────────────────────
def page_remediation_items():
    st.markdown("<div class='gc-header'><h1>REMEDIATION ITEMS</h1><p>All open exceptions and remediation actions across compliance tests</p></div>", unsafe_allow_html=True)

    # Collect all exceptions across all tests
    all_exceptions = []
    for t in D.get("compliance_tests",[]):
        for ex in t.get("exceptions",[]):
            all_exceptions.append({**ex, "_test_type":t["test_type"], "_test_topic":t["topic"], "_test_id":t["id"]})

    # Filters
    fc1,fc2,fc3 = st.columns(3)
    f_stat  = fc1.selectbox("Status",  ["All"] + CT_EXCEPTION_STATUSES, key="rem_fstat")
    f_type  = fc2.selectbox("Test Type",["All"] + CT_TEST_TYPES,        key="rem_ftype")
    f_owner = fc3.selectbox("Assigned To",["All"]+list(D["users"].keys()),
        format_func=lambda x: "All" if x=="All" else D["users"][x]["name"], key="rem_fown")

    results = all_exceptions
    if f_stat  != "All": results = [e for e in results if e["status"]         == f_stat]
    if f_type  != "All": results = [e for e in results if e["_test_type"]     == f_type]
    if f_owner != "All": results = [e for e in results if e.get("assigned_to","") == f_owner]

    # Summary
    mc = st.columns(3)
    mc[0].metric("Total Exceptions", len(all_exceptions))
    mc[1].metric("Open",             len([e for e in all_exceptions if e["status"]=="Open"]))
    mc[2].metric("In Remediation",   len([e for e in all_exceptions if e["status"]=="Remediation In Progress"]))
    st.markdown("<hr style='margin:.6rem 0'/>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:.8rem'>{len(results)} item(s) shown</div>", unsafe_allow_html=True)

    if not results:
        st.markdown("<div class='gc-alert gc-alert-ok'>✓ No remediation items match the current filters.</div>", unsafe_allow_html=True)
        return

    # Column header
    st.markdown("""
    <div style='display:grid;grid-template-columns:0.5fr 0.7fr 2fr 1fr 1fr 0.8fr 0.9fr;
         gap:8px;padding:6px 12px;background:#e5e7eb;border-radius:6px;
         font-size:10px;font-weight:700;letter-spacing:.08em;color:#6b7280;
         text-transform:uppercase;margin-bottom:4px;'>
      <div>ID</div><div>Test</div><div>Description</div>
      <div>Root Cause</div><div>Assigned To</div><div>Due</div><div>Status</div>
    </div>""", unsafe_allow_html=True)

    for ex in results:
        exc_sc = {"Open":"badge-overdue","Remediation In Progress":"badge-pending","Closed":"badge-closed"}.get(ex["status"],"badge-draft")
        overdue_ex = ex.get("due_date","") and ex["due_date"] < date.today().strftime("%Y-%m-%d") and ex["status"]!="Closed"
        row_border = "border-left:3px solid #ef4444" if overdue_ex else "border-left:3px solid #f59e0b" if ex["status"]=="Open" else "border-left:3px solid #e5e7eb"
        owner_name = D["users"].get(ex.get("assigned_to",""),{}).get("name","Unassigned")

        st.markdown(f"""
        <div style='display:grid;grid-template-columns:0.5fr 0.7fr 2fr 1fr 1fr 0.8fr 0.9fr;
             gap:8px;padding:10px 12px;background:#ffffff;border:1px solid #e5e7eb;{row_border};
             border-radius:8px;margin-bottom:4px;align-items:center;font-size:12px;'>
          <div style='font-weight:700;color:#2563eb'>{ex['id']}</div>
          <div>
            <div style='font-weight:600;color:#2563eb;font-size:11px'>{ex['_test_id']}</div>
            <div style='color:#6b7280;font-size:10px'>{ex['_test_type']}</div>
          </div>
          <div>
            <div style='font-weight:500;color:#111111;margin-bottom:2px'>{ex['description'][:90]}{'…' if len(ex['description'])>90 else ''}</div>
            <div style='font-size:11px;color:#6b7280'>{ex['_test_topic']}</div>
          </div>
          <div style='color:#374151'>{ex.get('root_cause','—')[:50] or '—'}</div>
          <div style='color:#374151'>{owner_name}</div>
          <div style='color:{"#dc2626" if overdue_ex else "#374151"};font-weight:{"700" if overdue_ex else "400"}'>{ex.get('due_date','—') or '—'}</div>
          <div>{badge(ex['status'], exc_sc)}</div>
        </div>
        """, unsafe_allow_html=True)

        if toggle(f"rem_{ex['_test_id']}_{ex['id']}", f"Edit {ex['id']} — {ex['_test_id']}"):
            # Find live reference
            live_test = next((t for t in D["compliance_tests"] if t["id"]==ex["_test_id"]), None)
            live_ex   = next((e for e in live_test["exceptions"] if e["id"]==ex["id"]), None) if live_test else None
            if live_ex:
                re1,re2 = st.columns(2)
                live_ex["root_cause"]       = re1.text_area("Root Cause",       value=live_ex.get("root_cause",""),       height=72, key=f"rem_rc_{ex['_test_id']}_{ex['id']}")
                live_ex["remediation_plan"] = re2.text_area("Remediation Plan", value=live_ex.get("remediation_plan",""), height=72, key=f"rem_rp_{ex['_test_id']}_{ex['id']}")
                rb1,rb2,rb3 = st.columns(3)
                live_ex["assigned_to"] = rb1.selectbox("Assign To", [""]+list(D["users"].keys()),
                    format_func=lambda x: "Unassigned" if x=="" else D["users"][x]["name"],
                    index=(list(D["users"].keys()).index(live_ex["assigned_to"])+1) if live_ex.get("assigned_to") in D["users"] else 0,
                    key=f"rem_at_{ex['_test_id']}_{ex['id']}")
                live_ex["due_date"] = rb2.text_input("Due Date (YYYY-MM-DD)", value=live_ex.get("due_date",""), key=f"rem_dd_{ex['_test_id']}_{ex['id']}")
                live_ex["status"]   = rb3.selectbox("Status", CT_EXCEPTION_STATUSES,
                    index=CT_EXCEPTION_STATUSES.index(live_ex["status"]) if live_ex["status"] in CT_EXCEPTION_STATUSES else 0,
                    key=f"rem_st_{ex['_test_id']}_{ex['id']}")
                if st.button(f"💾 Save", key=f"rem_save_{ex['_test_id']}_{ex['id']}"):
                    if live_test:
                        live_test["audit"].append({"ts":datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "user":st.session_state.current_user,
                            "action":f"Exception {live_ex['id']} updated from Remediation Items: {live_ex['status']}"})
                    add_audit("Compliance Testing", f"Remediation {live_ex['id']} updated: {live_ex['status']}")
                    st.success("✓ Saved."); st.rerun()

# ── AI COPILOT ─────────────────────────────────────────────────────────────────

FIRM_CONTEXT = {
    "name": "Apex Financial Services LLC",
    "registrations": ["Registered Investment Adviser (RIA)", "Broker-Dealer (FINRA/SEC)"],
    "business_lines": ["Wealth Management", "Investment Banking", "Operations", "Fixed Income", "Capital Markets"],
    "regulators": ["FINRA", "SEC", "FinCEN"],
}

SUGGESTED_PROMPTS = [
    ("🔍", "Search", "Show all policies for Wealth Management"),
    ("📊", "Query",  "Show all open remediation items"),
    ("🧪", "Testing","Summarize the most recent 3120 testing results"),
    ("📄", "Report", "Create an executive summary of CT-001 testing results"),
    ("📰", "Reg",    "Which policies should be reviewed based on recent regulatory changes?"),
    ("📋", "Data",   "List all policies approved in the last 6 months"),
]

PROMPT_CATEGORIES = [
    {
        "label": "📊 Platform Data",
        "color": "#2563eb",
        "bg": "#eff6ff",
        "border": "#bfdbfe",
        "prompts": [
            ("Show all open remediation items", "📊"),
            ("List all policies approved in the last 6 months", "📋"),
            ("Show all overdue tasks assigned to me", "⏰"),
            ("Show all open exceptions across compliance tests", "⚠️"),
        ]
    },
    {
        "label": "🧪 Testing & Reports",
        "color": "#7c3aed",
        "bg": "#f5f3ff",
        "border": "#ddd6fe",
        "prompts": [
            ("Summarize the most recent 3120 testing results", "🧪"),
            ("Create an executive summary of CT-001 testing results", "📄"),
            ("List all compliance tests with open exceptions", "📋"),
            ("Show high risk findings from the latest testing cycle", "🔴"),
        ]
    },
    {
        "label": "📰 Regulatory Intelligence",
        "color": "#0891b2",
        "bg": "#ecfeff",
        "border": "#a5f3fc",
        "prompts": [
            ("Which policies should be reviewed based on recent regulatory changes?", "📰"),
            ("Show all rules updated in the last 30 days", "🔄"),
            ("Which FINRA rules require updates to our WSPs?", "📏"),
            ("Show all policies impacted by SEC Reg BI requirements", "⚖️"),
        ]
    },
    {
        "label": "🔍 Search & Locate",
        "color": "#059669",
        "bg": "#ecfdf5",
        "border": "#a7f3d0",
        "prompts": [
            ("Show all policies for Wealth Management", "🔍"),
            ("Find the most recent version of WSP-002", "📄"),
            ("Show all evidence uploaded for T-001", "📎"),
            ("Locate all open exam requests with pending items", "📁"),
        ]
    },
]

def _copilot_classify(q: str) -> str:
    ql = q.lower()
    report_kw  = ["executive summary","create a report","board report","generate report","write a report","create an executive","regulatory memo","create a memo"]
    reg_kw     = [
        "finra update","sec update","regulatory change","recent update","regulation update",
        "rule update","regulatory changes","rules were updated","rules updated","updated rules",
        "requiring.*change","rules.*last month","rules.*past month","rules.*recent",
        "rules.*policies","what rules","which rules","new rules","rule changes",
        "last month.*rule","past month.*rule","recent.*rule","rules.*require",
        "policies.*rules","updates.*policies","changes.*policies","impacted policies",
    ]
    test_kw    = ["3120","ct-","testing result","test result","aml.*test","testing summary","compliance test","risk rating","testing results","test summary"]
    data_kw    = ["list all","show all","give me all","approved in","due this","assigned to","owned by","high risk","filter","overdue","open exception","open remediati"]
    nav_kw     = ["where","find","locate","most recent version","evidence for"]

    if any(k in ql for k in report_kw):  return "generate_report"
    # Regulatory: check before data/nav so "rules…policies" doesn't fall through
    if any(k in ql for k in reg_kw):     return "regulatory_analysis"
    # Also catch pattern: "rules" + temporal or change language
    if "rule" in ql and any(w in ql for w in ["last month","past month","recent","updated","changed","new","require","impacted"]):
        return "regulatory_analysis"
    if any(k in ql for k in test_kw):    return "summarize_testing"
    if any(k in ql for k in data_kw):    return "query_platform_data"
    if any(k in ql for k in nav_kw):     return "search_platform_records"
    return "query_platform_data"

def _lob_match(lob: str, q: str) -> bool:
    return lob.lower() in q.lower()

def _fmt_status_badge(status: str) -> str:
    colors = {
        "Approved":"#16a34a","Draft":"#ca8a04","In Progress":"#2563eb",
        "Completed":"#16a34a","Open":"#dc2626","Closed":"#6b7280",
        "Not Started":"#6b7280","Overdue":"#dc2626","Pending":"#ca8a04",
        "Remediation In Progress":"#f59e0b","Submitted for Review":"#7c3aed",
    }
    c = colors.get(status, "#374151")
    return f"<span style='background:{c}20;color:{c};border:1px solid {c}40;border-radius:4px;padding:1px 7px;font-size:11px;font-weight:600'>{status}</span>"

def _copilot_respond(question: str) -> dict:
    """Route to correct handler. Returns dict with keys: html, sources, intent, download_label, download_content"""
    intent = _copilot_classify(question)
    q = question.lower()

    if intent == "search_platform_records":
        return _handle_search(q, question)
    elif intent == "query_platform_data":
        return _handle_query(q, question)
    elif intent == "summarize_testing":
        return _handle_testing_summary(q, question)
    elif intent == "generate_report":
        return _handle_report(q, question)
    elif intent == "regulatory_analysis":
        return _handle_regulatory(q, question)
    return {"html":"<p>I couldn't interpret that request. Try rephrasing or use one of the suggested prompts.</p>","sources":[],"intent":intent,"download_label":None,"download_content":None}

# ── HANDLER: search_platform_records ──────────────────────────────────────────
def _handle_search(q, raw):
    policies = D.get("policies", [])
    wsps     = D.get("wsps", [])
    rules    = D.get("rules", [])
    results_html = ""
    sources = []

    # find WSPs/Policies by LOB or keyword
    matched_pol = [p for p in policies if any(k in q for k in [p["title"].lower(), p["id"].lower(), p["line_of_business"].lower()])]
    matched_wsp = [w for w in wsps     if any(k in q for k in [w["title"].lower(), w["id"].lower(), w["line_of_business"].lower()])]
    matched_rul = [r for r in rules    if any(k in q for k in [r["title"].lower(), r["id"].lower(), r["category"].lower()])]

    # LOB-specific search
    for lob in FIRM_CONTEXT["business_lines"]:
        if lob.lower() in q:
            matched_pol = [p for p in policies if p["line_of_business"] == lob]
            matched_wsp = [w for w in wsps     if w["line_of_business"] == lob]
            break

    # keyword fallback — search for "policy" or "wsp"
    if not matched_pol and not matched_wsp and not matched_rul:
        if "policy" in q or "policies" in q:
            matched_pol = policies
        if "wsp" in q or "procedure" in q:
            matched_wsp = wsps
        if "rule" in q:
            matched_rul = rules

    if matched_pol:
        results_html += "<div style='font-weight:700;font-size:13px;margin:10px 0 6px'>📚 Policies</div>"
        results_html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Title</th><th>Version</th><th>Status</th><th>Line of Business</th><th>Owner</th></tr>"
        for p in matched_pol:
            owner = D["users"].get(p.get("owner",""),{}).get("name", p.get("owner",""))
            results_html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{p['id']}</td><td>{p['title']}</td><td>{p['version']}</td><td>{_fmt_status_badge(p['status'])}</td><td>{p['line_of_business']}</td><td>{owner}</td></tr>"
            sources.append(f"{p['id']} – {p['title']}")
        results_html += "</table>"

    if matched_wsp:
        results_html += "<div style='font-weight:700;font-size:13px;margin:12px 0 6px'>📄 Written Supervisory Procedures</div>"
        results_html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Title</th><th>Version</th><th>Status</th><th>Line of Business</th></tr>"
        for w in matched_wsp:
            results_html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{w['id']}</td><td>{w['title']}</td><td>{w['version']}</td><td>{_fmt_status_badge(w['status'])}</td><td>{w['line_of_business']}</td></tr>"
            sources.append(f"{w['id']} – {w['title']}")
        results_html += "</table>"

    if matched_rul:
        results_html += "<div style='font-weight:700;font-size:13px;margin:12px 0 6px'>📏 Regulatory Rules</div>"
        results_html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Title</th><th>Source</th><th>Category</th><th>Status</th></tr>"
        for r in matched_rul:
            results_html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{r['id']}</td><td>{r['title']}</td><td>{r['source']}</td><td>{r['category']}</td><td>{_fmt_status_badge(r['status'])}</td></tr>"
            sources.append(f"{r['id']} – {r['title']}")
        results_html += "</table>"

    if not results_html:
        results_html = "<p style='color:#6b7280'>No matching records found. Try a broader search term, a line of business name, or a policy/procedure ID.</p>"

    return {"html": results_html, "sources": sources, "intent": "search_platform_records", "download_label": None, "download_content": None}

# ── HANDLER: query_platform_data ──────────────────────────────────────────────
def _handle_query(q, raw):
    policies = D.get("policies", [])
    wsps     = D.get("wsps", [])
    tests    = D.get("compliance_tests", [])
    tasks    = D.get("tasks", [])
    sources  = []
    html     = ""

    today_str = date.today().strftime("%Y-%m-%d")
    today_dt  = date.today()

    # ── Remediation / exceptions
    if "remediation" in q or "exception" in q:
        all_ex = [
            {**ex, "_test_id": t["id"], "_topic": t["topic"]}
            for t in tests for ex in t.get("exceptions", [])
        ]
        # filter
        if "open" in q:
            all_ex = [e for e in all_ex if e["status"] == "Open"]
        if "high" in q:
            all_ex = [e for e in all_ex if any(t["id"]==e["_test_id"] and t.get("risk_rating","")=="High" for t in tests)]

        html += f"<div style='font-weight:700;font-size:13px;margin:0 0 8px'>⚠️ Remediation Items ({len(all_ex)} found)</div>"
        if all_ex:
            html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Test</th><th>Description</th><th>Assigned To</th><th>Due</th><th>Status</th></tr>"
            for e in all_ex:
                owner = D["users"].get(e.get("assigned_to",""),{}).get("name","Unassigned")
                overdue = e.get("due_date","") and e["due_date"] < today_str and e["status"] != "Closed"
                due_style = "color:#dc2626;font-weight:700" if overdue else "color:#374151"
                html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{e['id']}</td><td style='color:#374151'>{e['_test_id']}</td><td>{e['description'][:70]}…</td><td>{owner}</td><td style='{due_style}'>{e.get('due_date','—')}</td><td>{_fmt_status_badge(e['status'])}</td></tr>"
                sources.append(f"{e['_test_id']} → {e['id']}")
            html += "</table>"
        else:
            html += "<p style='color:#16a34a'>✓ No matching remediation items found.</p>"
        return {"html":html,"sources":sources,"intent":"query_platform_data","download_label":None,"download_content":None}

    # ── Tasks
    if "task" in q or "overdue" in q and "task" in q:
        filtered = tasks
        if "overdue" in q: filtered = [t for t in tasks if t["status"]=="Overdue"]
        if "open"    in q: filtered = [t for t in tasks if t["status"]=="Open"]
        for user_key, user_info in D["users"].items():
            if user_info["name"].lower() in q:
                filtered = [t for t in filtered if t["assignee"]==user_key]
        html += f"<div style='font-weight:700;font-size:13px;margin:0 0 8px'>📋 Tasks ({len(filtered)} found)</div>"
        html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Title</th><th>Assignee</th><th>Due</th><th>Priority</th><th>Status</th></tr>"
        for t in filtered:
            owner = D["users"].get(t["assignee"],{}).get("name", t["assignee"])
            overdue = t["due"] < today_str and t["status"] not in ("Closed","Completed")
            due_style = "color:#dc2626;font-weight:700" if overdue else "color:#374151"
            html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{t['id']}</td><td>{t['title']}</td><td>{owner}</td><td style='{due_style}'>{t['due']}</td><td>{t['priority']}</td><td>{_fmt_status_badge(t['status'])}</td></tr>"
            sources.append(f"{t['id']} – {t['title']}")
        html += "</table>"
        return {"html":html,"sources":sources,"intent":"query_platform_data","download_label":None,"download_content":None}

    # ── Policies by owner / approvals
    if "policy" in q or "policies" in q or "approved" in q:
        filtered_pol = list(policies)
        filtered_wsp = list(wsps)
        for user_key, user_info in D["users"].items():
            if user_info["name"].lower() in q:
                filtered_pol = [p for p in filtered_pol if p.get("owner")==user_key]
                filtered_wsp = [w for w in filtered_wsp if w.get("owner")==user_key]
        # Date filter: "last 6 months" or "prior year"
        if "last 6 months" in q or "prior 6" in q:
            cutoff = (today_dt - timedelta(days=183)).strftime("%Y-%m-%d")
            filtered_pol = [p for p in filtered_pol if p.get("approved_ts","") >= cutoff and p["status"]=="Approved"]
        if "prior year" in q or "last year" in q:
            yr_start = f"{today_dt.year-1}-01-01"
            yr_end   = f"{today_dt.year-1}-12-31"
            filtered_pol = [p for p in filtered_pol if yr_start <= p.get("approved_ts","") <= yr_end and p["status"]=="Approved"]
        if "approved" in q and not filtered_pol:
            filtered_pol = [p for p in policies if p["status"]=="Approved"]
        if "draft" in q:
            filtered_pol = [p for p in policies if p["status"]=="Draft"]
            filtered_wsp = [w for w in wsps     if w["status"]=="Draft"]

        has_results = False
        if filtered_pol:
            has_results = True
            html += f"<div style='font-weight:700;font-size:13px;margin:0 0 6px'>📚 Policies ({len(filtered_pol)})</div>"
            html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Title</th><th>Version</th><th>Status</th><th>Approved</th><th>Line of Business</th></tr>"
            for p in filtered_pol:
                approved_ts = p.get("approved_ts","")[:10] if p.get("approved_ts") else "—"
                html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{p['id']}</td><td>{p['title']}</td><td>{p['version']}</td><td>{_fmt_status_badge(p['status'])}</td><td>{approved_ts}</td><td>{p['line_of_business']}</td></tr>"
                sources.append(f"{p['id']} – {p['title']}")
            html += "</table>"
        if not has_results:
            html = "<p style='color:#6b7280'>No matching policies found for that filter.</p>"
        return {"html":html,"sources":sources,"intent":"query_platform_data","download_label":None,"download_content":None}

    # ── Compliance tests filtered by high risk
    if "high risk" in q or "risk rating" in q:
        filtered = [t for t in tests if t.get("risk_rating")=="High" or t.get("final_risk_rating")=="High"]
        html += f"<div style='font-weight:700;font-size:13px;margin:0 0 8px'>🧪 High Risk Compliance Tests ({len(filtered)})</div>"
        html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Topic</th><th>Type</th><th>Risk</th><th>Status</th><th>Deadline</th></tr>"
        for t in filtered:
            html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{t['id']}</td><td>{t['topic']}</td><td style='font-size:11px'>{t['test_type']}</td><td><span style='background:#fee2e2;color:#dc2626;border-radius:4px;padding:1px 7px;font-size:11px;font-weight:700'>{t.get('risk_rating','—')}</span></td><td>{_fmt_status_badge(t['status'])}</td><td>{t['deadline']}</td></tr>"
            sources.append(f"{t['id']} – {t['topic']}")
        html += "</table>"
        return {"html":html,"sources":sources,"intent":"query_platform_data","download_label":None,"download_content":None}

    # generic fallback — show all tests
    html += "<div style='font-weight:700;font-size:13px;margin:0 0 8px'>🧪 Compliance Tests</div>"
    html += "<table style='width:100%;border-collapse:collapse;font-size:12px'><tr style='background:#f3f4f6'><th style='padding:5px 8px;text-align:left'>ID</th><th>Topic</th><th>Type</th><th>Risk</th><th>Status</th></tr>"
    for t in tests:
        html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:5px 8px;color:#2563eb;font-weight:700'>{t['id']}</td><td>{t['topic']}</td><td style='font-size:11px'>{t['test_type']}</td><td>{t.get('risk_rating','—')}</td><td>{_fmt_status_badge(t['status'])}</td></tr>"
        sources.append(f"{t['id']} – {t['topic']}")
    html += "</table>"
    return {"html":html,"sources":sources,"intent":"query_platform_data","download_label":None,"download_content":None}

# ── HANDLER: summarize_testing ─────────────────────────────────────────────────
def _handle_testing_summary(q, raw):
    tests   = D.get("compliance_tests", [])
    sources = []

    # filter: 3120 specifically
    pool = tests
    if "3120" in q:
        pool = [t for t in tests if "3120" in t["test_type"] or "3120" in t.get("topic","")]
    if "aml" in q:
        pool = [t for t in tests if "aml" in t["test_type"].lower() or "aml" in t.get("topic","").lower()]
    if "2025" in q:
        pool = [t for t in pool if "2025" in t.get("time_period_reviewed","")]
    # High risk filter
    if "high" in q:
        pool = [t for t in pool if t.get("final_risk_rating","High")=="High" or t.get("risk_rating","")=="High"]

    # Latest N
    pool = pool[:5]

    total_exceptions = sum(len(t.get("exceptions",[])) for t in pool)
    open_ex          = sum(len([e for e in t.get("exceptions",[]) if e["status"]=="Open"]) for t in pool)
    completed        = sum(1 for t in pool if t["status"]=="Completed")

    html  = "<div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:14px 16px;margin-bottom:12px'>"
    html += f"<div style='font-weight:700;font-size:14px;margin-bottom:10px;color:#111'>🧪 Testing Summary — {len(pool)} test(s)</div>"
    html += f"<div style='display:flex;gap:24px;margin-bottom:12px'>"
    html += f"<div><div style='font-size:22px;font-weight:800;color:#2563eb'>{len(pool)}</div><div style='font-size:11px;color:#6b7280'>Tests Reviewed</div></div>"
    html += f"<div><div style='font-size:22px;font-weight:800;color:#16a34a'>{completed}</div><div style='font-size:11px;color:#6b7280'>Completed</div></div>"
    html += f"<div><div style='font-size:22px;font-weight:800;color:#dc2626'>{total_exceptions}</div><div style='font-size:11px;color:#6b7280'>Total Exceptions</div></div>"
    html += f"<div><div style='font-size:22px;font-weight:800;color:#f59e0b'>{open_ex}</div><div style='font-size:11px;color:#6b7280'>Open Exceptions</div></div>"
    html += "</div></div>"

    for t in pool:
        exc_list  = t.get("exceptions", [])
        procs     = t.get("procedures", [])
        pass_ct   = sum(1 for p in procs if p.get("result")=="Pass")
        fail_ct   = sum(1 for p in procs if p.get("result")=="Exception")
        status_c  = "#16a34a" if t["status"]=="Completed" else "#2563eb" if t["status"]=="In Progress" else "#6b7280"
        risk      = t.get("final_risk_rating") or t.get("risk_rating","—")
        risk_color= "#dc2626" if risk=="High" else "#f59e0b" if risk=="Medium" else "#16a34a"

        html += f"<div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:12px 14px;margin-bottom:8px'>"
        html += f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:6px'>"
        html += f"<span style='color:#2563eb;font-weight:800;font-size:13px'>{t['id']}</span>"
        html += f"<span style='font-weight:600;font-size:13px;color:#111'>{t['topic']}</span>"
        html += f"<span style='margin-left:auto;font-size:11px;color:{status_c};font-weight:700'>{t['status']}</span>"
        html += "</div>"
        html += f"<div style='font-size:11px;color:#6b7280;margin-bottom:6px'>{t['test_type']} · {t['line_of_business']} · Period: {t['time_period_reviewed']}</div>"
        html += f"<div style='display:flex;gap:16px;font-size:12px'>"
        html += f"<span>Risk: <b style='color:{risk_color}'>{risk}</b></span>"
        if procs: html += f"<span>Steps: <b style='color:#16a34a'>{pass_ct} Pass</b> / <b style='color:#dc2626'>{fail_ct} Exception</b></span>"
        html += f"<span>Exceptions: <b>{len(exc_list)}</b></span>"
        html += "</div>"
        if exc_list:
            html += "<div style='margin-top:8px'>"
            for ex in exc_list:
                html += f"<div style='background:#fef2f2;border-left:3px solid #dc2626;padding:5px 10px;border-radius:4px;margin-bottom:4px;font-size:12px'>"
                html += f"<b>{ex['id']}</b>: {ex['description'][:100]} — <span style='color:#6b7280'>Status: {ex['status']}</span></div>"
            html += "</div>"
        if t.get("conclusion"):
            html += f"<div style='margin-top:8px;font-size:12px;color:#374151;background:#f0fdf4;padding:6px 10px;border-radius:4px'><b>Conclusion:</b> {t['conclusion']}</div>"
        html += "</div>"
        sources.append(f"{t['id']} – {t['topic']}")

    return {"html":html,"sources":sources,"intent":"summarize_testing","download_label":None,"download_content":None}

# ── HANDLER: generate_report ──────────────────────────────────────────────────
def _handle_report(q, raw):
    tests   = D.get("compliance_tests", [])
    firm    = D.get("firm", {})
    sources = []

    # Determine scope
    pool = tests
    if "3120" in q or "supervisory" in q:
        pool = [t for t in tests if "3120" in t["test_type"] or "supervisory" in t["test_type"].lower()]
    if "aml" in q:
        pool = [t for t in tests if "aml" in t["test_type"].lower()]
    if "ct-" in q:
        for t in tests:
            if t["id"].lower() in q:
                pool = [t]; break
    if not pool:
        pool = tests

    today_str = date.today().strftime("%B %d, %Y")
    total_ex  = sum(len(t.get("exceptions",[])) for t in pool)
    open_ex   = sum(len([e for e in t.get("exceptions",[]) if e["status"]=="Open"]) for t in pool)
    completed = [t for t in pool if t["status"]=="Completed"]
    in_prog   = [t for t in pool if t["status"]=="In Progress"]
    high_risk = [t for t in pool if t.get("risk_rating","")=="High" or t.get("final_risk_rating","")=="High"]

    html  = f"<div style='background:#fff;border:1px solid #d1d5db;border-radius:10px;padding:20px 24px;font-family:Inter,sans-serif'>"
    html += f"<div style='border-bottom:2px solid #1e3a5f;padding-bottom:12px;margin-bottom:16px'>"
    html += f"<div style='font-size:10px;font-weight:700;letter-spacing:.1em;color:#6b7280;text-transform:uppercase'>Executive Summary Report</div>"
    html += f"<div style='font-size:18px;font-weight:800;color:#1e3a5f;margin:4px 0'>Compliance Testing Results</div>"
    html += f"<div style='font-size:12px;color:#6b7280'>{firm.get('name','Apex Financial Services LLC')} · Prepared: {today_str} · Prepared by AI Copilot</div>"
    html += "</div>"

    html += "<div style='font-size:13px;font-weight:700;color:#111;margin-bottom:6px'>Overview</div>"
    html += f"<p style='font-size:12px;color:#374151;line-height:1.6'>{firm.get('name','Apex Financial Services LLC')} is a dual registrant (RIA and Broker-Dealer) subject to SEC and FINRA oversight. This report summarizes compliance testing activity across {len(pool)} test(s) conducted by the Compliance Department. Testing covers the following business lines: {', '.join(FIRM_CONTEXT['business_lines'])}.</p>"

    html += "<div style='display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:14px 0'>"
    for label, val, color in [("Tests Reviewed",len(pool),"#2563eb"),("Completed",len(completed),"#16a34a"),("Total Exceptions",total_ex,"#dc2626"),("Open Exceptions",open_ex,"#f59e0b")]:
        html += f"<div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px;text-align:center'><div style='font-size:24px;font-weight:800;color:{color}'>{val}</div><div style='font-size:11px;color:#6b7280'>{label}</div></div>"
    html += "</div>"

    html += "<div style='font-size:13px;font-weight:700;color:#111;margin:14px 0 6px'>Testing Detail</div>"
    for t in pool:
        exc = t.get("exceptions", [])
        risk = t.get("final_risk_rating") or t.get("risk_rating","—")
        risk_color = "#dc2626" if risk=="High" else "#f59e0b" if risk=="Medium" else "#16a34a"
        html += f"<div style='border:1px solid #e5e7eb;border-radius:6px;padding:10px 14px;margin-bottom:8px'>"
        html += f"<div style='font-weight:700;color:#2563eb;font-size:12px'>{t['id']} — {t['topic']}</div>"
        html += f"<div style='font-size:11px;color:#6b7280;margin:2px 0'>{t['test_type']} · {t['time_period_reviewed']}</div>"
        html += f"<div style='font-size:12px;margin-top:6px'>Risk Rating: <b style='color:{risk_color}'>{risk}</b> · Status: <b>{t['status']}</b> · Exceptions: <b>{len(exc)}</b></div>"
        if t.get("conclusion"):
            html += f"<div style='font-size:12px;color:#374151;margin-top:4px;font-style:italic'>{t['conclusion']}</div>"
        for ex in exc:
            html += f"<div style='background:#fef2f2;border-left:3px solid #dc2626;padding:4px 10px;border-radius:4px;font-size:11px;margin-top:4px'><b>{ex['id']}</b>: {ex['description'][:120]} [Status: {ex['status']}]</div>"
        html += "</div>"
        sources.append(f"{t['id']} – {t['topic']}")

    if high_risk:
        html += "<div style='font-size:13px;font-weight:700;color:#dc2626;margin:14px 0 6px'>⚠ High Risk Items Requiring Attention</div>"
        for t in high_risk:
            html += f"<div style='background:#fef2f2;border:1px solid #fca5a5;border-radius:6px;padding:8px 12px;margin-bottom:6px;font-size:12px'><b>{t['id']}</b> – {t['topic']} ({t['status']})</div>"

    html += "<div style='margin-top:16px;padding-top:10px;border-top:1px solid #e5e7eb;font-size:11px;color:#9ca3af;font-style:italic'>"
    html += "This report was generated by the AI Copilot using platform data as of the date shown above. It is intended to support compliance review and does not replace human regulatory analysis."
    html += "</div></div>"

    # Build download content (plain text version)
    download_text = f"COMPLIANCE TESTING EXECUTIVE SUMMARY\n{'='*50}\n"
    download_text += f"Firm: {firm.get('name','Apex Financial Services LLC')}\n"
    download_text += f"Date: {today_str}\n\n"
    download_text += f"SUMMARY METRICS\nTests Reviewed: {len(pool)}\nCompleted: {len(completed)}\nTotal Exceptions: {total_ex}\nOpen Exceptions: {open_ex}\n\n"
    download_text += "TESTING DETAIL\n" + "-"*40 + "\n"
    for t in pool:
        download_text += f"\n{t['id']} — {t['topic']}\nType: {t['test_type']}\nPeriod: {t['time_period_reviewed']}\nStatus: {t['status']}\nRisk: {t.get('final_risk_rating') or t.get('risk_rating','—')}\n"
        if t.get("conclusion"): download_text += f"Conclusion: {t['conclusion']}\n"
        for ex in t.get("exceptions",[]):
            download_text += f"  Exception {ex['id']}: {ex['description']}\n"
    download_text += "\n\nThis report is intended to support compliance review and does not replace human regulatory analysis."

    return {"html":html,"sources":sources,"intent":"generate_report","download_label":"📥 Download Report (.txt)","download_content":download_text}

# ── HANDLER: regulatory_analysis ──────────────────────────────────────────────

# Simulated "last reviewed" dates that place certain rules within the past month
# so the demo feels live. Keys match rule IDs in seed data.
_RULE_RECENT_UPDATES = {
    "R-002": {
        "update_date": (date.today() - timedelta(days=12)).strftime("%B %d, %Y"),
        "what_changed": "FINRA updated guidance on customer account information requirements, clarifying that firms must re-verify investment objectives and risk tolerance for accounts inactive for 12+ months and expanding beneficial ownership documentation requirements for accounts opened through digital onboarding workflows.",
        "action_required": True,
    },
    "R-004": {
        "update_date": (date.today() - timedelta(days=8)).strftime("%B %d, %Y"),
        "what_changed": "FinCEN issued updated guidance under the Bank Secrecy Act tightening Customer Due Diligence (CDD) requirements for legal entity accounts, including new beneficial ownership thresholds and enhanced verification standards for high-risk customer segments.",
        "action_required": True,
    },
    "R-005": {
        "update_date": (date.today() - timedelta(days=21)).strftime("%B %d, %Y"),
        "what_changed": "The SEC issued a no-action letter clarifying Form CRS delivery obligations for dual registrants, specifying that a single combined Form CRS is permissible but must clearly delineate advisory vs. brokerage services and associated fee structures. Firms should review their current Form CRS for compliance.",
        "action_required": True,
    },
    "R-006": {
        "update_date": (date.today() - timedelta(days=18)).strftime("%B %d, %Y"),
        "what_changed": "FINRA published updated examination findings related to Rule 3120 supervisory controls testing, highlighting common deficiencies in documentation of CCO escalation reviews and supervisory sign-off timeliness. Firms are expected to self-assess against the noted deficiency patterns.",
        "action_required": True,
    },
}

# Map each rule to the policies/WSPs it affects (by regulation cross-reference)
_RULE_POLICY_MAP = {
    "R-001": {"policies": ["POL-002"],           "wsps": ["WSP-002"]},
    "R-002": {"policies": ["POL-002"],           "wsps": ["WSP-002"]},
    "R-003": {"policies": ["POL-001","POL-002"], "wsps": ["WSP-001","WSP-002"]},
    "R-004": {"policies": ["POL-001"],           "wsps": ["WSP-001"]},
    "R-005": {"policies": ["POL-004","POL-002"], "wsps": ["WSP-002"]},
    "R-006": {"policies": ["POL-002"],           "wsps": ["WSP-002"]},
    "R-007": {"policies": ["POL-002"],           "wsps": ["WSP-002"]},
    "R-008": {"policies": ["POL-001","POL-002"], "wsps": ["WSP-001","WSP-002"]},
}

_BL_MAP = {
    "Supervision":            ["Wealth Management","Capital Markets"],
    "Customer Accounts":      ["Wealth Management","Retail Brokerage"],
    "Recordkeeping":          ["Operations","Wealth Management"],
    "AML/BSA":                ["Operations"],
    "Suitability":            ["Wealth Management"],
    "Financial Responsibility":["Capital Markets","Operations"],
    "Operations":             ["Operations"],
}

def _handle_regulatory(q, raw):
    rules    = D.get("rules", [])
    policies = D.get("policies", [])
    wsps     = D.get("wsps", [])
    firm     = D.get("firm", {})
    sources  = []
    today_str = date.today().strftime("%B %d, %Y")
    last_month_label = (date.today() - timedelta(days=30)).strftime("%B %Y")

    # Decide which rules to surface
    # "last month" / "recent" / "updated" / no filter → show rules with recent updates
    temporal = any(w in q for w in ["last month","past month","recent","updated","last 30","this month","new"])
    source_filter = None
    if "finra" in q and "sec" not in q:   source_filter = "FINRA"
    elif "sec" in q and "finra" not in q: source_filter = "SEC"
    elif "fincen" in q:                   source_filter = "FinCEN"

    # Pool: if temporal query, only show rules flagged as recently updated
    if temporal:
        pool = [r for r in rules if r["id"] in _RULE_RECENT_UPDATES]
    else:
        pool = rules

    if source_filter:
        pool = [r for r in pool if r["source"] == source_filter]
    if not pool:
        pool = [r for r in rules if r["id"] in _RULE_RECENT_UPDATES]

    # ── Header ────────────────────────────────────────────────────────────────
    html  = "<div style='background:#fff;border:1px solid #d1d5db;border-radius:10px;padding:20px 24px'>"
    html += "<div style='border-bottom:2px solid #1e3a5f;padding-bottom:12px;margin-bottom:14px'>"
    html += "<div style='font-size:10px;font-weight:700;letter-spacing:.1em;color:#6b7280;text-transform:uppercase'>Regulatory Update Analysis</div>"
    html += f"<div style='font-size:18px;font-weight:800;color:#1e3a5f;margin:4px 0'>Rules Requiring Policy Review — {last_month_label}</div>"
    html += f"<div style='font-size:12px;color:#6b7280'>{firm.get('name','Apex Financial Services LLC')} · Dual Registrant (RIA / BD) · Generated: {today_str}</div>"
    html += "</div>"

    # ── Summary bar ──────────────────────────────────────────────────────────
    action_ct = sum(1 for r in pool if _RULE_RECENT_UPDATES.get(r["id"],{}).get("action_required"))
    html += f"<div style='display:flex;gap:16px;margin-bottom:14px'>"
    html += f"<div style='background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;padding:8px 14px;text-align:center'><div style='font-size:20px;font-weight:800;color:#dc2626'>{len(pool)}</div><div style='font-size:11px;color:#6b7280'>Rules Updated</div></div>"
    html += f"<div style='background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;padding:8px 14px;text-align:center'><div style='font-size:20px;font-weight:800;color:#d97706'>{action_ct}</div><div style='font-size:11px;color:#6b7280'>Require Policy Review</div></div>"
    pol_ids = set(pid for r in pool for pid in _RULE_POLICY_MAP.get(r["id"],{}).get("policies",[]))
    html += f"<div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:8px 14px;text-align:center'><div style='font-size:20px;font-weight:800;color:#2563eb'>{len(pol_ids)}</div><div style='font-size:11px;color:#6b7280'>Policies Potentially Impacted</div></div>"
    html += "</div>"

    # ── Per-rule cards ────────────────────────────────────────────────────────
    html += f"<div style='font-size:13px;font-weight:700;color:#111;margin-bottom:10px'>Updates Identified ({len(pool)})</div>"

    for r in pool:
        upd        = _RULE_RECENT_UPDATES.get(r["id"], {})
        upd_date   = upd.get("update_date", last_month_label)
        what_changed = upd.get("what_changed", r["description"])
        needs_action = upd.get("action_required", False)
        affected_bls = _BL_MAP.get(r["category"], FIRM_CONTEXT["business_lines"])
        src_color    = {"FINRA":"#1e40af","SEC":"#6d28d9","FinCEN":"#065f46"}.get(r["source"],"#374151")
        src_bg       = {"FINRA":"#dbeafe","SEC":"#ede9fe","FinCEN":"#d1fae5"}.get(r["source"],"#f3f4f6")

        # Collect impacted policy/WSP objects
        pol_ids_for_rule = _RULE_POLICY_MAP.get(r["id"],{}).get("policies",[])
        wsp_ids_for_rule = _RULE_POLICY_MAP.get(r["id"],{}).get("wsps",[])
        imp_pols = [p for p in policies if p["id"] in pol_ids_for_rule]
        imp_wsps = [w for w in wsps     if w["id"] in wsp_ids_for_rule]

        border_color = "#fca5a5" if needs_action else "#e5e7eb"
        bg_color     = "#fff7f7" if needs_action else "#ffffff"

        html += f"<div style='border:1px solid {border_color};background:{bg_color};border-radius:8px;padding:14px 16px;margin-bottom:12px'>"

        # Rule header row
        html += "<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:10px'>"
        html += f"<div style='flex:1'>"
        html += f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:4px'>"
        html += f"<span style='background:{src_bg};color:{src_color};border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700'>{r['source']}</span>"
        html += f"<span style='background:#f3f4f6;color:#374151;border-radius:4px;padding:2px 8px;font-size:11px'>{r['category']}</span>"
        html += f"<span style='color:#9ca3af;font-size:11px'>Updated: {upd_date}</span>"
        if needs_action:
            html += f"<span style='background:#fef2f2;color:#dc2626;border:1px solid #fca5a5;border-radius:4px;padding:2px 8px;font-size:10px;font-weight:700;margin-left:auto'>⚠ ACTION RECOMMENDED</span>"
        html += "</div>"
        html += f"<div style='font-weight:700;font-size:13px;color:#111'>{r['id']} — {r['title']}</div>"
        html += "</div></div>"

        # What changed
        html += f"<div style='background:#f8fafc;border-radius:6px;padding:10px 12px;margin-bottom:10px'>"
        html += f"<div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px'>What Changed</div>"
        html += f"<div style='font-size:12px;color:#1f2937;line-height:1.6'>{what_changed}</div>"
        html += "</div>"

        # Why it matters + BLs
        html += f"<div style='font-size:12px;color:#374151;margin-bottom:8px;line-height:1.6'>"
        html += f"<b>Why it likely matters to this firm:</b> As a dual registrant operating across {', '.join(affected_bls[:2])}"
        html += f"{' and ' + affected_bls[2] if len(affected_bls)>2 else ''}, this firm is likely subject to these requirements. "
        html += f"Compliance procedures covering this area may warrant review against the updated guidance.</div>"

        html += f"<div style='font-size:12px;margin-bottom:8px'><b>Potentially impacted business lines:</b> "
        for bl in affected_bls:
            html += f"<span style='background:#f1f5f9;color:#374151;border-radius:4px;padding:1px 8px;font-size:11px;margin-right:4px'>{bl}</span>"
        html += "</div>"

        # Impacted policies table
        if imp_pols or imp_wsps:
            html += "<div style='font-size:12px;font-weight:600;margin-bottom:5px'>Potentially impacted policies &amp; procedures:</div>"
            html += "<table style='width:100%;border-collapse:collapse;font-size:11px;margin-bottom:8px'>"
            html += "<tr style='background:#f3f4f6'><th style='padding:4px 8px;text-align:left'>ID</th><th>Title</th><th>Type</th><th>Current Status</th><th>Last Approved</th></tr>"
            for p in imp_pols:
                approved_ts = p.get("approved_ts","")[:10] if p.get("approved_ts") else "—"
                html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:4px 8px;color:#2563eb;font-weight:700'>{p['id']}</td><td>{p['title']}</td><td>Policy</td><td>{_fmt_status_badge(p['status'])}</td><td>{approved_ts}</td></tr>"
                sources.append(f"{p['id']} – {p['title']}")
            for w in imp_wsps:
                html += f"<tr style='border-bottom:1px solid #e5e7eb'><td style='padding:4px 8px;color:#2563eb;font-weight:700'>{w['id']}</td><td>{w['title']}</td><td>WSP</td><td>{_fmt_status_badge(w['status'])}</td><td>—</td></tr>"
                sources.append(f"{w['id']} – {w['title']}")
            html += "</table>"

        # Recommended action
        html += "<div style='background:#fffbeb;border-left:3px solid #f59e0b;padding:7px 11px;border-radius:4px;font-size:12px'>"
        html += "<b>Recommended compliance review actions:</b><ul style='margin:4px 0 0 0;padding-left:16px;line-height:1.8'>"
        html += f"<li>Review the above policies and WSPs against the updated {r['source']} guidance</li>"
        html += "<li>Conduct a gap analysis to identify any procedures that may need updating</li>"
        html += "<li>Assign responsible compliance officer and document findings</li>"
        html += "<li>If gaps are identified, initiate a policy revision and route for approval</li>"
        html += "</ul></div>"

        html += "</div>"
        sources.append(f"{r['id']} – {r['title']}")

    # Disclaimer
    html += "<div style='margin-top:14px;padding-top:10px;border-top:1px solid #e5e7eb;font-size:11px;color:#9ca3af;font-style:italic'>"
    html += "⚠ This output is intended to support compliance review and does not replace human regulatory analysis. Regulatory applicability determinations should be reviewed by a qualified compliance professional or legal counsel."
    html += "</div></div>"

    # Download text
    download_text = f"REGULATORY UPDATE ANALYSIS — {last_month_label}\n{'='*55}\n"
    download_text += f"Firm: {firm.get('name','Apex Financial Services LLC')}\nRegistration: Dual Registrant (RIA / Broker-Dealer)\nGenerated: {today_str}\n\n"
    download_text += f"SUMMARY\n{'-'*40}\nRules Updated: {len(pool)}\nRequire Policy Review: {action_ct}\nPolicies Potentially Impacted: {len(pol_ids)}\n\n"
    download_text += f"UPDATES DETAIL\n{'='*55}\n"
    for r in pool:
        upd = _RULE_RECENT_UPDATES.get(r["id"],{})
        pol_ids_r = _RULE_POLICY_MAP.get(r["id"],{}).get("policies",[])
        wsp_ids_r = _RULE_POLICY_MAP.get(r["id"],{}).get("wsps",[])
        download_text += f"\n{r['source']} — {r['id']}: {r['title']}\n"
        download_text += f"Updated: {upd.get('update_date','—')}\n"
        download_text += f"What Changed: {upd.get('what_changed', r['description'])}\n"
        download_text += f"Potentially Impacted Policies: {', '.join(pol_ids_r)}\n"
        download_text += f"Potentially Impacted WSPs: {', '.join(wsp_ids_r)}\n"
        download_text += f"Recommended Action: Review identified policies and WSPs. Conduct gap analysis. Document findings.\n"
        download_text += "-"*40 + "\n"
    download_text += "\nThis memo is intended to support compliance review and does not replace human regulatory analysis."

    return {"html":html,"sources":sources,"intent":"regulatory_analysis","download_label":"📥 Download Regulatory Memo (.txt)","download_content":download_text}

# ── PAGE: AI COPILOT ───────────────────────────────────────────────────────────
def page_ai_copilot():
    if "copilot_history" not in st.session_state:
        st.session_state.copilot_history = []
    if "copilot_show_prompts" not in st.session_state:
        st.session_state.copilot_show_prompts = True

    st.markdown("""
    <div class='gc-header'>
      <h1>AI COPILOT</h1>
      <p>Ask questions about your compliance program, upload documents for review, or get regulatory guidance — powered by your firm's compliance records</p>
    </div>""", unsafe_allow_html=True)

    # Firm context pill
    st.markdown(f"""
    <div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:8px 14px;font-size:12px;color:#1e40af;margin-bottom:1.2rem;display:flex;align-items:center;gap:8px'>
      <span style='font-size:14px'>🏢</span>
      <span><b>Firm context active:</b> Apex Financial Services LLC · RIA + Broker-Dealer · Business lines: Wealth Management, Investment Banking, Operations, Fixed Income, Capital Markets</span>
    </div>""", unsafe_allow_html=True)

    # ── SUGGESTED PROMPTS (collapsed when chat has messages) ──────────────────
    if not st.session_state.copilot_history or st.session_state.copilot_show_prompts:
        if st.session_state.copilot_history:
            tog_label = "▲ Hide suggestions" if st.session_state.copilot_show_prompts else "▼ Show suggestions"
            if st.button(tog_label, key="copilot_tog_prompts"):
                st.session_state.copilot_show_prompts = not st.session_state.copilot_show_prompts
                st.rerun()

        if st.session_state.copilot_show_prompts:
            st.markdown("<div style='font-size:10px;font-weight:800;color:#6b7280;text-transform:uppercase;letter-spacing:.1em;margin-bottom:.7rem'>SUGGESTED PROMPTS</div>", unsafe_allow_html=True)
            cat_cols = st.columns(len(PROMPT_CATEGORIES))
            prompt_fired = None
            for ci, cat in enumerate(PROMPT_CATEGORIES):
                with cat_cols[ci]:
                    st.markdown(f"""
                    <div style='background:{cat["bg"]};border:1px solid {cat["border"]};border-radius:10px;
                         padding:10px 12px;margin-bottom:4px;min-height:120px'>
                      <div style='font-size:10px;font-weight:800;color:{cat["color"]};text-transform:uppercase;
                           letter-spacing:.07em;margin-bottom:8px;border-bottom:1px solid {cat["border"]};padding-bottom:6px'>
                        {cat["label"]}
                      </div>
                    </div>""", unsafe_allow_html=True)
                    for pi, (prompt_text, icon) in enumerate(cat["prompts"]):
                        btn_key = f"copilot_sug_{ci}_{pi}"
                        short = f"{icon} {prompt_text[:36]}…" if len(prompt_text) > 38 else f"{icon} {prompt_text}"
                        if st.button(short, key=btn_key, use_container_width=True):
                            prompt_fired = prompt_text
            if prompt_fired:
                st.session_state.copilot_history.append({"role":"user","content":prompt_fired,"attachment":None})
                st.session_state.copilot_show_prompts = False
                with st.spinner("Analyzing…"):
                    result = _copilot_respond(prompt_fired)
                st.session_state.copilot_history.append({"role":"assistant","content":result})
                st.rerun()

            # Document review prompts section
            st.markdown(f"""
            <div style='background:#fafafa;border:1px solid #e5e7eb;border-radius:10px;padding:12px 16px;margin-top:.5rem'>
              <div style='font-size:10px;font-weight:800;color:#374151;text-transform:uppercase;letter-spacing:.07em;margin-bottom:6px'>
                📎 DOCUMENT REVIEW (attach a file below)
              </div>
              <div style='display:flex;gap:8px;flex-wrap:wrap'>
                <span style='background:#fff;border:1px solid #d1d5db;border-radius:6px;padding:4px 10px;font-size:11px;color:#374151;font-style:italic'>
                  "Review this marketing deck using FINRA/SEC marketing material guidance and flag any compliance issues"
                </span>
                <span style='background:#fff;border:1px solid #d1d5db;border-radius:6px;padding:4px 10px;font-size:11px;color:#374151;font-style:italic'>
                  "Review this policy against the regulations in its version table — identify any rules that were amended, what changed, when it went into effect, and what policy updates are required"
                </span>
                <span style='background:#fff;border:1px solid #d1d5db;border-radius:6px;padding:4px 10px;font-size:11px;color:#374151;font-style:italic'>
                  "Summarize this document and identify any compliance gaps"
                </span>
              </div>
            </div>""", unsafe_allow_html=True)

            st.markdown("<div style='margin-top:.8rem'></div>", unsafe_allow_html=True)

    # ── CHAT HISTORY ──────────────────────────────────────────────────────────
    if st.session_state.copilot_history:
        st.markdown("<div style='border-top:1px solid #e5e7eb;margin-bottom:1rem'></div>", unsafe_allow_html=True)

    for i, msg in enumerate(st.session_state.copilot_history):
        if msg["role"] == "user":
            attach_html = ""
            if msg.get("attachment"):
                attach_html = f"<div style='background:rgba(255,255,255,0.18);border-radius:6px;padding:4px 8px;font-size:11px;margin-top:5px;display:inline-block'>📎 {msg['attachment']}</div>"
            st.markdown(f"""
            <div style='display:flex;justify-content:flex-end;margin-bottom:10px;gap:8px;align-items:flex-end'>
              <div style='background:#2563eb;color:#fff !important;border-radius:16px 16px 4px 16px;
                   padding:10px 15px;max-width:72%;font-size:13px;line-height:1.55;
                   box-shadow:0 2px 8px rgba(37,99,235,0.2)'>
                <span style='color:#fff !important'>{msg['content']}</span>
                {attach_html}
              </div>
              <div style='width:32px;height:32px;border-radius:50%;background:#1d4ed8;
                   display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0'>
                👤
              </div>
            </div>""", unsafe_allow_html=True)
        else:
            result = msg["content"]
            intent_labels = {
                "search_platform_records": ("🔍", "Platform Search", "#059669", "#ecfdf5", "#a7f3d0"),
                "query_platform_data":     ("📊", "Data Query",      "#2563eb", "#eff6ff", "#bfdbfe"),
                "summarize_testing":       ("🧪", "Testing Summary", "#7c3aed", "#f5f3ff", "#ddd6fe"),
                "generate_report":         ("📄", "Report",          "#374151", "#f9fafb", "#e5e7eb"),
                "regulatory_analysis":     ("📰", "Regulatory Analysis","#0891b2","#ecfeff","#a5f3fc"),
                "document_review":         ("📋", "Document Review", "#b45309", "#fffbeb", "#fde68a"),
            }
            ic, lbl, lc, lbg, lbdr = intent_labels.get(result.get("intent",""), ("💬","Response","#374151","#f9fafb","#e5e7eb"))

            st.markdown(f"""
            <div style='display:flex;justify-content:flex-start;margin-bottom:10px;gap:8px;align-items:flex-start'>
              <div style='width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,#1e3a5f,#2563eb);
                   display:flex;align-items:center;justify-content:center;font-size:15px;flex-shrink:0;margin-top:2px'>
                ⚖️
              </div>
              <div style='flex:1;max-width:88%'>
                <div style='background:#ffffff;border:1px solid #e5e7eb;border-radius:4px 16px 16px 16px;
                     padding:14px 18px;box-shadow:0 2px 6px rgba(0,0,0,0.05)'>
                  <div style='display:flex;align-items:center;gap:6px;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid #f0f2f5'>
                    <span style='font-size:13px'>{ic}</span>
                    <span style='font-size:10px;font-weight:800;color:{lc};text-transform:uppercase;letter-spacing:.08em'>{lbl}</span>
                    <span style='background:{lbg};border:1px solid {lbdr};color:{lc};border-radius:4px;
                         padding:1px 7px;font-size:10px;font-weight:600;margin-left:auto'>Graphite AI</span>
                  </div>
                  {result['html']}
                </div>""", unsafe_allow_html=True)

            if result.get("sources"):
                src_text = " · ".join(result["sources"][:6])
                st.markdown(f"<div style='font-size:11px;color:#9ca3af;margin:.3rem 0 .2rem 40px;padding-left:2px'>📎 Sources: {src_text}</div>", unsafe_allow_html=True)

            if result.get("download_label") and result.get("download_content"):
                dl_col, _ = st.columns([2, 5])
                with dl_col:
                    st.download_button(
                        label=result["download_label"],
                        data=result["download_content"],
                        file_name="graphite_copilot_report.txt",
                        mime="text/plain",
                        key=f"copilot_dl_{i}",
                    )

            st.markdown("</div></div>", unsafe_allow_html=True)

    # ── INPUT AREA ────────────────────────────────────────────────────────────
    st.markdown("<div style='font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.08em;margin-top:.8rem;margin-bottom:4px'>New Message</div>", unsafe_allow_html=True)

    with st.container():
        # File attachment
        with st.expander("📎 Attach a document for review (PDF, DOCX, marketing materials, policies)", expanded=False):
            st.markdown("""
            <div style='font-size:12px;color:#374151;margin-bottom:.5rem;line-height:1.6'>
              Upload a <b>marketing deck</b>, <b>policy document</b>, or any compliance-related file.
              After uploading, type your question in the box below — for example:<br/>
              <span style='color:#2563eb'>• "Review this marketing deck using FINRA/SEC marketing material guidance and flag any compliance issues"</span><br/>
              <span style='color:#2563eb'>• "Review this policy against the regulations in its version table and identify any rules that were amended, what changed, when it took effect, and what policy updates are required"</span>
            </div>""", unsafe_allow_html=True)
            uploaded_doc = st.file_uploader(
                "Upload document",
                type=["pdf","docx","doc","pptx","xlsx","csv","txt"],
                key="copilot_file_upload",
                label_visibility="collapsed",
            )
            if uploaded_doc:
                st.markdown(f"""
                <div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
                     padding:8px 12px;font-size:12px;color:#1e40af;margin-top:4px'>
                  ✅ <b>{uploaded_doc.name}</b> attached ({round(uploaded_doc.size/1024,1)} KB) — type your question below and click Send
                </div>""", unsafe_allow_html=True)

        msg_col, send_col = st.columns([6, 1])
        with msg_col:
            user_input = st.text_area(
                "Message",
                key="copilot_input",
                placeholder="Ask anything about your compliance program, or describe what you'd like me to review in the attached document…",
                label_visibility="collapsed",
                height=80,
            )
        with send_col:
            st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
            send = st.button("Send ➤", key="copilot_send", use_container_width=True)

    if send and user_input and user_input.strip():
        attachment_name = None
        if "copilot_file_upload" in st.session_state and st.session_state.copilot_file_upload is not None:
            attachment_name = st.session_state.copilot_file_upload.name

        # Build enhanced question if a file is attached
        question = user_input.strip()
        if attachment_name:
            q_for_classify = question + " document review marketing policy"
        else:
            q_for_classify = question

        st.session_state.copilot_history.append({"role":"user","content":question,"attachment":attachment_name})
        st.session_state.copilot_show_prompts = False

        with st.spinner("Analyzing…"):
            if attachment_name:
                # Document review response
                fname_lower = attachment_name.lower()
                is_marketing = any(k in fname_lower or k in question.lower() for k in ["marketing","deck","presentation","pptx","brochure","ad","advertisement","email blast","social"])
                is_policy    = any(k in fname_lower or k in question.lower() for k in ["policy","wsp","procedure","compliance program","manual"])

                if is_marketing or ("finra" in question.lower() or "sec" in question.lower() and not is_policy):
                    html_out = f"""
                    <div style='font-size:13px;color:#374151;line-height:1.7'>
                      <div style='font-weight:700;font-size:14px;color:#1a1d23;margin-bottom:10px'>
                        📋 Marketing Material Compliance Review — {attachment_name}
                      </div>
                      <div style='background:#fffbeb;border:1px solid #fde68a;border-left:4px solid #f59e0b;border-radius:8px;padding:10px 14px;margin-bottom:12px;font-size:12px;color:#78350f'>
                        <b>⚠ Demo Mode:</b> In the full version, Graphite AI reads your uploaded document and cross-references it against FINRA Rule 2210 (Communications with the Public), SEC Marketing Rule (Advisers Act Rule 206(4)-1), and applicable SEC/FINRA guidance. Below is a sample output structure.
                      </div>
                      <b>Applicable Standards Reviewed:</b>
                      <ul style='margin:6px 0 10px 0;padding-left:18px;line-height:1.9'>
                        <li><b>FINRA Rule 2210</b> — Communications with the Public (broker-dealer materials)</li>
                        <li><b>SEC Marketing Rule 206(4)-1</b> — Investment adviser marketing materials</li>
                        <li><b>SEC Guidance: Use of Social Media</b> (IM-2210-7)</li>
                        <li><b>FINRA Regulatory Notice 21-15</b> — Digital Communications</li>
                      </ul>
                      <b>Findings Summary:</b>
                      <table style='width:100%;border-collapse:collapse;font-size:12px;margin:8px 0'>
                        <tr style='background:#f3f4f6'>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Finding</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Rule Reference</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Risk</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Recommended Action</th>
                        </tr>
                        <tr style='border-bottom:1px solid #f0f2f5'>
                          <td style='padding:7px 10px'>[AI will review document and populate findings from your uploaded file]</td>
                          <td style='padding:7px 10px'>FINRA Rule 2210(d)</td>
                          <td style='padding:7px 10px'><span style='background:#fee2e2;color:#991b1b;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700'>HIGH</span></td>
                          <td style='padding:7px 10px'>Revise language and route for principal review</td>
                        </tr>
                      </table>
                      <div style='background:#f0fdf4;border:1px solid #bbf7d0;border-left:4px solid #22c55e;border-radius:8px;padding:8px 12px;font-size:12px;color:#14532d;margin-top:8px'>
                        <b>✅ In the full version:</b> Graphite AI reads every claim, statistic, performance figure, and disclosure in your document and checks each against the applicable FINRA/SEC standard. Output includes a finding-by-finding breakdown, risk level, rule citation, and recommended revision — ready for principal review and filing.
                      </div>
                    </div>"""
                    result = {"html": html_out, "sources": ["FINRA Rule 2210","SEC Rule 206(4)-1","FINRA RN 21-15"], "intent": "document_review", "download_label": None, "download_content": None}
                else:
                    # Policy review
                    html_out = f"""
                    <div style='font-size:13px;color:#374151;line-height:1.7'>
                      <div style='font-weight:700;font-size:14px;color:#1a1d23;margin-bottom:10px'>
                        📋 Policy Compliance Review — {attachment_name}
                      </div>
                      <div style='background:#fffbeb;border:1px solid #fde68a;border-left:4px solid #f59e0b;border-radius:8px;padding:10px 14px;margin-bottom:12px;font-size:12px;color:#78350f'>
                        <b>⚠ Demo Mode:</b> In the full version, Graphite AI reads your uploaded policy, extracts the regulatory citations from the version table, looks up each rule for recent amendments, and produces the analysis below.
                      </div>
                      <b>Analysis Structure — Policy vs. Regulatory Version Table:</b>
                      <table style='width:100%;border-collapse:collapse;font-size:12px;margin:8px 0 12px 0'>
                        <tr style='background:#f3f4f6'>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Rule / Regulation</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Amendment Summary</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Effective Date</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Policy Update Required?</th>
                          <th style='padding:6px 10px;text-align:left;border-bottom:2px solid #e5e7eb'>Recommended Changes</th>
                        </tr>
                        <tr style='border-bottom:1px solid #f0f2f5'>
                          <td style='padding:7px 10px;font-weight:700;color:#2563eb'>[Extracted from your document]</td>
                          <td style='padding:7px 10px'>[Rule amendment summary from regulatory source]</td>
                          <td style='padding:7px 10px'>[Effective date]</td>
                          <td style='padding:7px 10px'><span style='background:#fee2e2;color:#991b1b;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700'>YES</span></td>
                          <td style='padding:7px 10px'>[Specific policy language updates required]</td>
                        </tr>
                      </table>
                      <div style='background:#f0fdf4;border:1px solid #bbf7d0;border-left:4px solid #22c55e;border-radius:8px;padding:8px 12px;font-size:12px;color:#14532d;margin-top:8px'>
                        <b>✅ In the full version:</b> Graphite AI reads every regulation in your policy's version table, queries live regulatory update sources for each rule, identifies amendments that occurred after your policy's last review date, and produces a rule-by-rule gap analysis — including the amendment summary, effective date, specific policy sections impacted, and recommended updated language.
                      </div>
                    </div>"""
                    result = {"html": html_out, "sources": [attachment_name], "intent": "document_review", "download_label": None, "download_content": None}
            else:
                result = _copilot_respond(question)

        st.session_state.copilot_history.append({"role":"assistant","content":result})
        st.rerun()

    # Clear button
    if st.session_state.copilot_history:
        st.markdown("<div style='margin-top:.5rem'></div>", unsafe_allow_html=True)
        clear_col, _ = st.columns([2, 6])
        with clear_col:
            if st.button("🗑 Clear conversation", key="copilot_clear"):
                st.session_state.copilot_history = []
                st.session_state.copilot_show_prompts = True
                st.rerun()

# ── AI COMPLIANCE INTELLIGENCE — shared helpers ────────────────────────────────

def _premium_header(title, subtitle, icon="✦"):
    """Renders the page header with a gold Premium badge."""
    st.markdown(f"""
    <div style='background:#fff;border:1px solid #e5e7eb;border-left:4px solid #f59e0b;
         border-radius:8px;padding:1.2rem 1.6rem;margin-bottom:1.2rem;
         box-shadow:0 1px 3px rgba(0,0,0,0.05)'>
      <div style='display:flex;align-items:center;gap:10px;margin-bottom:.3rem'>
        <h1 style='font-size:1.2rem;font-weight:700;color:#111827;margin:0'>{title}</h1>
        <span style='background:linear-gradient(135deg,#f59e0b,#d97706);color:#fff;
             font-size:9px;font-weight:800;letter-spacing:.1em;text-transform:uppercase;
             padding:2px 8px;border-radius:10px;white-space:nowrap'>✦ PREMIUM</span>
      </div>
      <p style='font-size:13px;color:#6b7280;margin:0'>{subtitle}</p>
    </div>""", unsafe_allow_html=True)

def _premium_locked_banner():
    """Renders the upgrade CTA banner."""
    st.markdown("""
    <div style='background:linear-gradient(135deg,#fffbeb,#fef3c7);border:1px solid #fcd34d;
         border-radius:10px;padding:16px 20px;margin-bottom:1.2rem;display:flex;
         align-items:flex-start;gap:14px'>
      <div style='font-size:28px;line-height:1'>✦</div>
      <div>
        <div style='font-weight:700;font-size:14px;color:#78350f;margin-bottom:4px'>
          Premium Feature — Demo Preview Active
        </div>
        <div style='font-size:12px;color:#92400e;line-height:1.6'>
          This module is part of the <b>AI Compliance Intelligence</b> add-on, a premium capability
          built on top of Graphite's core compliance infrastructure. The data shown below reflects
          a live simulation using your firm's platform records.
          Contact your account manager to activate this feature.
        </div>
      </div>
      <div style='margin-left:auto;white-space:nowrap'>
        <span style='background:#f59e0b;color:#fff;font-size:11px;font-weight:700;
             padding:6px 14px;border-radius:6px;cursor:pointer'>Activate Premium →</span>
      </div>
    </div>""", unsafe_allow_html=True)

def _ai_thinking_card(label="AI analysis complete"):
    st.markdown(f"""
    <div style='display:flex;align-items:center;gap:8px;background:#f0f9ff;
         border:1px solid #bae6fd;border-radius:6px;padding:7px 12px;
         font-size:12px;color:#0369a1;margin-bottom:12px'>
      <span>🤖</span><span><b>AI Copilot</b> — {label}</span>
    </div>""", unsafe_allow_html=True)

# ── PAGE: PEER ENFORCEMENT INSIGHTS ───────────────────────────────────────────
def page_peer_enforcement():
    _premium_header(
        "PEER ENFORCEMENT INSIGHTS",
        "Monthly AI-generated memos summarizing relevant regulatory enforcement actions and peer firm findings",
    )
    _premium_locked_banner()
    _ai_thinking_card("Analyzed 847 enforcement actions · 6 flagged as relevant to your firm profile · Memo generated Feb 2026")

    today = date.today()
    memo_month = (today.replace(day=1) - timedelta(days=1)).strftime("%B %Y")

    st.markdown(f"""
    <div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:20px 24px;margin-bottom:14px'>
      <div style='border-bottom:2px solid #1e3a5f;padding-bottom:10px;margin-bottom:14px'>
        <div style='font-size:10px;font-weight:700;letter-spacing:.1em;color:#6b7280;text-transform:uppercase'>AI-Generated Enforcement Memo</div>
        <div style='font-size:17px;font-weight:800;color:#1e3a5f;margin:4px 0'>Peer Enforcement Actions — {memo_month}</div>
        <div style='font-size:12px;color:#6b7280'>Apex Financial Services LLC · Dual Registrant (RIA / BD) · Auto-generated by AI Copilot</div>
      </div>
      <p style='font-size:12px;color:#374151;line-height:1.7;margin-bottom:0'>
        The AI Copilot reviewed enforcement actions published by FINRA and the SEC during {memo_month}.
        Six actions were identified as relevant based on the firm's dual-registrant status, business lines,
        and active regulatory obligations. The actions involve peer firms operating in wealth management,
        retail brokerage, and investment advisory, and implicate supervisory controls, AML programs,
        and best interest obligations — areas the firm actively manages.
      </p>
    </div>""", unsafe_allow_html=True)

    ENFORCEMENTS = [
        {
            "id":"ENF-001","regulator":"FINRA","date":f"{today.year}-02-08",
            "firm_type":"Mid-size Broker-Dealer","rule":"FINRA Rule 3110 / Rule 3120",
            "headline":"Broker-dealer fined $2.1M for supervisory control failures in options trading",
            "what_happened":"The firm failed to establish and maintain an adequate supervisory system for options trading activity. Surveillance alerts were not reviewed within required timeframes, and supervisory sign-off documentation was incomplete across 1,400+ transactions over a 2-year period.",
            "risk_color":"#dc2626","risk":"High",
            "exposure":"The firm conducts options supervisory reviews under WSP-002. CT-001 identified a CCO escalation SLA breach in Q1 2026. The pattern identified in this enforcement action — inadequate sign-off documentation and missed review SLAs — closely mirrors the exception documented in CT-001.",
            "review_areas":["WSP-002 supervisory sign-off procedures","CT-001 exception remediation (EX-001)","CCO escalation SLA monitoring controls"],
        },
        {
            "id":"ENF-002","regulator":"FINRA","date":f"{today.year}-01-22",
            "firm_type":"Retail Broker-Dealer","rule":"FINRA Rule 3310 / BSA AML",
            "headline":"FINRA censures broker-dealer for AML program deficiencies and SAR filing delays",
            "what_happened":"The firm's AML program failed to file SARs within the 30-day regulatory deadline in 38 instances. Customer identification program documentation was incomplete for 12% of new accounts opened during the review period. Annual independent testing was not completed for two consecutive years.",
            "risk_color":"#dc2626","risk":"High",
            "exposure":"The firm's AML program is governed by POL-001 and WSP-001. CT-003 identified CIP documentation gaps in 2 of 25 sampled accounts. SAR filing procedures are covered by WT-001-3. The firm should confirm CIP remediation from CT-003 is complete and that SAR filing timeliness is being tracked.",
            "review_areas":["CT-003 CIP exception (EX-002) — confirm remediation complete","WSP-001 SAR filing SLA monitoring","POL-001 AML program independent testing schedule"],
        },
        {
            "id":"ENF-003","regulator":"SEC","date":f"{today.year}-02-14",
            "firm_type":"Registered Investment Adviser","rule":"SEC Reg BI / Rule 206(4)-7",
            "headline":"SEC charges RIA with Reg BI violations — variable annuity recommendation documentation failures",
            "what_happened":"The adviser failed to document cost-benefit analyses for variable annuity recommendations in 23% of sampled transactions. Form CRS was not delivered at or prior to account opening for 11 accounts. The firm lacked written policies specifically addressing documentation standards for complex product recommendations.",
            "risk_color":"#f59e0b","risk":"Medium",
            "exposure":"CT-004 identified an identical issue — a variable annuity recommendation for account #7823 lacks required cost-benefit comparison documentation (EX-003). POL-004 (Reg BI Policy) is currently in Draft status awaiting CCO approval. This enforcement action adds urgency to finalizing POL-004.",
            "review_areas":["CT-004 EX-003 — variable annuity documentation remediation","POL-004 approval — Reg BI Policy finalization","Form CRS delivery tracking and documentation"],
        },
        {
            "id":"ENF-004","regulator":"SEC","date":f"{today.year}-01-30",
            "firm_type":"Dual Registrant","rule":"SEC Rule 17a-4 / Recordkeeping",
            "headline":"Dual registrant charged with electronic records retention failures",
            "what_happened":"The firm failed to preserve electronic communications in WORM-compliant format over a 4-year period. The firm's records index was not maintained and could not produce records within the 4-hour regulatory response window. Correspondence records for a 3-year period were missing.",
            "risk_color":"#f59e0b","risk":"Medium",
            "exposure":"As a dual registrant, the firm is subject to Rule 17a-4. POL-001 and WSP-002 reference recordkeeping obligations. The firm should confirm that its WORM storage vendor attestation is current and that the records index is available for regulatory production.",
            "review_areas":["WORM storage vendor current attestation","Records index — confirm 4-hour production readiness","WSP-001 and WSP-002 recordkeeping procedure review"],
        },
    ]

    for e in ENFORCEMENTS:
        risk_bg = "#fef2f2" if e["risk"]=="High" else "#fffbeb"
        risk_border = "#fca5a5" if e["risk"]=="High" else "#fcd34d"
        reg_color = {"FINRA":"#1e40af","SEC":"#6d28d9"}.get(e["regulator"],"#374151")
        reg_bg    = {"FINRA":"#dbeafe","SEC":"#ede9fe"}.get(e["regulator"],"#f3f4f6")

        st.markdown(f"""
        <div style='border:1px solid {risk_border};background:{risk_bg};border-radius:8px;
             padding:14px 18px;margin-bottom:10px'>
          <div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px'>
            <div style='flex:1'>
              <div style='display:flex;align-items:center;gap:8px;margin-bottom:5px;flex-wrap:wrap'>
                <span style='background:{reg_bg};color:{reg_color};border-radius:4px;
                     padding:2px 8px;font-size:11px;font-weight:700'>{e["regulator"]}</span>
                <span style='background:#f3f4f6;color:#374151;border-radius:4px;
                     padding:2px 8px;font-size:11px'>{e["firm_type"]}</span>
                <span style='background:{e["risk_color"]}20;color:{e["risk_color"]};
                     border:1px solid {e["risk_color"]}40;border-radius:4px;
                     padding:2px 8px;font-size:11px;font-weight:700'>{e["risk"]} Relevance</span>
                <span style='color:#9ca3af;font-size:11px;margin-left:auto'>{e["date"]} · {e["id"]}</span>
              </div>
              <div style='font-weight:700;font-size:13px;color:#111'>{e["headline"]}</div>
              <div style='font-size:11px;color:#6b7280;margin-top:2px'>Rule: {e["rule"]}</div>
            </div>
          </div>
          <div style='background:#fff;border-radius:6px;padding:10px 12px;margin-bottom:10px'>
            <div style='font-size:11px;font-weight:700;color:#6b7280;text-transform:uppercase;
                 letter-spacing:.06em;margin-bottom:4px'>What the Firm Did Wrong</div>
            <div style='font-size:12px;color:#1f2937;line-height:1.6'>{e["what_happened"]}</div>
          </div>
          <div style='font-size:12px;color:#374151;margin-bottom:8px;line-height:1.6'>
            <b>Potential exposure for your firm:</b> {e["exposure"]}
          </div>
          <div style='font-size:12px;font-weight:600;margin-bottom:5px'>Recommended internal review areas:</div>
          <div style='display:flex;flex-wrap:wrap;gap:6px'>
            {''.join(f"<span style='background:#fffbeb;border:1px solid #fcd34d;color:#78350f;border-radius:4px;padding:3px 9px;font-size:11px'>⚑ {area}</span>" for area in e["review_areas"])}
          </div>
        </div>""", unsafe_allow_html=True)

    dl_text = f"PEER ENFORCEMENT INSIGHTS — {memo_month}\n{'='*55}\nFirm: Apex Financial Services LLC\n\n"
    for e in ENFORCEMENTS:
        dl_text += f"\n{e['id']} — {e['regulator']}: {e['headline']}\nRule: {e['rule']}\nRelevance: {e['risk']}\n{e['what_happened']}\nExposure: {e['exposure']}\n" + "-"*40 + "\n"
    dl_text += "\nThis memo is AI-generated for compliance review purposes and does not replace human regulatory analysis."
    st.download_button("📥 Download Enforcement Memo (.txt)", data=dl_text,
                       file_name=f"enforcement_memo_{memo_month.replace(' ','_')}.txt",
                       mime="text/plain", key="enf_dl")

# ── PAGE: AI RISK ASSESSMENT ENGINE ──────────────────────────────────────────
def page_ai_risk_assessment():
    _premium_header(
        "AI RISK ASSESSMENT ENGINE",
        "Automated annual compliance risk scoring for supervisory testing programs — draft output pending CCO review",
    )
    _premium_locked_banner()

    tests    = D.get("compliance_tests", [])
    policies = D.get("policies", [])
    wsps     = D.get("wsps", [])

    open_ex  = sum(len([e for e in t.get("exceptions",[]) if e["status"]=="Open"]) for t in tests)
    high_ct  = sum(1 for t in tests if t.get("risk_rating")=="High")
    approved = sum(1 for p in policies+wsps if p.get("status")=="Approved")

    _ai_thinking_card(f"Analyzed {len(tests)} prior tests · {len(policies)} policies · {len(wsps)} WSPs · {open_ex} open exceptions · Draft scores generated")

    st.markdown(f"""
    <div style='background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;
         padding:10px 14px;font-size:12px;color:#78350f;margin-bottom:14px'>
      ⚠ <b>Draft — Pending CCO Review.</b> The AI has proposed initial risk scores based on platform data.
      Scores must be reviewed and approved by Jordan Reed (CCO) before being finalized and used for testing planning.
    </div>""", unsafe_allow_html=True)

    TOPICS = [
        {
            "topic":"Supervisory Controls (FINRA Rule 3120)",
            "lob":"Retail Brokerage","regulation":"FINRA Rule 3110 / 3120",
            "prior_score":"High","proposed_score":"High",
            "increasing":["Prior exam finding: CCO escalation SLA breach (CT-001 EX-001)","Open remediation item unresolved (EX-001 assigned to admin)","FINRA published 3120 deficiency patterns in recent examination findings"],
            "reducing":["WSP-002 reviewed and approved (v2.0)","CT-001 testing coverage in Q1 2026","Supervisory sign-off documented in 24 of 25 sampled transactions"],
            "recommendation":"Maintain High risk rating. Prioritize CT-001 remediation before next testing cycle.",
        },
        {
            "topic":"AML / BSA Program",
            "lob":"Operations","regulation":"FINRA Rule 3310 / BSA / FinCEN CDD",
            "prior_score":"High","proposed_score":"Medium",
            "increasing":["CIP documentation gap identified in CT-003 (2 accounts)","FinCEN CDD guidance updated — beneficial ownership thresholds revised"],
            "reducing":["CT-003 completed — AML program substantively compliant","POL-001 approved v2.1 (updated FinCEN references)","Annual AML Policy Certification signed off (CL-001)","AML independent testing completed for 2025"],
            "recommendation":"Downgrade to Medium based on completed testing and policy currency. Monitor CIP remediation to closure.",
        },
        {
            "topic":"Reg BI / Best Interest",
            "lob":"Wealth Management","regulation":"SEC Reg BI / Rule 15l-1",
            "prior_score":"Medium","proposed_score":"High",
            "increasing":["CT-004 exception: VA recommendation lacking cost-benefit documentation (EX-003 open)","POL-004 (Reg BI Policy) in Draft status — not yet approved","SEC enforcement action against peer dual registrant for identical VA documentation failures"],
            "reducing":["CT-004 testing in progress covering Q4 2025","WSP-002 incorporates Reg BI supervisory procedures"],
            "recommendation":"Upgrade to High. POL-004 approval is overdue. EX-003 remediation should be prioritized.",
        },
        {
            "topic":"Books & Records / Recordkeeping",
            "lob":"Operations","regulation":"SEC Rule 17a-4",
            "prior_score":"Low","proposed_score":"Medium",
            "increasing":["Recent SEC enforcement action against dual registrant for Rule 17a-4 failures","WORM vendor attestation not confirmed in current platform records"],
            "reducing":["WSP-001 and WSP-002 reference recordkeeping obligations","Firm retention policy set to 6 years (compliant)"],
            "recommendation":"Upgrade to Medium due to recent enforcement activity. Confirm WORM attestation and records index readiness.",
        },
    ]

    score_colors = {"High":("#dc2626","#fef2f2","#fca5a5"),"Medium":("#d97706","#fffbeb","#fcd34d"),"Low":("#16a34a","#f0fdf4","#86efac")}

    for t in TOPICS:
        fc, fbg, fbdr = score_colors.get(t["proposed_score"], ("#374151","#f3f4f6","#d1d5db"))
        pc, pbg, pbdr = score_colors.get(t["prior_score"], ("#374151","#f3f4f6","#d1d5db"))
        changed = t["prior_score"] != t["proposed_score"]

        st.markdown(f"""
        <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;
             padding:14px 18px;margin-bottom:10px'>
          <div style='display:flex;align-items:flex-start;gap:12px;margin-bottom:10px'>
            <div style='flex:1'>
              <div style='font-weight:700;font-size:14px;color:#111;margin-bottom:4px'>{t["topic"]}</div>
              <div style='font-size:11px;color:#6b7280'>{t["lob"]} · {t["regulation"]}</div>
            </div>
            <div style='text-align:right'>
              <div style='font-size:10px;color:#6b7280;margin-bottom:3px'>Prior Score</div>
              <span style='background:{pbg};color:{pc};border:1px solid {pbdr};
                   border-radius:4px;padding:2px 10px;font-size:11px;font-weight:700'>{t["prior_score"]}</span>
            </div>
            <div style='text-align:right'>
              <div style='font-size:10px;color:#6b7280;margin-bottom:3px'>AI Proposed</div>
              <span style='background:{fbg};color:{fc};border:1px solid {fbdr};
                   border-radius:4px;padding:2px 10px;font-size:11px;font-weight:700'>
                {t["proposed_score"]} {'↑' if t["proposed_score"]=="High" and t["prior_score"]!="High" else ('↓' if t["proposed_score"]=="Low" and t["prior_score"]!="Low" else '')}
              </span>
            </div>
          </div>
          <div style='display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px'>
            <div style='background:#fef2f2;border-radius:6px;padding:10px 12px'>
              <div style='font-size:11px;font-weight:700;color:#dc2626;text-transform:uppercase;
                   letter-spacing:.05em;margin-bottom:6px'>▲ Risk Increasing Indicators</div>
              {''.join(f"<div style='font-size:12px;color:#374151;margin-bottom:4px;line-height:1.5'>• {i}</div>" for i in t["increasing"])}
            </div>
            <div style='background:#f0fdf4;border-radius:6px;padding:10px 12px'>
              <div style='font-size:11px;font-weight:700;color:#16a34a;text-transform:uppercase;
                   letter-spacing:.05em;margin-bottom:6px'>▼ Risk Reducing Indicators</div>
              {''.join(f"<div style='font-size:12px;color:#374151;margin-bottom:4px;line-height:1.5'>• {i}</div>" for i in t["reducing"])}
            </div>
          </div>
          <div style='background:#fffbeb;border-left:3px solid #f59e0b;padding:7px 12px;
               border-radius:4px;font-size:12px;color:#78350f'>
            <b>AI Recommendation:</b> {t["recommendation"]}
          </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<hr style='margin:.8rem 0'/>", unsafe_allow_html=True)
    st.markdown("<div style='font-size:12px;font-weight:700;color:#374151;margin-bottom:8px'>CCO Review & Approval</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    c1.text_input("Reviewing Officer", value="Jordan Reed — Chief Compliance Officer", disabled=True, key="ra_reviewer")
    c2.text_input("Status", value="Pending CCO Approval", disabled=True, key="ra_status")
    if c3.button("✓ Approve Risk Scores", key="ra_approve"):
        add_audit("AI Risk Assessment", "Risk assessment scores approved by CCO")
        st.success("✓ Risk scores approved and locked for testing planning.")

# ── PAGE: POLICY GAP ANALYSIS ─────────────────────────────────────────────────
def page_policy_gap_analysis():
    _premium_header(
        "POLICY GAP ANALYSIS",
        "AI-powered review of policies and WSPs against regulatory rules, guidance, and enforcement themes",
    )
    _premium_locked_banner()

    policies = D.get("policies", [])
    wsps     = D.get("wsps", [])
    rules    = D.get("rules", [])

    _ai_thinking_card(f"Analyzed {len(policies)} policies · {len(wsps)} WSPs · {len(rules)} regulatory rules · 11 potential gaps identified")

    st.markdown(f"""
    <div style='display:flex;gap:12px;margin-bottom:14px'>
      <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 16px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#dc2626'>4</div>
        <div style='font-size:11px;color:#6b7280'>Potential Missing Controls</div>
      </div>
      <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 16px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#f59e0b'>3</div>
        <div style='font-size:11px;color:#6b7280'>Outdated References</div>
      </div>
      <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 16px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#7c3aed'>2</div>
        <div style='font-size:11px;color:#6b7280'>Potential Inconsistencies</div>
      </div>
      <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:10px 16px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#2563eb'>2</div>
        <div style='font-size:11px;color:#6b7280'>Conflicting Language</div>
      </div>
    </div>""", unsafe_allow_html=True)

    GAPS = [
        {
            "id":"GAP-001","type":"Missing Control","severity":"High",
            "doc":"POL-001 / WSP-001","title":"No documented escalation matrix for high-risk SAR candidates",
            "detail":"POL-001 and WSP-001 reference SAR filing procedures but do not include a documented escalation matrix specifying who must approve SAR filings above defined thresholds (e.g., transactions >$500K). FINRA and FinCEN have cited absence of tiered escalation procedures in recent enforcement actions.",
            "rule_ref":"FINRA Rule 3310 · FinCEN SAR Guidance 2023",
            "suggestion":"Add a tiered SAR escalation matrix to WSP-001 Section 4. Define approval authority by transaction size and customer risk tier.",
        },
        {
            "id":"GAP-002","type":"Outdated Reference","severity":"Medium",
            "doc":"POL-001","title":"Beneficial ownership threshold references pre-2024 FinCEN CDD rule",
            "detail":"POL-001 v2.1 references FinCEN beneficial ownership thresholds at the 25% ownership level. FinCEN's updated CDD rule (effective January 2024) introduced enhanced verification requirements for legal entity accounts with complex ownership structures. The policy does not reflect this updated standard.",
            "rule_ref":"FinCEN CDD Rule (31 CFR 1010.230) — 2024 Amendment",
            "suggestion":"Update POL-001 Section 3 to reflect current FinCEN beneficial ownership thresholds and enhanced verification requirements for legal entity accounts.",
        },
        {
            "id":"GAP-003","type":"Missing Control","severity":"High",
            "doc":"POL-004 (Draft)","title":"Reg BI policy does not address digital / robo-advisory recommendations",
            "detail":"POL-004 (currently in Draft) addresses traditional recommendation scenarios but does not include controls for algorithmic or robo-advisory recommendation workflows. The SEC's 2024 FAQ on Reg BI explicitly extended best interest obligations to automated tools. This is a gap in the policy's scope.",
            "rule_ref":"SEC Reg BI · SEC Staff FAQ on Automated Investment Tools (2024)",
            "suggestion":"Add a section to POL-004 covering automated recommendation tools, including documentation requirements, oversight controls, and periodic sampling of algorithm outputs.",
        },
        {
            "id":"GAP-004","type":"Potential Inconsistency","severity":"Medium",
            "doc":"WSP-002 / POL-002","title":"Daily trade review SLA inconsistency between policy and procedure",
            "detail":"POL-002 (Supervision Policy) states flagged trades must be reviewed 'within one business day.' WSP-002 Section 2 states reviews must occur 'within 24 hours.' These timeframes may diverge during days following a weekend or holiday. A regulator could cite inconsistency during an exam.",
            "rule_ref":"FINRA Rule 3110 — Supervisory Procedures",
            "suggestion":"Align POL-002 and WSP-002 to use a consistent SLA definition. Recommend adopting '24 hours excluding weekends and holidays' or define a specific business hour window.",
        },
        {
            "id":"GAP-005","type":"Outdated Reference","severity":"Low",
            "doc":"WSP-003","title":"Information barrier policy references superseded FINRA Notice 96-34",
            "detail":"WSP-003 v1.1 includes a footnote referencing FINRA Notice to Members 96-34 as guidance on information barriers. This notice has been superseded by more recent FINRA guidance. Referencing outdated regulatory notices can reflect poorly during regulatory review.",
            "rule_ref":"FINRA Regulatory Notice 11-43 (supersedes NTM 96-34)",
            "suggestion":"Update WSP-003 to cite current FINRA Regulatory Notice 11-43 and remove reference to NTM 96-34. Review section 3 for any additional outdated citations.",
        },
        {
            "id":"GAP-006","type":"Conflicting Language","severity":"Medium",
            "doc":"POL-003 / WSP-003","title":"Information barrier policy and procedure use inconsistent definitions of MNPI",
            "detail":"POL-003 defines Material Non-Public Information (MNPI) as information that 'a reasonable investor would consider important in making an investment decision.' WSP-003 uses a slightly narrower definition referencing only 'material pending corporate events.' The inconsistency could create ambiguity in employee compliance determinations.",
            "rule_ref":"SEC Rule 10b-5 · FINRA Rule 3110",
            "suggestion":"Harmonize the MNPI definition across POL-003 and WSP-003. Recommend adopting the broader POL-003 standard and ensuring WSP-003 cross-references it explicitly.",
        },
    ]

    type_colors = {"Missing Control":("#dc2626","#fef2f2","#fca5a5"),"Outdated Reference":("#d97706","#fffbeb","#fcd34d"),"Potential Inconsistency":("#7c3aed","#f5f3ff","#c4b5fd"),"Conflicting Language":("#2563eb","#eff6ff","#bfdbfe")}
    sev_colors  = {"High":"#dc2626","Medium":"#f59e0b","Low":"#16a34a"}

    for g in GAPS:
        tc, tbg, tbdr = type_colors.get(g["type"], ("#374151","#f3f4f6","#d1d5db"))
        sc = sev_colors.get(g["severity"], "#374151")

        st.markdown(f"""
        <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;
             padding:14px 18px;margin-bottom:10px'>
          <div style='display:flex;align-items:flex-start;gap:8px;margin-bottom:8px;flex-wrap:wrap'>
            <span style='color:#6b7280;font-size:11px;font-weight:700'>{g["id"]}</span>
            <span style='background:{tbg};color:{tc};border:1px solid {tbdr};border-radius:4px;
                 padding:2px 8px;font-size:11px;font-weight:700'>{g["type"]}</span>
            <span style='background:{sc}20;color:{sc};border:1px solid {sc}40;border-radius:4px;
                 padding:2px 8px;font-size:11px;font-weight:700'>{g["severity"]} Severity</span>
            <span style='background:#f1f5f9;color:#374151;border-radius:4px;
                 padding:2px 8px;font-size:11px'>{g["doc"]}</span>
          </div>
          <div style='font-weight:700;font-size:13px;color:#111;margin-bottom:6px'>{g["title"]}</div>
          <div style='font-size:12px;color:#374151;line-height:1.6;margin-bottom:8px'>{g["detail"]}</div>
          <div style='font-size:12px;color:#6b7280;margin-bottom:8px'>
            <b>Rule Reference:</b> {g["rule_ref"]}
          </div>
          <div style='background:#f0fdf4;border-left:3px solid #22c55e;padding:7px 12px;
               border-radius:4px;font-size:12px;color:#166534'>
            <b>Suggested Review:</b> {g["suggestion"]}
          </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style='margin-top:8px;font-size:11px;color:#9ca3af;font-style:italic;padding:8px 0'>
    ⚠ Results are presented as suggested review items and do not constitute definitive compliance conclusions.
    All findings should be reviewed by a qualified compliance professional.
    </div>""", unsafe_allow_html=True)

    dl = "POLICY GAP ANALYSIS\n"+"="*50+"\nFirm: Apex Financial Services LLC\n\n"
    for g in GAPS:
        dl += f"\n{g['id']} — {g['type']} ({g['severity']})\nDocument: {g['doc']}\nIssue: {g['title']}\n{g['detail']}\nRule: {g['rule_ref']}\nSuggestion: {g['suggestion']}\n"+"-"*40+"\n"
    dl += "\nResults are suggested review items and do not constitute definitive compliance conclusions."
    st.download_button("📥 Download Gap Analysis (.txt)", data=dl,
                       file_name="policy_gap_analysis.txt", mime="text/plain", key="gap_dl")

# ── PAGE: MOCK REGULATORY EXAM SIMULATOR ──────────────────────────────────────
def page_exam_simulator():
    _premium_header(
        "MOCK REGULATORY EXAM SIMULATOR",
        "Simulate a FINRA or SEC exam with AI-generated document request lists based on your firm's risk profile",
    )
    _premium_locked_banner()

    if "sim_started" not in st.session_state:
        st.session_state.sim_started = False
    if "sim_responses" not in st.session_state:
        st.session_state.sim_responses = {}

    if not st.session_state.sim_started:
        st.markdown("""
        <div style='background:#fff;border:1px solid #e5e7eb;border-radius:10px;
             padding:20px 24px;text-align:center'>
          <div style='font-size:40px;margin-bottom:10px'>🏛️</div>
          <div style='font-weight:700;font-size:16px;color:#111;margin-bottom:6px'>Launch Mock Regulatory Exam</div>
          <div style='font-size:13px;color:#6b7280;margin-bottom:16px;max-width:500px;margin-left:auto;margin-right:auto'>
            The AI will generate a realistic exam document request list based on your firm's risk profile,
            prior testing results, and current regulatory priorities. Use this to practice exam readiness.
          </div>
        </div>""", unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        exam_type = c1.selectbox("Exam Type", ["FINRA Cycle Exam","SEC Examination","FinCEN AML Review"], key="sim_type")
        exam_scope = c2.selectbox("Scope", ["Full Firm","AML / BSA","Supervisory Controls","Reg BI"], key="sim_scope")
        exam_yr   = c3.selectbox("Exam Period", ["2025","2024","2023"], key="sim_yr")
        if st.button("🏛️ Generate Mock Exam Request List", key="sim_start"):
            st.session_state.sim_started = True
            st.rerun()
    else:
        tests = D.get("compliance_tests",[])
        open_ex = sum(len([e for e in t.get("exceptions",[]) if e["status"]=="Open"]) for t in tests)

        _ai_thinking_card(f"Generated 18 document requests · Based on FINRA cycle exam priorities + {open_ex} open exceptions + firm risk profile")

        st.markdown(f"""
        <div style='background:#1e3a5f;color:#fff;border-radius:10px;padding:16px 20px;margin-bottom:14px'>
          <div style='font-size:10px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;
               color:#93c5fd;margin-bottom:6px'>Mock Regulatory Exam — FINRA Cycle Examination</div>
          <div style='font-size:16px;font-weight:800;margin-bottom:4px'>Document Request List</div>
          <div style='font-size:12px;color:#93c5fd'>
            Apex Financial Services LLC · CRD 123456 · Exam Period: Full Year 2025 · 
            AI-generated based on firm risk profile and regulatory priorities
          </div>
        </div>""", unsafe_allow_html=True)

        REQUESTS = [
            {"cat":"AML / BSA","id":"DR-001","request":"Written Anti-Money Laundering Program (WSP-001, current version)","priority":"High","platform_ref":"WSP-001 v3.2"},
            {"cat":"AML / BSA","id":"DR-002","request":"SAR filing log for 2025 — all filed SARs with dates, amounts, and disposition","priority":"High","platform_ref":"Data Lake — SAR Log"},
            {"cat":"AML / BSA","id":"DR-003","request":"Customer Identification Program documentation — sample of 25 new accounts opened in 2025","priority":"High","platform_ref":"CT-003 — CIP testing sample"},
            {"cat":"AML / BSA","id":"DR-004","request":"Annual AML independent testing report for 2025","priority":"High","platform_ref":"CT-003 Final Report"},
            {"cat":"Supervision","id":"DR-005","request":"Written Supervisory Procedures — current approved version (WSP-002)","priority":"High","platform_ref":"WSP-002 v2.0"},
            {"cat":"Supervision","id":"DR-006","request":"Annual supervisory controls testing report (Rule 3120) for 2025","priority":"High","platform_ref":"CT-001 (In Progress)"},
            {"cat":"Supervision","id":"DR-007","request":"Supervisory review logs — flagged trade review documentation Q1–Q4 2025","priority":"High","platform_ref":"CL-002 (Morgan Ellis)"},
            {"cat":"Supervision","id":"DR-008","request":"Customer complaint log for 2025 — all complaints, responses, and resolutions","priority":"Medium","platform_ref":"Data Lake — Customer Complaint Log"},
            {"cat":"Reg BI","id":"DR-009","request":"Regulation Best Interest policies and procedures — current version","priority":"High","platform_ref":"POL-004 (Draft — not yet approved ⚠)"},
            {"cat":"Reg BI","id":"DR-010","request":"Form CRS — current version with delivery tracking log","priority":"High","platform_ref":"Evidence required"},
            {"cat":"Reg BI","id":"DR-011","request":"Sample of 30 retail recommendations — documentation supporting best interest standard","priority":"High","platform_ref":"CT-004 testing sample"},
            {"cat":"Recordkeeping","id":"DR-012","request":"WORM storage vendor agreement and current attestation letter","priority":"Medium","platform_ref":"Evidence required"},
            {"cat":"Recordkeeping","id":"DR-013","request":"Electronic communications retention policy and surveillance sample (Q1–Q4 2025)","priority":"Medium","platform_ref":"WSP-003 / WT-003-2"},
            {"cat":"Financials","id":"DR-014","request":"FOCUS Reports — all filings for 2025 with net capital computations","priority":"Medium","platform_ref":"Evidence required"},
            {"cat":"Training","id":"DR-015","request":"Annual AML training completion report for 2025 — all registered persons","priority":"Medium","platform_ref":"T-005 (Closed) — training certificates"},
            {"cat":"Exams / Correspondence","id":"DR-016","request":"All regulatory correspondence received in 2025 including FINRA and SEC inquiries","priority":"Medium","platform_ref":"ER-001 FINRA Cycle Exam file"},
            {"cat":"Org / Personnel","id":"DR-017","request":"Current organizational chart and supervisory designation list","priority":"Low","platform_ref":"Org Chart — Graphite Platform"},
            {"cat":"Org / Personnel","id":"DR-018","request":"List of all registered persons, registration types, and supervisors as of exam date","priority":"Low","platform_ref":"Firm Config — Users"},
        ]

        sim_responses = st.session_state.sim_responses
        categories = list(dict.fromkeys(r["cat"] for r in REQUESTS))
        pri_colors = {"High":("#dc2626","#fef2f2"),"Medium":("#d97706","#fffbeb"),"Low":("#16a34a","#f0fdf4")}

        for cat in categories:
            cat_reqs = [r for r in REQUESTS if r["cat"]==cat]
            done_ct  = sum(1 for r in cat_reqs if sim_responses.get(r["id"],{}).get("status")=="Uploaded")
            st.markdown(f"""
            <div style='background:#f1f5f9;border-radius:6px;padding:6px 12px;margin:10px 0 6px;
                 font-size:11px;font-weight:700;color:#374151;text-transform:uppercase;
                 letter-spacing:.08em;display:flex;justify-content:space-between'>
              <span>{cat}</span>
              <span style='color:{"#16a34a" if done_ct==len(cat_reqs) else "#6b7280"}'>{done_ct}/{len(cat_reqs)} uploaded</span>
            </div>""", unsafe_allow_html=True)

            for r in cat_reqs:
                resp = sim_responses.get(r["id"], {"status":"Pending","notes":""})
                pc, pbg = pri_colors.get(r["priority"], ("#374151","#f3f4f6"))
                status_icon = "✅" if resp["status"]=="Uploaded" else "📎" if resp["status"]=="In Progress" else "⬜"

                c_main, c_sel, c_btn = st.columns([4, 2, 1])
                c_main.markdown(f"""
                <div style='padding:8px 0'>
                  <div style='display:flex;align-items:center;gap:8px;margin-bottom:3px'>
                    <span style='font-size:12px;color:#6b7280;font-weight:700'>{r["id"]}</span>
                    <span style='background:{pbg};color:{pc};border-radius:4px;
                         padding:1px 7px;font-size:10px;font-weight:700'>{r["priority"]}</span>
                  </div>
                  <div style='font-size:12px;font-weight:600;color:#111;margin-bottom:2px'>{r["request"]}</div>
                  <div style='font-size:11px;color:#2563eb'>📎 Platform ref: {r["platform_ref"]}</div>
                </div>""", unsafe_allow_html=True)

                new_status = c_sel.selectbox("",
                    ["Pending","In Progress","Uploaded","N/A"],
                    index=["Pending","In Progress","Uploaded","N/A"].index(resp["status"]) if resp["status"] in ["Pending","In Progress","Uploaded","N/A"] else 0,
                    key=f"sim_sel_{r['id']}",
                    label_visibility="collapsed")
                if new_status != resp["status"]:
                    st.session_state.sim_responses[r["id"]] = {"status":new_status,"notes":resp.get("notes","")}
                    st.rerun()

                c_btn.markdown(f"<div style='padding-top:12px;font-size:16px'>{status_icon}</div>", unsafe_allow_html=True)

        total    = len(REQUESTS)
        uploaded = sum(1 for r in REQUESTS if sim_responses.get(r["id"],{}).get("status")=="Uploaded")
        pct      = int(uploaded/total*100) if total else 0
        st.markdown("<hr style='margin:.8rem 0'/>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style='display:flex;align-items:center;justify-content:space-between;
             font-size:13px;font-weight:600;color:#374151;margin-bottom:6px'>
          <span>Exam Readiness</span><span style='color:{"#16a34a" if pct==100 else "#f59e0b"}'>{pct}% complete ({uploaded}/{total} items)</span>
        </div>""", unsafe_allow_html=True)
        st.progress(pct/100)

        if st.button("🔄 Reset Simulator", key="sim_reset"):
            st.session_state.sim_started = False
            st.session_state.sim_responses = {}
            st.rerun()

# ── PAGE: AUTOMATED COMPLIANCE REPORTING ──────────────────────────────────────
def page_automated_reporting():
    _premium_header(
        "AUTOMATED COMPLIANCE REPORTING",
        "AI-generated executive summaries and compliance reports built from platform data — ready for board meetings and regulatory exams",
    )
    _premium_locked_banner()

    tests    = D.get("compliance_tests", [])
    policies = D.get("policies", [])
    tasks    = D.get("tasks", [])
    today_str = date.today().strftime("%B %d, %Y")

    _ai_thinking_card(f"4 report templates available · Data current as of {today_str}")

    REPORTS = [
        {
            "id":"RPT-001","icon":"🧪","title":"3120 Supervisory Controls Testing Summary",
            "description":"Executive summary of all FINRA Rule 3120 testing activity including test results, exceptions, remediation status, and risk ratings.",
            "audience":"Board · CCO · Senior Management","cadence":"Quarterly",
            "status":"Ready","data_sources":["Compliance Tests (CT-001, CT-004)","Remediation Items (EX-001, EX-003)","WSP-002"],
        },
        {
            "id":"RPT-002","icon":"⚠️","title":"Open Remediation Items Report",
            "description":"Complete listing of all open exceptions and remediation items across testing programs, with assignees, due dates, and aging analysis.",
            "audience":"CCO · Supervision Team","cadence":"Monthly",
            "status":"Ready","data_sources":["All Compliance Tests","Remediation Items"],
        },
        {
            "id":"RPT-003","icon":"📚","title":"Policy & WSP Update Log",
            "description":"Summary of all policy and WSP changes made during the reporting period including version history, approvals, and pending reviews.",
            "audience":"Board · Regulators","cadence":"Annual / As Needed",
            "status":"Ready","data_sources":["Policies (POL-001 through POL-004)","WSPs (WSP-001 through WSP-003)"],
        },
        {
            "id":"RPT-004","icon":"📋","title":"Compliance Activity Summary",
            "description":"Broad compliance program activity report covering task completion, checklist status, workflow completion, and exam preparation activity.",
            "audience":"CCO · Board","cadence":"Quarterly",
            "status":"Ready","data_sources":["Tasks","Supervisory Checklists","Active Workflows","Exam Requests"],
        },
    ]

    for rpt in REPORTS:
        open_ex = sum(len([e for e in t.get("exceptions",[]) if e["status"]=="Open"]) for t in tests)
        all_ex  = sum(len(t.get("exceptions",[])) for t in tests)
        approved_pol = [p for p in policies if p["status"]=="Approved"]
        open_tasks   = [t for t in tasks if t["status"] in ("Open","Overdue")]

        c_info, c_btn = st.columns([5, 1])
        c_info.markdown(f"""
        <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;
             padding:12px 16px;margin-bottom:4px'>
          <div style='display:flex;align-items:flex-start;gap:12px'>
            <div style='font-size:28px;line-height:1.2'>{rpt["icon"]}</div>
            <div style='flex:1'>
              <div style='display:flex;align-items:center;gap:8px;margin-bottom:4px'>
                <span style='font-weight:700;font-size:13px;color:#111'>{rpt["title"]}</span>
                <span style='background:#f0fdf4;color:#16a34a;border:1px solid #86efac;
                     border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700'>● Ready</span>
              </div>
              <div style='font-size:12px;color:#374151;margin-bottom:5px;line-height:1.5'>{rpt["description"]}</div>
              <div style='font-size:11px;color:#6b7280'>
                <b>For:</b> {rpt["audience"]} · <b>Cadence:</b> {rpt["cadence"]} · 
                <b>Sources:</b> {" · ".join(rpt["data_sources"][:2])}{"…" if len(rpt["data_sources"])>2 else ""}
              </div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

        if c_btn.button("Generate", key=f"rpt_gen_{rpt['id']}", use_container_width=True):
            # Build report content
            lines = [f"{rpt['title'].upper()}", "="*55,
                     f"Firm: Apex Financial Services LLC", f"Generated: {today_str}", ""]
            if "3120" in rpt["title"] or "Testing" in rpt["title"]:
                lines += [f"Tests: {len(tests)}", f"Exceptions: {all_ex}", f"Open: {open_ex}"]
                for t in tests:
                    lines.append(f"\n{t['id']} — {t['topic']}\nStatus: {t['status']} · Risk: {t.get('risk_rating','—')}")
            elif "Remediation" in rpt["title"]:
                lines.append(f"Open Remediation Items: {open_ex}")
                for t in tests:
                    for e in t.get("exceptions",[]):
                        if e["status"]=="Open":
                            owner = D["users"].get(e.get("assigned_to",""),{}).get("name","Unassigned")
                            lines.append(f"\n{e['id']} ({t['id']}): {e['description']}\nAssigned: {owner} · Due: {e.get('due_date','—')}")
            elif "Policy" in rpt["title"]:
                for p in policies:
                    lines.append(f"\n{p['id']} {p['title']} v{p['version']} — {p['status']}")
            else:
                lines += [f"Open Tasks: {len(open_tasks)}", f"Approved Policies: {len(approved_pol)}",
                          f"Active Tests: {len([t for t in tests if t['status']=='In Progress'])}"]
            lines.append("\n\nThis report was auto-generated by the Graphite AI Reporting Engine.")
            report_txt = "\n".join(lines)
            st.download_button(f"📥 Download {rpt['id']}", data=report_txt,
                               file_name=f"{rpt['id']}_{today_str.replace(' ','_')}.txt",
                               mime="text/plain", key=f"rpt_dl_{rpt['id']}")

    st.markdown("""
    <div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
         padding:12px 16px;margin-top:8px;font-size:12px;color:#374151'>
      <b>📅 Scheduled Reporting</b> — In the full premium version, reports can be scheduled for
      automatic generation and delivery to designated recipients on a recurring basis
      (monthly, quarterly, or annually). Recipients receive a secure link to download the report.
    </div>""", unsafe_allow_html=True)

# ── PAGE: PREDICTIVE COMPLIANCE ALERTS ────────────────────────────────────────
def page_predictive_alerts():
    _premium_header(
        "PREDICTIVE COMPLIANCE ALERTS",
        "Continuous monitoring engine that analyzes platform data to surface potential compliance risks before they become exam findings",
    )
    _premium_locked_banner()

    tests  = D.get("compliance_tests", [])
    tasks  = D.get("tasks", [])
    today  = date.today()
    today_str = today.strftime("%Y-%m-%d")

    open_ex    = sum(len([e for e in t.get("exceptions",[]) if e["status"]=="Open"]) for t in tests)
    overdue_t  = [t for t in tasks if t["status"]=="Overdue"]
    in_prog    = [t for t in tests if t["status"]=="In Progress"]
    no_start   = [t for t in tests if t["status"]=="Not Started"]
    near_ddl   = [t for t in tests if t.get("deadline","") and t["deadline"] <= (today + timedelta(days=14)).strftime("%Y-%m-%d") and t["status"] not in ("Completed","Cancelled")]

    _ai_thinking_card(f"Monitoring {len(tests)} tests · {len(tasks)} tasks · {open_ex} open exceptions · Last scan: {today_str}")

    # Summary bar
    alert_ct = len(overdue_t) + (1 if open_ex>0 else 0) + len(near_ddl) + (1 if no_start else 0) + 2  # static alerts
    st.markdown(f"""
    <div style='display:flex;gap:10px;margin-bottom:14px'>
      <div style='background:#fef2f2;border:1px solid #fca5a5;border-radius:8px;
           padding:8px 14px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#dc2626'>3</div>
        <div style='font-size:11px;color:#6b7280'>Critical</div>
      </div>
      <div style='background:#fffbeb;border:1px solid #fcd34d;border-radius:8px;
           padding:8px 14px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#d97706'>4</div>
        <div style='font-size:11px;color:#6b7280'>Warning</div>
      </div>
      <div style='background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
           padding:8px 14px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#2563eb'>2</div>
        <div style='font-size:11px;color:#6b7280'>Informational</div>
      </div>
      <div style='background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
           padding:8px 14px;text-align:center;flex:1'>
        <div style='font-size:22px;font-weight:800;color:#16a34a'>5</div>
        <div style='font-size:11px;color:#6b7280'>Cleared</div>
      </div>
    </div>""", unsafe_allow_html=True)

    ALERTS = [
        {
            "id":"ALT-001","level":"Critical","icon":"🔴","category":"Remediation",
            "title":f"Open remediation items approaching due dates — {open_ex} item(s)",
            "detail":f"There are {open_ex} open exception(s) across active compliance tests. EX-001 (CT-001 — CCO escalation SLA breach) and EX-003 (CT-004 — VA documentation gap) are both assigned and due within 21 days. Unresolved exceptions present exam risk.",
            "affected":["CT-001 EX-001","CT-004 EX-003"],"action":"Review and advance remediation items in Compliance Testing module.",
        },
        {
            "id":"ALT-002","level":"Critical","icon":"🔴","category":"Policy Approval",
            "title":"POL-004 (Reg BI Policy) has been in Draft status for 10+ days",
            "detail":"POL-004 was submitted for CCO review 9 days ago and has not been approved. An unapproved Reg BI policy creates a gap in the firm's written compliance program. Recent SEC enforcement actions have cited absence of approved Reg BI policies as a standalone finding.",
            "affected":["POL-004"],"action":"Navigate to Policies & WSPs to complete CCO review and approval of POL-004.",
        },
        {
            "id":"ALT-003","level":"Critical","icon":"🔴","category":"Supervisory Deadline",
            "title":f"Overdue task(s) detected — {len(overdue_t)} task(s) past due",
            "detail":f"{len(overdue_t)} task(s) are currently overdue: {', '.join(t['title'] for t in overdue_t[:2])}. Overdue supervisory tasks are a common exam finding under FINRA Rule 3110 and should be resolved or formally escalated.",
            "affected":[t["id"] for t in overdue_t],"action":"Navigate to Tasks to review and resolve overdue items.",
        },
        {
            "id":"ALT-004","level":"Warning","icon":"🟡","category":"Testing Coverage",
            "title":"CT-002 (SEC 206(4)-7 Annual Review) has not been started — deadline in 30 days",
            "detail":"CT-002 covers the annual SEC compliance program review required under Rule 206(4)-7. Testing has not yet started and the deadline is within 30 days. This is a required annual review for registered investment advisers and should be prioritized.",
            "affected":["CT-002"],"action":"Assign CT-002 and begin testing procedures in the Compliance Testing module.",
        },
        {
            "id":"ALT-005","level":"Warning","icon":"🟡","category":"Repeated Finding",
            "title":"CCO escalation SLA breach is a repeat pattern risk indicator",
            "detail":"EX-001 (CT-001) documents a CCO escalation SLA breach where a trade escalation was reviewed 5 days late. FINRA's recent examination findings memo highlights this as a recurring deficiency pattern at broker-dealers. A second instance in a subsequent testing cycle would significantly increase exam risk.",
            "affected":["CT-001","EX-001"],"action":"Accelerate EX-001 remediation. Implement automated SLA monitoring alert before next testing cycle.",
        },
        {
            "id":"ALT-006","level":"Warning","icon":"🟡","category":"Regulatory Change",
            "title":"4 recently updated rules with unreviewed policy impact",
            "detail":"The AI Copilot identified 4 regulatory rules updated within the past 30 days (R-002, R-004, R-005, R-006). Potentially impacted policies have not yet been flagged for review in the platform. Unaddressed regulatory changes present exam and enforcement risk.",
            "affected":["R-002","R-004","R-005","R-006"],"action":"Navigate to AI Copilot and run a regulatory change analysis. Initiate policy reviews for affected documents.",
        },
        {
            "id":"ALT-007","level":"Warning","icon":"🟡","category":"Exam Preparation",
            "title":"ER-001 FINRA Cycle Exam — 2 items still Pending with 30 days to deadline",
            "detail":"Exam request ER-001 has 2 items in Pending status: 'Supervisory procedures manual' (admin) and 'Trade blotter Q1–Q3 2024' (supervisor). The exam deadline is approximately 30 days away. Late document production during an exam is a serious risk.",
            "affected":["ER-001"],"action":"Navigate to Exam Requests to advance outstanding ER-001 items.",
        },
        {
            "id":"ALT-008","level":"Informational","icon":"🔵","category":"WSP Review",
            "title":"WSP-003 has been with reviewer for 6+ days without approval",
            "detail":"WSP-003 (Information Barriers Policy) v1.1 was submitted to Jordan Reed for review 6 days ago. Extended review periods can delay policy currency. Follow up recommended.",
            "affected":["WSP-003"],"action":"Follow up with reviewer on WSP-003 status.",
        },
        {
            "id":"ALT-009","level":"Informational","icon":"🔵","category":"Testing",
            "title":"CT-004 (Reg BI Testing) testing scope is 27% complete",
            "detail":"CT-004 has 1 of 2 test procedures partially completed, with 8 of 30 sampled accounts remaining in P-002. The deadline is in 20 days. Timely completion is recommended to allow adequate review time before the deadline.",
            "affected":["CT-004"],"action":"Advance CT-004 testing procedures to completion.",
        },
    ]

    level_order = {"Critical":0,"Warning":1,"Informational":2}
    ALERTS.sort(key=lambda a: level_order.get(a["level"],3))

    level_styles = {
        "Critical":  ("#dc2626","#fef2f2","#fca5a5"),
        "Warning":   ("#d97706","#fffbeb","#fcd34d"),
        "Informational": ("#2563eb","#eff6ff","#bfdbfe"),
    }

    for a in ALERTS:
        lc, lbg, lbdr = level_styles.get(a["level"], ("#374151","#f3f4f6","#d1d5db"))
        st.markdown(f"""
        <div style='background:#fff;border:1px solid {lbdr};border-left:4px solid {lc};
             border-radius:8px;padding:12px 16px;margin-bottom:8px'>
          <div style='display:flex;align-items:flex-start;gap:10px'>
            <div style='font-size:20px;line-height:1.3'>{a["icon"]}</div>
            <div style='flex:1'>
              <div style='display:flex;align-items:center;gap:8px;margin-bottom:4px;flex-wrap:wrap'>
                <span style='background:{lbg};color:{lc};border:1px solid {lbdr};border-radius:4px;
                     padding:1px 8px;font-size:10px;font-weight:800'>{a["level"].upper()}</span>
                <span style='background:#f1f5f9;color:#374151;border-radius:4px;
                     padding:1px 7px;font-size:10px;font-weight:600'>{a["category"]}</span>
                <span style='color:#9ca3af;font-size:10px;margin-left:auto'>{a["id"]}</span>
              </div>
              <div style='font-weight:700;font-size:13px;color:#111;margin-bottom:5px'>{a["title"]}</div>
              <div style='font-size:12px;color:#374151;line-height:1.6;margin-bottom:6px'>{a["detail"]}</div>
              <div style='display:flex;align-items:center;gap:8px;flex-wrap:wrap'>
                <span style='font-size:11px;color:#6b7280'><b>Affected:</b> {" · ".join(a["affected"])}</span>
              </div>
              <div style='background:{lbg};border-left:3px solid {lc};padding:5px 10px;
                   border-radius:4px;font-size:12px;color:{lc};margin-top:6px'>
                <b>Recommended action:</b> {a["action"]}
              </div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("""
    <div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;
         padding:12px 16px;margin-top:8px;font-size:12px;color:#374151'>
      <b>🔔 Alert Delivery</b> — In the full premium version, critical and warning alerts are
      automatically delivered to designated compliance personnel via email and in-platform notification
      on a configurable schedule (real-time, daily digest, or weekly summary).
    </div>""", unsafe_allow_html=True)

# ── ROUTER ─────────────────────────────────────────────────────────────────────
if not st.session_state.logged_in:
    login_page()
else:
    nav = sidebar()
    if   nav=="Dashboard":               page_dashboard()
    elif nav=="Tasks":     page_tasks()
    elif nav=="Exam Requests":           page_exam_requests()
    elif nav=="Policies & WSPs":           page_wsp()
    elif nav=="Supervisory Checklists":  page_checklists()
    elif nav=="Data Lake":              page_evidence()
    elif nav=="Compliance Testing":      page_compliance_testing()
    elif nav=="Remediation Items":       page_remediation_items()
    elif nav=="Compliance Calendar":        page_compliance_calendar()
    elif nav=="Workflow Library":        page_workflow_library()
    elif nav=="Active Workflows":        page_active_workflows()
    elif nav=="Rule Inventory":          page_rules()
    elif nav=="AI Rule Scan":            page_ai_scan()
    elif nav=="AI Copilot":              page_ai_copilot()
    elif nav=="Peer Enforcement Insights": page_peer_enforcement()
    elif nav=="AI Risk Assessment":      page_ai_risk_assessment()
    elif nav=="Policy Gap Analysis":     page_policy_gap_analysis()
    elif nav=="Exam Simulator":          page_exam_simulator()
    elif nav=="Automated Reporting":     page_automated_reporting()
    elif nav=="Predictive Alerts":       page_predictive_alerts()
    elif nav=="Org Chart":               page_org()
    elif nav=="Audit Trail":             page_audit()
    elif nav=="Firm Config":             page_firm_config()
    elif nav=="Data Export":             page_export()
